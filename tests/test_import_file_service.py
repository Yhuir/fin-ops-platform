from __future__ import annotations

import sys
import unittest
from io import BytesIO
from pathlib import Path

from fin_ops_platform.domain.enums import BatchType
from fin_ops_platform.services.import_file_service import (
    BankStatementMappingRequired,
    FileImportService,
    UploadedImportFile,
    aggregate_invoice_line_rows,
    is_company_identity,
    parse_bank_statement_rows,
)
from fin_ops_platform.services.imports import ImportNormalizationService
from openpyxl import Workbook, load_workbook

TESTS_DIR = Path(__file__).resolve().parent
if str(TESTS_DIR) not in sys.path:
    sys.path.insert(0, str(TESTS_DIR))

from mock_import_files import (
    BOCOM_JAN,
    CEB_JAN,
    INVOICE_JAN,
    PINGAN_JAN,
    icbc_history_file,
    invoice_export_file,
    invoice_summary_file,
)


def repeat_last_xlsx_row(content: bytes, *, total_data_rows: int) -> bytes:
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    last_row = [cell.value for cell in sheet[sheet.max_row]]
    for _ in range(total_data_rows - 1):
        sheet.append(last_row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def append_xlsx_note(content: bytes, note: str) -> bytes:
    workbook = load_workbook(BytesIO(content))
    workbook.properties.subject = note
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def invoice_export_second_sheet_file() -> bytes:
    source = load_workbook(BytesIO(INVOICE_JAN.content))
    workbook = Workbook()
    summary = workbook.active
    summary.title = "导出摘要"
    summary["A1"] = "此工作表不是发票明细。"
    data_sheet = workbook.create_sheet("合并数据")
    for row in source.active.iter_rows(values_only=True):
        data_sheet.append(list(row))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class FakeImportIdStore:
    def __init__(self) -> None:
        self._existing_session_ids = {"import_session_0001", "import_session_0002"}
        self._existing_file_ids = {"import_file_0001", "import_file_0002"}
        self._stored_refs: list[str] = []
        self._stored_files: dict[str, bytes] = {}

    def import_session_exists(self, session_id: str) -> bool:
        return session_id in self._existing_session_ids

    def import_file_exists(self, file_id: str) -> bool:
        return file_id in self._existing_file_ids

    def store_import_file(
        self,
        *,
        session_id: str,
        file_id: str,
        file_name: str,
        content: bytes,
        imported_by: str | None = None,
    ) -> str:
        self._existing_session_ids.add(session_id)
        self._existing_file_ids.add(file_id)
        ref = f"stored://{session_id}/{file_id}/{file_name}"
        self._stored_refs.append(ref)
        self._stored_files[ref] = content
        return ref

    def read_import_file(self, stored_file_path: str) -> bytes:
        return self._stored_files[stored_file_path]


class FakeImportEntityRegistry:
    def __init__(self) -> None:
        self._existing_batch_ids = {"batch_import_0001"}
        self._existing_invoice_ids = {"inv_imported_0001"}
        self._existing_transaction_ids = {"txn_imported_0001"}

    def import_batch_exists(self, batch_id: str) -> bool:
        return batch_id in self._existing_batch_ids

    def invoice_exists(self, invoice_id: str) -> bool:
        return invoice_id in self._existing_invoice_ids

    def transaction_exists(self, transaction_id: str) -> bool:
        return transaction_id in self._existing_transaction_ids


class FailingSubmittedEtcIdentityRepository:
    def find_invoices_by_identity_keys(self, *, canonical_keys: list[str], suspected_keys: list[str]) -> list[object]:
        return []

    def find_submitted_etc_invoice_by_identity(self, **kwargs: object) -> object | None:
        raise RuntimeError("submitted etc lookup failed")


class ImportFileServiceTests(unittest.TestCase):
    def test_ccb_current_export_header_uses_metadata_account_and_unit_aliases(self) -> None:
        parsed = parse_bank_statement_rows(
            [
                ["中国建设银行"],
                ["账　　号", "53001905038050548106", "账户名称", "云南溯源科技有限公司"],
                [
                    "交易时间",
                    "借方发生额/元(支取)",
                    "贷方发生额/元(收入)",
                    "余额",
                    "币种",
                    "对方户名",
                    "对方账号",
                    "对方开户机构",
                    "记账日期",
                    "摘要",
                    "备注",
                    "账户明细编号-交易流水号",
                ],
                [
                    "20260801 15:24:03",
                    "2100.00",
                    "0",
                    "131301.08",
                    "人民币元",
                    "张丽芬",
                    "62166022700000810872",
                    "中国银行昭通支行",
                    "20260801",
                    "电子转账",
                    "住宿费",
                    "13835-530905038F4BFMPXRK8",
                ],
            ]
        )

        self.assertEqual(parsed.template_code, "bank_statement")
        self.assertEqual(parsed.detected_bank_name, "建设银行")
        self.assertEqual(parsed.rows[0]["account_no"], "53001905038050548106")
        self.assertEqual(parsed.rows[0]["debit_amount"], "2100.00")
        self.assertEqual(parsed.rows[0]["account_detail_no"], "13835-530905038F4BFMPXRK8")

    def test_manual_mapping_is_reused_for_same_header_signature(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["交易日", "付出金额", "收到数额", "对方名称", "流水标识"])
        sheet.append(["2026-08-01", "0", "88.50", "测试客户", "BANK-001"])
        buffer = BytesIO()
        workbook.save(buffer)
        content = buffer.getvalue()
        file_store = FakeImportIdStore()
        service = FileImportService(ImportNormalizationService(), file_store=file_store)

        first = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name="unknown-header.xlsx",
                    content=content,
                    batch_type_override="bank_transaction",
                    selected_bank_mapping_id="bank_mapping_test_0093",
                    selected_bank_name="平安银行",
                    selected_bank_last4="0093",
                )
            ],
        )
        first_file = first.files[0]
        self.assertEqual(first_file.status, "unrecognized_template")
        self.assertTrue(first_file.header_signature)
        with self.assertRaises(BankStatementMappingRequired):
            parse_bank_statement_rows([
                ["交易日", "付出金额", "收到数额", "对方名称", "流水标识"],
                ["2026-08-01", "0", "88.50", "测试客户", "BANK-001"],
            ])

        retried = service.retry_session_files(
            session_id=first.id,
            selected_file_ids=[first_file.id],
            overrides={first_file.id: {"field_mapping": {"credit_amount": "2", "bank_serial_no": "4"}}},
        )
        self.assertEqual(retried.files[0].status, "preview_ready")
        self.assertEqual(retried.files[0].mapping_source, "manual")

        second = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name="same-header.xlsx",
                    content=content,
                    batch_type_override="bank_transaction",
                    selected_bank_mapping_id="bank_mapping_test_0093",
                    selected_bank_name="平安银行",
                    selected_bank_last4="0093",
                )
            ],
        )
        self.assertEqual(second.files[0].status, "preview_ready")
        self.assertEqual(second.files[0].mapping_source, "saved")
        self.assertEqual(second.files[0].normalized_rows[0]["amount"], "88.50")

    def test_invoice_export_aggregates_distinct_lines_and_discount_into_one_invoice(self) -> None:
        header = {
            "digital_invoice_no": "26117000001052654674",
            "invoice_no": "1052654674",
            "seller_tax_no": "915300000000000001",
            "buyer_tax_no": "915300007194052520",
            "seller_name": "供应商",
            "buyer_name": "云南溯源科技有限公司",
            "invoice_date": "2026-07-01",
            "invoice_status_from_source": "正常",
        }

        rows = aggregate_invoice_line_rows(
            [
                {**header, "taxable_item_name": "服务", "amount": "39.58", "tax_amount": "5.15", "total_with_tax": "44.73", "tax_rate": "13%"},
                {**header, "taxable_item_name": "折扣", "amount": "-1.77", "tax_amount": "-0.23", "total_with_tax": "-2.00", "tax_rate": "13%"},
            ]
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["amount"], "37.81")
        self.assertEqual(rows[0]["tax_amount"], "4.92")
        self.assertEqual(rows[0]["total_with_tax"], "42.73")
        self.assertEqual(rows[0]["source_line_count"], 2)

    def test_invoice_export_keeps_identical_repeated_rows_for_duplicate_audit(self) -> None:
        row = {
            "digital_invoice_no": "26117000001052654674",
            "taxable_item_name": "服务",
            "amount": "39.58",
            "tax_amount": "5.15",
            "total_with_tax": "44.73",
            "tax_rate": "13%",
        }

        self.assertEqual(aggregate_invoice_line_rows([row, dict(row)]), [row, row])

    def test_import_rows_use_batch_scoped_ids_without_legacy_counter_state(self) -> None:
        service = ImportNormalizationService.from_snapshot({"batch_counter": 54, "row_counter": 999})

        preview = service.preview_import(
            batch_type=BatchType.INPUT_INVOICE,
            source_name="invoice.xlsx",
            imported_by="tester",
            rows=[
                {
                    "digital_invoice_no": "26117000001052654674",
                    "counterparty_name": "供应商",
                    "seller_name": "供应商",
                    "invoice_date": "2026-07-01",
                    "amount": "42.73",
                    "tax_amount": "4.92",
                    "total_with_tax": "42.73",
                }
            ],
        )

        self.assertRegex(
            preview.row_results[0].id,
            r"^batch_row:batch_import_[0-9a-f]{32}:00001$",
        )
        self.assertNotIn("row_counter", service.snapshot())

    def test_confirmed_session_persistence_payload_excludes_unrelated_fact_domains_and_sessions(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        invoice_session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=INVOICE_JAN.name, content=INVOICE_JAN.content)],
        )
        service.confirm_session(
            session_id=invoice_session.id,
            selected_file_ids=[invoice_session.files[0].id],
        )
        bank_session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=PINGAN_JAN.name, content=PINGAN_JAN.content)],
        )
        service.confirm_session(
            session_id=bank_session.id,
            selected_file_ids=[bank_session.files[0].id],
        )

        payload = service.confirmed_session_persistence_payload(
            session_id=bank_session.id,
            selected_file_ids=[bank_session.files[0].id],
        )

        self.assertEqual(set(payload), {"imports", "file_imports"})
        self.assertEqual(set(payload["file_imports"]["sessions"]), {bank_session.id})
        self.assertEqual(set(payload["imports"]["batches"]), {bank_session.files[0].batch_id})
        self.assertEqual(payload["imports"]["invoices"], [])
        self.assertTrue(payload["imports"]["transactions"])
        self.assertTrue(
            all(transaction.source_batch_id == bank_session.files[0].batch_id for transaction in payload["imports"]["transactions"])
        )

    def test_preview_session_persistence_payload_excludes_unrelated_sessions_and_canonical_facts(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        invoice_session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=INVOICE_JAN.name, content=INVOICE_JAN.content)],
        )
        bank_session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=PINGAN_JAN.name, content=PINGAN_JAN.content)],
        )

        payload = service.preview_session_persistence_payload(bank_session.id)

        self.assertEqual(set(payload["file_imports"]["sessions"]), {bank_session.id})
        self.assertEqual(set(payload["imports"]["batches"]), {bank_session.files[0].preview_batch_id})
        self.assertNotIn(invoice_session.files[0].preview_batch_id, payload["imports"]["batches"])
        self.assertNotIn("invoices", payload["imports"])
        self.assertNotIn("transactions", payload["imports"])
        self.assertGreaterEqual(payload["imports"]["batch_counter"], 2)
        self.assertEqual(payload["file_imports"]["session_counter"], 2)
        self.assertEqual(payload["file_imports"]["file_counter"], 2)

    def test_company_identity_name_keywords_use_yunnan_and_generic_suyuan_names(self) -> None:
        self.assertTrue(is_company_identity(None, "云南溯源科技有限公司"))
        self.assertTrue(is_company_identity(None, "溯源科技有限公司"))
        self.assertTrue(is_company_identity("91330106589876543T", "无关公司名称"))

    def test_preview_marks_corrupt_excel_as_file_level_error_instead_of_raising(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name="损坏发票.xlsx", content=b"not-a-real-xlsx")],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "unrecognized_template")
        self.assertEqual(preview_file.row_count, 0)
        self.assertIn("不是有效的 Excel 工作簿", preview_file.message)

    def test_preview_accepts_invoice_summary_header_aliases(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = invoice_summary_file()

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=upload.name,
                    content=upload.content,
                    template_code_override="invoice_export",
                    batch_type_override="input_invoice",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "invoice_export")
        self.assertEqual(preview_file.batch_type.value, "input_invoice")
        self.assertEqual(preview_file.row_count, 1)
        self.assertEqual(preview_file.success_count, 1)
        normalized = preview_file.normalized_rows[0]
        self.assertEqual(normalized["digital_invoice_no"], "26312000002781821596")
        self.assertEqual(normalized["buyer_name"], "云南溯源科技有限公司")
        self.assertEqual(normalized["buyer_tax_no"], "915300007194052520")
        self.assertEqual(normalized["seller_name"], "阿法拉伐（上海）技术有限公司")
        self.assertEqual(normalized["seller_tax_no"], "91310000607371000G")
        self.assertEqual(normalized["counterparty_name"], "阿法拉伐（上海）技术有限公司")
        self.assertEqual(normalized["taxable_item_name"], "*通用设备*垫片板式换热器")
        self.assertEqual(normalized["invoice_kind"], "数电票(专用发票)")
        self.assertEqual(normalized["total_with_tax"], "14384.00")

    def test_preview_detects_invoice_summary_without_template_override(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = invoice_summary_file()

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=upload.name, content=upload.content)],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "invoice_export")
        self.assertEqual(preview_file.batch_type.value, "input_invoice")

    def test_preview_detects_invoice_data_when_summary_sheet_is_first(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name="进项发票合并.xlsx",
                    content=invoice_export_second_sheet_file(),
                    batch_type_override="input_invoice",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "invoice_export")
        self.assertEqual(preview_file.row_count, 1)
        self.assertEqual(preview_file.normalized_rows[0]["digital_invoice_no"], "25502000000145098656")

    def test_preview_uses_process_independent_opaque_ids(self) -> None:
        file_store = FakeImportIdStore()
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service, file_store=file_store)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=INVOICE_JAN.name, content=INVOICE_JAN.content)],
        )

        self.assertRegex(session.id, r"^import_session_[0-9a-f]{32}$")
        self.assertRegex(session.files[0].id, r"^import_file_[0-9a-f]{32}$")
        self.assertRegex(session.files[0].preview_batch_id or "", r"^batch_import_[0-9a-f]{32}$")
        self.assertEqual(session.files[0].status, "preview_ready")
        self.assertTrue(session.files[0].stored_file_path)

        second_import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        second_service = FileImportService(
            second_import_service,
            file_store=FakeImportIdStore(),
        )
        second = second_service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=INVOICE_JAN.name, content=INVOICE_JAN.content)],
        )
        self.assertNotEqual(second.id, session.id)
        self.assertNotEqual(second.files[0].id, session.files[0].id)
        self.assertNotEqual(second.files[0].preview_batch_id, session.files[0].preview_batch_id)
        service.confirm_session(session_id=session.id, selected_file_ids=[session.files[0].id])
        second_service.confirm_session(session_id=second.id, selected_file_ids=[second.files[0].id])
        self.assertNotEqual(
            second_import_service.list_invoices()[0].id,
            import_service.list_invoices()[0].id,
        )

    def test_preview_persists_selected_bank_mapping_and_marks_conflict_against_detected_account(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=PINGAN_JAN.name,
                    content=PINGAN_JAN.content,
                    selected_bank_mapping_id="bank_mapping_pingan_override",
                    selected_bank_name="建设银行",
                    selected_bank_last4="8826",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.selected_bank_mapping_id, "bank_mapping_pingan_override")
        self.assertEqual(preview_file.selected_bank_name, "建设银行")
        self.assertEqual(preview_file.selected_bank_last4, "8826")
        self.assertTrue(preview_file.bank_selection_conflict)
        self.assertEqual(preview_file.detected_last4, "0093")
        self.assertIn("建设银行", preview_file.conflict_message)
        self.assertIn("0093", preview_file.conflict_message)

    def test_preview_does_not_mark_bank_name_alias_as_conflict_when_last4_matches(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=CEB_JAN.name,
                    content=CEB_JAN.content,
                    selected_bank_mapping_id="bank_mapping_ceb_8826",
                    selected_bank_name="光大",
                    selected_bank_last4="8826",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.detected_bank_name, "光大银行")
        self.assertEqual(preview_file.detected_last4, "8826")
        self.assertFalse(preview_file.bank_selection_conflict)
        self.assertIsNone(preview_file.conflict_message)

    def test_preview_does_not_mark_bank_short_name_as_conflict_when_last4_matches(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=CEB_JAN.name,
                    content=CEB_JAN.content,
                    selected_bank_mapping_id="bank_mapping_ceb_8826",
                    selected_bank_name="中国光大银行股份有限公司",
                    selected_bank_short_name="光大",
                    selected_bank_last4="8826",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.selected_bank_short_name, "光大")
        self.assertEqual(preview_file.detected_bank_name, "光大银行")
        self.assertEqual(preview_file.detected_last4, "8826")
        self.assertFalse(preview_file.bank_selection_conflict)
        self.assertIsNone(preview_file.conflict_message)

    def test_preview_does_not_mark_bank_legal_name_as_conflict_when_last4_matches(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=CEB_JAN.name,
                    content=CEB_JAN.content,
                    selected_bank_mapping_id="bank_mapping_ceb_8826",
                    selected_bank_name="中国光大银行股份有限公司",
                    selected_bank_last4="8826",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.detected_bank_name, "光大银行")
        self.assertEqual(preview_file.detected_last4, "8826")
        self.assertFalse(preview_file.bank_selection_conflict)
        self.assertIsNone(preview_file.conflict_message)

    def test_preview_does_not_detect_icbc_last4_from_filename(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = icbc_history_file(name="historydetail1410.xlsx")

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=upload.name,
                    content=upload.content,
                    selected_bank_mapping_id="bank_mapping_icbc_6386",
                    selected_bank_name="工商银行",
                    selected_bank_short_name="工行",
                    selected_bank_last4="6386",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "bank_statement")
        self.assertEqual(preview_file.detected_bank_name, "工商银行")
        self.assertIsNone(preview_file.detected_last4)
        self.assertFalse(preview_file.bank_selection_conflict)
        self.assertIsNone(preview_file.conflict_message)

    def test_preview_uses_selected_bank_account_when_icbc_file_has_no_account(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = icbc_history_file(name="historydetail1410.xlsx")

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=upload.name,
                    content=upload.content,
                    selected_bank_mapping_id="bank_mapping_icbc_6386",
                    selected_bank_name="工商银行",
                    selected_bank_short_name="工行",
                    selected_bank_last4="6386",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.success_count, 1)
        self.assertEqual(preview_file.error_count, 0)
        self.assertIsNone(preview_file.detected_last4)
        self.assertEqual(preview_file.normalized_rows[0]["account_no"], "bank_mapping_icbc_6386")
        self.assertEqual(preview_file.normalized_rows[0]["imported_bank_last4"], "6386")

    def test_preview_files_audit_counts_cross_file_invoice_duplicates(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        jan_upload = invoice_export_file("jan.xlsx")
        feb_upload = invoice_export_file("feb.xlsx")
        feb_upload = type(feb_upload)(feb_upload.name, append_xlsx_note(feb_upload.content, "第二次导出"))

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(file_name=jan_upload.name, content=jan_upload.content),
                UploadedImportFile(file_name=feb_upload.name, content=feb_upload.content),
            ],
        )

        self.assertEqual(session.audit.original_count, 2)
        self.assertEqual(session.audit.unique_count, 1)
        self.assertEqual(session.audit.duplicate_across_files_count, 1)
        self.assertEqual(session.audit.duplicate_count, 1)
        self.assertEqual(session.audit.importable_count, 1)
        self.assertEqual(len(session.duplicate_groups), 1)
        self.assertEqual(session.duplicate_groups[0].duplicate_type, "duplicate_across_files")
        self.assertEqual(session.files[0].audit.importable_count, 1)
        self.assertEqual(session.files[1].audit.duplicate_across_files_count, 1)

    def test_preview_bounds_large_invoice_duplicate_group_to_one_confirmable_row(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = invoice_export_file("large-invoice-duplicates.xlsx")
        content = repeat_last_xlsx_row(upload.content, total_data_rows=240)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=upload.name, content=content)],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.row_count, 240)
        self.assertEqual(preview_file.success_count, 240)
        self.assertEqual(preview_file.error_count, 0)
        self.assertEqual(session.audit.original_count, 240)
        self.assertEqual(session.audit.unique_count, 1)
        self.assertEqual(session.audit.duplicate_in_file_count, 239)
        self.assertEqual(session.audit.duplicate_across_files_count, 0)
        self.assertEqual(session.audit.importable_count, 1)
        self.assertEqual(session.audit.confirmable_count, 1)
        self.assertEqual(session.audit.skipped_count, 239)
        self.assertEqual(preview_file.audit.duplicate_in_file_count, 239)
        self.assertEqual(len(session.duplicate_groups), 1)
        duplicate_group = session.duplicate_groups[0]
        self.assertEqual(duplicate_group.record_type, "invoice")
        self.assertEqual(duplicate_group.duplicate_type, "duplicate_in_file")
        self.assertEqual(len(duplicate_group.rows), 240)
        self.assertEqual(duplicate_group.rows[-1]["row_no"], 240)

        review = service.review_rows(
            session_id=session.id,
            kind="duplicates",
            offset=20,
            limit=25,
        )
        self.assertEqual(review["total"], 240)
        self.assertEqual(len(review["rows"]), 25)
        self.assertEqual(review["offset"], 20)
        self.assertTrue(review["has_more"])
        self.assertEqual(review["rows"][0]["record_type"], "invoice")
        self.assertTrue(review["rows"][0]["invoice_no"])
        self.assertTrue(review["rows"][0]["invoice_date"])
        self.assertTrue(review["rows"][0]["seller_name"])
        self.assertTrue(review["rows"][0]["buyer_name"])
        self.assertTrue(review["rows"][0]["total_with_tax"])

    def test_preview_files_audit_counts_cross_file_bank_transaction_identity_duplicates(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        jan_upload = icbc_history_file(name="historydetail-jan.xlsx")
        feb_upload = icbc_history_file(name="historydetail-feb.xlsx")
        feb_upload = type(feb_upload)(feb_upload.name, append_xlsx_note(feb_upload.content, "第二次导出"))

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=jan_upload.name,
                    content=jan_upload.content,
                    selected_bank_mapping_id="bank_mapping_icbc_6386",
                    selected_bank_name="工商银行",
                    selected_bank_short_name="工行",
                    selected_bank_last4="6386",
                ),
                UploadedImportFile(
                    file_name=feb_upload.name,
                    content=feb_upload.content,
                    selected_bank_mapping_id="bank_mapping_icbc_6386",
                    selected_bank_name="工商银行",
                    selected_bank_short_name="工行",
                    selected_bank_last4="6386",
                ),
            ],
        )

        self.assertEqual(session.audit.original_count, 2)
        self.assertEqual(session.audit.unique_count, 1)
        self.assertEqual(session.audit.duplicate_across_files_count, 1)
        self.assertEqual(session.audit.duplicate_count, 1)
        self.assertEqual(session.audit.importable_count, 0)
        self.assertEqual(session.audit.suspected_duplicate_count, 2)
        self.assertEqual(session.audit.skipped_count, 2)
        self.assertEqual(len(session.duplicate_groups), 1)
        self.assertEqual(session.duplicate_groups[0].record_type, "bank_transaction")
        self.assertEqual(session.duplicate_groups[0].duplicate_type, "duplicate_across_files")
        duplicate_row = session.duplicate_groups[0].rows[1]
        self.assertEqual(duplicate_row["file_name"], "historydetail-feb.xlsx")
        self.assertEqual(duplicate_row["identity_kind"], "suspected")
        self.assertEqual(duplicate_row["decision"], "created")
        self.assertEqual(duplicate_row["account_no"], "bank_mapping_icbc_6386")
        self.assertEqual(duplicate_row["trade_time"], "2026-01-03 09:12:00")
        self.assertEqual(duplicate_row["direction"], "outflow")
        self.assertEqual(duplicate_row["amount"], "6180.00")
        self.assertEqual(duplicate_row["counterparty_name"], "重庆高新技术产业开发区国家税务局")
        self.assertEqual(session.files[0].audit.suspected_duplicate_count, 1)
        self.assertEqual(session.files[1].audit.duplicate_across_files_count, 1)

    def test_preview_bounds_large_bank_duplicate_group_to_one_confirmable_row(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = icbc_history_file(name="large-historydetail.xlsx", account_no="6222020200006386")
        content = repeat_last_xlsx_row(upload.content, total_data_rows=240)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=upload.name,
                    content=content,
                    selected_bank_mapping_id="bank_mapping_icbc_6386",
                    selected_bank_name="工商银行",
                    selected_bank_short_name="工行",
                    selected_bank_last4="6386",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.row_count, 240)
        self.assertEqual(preview_file.success_count, 240)
        self.assertEqual(preview_file.error_count, 0)
        self.assertEqual(session.audit.original_count, 240)
        self.assertEqual(session.audit.unique_count, 1)
        self.assertEqual(session.audit.duplicate_in_file_count, 239)
        self.assertEqual(session.audit.duplicate_across_files_count, 0)
        self.assertEqual(session.audit.importable_count, 0)
        self.assertEqual(session.audit.suspected_duplicate_count, 240)
        self.assertEqual(session.audit.confirmable_count, 0)
        self.assertEqual(session.audit.skipped_count, 240)
        self.assertEqual(preview_file.audit.duplicate_in_file_count, 239)
        self.assertEqual(len(session.duplicate_groups), 1)
        duplicate_group = session.duplicate_groups[0]
        self.assertEqual(duplicate_group.record_type, "bank_transaction")
        self.assertEqual(duplicate_group.duplicate_type, "duplicate_in_file")
        self.assertEqual(len(duplicate_group.rows), 240)
        self.assertEqual(duplicate_group.rows[-1]["row_no"], 240)
        self.assertEqual(duplicate_group.rows[-1]["account_no"], "6222020200006386")
        self.assertEqual(duplicate_group.rows[-1]["direction"], "outflow")
        self.assertEqual(duplicate_group.rows[-1]["amount"], "6180.00")

    def test_preview_blocks_identical_file_content_before_row_parsing(self) -> None:
        service = FileImportService(ImportNormalizationService())
        first_upload = invoice_export_file("first.xlsx")
        second_upload = type(first_upload)("renamed.xlsx", first_upload.content)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(file_name=first_upload.name, content=first_upload.content),
                UploadedImportFile(file_name=second_upload.name, content=second_upload.content),
            ],
        )

        self.assertEqual(session.files[0].status, "preview_ready")
        self.assertEqual(session.files[1].status, "duplicate_file")
        self.assertEqual(session.files[1].duplicate_file_name, "first.xlsx")
        self.assertEqual(session.status, "preview_ready_with_errors")

    def test_preview_blocks_file_content_already_confirmed_under_another_name(self) -> None:
        service = FileImportService(ImportNormalizationService())
        upload = invoice_export_file("first.xlsx")
        first = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=upload.name, content=upload.content)],
        )
        service.confirm_session(session_id=first.id, selected_file_ids=[first.files[0].id])

        second = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name="renamed.xlsx", content=upload.content)],
        )

        self.assertEqual(second.files[0].status, "duplicate_file")
        self.assertEqual(second.files[0].duplicate_file_name, "first.xlsx")
        self.assertEqual(second.status, "preview_ready_with_errors")

    def test_bank_source_control_mismatch_blocks_preview(self) -> None:
        rows = [
            ["交易时间", "借方发生额", "贷方发生额", "对方户名"],
            ["2026-08-01 10:00:00", "10.00", "", "供应商"],
            ["借方交易笔数", "2", "借方交易金额", "20.00", "贷方交易笔数", "0", "贷方交易金额", "0.00"],
        ]
        parsed = parse_bank_statement_rows(rows)

        self.assertEqual(parsed.source_control.status, "mismatch")
        self.assertEqual(set(parsed.source_control.mismatch_fields), {"row_count", "debit_total"})

    def test_confirm_session_rejects_stale_preview_when_existing_records_change(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = invoice_export_file("jan.xlsx")
        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=upload.name, content=upload.content)],
        )
        competing_preview = import_service.preview_import(
            batch_type=session.files[0].batch_type,
            source_name="competing.json",
            imported_by="user_finance_02",
            rows=[session.files[0].row_results[0].raw_payload],
        )
        import_service.confirm_import(competing_preview.id)

        with self.assertRaisesRegex(ValueError, "preview_stale"):
            service.confirm_session(session_id=session.id, selected_file_ids=[session.files[0].id])

    def test_replay_confirmed_session_files_reuses_verified_source_and_creates_new_audit_session(self) -> None:
        file_store = FakeImportIdStore()
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service, file_store=file_store)
        source = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=INVOICE_JAN.name, content=INVOICE_JAN.content)],
        )
        service.confirm_session(session_id=source.id, selected_file_ids=[source.files[0].id])

        replay = service.replay_confirmed_session_files(
            source_session_id=source.id,
            selected_file_ids=[source.files[0].id],
            imported_by="system-repair",
        )

        self.assertNotEqual(replay.id, source.id)
        self.assertNotEqual(replay.files[0].id, source.files[0].id)
        self.assertEqual(replay.files[0].stored_file_path, source.files[0].stored_file_path)
        self.assertEqual(replay.files[0].content_sha256, source.files[0].content_sha256)
        self.assertEqual(replay.files[0].status, "preview_ready")
        self.assertEqual(replay.files[0].duplicate_count, 1)
        self.assertEqual(source.files[0].status, "confirmed")

        service.confirm_session(session_id=replay.id, selected_file_ids=[replay.files[0].id])
        repeated = service.replay_confirmed_session_files(
            source_session_id=source.id,
            selected_file_ids=[source.files[0].id],
            imported_by="system-repair",
        )
        self.assertEqual(repeated.files[0].status, "preview_ready")
        self.assertEqual(repeated.files[0].duplicate_count, 1)

    def test_replay_confirmed_session_files_rejects_changed_source_content(self) -> None:
        file_store = FakeImportIdStore()
        service = FileImportService(
            ImportNormalizationService(id_registry=FakeImportEntityRegistry()),
            file_store=file_store,
        )
        source = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=INVOICE_JAN.name, content=INVOICE_JAN.content)],
        )
        service.confirm_session(session_id=source.id, selected_file_ids=[source.files[0].id])
        stored_path = str(source.files[0].stored_file_path)
        file_store._stored_files[stored_path] = b"changed"

        with self.assertRaisesRegex(ValueError, "checksum changed"):
            service.replay_confirmed_session_files(
                source_session_id=source.id,
                selected_file_ids=[source.files[0].id],
                imported_by="system-repair",
            )

    def test_confirm_session_rolls_back_when_import_confirm_fails(self) -> None:
        import_service = ImportNormalizationService(
            id_registry=FakeImportEntityRegistry(),
            fact_repository=FailingSubmittedEtcIdentityRepository(),
        )
        service = FileImportService(import_service)
        upload = invoice_export_file("jan.xlsx")
        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=upload.name, content=upload.content)],
        )

        with self.assertRaisesRegex(RuntimeError, "submitted etc lookup failed"):
            service.confirm_session(session_id=session.id, selected_file_ids=[session.files[0].id])

        restored = service.get_session(session.id)
        self.assertEqual(restored.status, "preview_ready")
        self.assertEqual(restored.files[0].status, "preview_ready")
        self.assertIsNone(restored.files[0].batch_id)
        self.assertEqual(import_service.list_invoices(), [])

    def test_preview_detects_icbc_last4_from_explicit_file_account(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)
        upload = icbc_history_file(name="historydetail1410.xlsx", account_no="6222020200006386")

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=upload.name,
                    content=upload.content,
                    selected_bank_mapping_id="bank_mapping_icbc_6386",
                    selected_bank_name="工商银行",
                    selected_bank_short_name="工行",
                    selected_bank_last4="6386",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "bank_statement")
        self.assertEqual(preview_file.detected_bank_name, "工商银行")
        self.assertEqual(preview_file.detected_last4, "6386")
        self.assertFalse(preview_file.bank_selection_conflict)
        self.assertIsNone(preview_file.conflict_message)

    def test_preview_accepts_ceb_xlsx_statement_with_yuan_amount_headers(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["中国光大银行对公账户对账单"])
        sheet.append(["查询日期：2026-04-24 11:19:56"])
        sheet.append(["交易日期：20260101-20260423", "", "借贷方向：全部"])
        sheet.append(["账号：39610188000598826", "", "账户名称：云南溯源科技有限公司"])
        sheet.append(["借方笔数：1", "", "借方发生额汇总：23,053.31"])
        sheet.append(["贷方笔数：0", "", "贷方发生额汇总：0.00"])
        sheet.append(["交易日期", "交易时间", "借方发生额（元）", "贷方发生额（元）", "账户余额（元）", "对方账号", "对方名称", "摘要"])
        sheet.append(["2026-04-23", "11:18:17", "23,053.31", "", "3,518.86", "2502046609100018276", "云南辰飞机电工程有限公司", "货款"])
        buffer = BytesIO()
        workbook.save(buffer)
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name="光大银行EXCEL账户明细_39610188000598826_20260101-20260423_260424111837.xlsx",
                    content=buffer.getvalue(),
                    selected_bank_mapping_id="bank_mapping_ceb_8826",
                    selected_bank_name="光大银行",
                    selected_bank_short_name="光大",
                    selected_bank_last4="8826",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "bank_statement")
        self.assertEqual(preview_file.detected_bank_name, "光大银行")
        self.assertEqual(preview_file.detected_last4, "8826")
        self.assertEqual(preview_file.row_count, 1)
        self.assertEqual(preview_file.success_count, 1)
        self.assertFalse(preview_file.bank_selection_conflict)

    def test_preview_accepts_ceb_xlsx_statement_with_income_expense_amount_headers(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["中国光大银行对公账户对账单"])
        sheet.append(["查询日期：2026-06-18 14:54:01"])
        sheet.append(["交易日期：20260101-20260617", "借贷方向：全部", "币种：人民币"])
        sheet.append(["账号：39610188000598826", "账户名称：云南溯源科技有限公司"])
        sheet.append(["借方（支出）笔数汇总：30", "借方（支出）金额汇总：6,519,680.73"])
        sheet.append(["贷方（收入）笔数汇总：13", "贷方（收入）金额汇总：6,435,123.74"])
        sheet.append([
            "交易日期",
            "交易时间",
            "借方金额（支出）",
            "贷方金额（收入）",
            "账户余额",
            "对方账号",
            "对方名称",
            "摘要",
            "对方银行",
            "凭证号",
            "流水号",
        ])
        sheet.append([
            "2026-06-16",
            "10:34:29",
            "",
            "748.00",
            "192,872.59",
            "8111901012800374721",
            "云南钢晟商贸有限公司",
            "货款",
            "中信银行股份有限公司",
            "901309005688",
            "202606169013090056880000000001",
        ])
        sheet.append([
            "2026-06-10",
            "17:37:20",
            "-17,626.82",
            "",
            "195,920.59",
            "9902001850796325",
            "云南溯源科技有限公司",
            "网银跨行汇款失败，原交易流水号：901317033905，收款行拒绝原因：账号、户名不符",
            "中国民生银行",
            "",
            "901a66024449",
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name="光大明细20260101-20260617.xlsx",
                    content=buffer.getvalue(),
                    selected_bank_mapping_id="bank_mapping_ceb_8826",
                    selected_bank_name="光大银行",
                    selected_bank_short_name="光大",
                    selected_bank_last4="8826",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "bank_statement")
        self.assertEqual(preview_file.detected_bank_name, "光大银行")
        self.assertEqual(preview_file.detected_last4, "8826")
        self.assertEqual(preview_file.row_count, 2)
        self.assertEqual(preview_file.success_count, 2)
        self.assertEqual(preview_file.error_count, 0)
        self.assertEqual(preview_file.normalized_rows[0]["txn_direction"], "inflow")
        self.assertEqual(preview_file.normalized_rows[0]["amount"], "748.00")
        self.assertEqual(preview_file.normalized_rows[0]["counterparty_bank_name"], "中信银行股份有限公司")
        self.assertEqual(preview_file.normalized_rows[1]["txn_direction"], "inflow")
        self.assertEqual(preview_file.normalized_rows[1]["amount"], "17626.82")
        self.assertFalse(preview_file.bank_selection_conflict)

    def test_parse_ccb_statement_accepts_customer_account_and_voucher_number_headers(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        parsed = service._parse_rows(
            rows=[
                [
                    "客户账号",
                    "账户名称",
                    "交易时间",
                    "借方发生额（支取）",
                    "贷方发生额（收入）",
                    "余额",
                    "币种",
                    "对方户名",
                    "对方账号",
                    "对方开户机构",
                    "记账日期",
                    "摘要",
                    "备注",
                    "账户明细编号-交易流水号",
                    "企业流水号",
                    "凭证种类",
                    "凭证号码",
                ],
                [
                    "53001905038050548106",
                    "云南溯源科技有限公司",
                    "2026010309:00:13",
                    "6868.55",
                    "0",
                    "154699",
                    "人民币元",
                    "刘树刚",
                    "6217003860012460901",
                    "",
                    "20260103",
                    "电子转账",
                    "代购公车款",
                    "13286-5309050388V2M1WGPI2",
                    "",
                    "电子转账凭证",
                    "108095854700",
                ],
            ],
        )

        self.assertEqual(parsed.template_code, "bank_statement")
        self.assertEqual(parsed.batch_type.value, "bank_transaction")
        self.assertEqual(parsed.rows[0]["account_no"], "53001905038050548106")
        self.assertEqual(parsed.rows[0]["voucher_no"], "108095854700")
        self.assertEqual(parsed.rows[0]["counterparty_name"], "刘树刚")

    def test_preview_accepts_bocom_transaction_detail_statement(self) -> None:
        import_service = ImportNormalizationService(id_registry=FakeImportEntityRegistry())
        service = FileImportService(import_service)

        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[
                UploadedImportFile(
                    file_name=BOCOM_JAN.name,
                    content=BOCOM_JAN.content,
                    selected_bank_mapping_id="bank_mapping_bocom_3847",
                    selected_bank_name="交通银行",
                    selected_bank_short_name="交行",
                    selected_bank_last4="3847",
                )
            ],
        )

        preview_file = session.files[0]
        self.assertEqual(preview_file.status, "preview_ready")
        self.assertEqual(preview_file.template_code, "bank_statement")
        self.assertEqual(preview_file.batch_type.value, "bank_transaction")
        self.assertEqual(preview_file.detected_bank_name, "交通银行")
        self.assertEqual(preview_file.detected_last4, "3847")
        self.assertEqual(preview_file.row_count, 2)
        self.assertEqual(preview_file.success_count, 2)
        self.assertEqual(preview_file.error_count, 0)
        self.assertFalse(preview_file.bank_selection_conflict)
        self.assertEqual(preview_file.normalized_rows[0]["account_no"], "531899991015003383847")
        self.assertEqual(preview_file.normalized_rows[0]["account_name"], "云南溯源科技有限公司")

    def test_discard_session_is_owned_idempotent_and_not_confirmable(self) -> None:
        service = FileImportService(ImportNormalizationService())
        session = service.preview_files(
            imported_by="user_finance_01",
            uploads=[UploadedImportFile(file_name=INVOICE_JAN.name, content=INVOICE_JAN.content)],
        )

        with self.assertRaises(PermissionError):
            service.discard_session(session_id=session.id, imported_by="another_user")

        discarded = service.discard_session(session_id=session.id, imported_by="user_finance_01")
        repeated = service.discard_session(session_id=session.id, imported_by="user_finance_01")

        self.assertEqual(discarded.status, "reverted")
        self.assertEqual(repeated.status, "reverted")
        self.assertEqual(discarded.files[0].status, "reverted")
        self.assertEqual(service.list_active_sessions(imported_by="user_finance_01", mode="invoice"), [])
        with self.assertRaises(ValueError):
            service.confirm_session(session_id=session.id, selected_file_ids=[session.files[0].id])


if __name__ == "__main__":
    unittest.main()
