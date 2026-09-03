from __future__ import annotations

import unittest
from dataclasses import replace
from io import BytesIO
from types import SimpleNamespace

from fin_ops_platform.services.oa_adapter import OAApplicationRecord
from fin_ops_platform.services.oa_payment_status_service import OAPaymentStatusRecord
from fin_ops_platform.services.oa_pending_payment_query_service import OaPendingPaymentQueryService
from fin_ops_platform.services.postgres_connection import PostgresConnection, PostgresSettings
from fin_ops_platform.services.postgres_repositories.oa_attachment_identity_bridge import (
    reconcile_oa_attachment_cache_identity_sources,
)
from fin_ops_platform.services.postgres_repositories.oa_pending_payment_query import (
    PostgresOaPendingPaymentQueryRepository,
)
from fin_ops_platform.services.postgres_repositories.oa_pending_payment_source_snapshot import (
    PostgresOaPendingPaymentSourceSnapshotRepository,
)
from fin_ops_platform.services.postgres_repositories.ops_tax_etc import PostgresOpsTaxEtcRepository
from fin_ops_platform.services.postgres_repositories.workbench_relation import (
    PostgresWorkbenchRelationRepository,
)
from fin_ops_platform.services.workbench_relation_command_service import (
    WorkbenchRelationCommandError,
    WorkbenchRelationCommandService,
)
from openpyxl import load_workbook

from tests.postgres_test_utils import (
    apply_test_migrations,
    require_postgres_test_database_url,
    truncate_test_database,
)


class OaPendingPaymentPostgresIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.database_url = require_postgres_test_database_url()
        apply_test_migrations(self.database_url)
        truncate_test_database(self.database_url)
        self.connection = PostgresConnection(
            PostgresSettings(database_url=self.database_url, pool_enabled=False)
        )

    def _source_snapshot(self) -> PostgresOaPendingPaymentSourceSnapshotRepository:
        return PostgresOaPendingPaymentSourceSnapshotRepository(
            self.connection,
            relation_command_service_for_transaction=lambda transaction: WorkbenchRelationCommandService(
                relation_repository=PostgresWorkbenchRelationRepository(transaction),
            ),
        )

    def test_attachment_identity_bridge_converges_when_cache_arrives_before_oa_facts(self) -> None:
        cache = PostgresOpsTaxEtcRepository(self.connection)
        cache.save_oa_attachment_invoice_cache_entry(
            "parser-cache-key-before",
            _attachment_cache_payload("parser-cache-key-before"),
        )

        self._source_snapshot().commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[_record_with_attachment_file()],
            admission_records=[_record_with_attachment_file()],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(
                    flow_id="flow-integration-1",
                    pay_status=0,
                )
            },
        )

        self.assertEqual(
            int(
                self.connection.fetch_one(
                    """
                    select count(*)::int as count
                    from app.oa_attachment_invoice_cache_sources
                    where cache_source_attachment_key = %s
                      and source_attachment_key = %s
                      and source_kind = 'attachment_identity_invoice'
                    """,
                    ("parser-cache-key-before", "actual-oa-attachment-key"),
                )["count"]
            ),
            1,
        )

    def test_attachment_identity_bridge_converges_when_cache_arrives_after_oa_facts(self) -> None:
        self._source_snapshot().commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[_record_with_attachment_file()],
            admission_records=[_record_with_attachment_file()],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(
                    flow_id="flow-integration-1",
                    pay_status=0,
                )
            },
        )

        PostgresOpsTaxEtcRepository(self.connection).save_oa_attachment_invoice_cache_entry(
            "parser-cache-key-after",
            _attachment_cache_payload("parser-cache-key-after"),
        )

        self.assertEqual(
            int(
                self.connection.fetch_one(
                    """
                    select count(*)::int as count
                    from app.oa_attachment_invoice_cache_sources
                    where cache_source_attachment_key = %s
                      and source_attachment_key = %s
                      and source_kind = 'attachment_identity_invoice'
                    """,
                    ("parser-cache-key-after", "actual-oa-attachment-key"),
                )["count"]
            ),
            1,
        )

    def test_attachment_identity_bridge_fails_closed_for_multiple_current_owners(self) -> None:
        self.connection.execute(
            """
            insert into app.oa_applications(
                oa_source_id, form_id, row_id, status, scope_month
            ) values (
                'ambiguous-source-a', 'ambiguous-form-a',
                'oa-ambiguous-a', 'completed', '2026-05-01'
            )
            """
        )
        self.connection.execute(
            """
            insert into app.oa_application_items(
                oa_application_id, oa_source_id, form_id, row_id
            )
            select id, oa_source_id, form_id, 'shared-current-item'
            from app.oa_applications
            where row_id = 'oa-ambiguous-a'
            """
        )
        self.connection.execute(
            """
            insert into app.oa_attachments(
                oa_application_id, oa_source_id, form_id, row_id,
                source_attachment_key, filename, normalized_payload
            )
            select
                id,
                oa_source_id,
                form_id,
                'shared-current-item',
                row_id || ':attachment',
                'invoice.pdf',
                jsonb_build_object(
                    'source_expense_item_id', 'shared-current-item',
                    'source_expense_row_index', '0',
                    'source_attachment_name', 'invoice.pdf'
                )
            from app.oa_applications
            where row_id = 'oa-ambiguous-a'
            """
        )
        PostgresOpsTaxEtcRepository(self.connection).save_oa_attachment_invoice_cache_entry(
            "ambiguous-cache-key",
            {
                "parser_version": "integration-v1",
                "cache_schema_version": "integration-v1",
                "invoices": [{
                    "source_attachment_key": "ambiguous-cache-key",
                    "source_expense_item_id": "shared-current-item",
                    "source_expense_row_index": "0",
                    "source_attachment_name": "invoice.pdf",
                    "evidence_type": "tax_invoice",
                    "digital_invoice_no": "26534000000060092012",
                    "total_with_tax": "100.00",
                }],
                "evidences": [],
                "artifacts": [],
            },
        )

        self.assertEqual(
            int(
                self.connection.fetch_one(
                    """
                    select count(*)::int as count
                    from app.oa_attachment_invoice_cache_sources
                    where cache_source_attachment_key = 'ambiguous-cache-key'
                      and source_kind = 'attachment_identity_invoice'
                    """
                )["count"]
            ),
            1,
        )

        self.connection.execute(
            """
            with inserted_app as (
                insert into app.oa_applications(
                    oa_source_id, form_id, row_id, status, scope_month
                ) values (
                    'unrelated-source', 'unrelated-form',
                    'oa-unrelated', 'completed', '2026-05-01'
                )
                returning id, oa_source_id, form_id, row_id
            ), inserted_item as (
                insert into app.oa_application_items(
                    oa_application_id, oa_source_id, form_id, row_id
                )
                select id, oa_source_id, form_id, 'unrelated-current-item'
                from inserted_app
            )
            insert into app.oa_attachments(
                oa_application_id, oa_source_id, form_id, row_id,
                source_attachment_key, filename, normalized_payload
            )
            select
                id,
                oa_source_id,
                form_id,
                'unrelated-current-item',
                'unrelated-actual-attachment',
                'unrelated.pdf',
                jsonb_build_object(
                    'source_expense_item_id', 'unrelated-current-item',
                    'source_expense_row_index', '0',
                    'source_attachment_name', 'unrelated.pdf'
                )
            from inserted_app
            """
        )
        PostgresOpsTaxEtcRepository(self.connection).save_oa_attachment_invoice_cache_entry(
            "unrelated-cache-key",
            {
                "parser_version": "integration-v1",
                "cache_schema_version": "integration-v1",
                "invoices": [{
                    "source_attachment_key": "unrelated-cache-key",
                    "source_expense_item_id": "unrelated-current-item",
                    "source_expense_row_index": "0",
                    "source_attachment_name": "unrelated.pdf",
                    "evidence_type": "tax_invoice",
                    "digital_invoice_no": "26534000000060092013",
                    "total_with_tax": "100.00",
                }],
                "evidences": [],
                "artifacts": [],
            },
        )
        unrelated_before = self.connection.fetch_one(
            """
            select source_attachment_key, source_expense_item_id,
                   source_expense_row_index, source_attachment_name, updated_at
            from app.oa_attachment_invoice_cache_sources
            where cache_source_attachment_key = 'unrelated-cache-key'
              and source_kind = 'attachment_identity_invoice'
            """
        )
        self.assertIsNotNone(unrelated_before)

        self.connection.execute(
            """
            with inserted_app as (
                insert into app.oa_applications(
                    oa_source_id, form_id, row_id, status, scope_month
                ) values (
                    'ambiguous-source-b', 'ambiguous-form-b',
                    'oa-ambiguous-b', 'completed', '2026-05-01'
                )
                returning id, oa_source_id, form_id, row_id
            ), inserted_item as (
                insert into app.oa_application_items(
                    oa_application_id, oa_source_id, form_id, row_id
                )
                select id, oa_source_id, form_id, 'shared-current-item'
                from inserted_app
            )
            insert into app.oa_attachments(
                oa_application_id, oa_source_id, form_id, row_id,
                source_attachment_key, filename, normalized_payload
            )
            select
                id,
                oa_source_id,
                form_id,
                'shared-current-item',
                row_id || ':attachment',
                'invoice.pdf',
                jsonb_build_object(
                    'source_expense_item_id', 'shared-current-item',
                    'source_expense_row_index', '0',
                    'source_attachment_name', 'invoice.pdf'
                )
            from inserted_app
            """
        )
        reconcile_oa_attachment_cache_identity_sources(
            self.connection,
            oa_row_ids=["oa-ambiguous-b"],
        )

        self.assertEqual(
            int(
                self.connection.fetch_one(
                    """
                    select count(*)::int as count
                    from app.oa_attachment_invoice_cache_sources
                    where cache_source_attachment_key = 'ambiguous-cache-key'
                      and source_kind = 'attachment_identity_invoice'
                    """
                )["count"]
            ),
            0,
        )
        unrelated_after = self.connection.fetch_one(
            """
            select source_attachment_key, source_expense_item_id,
                   source_expense_row_index, source_attachment_name, updated_at
            from app.oa_attachment_invoice_cache_sources
            where cache_source_attachment_key = 'unrelated-cache-key'
              and source_kind = 'attachment_identity_invoice'
            """
        )
        self.assertEqual(unrelated_after, unrelated_before)

    def test_attachment_identity_bridge_fails_closed_for_multiple_owners_in_same_oa(self) -> None:
        self.connection.execute(
            """
            with inserted_app as (
                insert into app.oa_applications(
                    oa_source_id, form_id, row_id, status, scope_month
                ) values (
                    'same-oa-ambiguous-source', 'same-oa-ambiguous-form',
                    'oa-same-ambiguous', 'completed', '2026-05-01'
                )
                returning id, oa_source_id, form_id
            ), inserted_item as (
                insert into app.oa_application_items(
                    oa_application_id, oa_source_id, form_id, row_id,
                    item_no, normalized_payload
                )
                select id, oa_source_id, form_id, 'same-oa-shared-item',
                       '0', '{"row_index":"0"}'::jsonb
                from inserted_app
                returning oa_application_id, oa_source_id, form_id, row_id
            )
            insert into app.oa_attachments(
                oa_application_id, oa_source_id, form_id, row_id,
                source_attachment_key, filename, normalized_payload
            )
            select item.oa_application_id, item.oa_source_id, item.form_id, item.row_id,
                   source.key, 'same-name.pdf',
                   jsonb_build_object(
                       'source_expense_item_id', item.row_id,
                       'source_expense_row_index', '0',
                       'source_attachment_name', 'same-name.pdf'
                   )
            from inserted_item item
            cross join (values ('same-oa-attachment-a'), ('same-oa-attachment-b')) source(key)
            """
        )

        PostgresOpsTaxEtcRepository(self.connection).save_oa_attachment_invoice_cache_entry(
            "same-oa-ambiguous-cache",
            {
                "parser_version": "integration-v1",
                "cache_schema_version": "integration-v1",
                "invoices": [{
                    "source_attachment_key": "non-current-parser-occurrence",
                    "source_expense_item_id": "same-oa-shared-item",
                    "source_expense_row_index": "0",
                    "source_attachment_name": "same-name.pdf",
                    "evidence_type": "tax_invoice",
                    "digital_invoice_no": "26534000000060092014",
                    "total_with_tax": "100.00",
                }],
                "evidences": [],
                "artifacts": [],
            },
        )

        self.assertEqual(
            int(
                self.connection.fetch_one(
                    """
                    select count(*)::int as count
                    from app.oa_attachment_invoice_cache_sources
                    where cache_source_attachment_key = 'same-oa-ambiguous-cache'
                      and source_kind = 'attachment_identity_invoice'
                    """
                )["count"]
            ),
            0,
        )

    def test_attachment_identity_bridge_preserves_distinct_same_oa_item_occurrences(self) -> None:
        self.connection.execute(
            """
            insert into app.oa_applications(
                oa_source_id, form_id, row_id, status, scope_month
            ) values (
                'same-physical-source', 'same-physical-form',
                'oa-same-physical', 'completed', '2026-05-01'
            )
            """
        )
        self.connection.execute(
            """
            insert into app.oa_application_items(
                oa_application_id, oa_source_id, form_id, row_id,
                item_no, normalized_payload
            )
            select id, oa_source_id, form_id, 'same-physical-item-0',
                   '0', '{"row_index":"0"}'::jsonb
            from app.oa_applications where row_id = 'oa-same-physical'
            union all
            select id, oa_source_id, form_id, 'same-physical-item-1',
                   '1', '{"row_index":"1"}'::jsonb
            from app.oa_applications where row_id = 'oa-same-physical'
            """
        )
        self.connection.execute(
            """
            insert into app.oa_attachments(
                oa_application_id, oa_source_id, form_id, row_id,
                source_attachment_key, filename, normalized_payload
            )
            select app.id, app.oa_source_id, app.form_id, item.row_id,
                   'same-physical-occurrence-' || item.item_no,
                   'same-physical.pdf',
                   jsonb_build_object(
                       'source_expense_item_id', item.row_id,
                       'source_expense_row_index', item.item_no,
                       'source_attachment_name', 'same-physical.pdf',
                       'physical_source_attachment_key', 'one-physical-file'
                   )
            from app.oa_applications app
            join app.oa_application_items item on item.oa_application_id = app.id
            where app.row_id = 'oa-same-physical'
            """
        )

        PostgresOpsTaxEtcRepository(self.connection).save_oa_attachment_invoice_cache_entry(
            "same-physical-cache",
            {
                "parser_version": "integration-v1",
                "cache_schema_version": "integration-v1",
                "invoices": [
                    {
                        "source_attachment_key": "same-physical-occurrence-0",
                        "source_expense_item_id": "same-physical-item-0",
                        "source_expense_row_index": "0",
                        "source_attachment_name": "same-physical.pdf",
                        "evidence_type": "tax_invoice",
                        "digital_invoice_no": "26534000000060092015",
                        "total_with_tax": "100.00",
                    },
                    {
                        "source_attachment_key": "same-physical-occurrence-1",
                        "source_expense_item_id": "same-physical-item-1",
                        "source_expense_row_index": "1",
                        "source_attachment_name": "same-physical.pdf",
                        "evidence_type": "tax_invoice",
                        "digital_invoice_no": "26534000000060092015",
                        "total_with_tax": "100.00",
                    },
                ],
                "evidences": [],
                "artifacts": [],
            },
        )

        identity_rows = self.connection.fetch_all(
            """
            select source_attachment_key, source_expense_item_id
            from app.oa_attachment_invoice_cache_sources
            where cache_source_attachment_key = 'same-physical-cache'
              and source_kind = 'attachment_identity_invoice'
            order by source_attachment_key
            """
        )
        self.assertEqual(
            [
                (row["source_attachment_key"], row["source_expense_item_id"])
                for row in identity_rows
            ],
            [
                ("same-physical-occurrence-0", "same-physical-item-0"),
                ("same-physical-occurrence-1", "same-physical-item-1"),
            ],
        )

    def test_canonical_page_query_observes_active_relation_changes_without_read_model_refresh(self) -> None:
        source_snapshot = self._source_snapshot()
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[_record()],
            admission_records=[_record()],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(
                    flow_id="flow-integration-1",
                    pay_status=0,
                )
            },
        )
        self.connection.execute(
            """
            insert into app.bank_transactions(
                legacy_mongo_id, account_no, txn_direction, counterparty_name_raw,
                amount, signed_amount, txn_date, txn_month, status, raw_payload
            )
            values
                (
                    'bank-direct-query', '622200001234', 'outflow', '集成测试供应商',
                    100, -100, '2026-05-20', '2026-05-01', 'pending',
                    '{"normalized_payload":{"bank_name":"建设银行","account_last4":"1234"}}'::jsonb
                ),
                (
                    'bank-turnover-inflow-60', '622200001234', 'inflow', '集成测试供应商',
                    60, 60, '2026-05-19', '2026-05-01', 'pending', '{}'::jsonb
                ),
                (
                    'bank-turnover-inflow-40', '622200001234', 'inflow', '集成测试供应商',
                    40, 40, '2026-05-18', '2026-05-01', 'pending', '{}'::jsonb
                )
            """
        )
        self.connection.execute(
            """
            insert into app.invoices(
                legacy_mongo_id, invoice_type, invoice_no, invoice_date, invoice_month,
                seller_name, amount, signed_amount, total_with_tax, status, raw_payload
            )
            values (
                'invoice-direct-query', 'input', 'INV-DIRECT-QUERY',
                '2026-05-18', '2026-05-01', '集成测试供应商',
                100, 100, 100, 'pending', '{}'::jsonb
            )
            """
        )
        self.connection.execute(
            """
            insert into app.workbench_pair_relations(
                case_id, relation_mode, status, version, month_scope,
                row_ids, row_types, raw_payload
            )
            values (
                'oa-direct-query', 'turnover_manual_closure', 'active', 1, '2026-05-01',
                array[
                    'oa-integration-1',
                    'bank-turnover-inflow-60',
                    'bank-turnover-inflow-40',
                    'bank-direct-query',
                    'invoice-direct-query'
                ],
                array['oa', 'bank', 'bank', 'bank', 'invoice'],
                '{
                    "normalized_payload": {
                        "case_id": "oa-direct-query",
                        "status": "active",
                        "relation_mode": "turnover_manual_closure",
                        "version": 1,
                        "row_ids": [
                            "oa-integration-1",
                            "bank-turnover-inflow-60",
                            "bank-turnover-inflow-40",
                            "bank-direct-query",
                            "invoice-direct-query"
                        ],
                        "row_types": ["oa", "bank", "bank", "bank", "invoice"]
                    }
                }'::jsonb
            )
            """
        )
        service = OaPendingPaymentQueryService(
            repository=PostgresOaPendingPaymentQueryRepository(self.connection)
        )
        query = {"month": ["2026-05"], "page": ["1"], "page_size": ["20"]}

        active = service.rows(query, tenant_id="default")
        active_candidates = service.bank_transaction_candidates(
            {
                "relation_status": ["matched"],
                "page": ["1"],
                "page_size": ["20"],
            },
            tenant_id="default",
        )

        self.assertEqual(active["pagination"]["total"], 1)
        self.assertEqual(active["rows"][0]["bankTransaction"]["relationCount"], 1)
        self.assertEqual(active["rows"][0]["bankTransaction"]["paidTotal"], "100.00")
        self.assertEqual(active["rows"][0]["bankTransaction"]["nonOutflowBankRelationCount"], 2)
        self.assertEqual(active["rows"][0]["paymentStatus"]["code"], "paid")
        self.assertEqual(
            [
                summary["bankTransactionId"]
                for summary in active["rows"][0]["bankTransaction"]["summaries"]
            ],
            ["bank-direct-query"],
        )
        self.assertEqual(active["rows"][0]["invoice"]["relationCount"], 1)
        self.assertEqual([row["id"] for row in active_candidates["rows"]], ["bank-direct-query"])
        self.assertNotIn("read_model_status", active)
        self.assertEqual(
            self.connection.fetch_one(
                "select count(*)::integer as count from job.outbox_events"
            )["count"],
            0,
        )

        self.connection.execute(
            """
            update app.workbench_pair_relations
            set status = 'withdrawn'
            where case_id = 'oa-direct-query'
            """
        )
        withdrawn = service.rows(query, tenant_id="default")
        withdrawn_candidates = service.bank_transaction_candidates(
            {
                "relation_status": ["unmatched"],
                "page": ["1"],
                "page_size": ["20"],
            },
            tenant_id="default",
        )

        self.assertEqual(withdrawn["pagination"]["total"], 1)
        self.assertEqual(withdrawn["rows"][0]["bankTransaction"]["relationCount"], 0)
        self.assertEqual(withdrawn["rows"][0]["invoice"]["relationCount"], 0)
        self.assertEqual([row["id"] for row in withdrawn_candidates["rows"]], ["bank-direct-query"])

    def test_export_reads_both_oa_fact_sources_without_queue_or_non_oa_fields(self) -> None:
        completed = _record()
        in_progress = _in_progress_record()
        self._source_snapshot().commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[completed],
            admission_records=[completed, in_progress],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(
                    flow_id="flow-integration-1",
                    pay_status=0,
                ),
                "flow-in-progress-1": OAPaymentStatusRecord(
                    flow_id="flow-in-progress-1",
                    pay_status=0,
                ),
            },
        )
        outbox_count_before = int(
            self.connection.fetch_one("select count(*)::integer as count from job.outbox_events")["count"]
        )
        service = OaPendingPaymentQueryService(
            repository=PostgresOaPendingPaymentQueryRepository(self.connection)
        )

        exported = service.export_sources(
            {"sources": ["completed,in_progress"]},
            tenant_id="default",
        )

        workbook = load_workbook(BytesIO(exported["content"]), read_only=True, data_only=False)
        self.assertEqual(workbook.sheetnames, ["已完成OA", "进行中OA"])
        self.assertEqual(workbook["已完成OA"]["A2"].value, "oa-integration-1")
        self.assertEqual(workbook["进行中OA"]["A2"].value, "oa-in-progress-1")
        headers = [cell.value for cell in workbook["已完成OA"][1]]
        workbook.close()
        self.assertEqual(exported["counts"], {"completed": 1, "in_progress": 1})
        self.assertTrue({"OA ID", "申请人", "申请金额", "申请事由"}.issubset(headers))
        self.assertTrue({"流水", "发票", "关联关系"}.isdisjoint(headers))
        self.assertEqual(
            int(self.connection.fetch_one("select count(*)::integer as count from job.outbox_events")["count"]),
            outbox_count_before,
        )

    def test_identical_canonical_commit_keeps_projection_and_status_rows_unchanged(self) -> None:
        source_snapshot = self._source_snapshot()
        payment_statuses = {
            "flow-integration-1": OAPaymentStatusRecord(
                flow_id="flow-integration-1",
                pay_status=0,
            )
        }

        first = source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[_record()],
            admission_records=[_record()],
            payment_statuses=payment_statuses,
        )
        before = self.connection.fetch_one(
            """
            select
                application.updated_at as application_updated_at,
                status.updated_at as status_updated_at,
                (select count(*) from job.outbox_events) as outbox_count
            from app.oa_applications application
            join app.oa_pending_payment_status_snapshots status
              on status.tenant_id = 'default'
             and status.flow_id = 'flow-integration-1'
            where application.row_id = 'oa-integration-1'
            """
        )

        second = source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[_record()],
            admission_records=[_record()],
            payment_statuses=payment_statuses,
        )
        after = self.connection.fetch_one(
            """
            select
                application.updated_at as application_updated_at,
                status.updated_at as status_updated_at,
                (select count(*) from job.outbox_events) as outbox_count
            from app.oa_applications application
            join app.oa_pending_payment_status_snapshots status
              on status.tenant_id = 'default'
             and status.flow_id = 'flow-integration-1'
            where application.row_id = 'oa-integration-1'
            """
        )

        self.assertEqual(first.completed_projection_changed_scopes, ("2026-05",))
        self.assertEqual(first.oa_pending_payment_changed_scopes, ("2026-05",))
        self.assertEqual(first.upserted_completed_count, 1)
        self.assertEqual(second.completed_projection_changed_scopes, ())
        self.assertEqual(second.oa_pending_payment_changed_scopes, ())
        self.assertEqual(second.upserted_completed_count, 0)
        self.assertEqual(after, before)

    def test_authoritative_snapshot_cancels_relation_when_one_of_multiple_completed_oa_disappears(self) -> None:
        source_snapshot = self._source_snapshot()
        retained = _record()
        removed = replace(
            retained,
            id="oa-integration-removed",
            detail_fields={
                "Mongo文档ID": "flow-integration-removed",
                "paymentFlowId": "flow-integration-removed",
            },
        )
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[retained, removed],
            admission_records=[retained, removed],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(flow_id="flow-integration-1", pay_status=0),
                "flow-integration-removed": OAPaymentStatusRecord(
                    flow_id="flow-integration-removed",
                    pay_status=0,
                ),
            },
        )
        self.connection.execute(
            """
            insert into app.bank_transactions(
                legacy_mongo_id, account_no, txn_direction, counterparty_name_raw,
                amount, signed_amount, txn_date, txn_month, status, raw_payload
            )
            values (
                'bank-oa-removal', '622200001234', 'outflow', '集成测试供应商',
                100, -100, '2026-05-20', '2026-05-01', 'pending', '{}'::jsonb
            )
            """
        )
        self.connection.execute(
            """
            insert into app.workbench_pair_relations(
                case_id, relation_mode, status, version, month_scope,
                row_ids, row_types, raw_payload
            )
            values (
                'oa-removal-case', 'manual_confirm', 'active', 1, '2026-05-01',
                array['oa-integration-removed', 'bank-oa-removal'],
                array['oa', 'bank'],
                '{
                    "normalized_payload": {
                        "case_id": "oa-removal-case",
                        "status": "active",
                        "relation_mode": "manual_confirm",
                        "version": 1,
                        "row_ids": ["oa-integration-removed", "bank-oa-removal"],
                        "row_types": ["oa", "bank"]
                    }
                }'::jsonb
            )
            """
        )

        result = source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[retained],
            admission_records=[retained],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(flow_id="flow-integration-1", pay_status=0)
            },
        )

        self.assertEqual(result.removed_stale_completed_count, 1)
        self.assertEqual(result.relation_cleanup[0]["changed_case_ids"], ["oa-removal-case"])
        self.assertIsNone(
            self.connection.fetch_one(
                "select row_id from app.oa_applications where row_id = 'oa-integration-removed'"
            )
        )
        self.assertEqual(
            self.connection.fetch_one(
                "select status from app.workbench_pair_relations where case_id = 'oa-removal-case'"
            )["status"],
            "cancelled",
        )

    def test_full_authoritative_snapshot_removes_disappeared_oa_payment_status_and_queues_external_delete(
        self,
    ) -> None:
        source_snapshot = self._source_snapshot()
        record = _record()
        status = OAPaymentStatusRecord(flow_id="flow-integration-1", pay_status=1)
        source_snapshot.commit_authoritative_snapshot(
            scope_key="all",
            tenant_id="default",
            projection_records=[record],
            admission_records=[record],
            authoritative_payment_flow_ids=["flow-integration-1"],
            payment_statuses={"flow-integration-1": status},
        )

        result = source_snapshot.commit_authoritative_snapshot(
            scope_key="all",
            tenant_id="default",
            projection_records=[],
            admission_records=[],
            authoritative_payment_flow_ids=[],
            payment_statuses={"flow-integration-1": status},
        )

        self.assertEqual(result.removed_payment_status_flow_ids, ("flow-integration-1",))
        self.assertIsNone(
            self.connection.fetch_one(
                "select row_id from app.oa_applications where row_id = 'oa-integration-1'"
            )
        )
        self.assertIsNone(
            self.connection.fetch_one(
                """
                select flow_id
                from app.oa_pending_payment_status_snapshots
                where tenant_id = 'default' and flow_id = 'flow-integration-1'
                """
            )
        )
        event = self.connection.fetch_one(
            """
            select payload
            from job.outbox_events
            where event_type = 'oa.payment_status.reconcile'
              and aggregate_type = 'oa_source_snapshot'
            order by created_at desc
            limit 1
            """
        )
        self.assertEqual(event["payload"]["operation"], "remove_missing_oa_statuses")
        self.assertEqual(event["payload"]["removed_flow_ids"], ["flow-integration-1"])

    def test_stale_matching_plan_cannot_recreate_relation_after_oa_disappears(self) -> None:
        source_snapshot = self._source_snapshot()
        retained = _record()
        removed = replace(
            retained,
            id="oa-stale-plan-removed",
            detail_fields={
                "Mongo文档ID": "flow-stale-plan-removed",
                "paymentFlowId": "flow-stale-plan-removed",
            },
        )
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[retained, removed],
            admission_records=[retained, removed],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(flow_id="flow-integration-1", pay_status=0),
                "flow-stale-plan-removed": OAPaymentStatusRecord(
                    flow_id="flow-stale-plan-removed",
                    pay_status=0,
                ),
            },
        )
        self.connection.execute(
            """
            insert into app.bank_transactions(
                legacy_mongo_id, account_no, txn_direction, counterparty_name_raw,
                amount, signed_amount, txn_date, txn_month, status, raw_payload
            )
            values (
                'bank-stale-plan', '622200001234', 'outflow', '集成测试供应商',
                100, -100, '2026-05-20', '2026-05-01', 'pending', '{}'::jsonb
            )
            """
        )
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[retained],
            admission_records=[retained],
            payment_statuses={
                "flow-integration-1": OAPaymentStatusRecord(flow_id="flow-integration-1", pay_status=0)
            },
        )
        stale_plan = SimpleNamespace(
            case_id="CASE-STALE-PLAN",
            row_ids=("oa-stale-plan-removed", "bank-stale-plan"),
            row_types=("oa", "bank"),
            relation_fingerprint="stale-plan-fingerprint",
            batch_hash="stale-plan-batch",
            rule_code="exact_amount",
            rule_version="1",
            amount_minor=10000,
            currency="CNY",
            scope_keys=("2026-05",),
            evidence_summary=(),
            target_case_id=None,
            oa_attachment_bindings=(),
        )

        with self.connection.transaction() as transaction:
            service = WorkbenchRelationCommandService(
                relation_repository=PostgresWorkbenchRelationRepository(transaction),
            )
            with self.assertRaises(WorkbenchRelationCommandError) as context:
                service.confirm_formal_relation_plans(
                    [stale_plan],
                    actor_id="system:workbench-matching",
                )

        self.assertEqual(
            context.exception.error_code,
            "workbench_relation_canonical_member_missing",
        )
        self.assertIsNone(
            self.connection.fetch_one(
                "select case_id from app.workbench_pair_relations where case_id = 'CASE-STALE-PLAN'"
            )
        )

    def test_admission_only_commit_preserves_stable_completed_fact_and_never_enqueues_shared_read_models(self) -> None:
        source_snapshot = self._source_snapshot()
        completed_record = _record()
        record = replace(_in_progress_record(), amount="", applicant="", reason="")
        payment_statuses = {
            "flow-integration-1": OAPaymentStatusRecord(
                flow_id="flow-integration-1",
                pay_status=0,
            ),
            "flow-in-progress-1": OAPaymentStatusRecord(
                flow_id="flow-in-progress-1",
                pay_status=0,
            )
        }

        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[completed_record],
            admission_records=[completed_record, record],
            payment_statuses=payment_statuses,
        )
        self.assertEqual(
            self.connection.fetch_one(
                "select count(*)::integer as count from job.outbox_events"
            )["count"],
            0,
        )
        before = self.connection.fetch_one(
            """
            select
                admission.updated_at as admission_updated_at,
                admission.amount as admission_amount,
                status.updated_at as status_updated_at,
                application.updated_at as application_updated_at,
                (select count(*) from job.outbox_events) as outbox_count,
                (select count(*) from app.oa_applications) as completed_projection_count
            from app.oa_pending_payment_admissions admission
            join app.oa_pending_payment_status_snapshots status
              on status.tenant_id = admission.tenant_id
             and status.flow_id = 'flow-in-progress-1'
            join app.oa_applications application
              on application.row_id = 'oa-integration-1'
            where admission.tenant_id = 'default'
              and admission.oa_id = 'oa-in-progress-1'
            """
        )
        before_outbox_ids = {
            str(row["id"])
            for row in self.connection.fetch_all("select id from job.outbox_events")
        }

        changed_admission = replace(record, reason="进行中 OA admission 隔离（已更新）")
        admission_only = source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[completed_record],
            admission_records=[completed_record, changed_admission],
            payment_statuses=payment_statuses,
        )
        after_admission_change = self.connection.fetch_one(
            """
            select
                admission.updated_at as admission_updated_at,
                admission.amount as admission_amount,
                status.updated_at as status_updated_at,
                application.updated_at as application_updated_at,
                (select count(*) from job.outbox_events) as outbox_count,
                (select count(*) from app.oa_applications) as completed_projection_count
            from app.oa_pending_payment_admissions admission
            join app.oa_pending_payment_status_snapshots status
              on status.tenant_id = admission.tenant_id
             and status.flow_id = 'flow-in-progress-1'
            join app.oa_applications application
              on application.row_id = 'oa-integration-1'
            where admission.tenant_id = 'default'
              and admission.oa_id = 'oa-in-progress-1'
            """
        )
        incremental_outbox_rows = [
            row
            for row in self.connection.fetch_all(
                """
                select id, event_type, scope_type, scope_key
                from job.outbox_events
                order by available_at, created_at, id
                """,
            )
            if str(row["id"]) not in before_outbox_ids
        ]

        identical = source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[completed_record],
            admission_records=[completed_record, changed_admission],
            payment_statuses=payment_statuses,
        )
        after_identical = self.connection.fetch_one(
            """
            select
                admission.updated_at as admission_updated_at,
                admission.amount as admission_amount,
                status.updated_at as status_updated_at,
                application.updated_at as application_updated_at,
                (select count(*) from job.outbox_events) as outbox_count,
                (select count(*) from app.oa_applications) as completed_projection_count
            from app.oa_pending_payment_admissions admission
            join app.oa_pending_payment_status_snapshots status
              on status.tenant_id = admission.tenant_id
             and status.flow_id = 'flow-in-progress-1'
            join app.oa_applications application
              on application.row_id = 'oa-integration-1'
            where admission.tenant_id = 'default'
              and admission.oa_id = 'oa-in-progress-1'
            """
        )

        self.assertEqual(admission_only.completed_projection_changed_scopes, ())
        self.assertEqual(admission_only.oa_pending_payment_changed_scopes, ("2026-05",))
        self.assertEqual(identical.completed_projection_changed_scopes, ())
        self.assertEqual(identical.oa_pending_payment_changed_scopes, ())
        self.assertEqual(after_admission_change["application_updated_at"], before["application_updated_at"])
        self.assertEqual(after_admission_change["status_updated_at"], before["status_updated_at"])
        self.assertNotEqual(after_admission_change["admission_updated_at"], before["admission_updated_at"])
        self.assertEqual(after_identical, after_admission_change)
        self.assertIsNone(before["admission_amount"])
        self.assertEqual(int(before["completed_projection_count"]), 1)
        self.assertEqual(
            [
                (row["event_type"], row["scope_type"], row["scope_key"])
                for row in incremental_outbox_rows
            ],
            [],
        )

    def test_in_progress_to_completed_removes_admission_and_reports_shared_fact_change(self) -> None:
        source_snapshot = self._source_snapshot()
        in_progress = _in_progress_record()
        completed = replace(in_progress, workflow_status="completed")
        payment_statuses = {
            "flow-in-progress-1": OAPaymentStatusRecord(
                flow_id="flow-in-progress-1",
                pay_status=0,
            )
        }
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[],
            admission_records=[in_progress],
            payment_statuses=payment_statuses,
        )

        transitioned = source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[completed],
            admission_records=[completed],
            payment_statuses=payment_statuses,
        )
        counts = self.connection.fetch_one(
            """
            select
                (select count(*) from app.oa_applications) as completed_projection_count,
                (select count(*) from app.oa_pending_payment_admissions) as admission_count
            """
        )

        self.assertEqual(transitioned.completed_projection_changed_scopes, ("2026-05",))
        self.assertEqual(transitioned.oa_pending_payment_changed_scopes, ("2026-05",))
        self.assertEqual(transitioned.admission_count, 0)
        self.assertEqual(int(counts["completed_projection_count"]), 1)
        self.assertEqual(int(counts["admission_count"]), 0)

    def test_targeted_refresh_updates_existing_pending_owner_without_advancing_watermark(
        self,
    ) -> None:
        record = replace(
            _in_progress_record(),
            apply_type="日常报销",
            attachment_file_count=1,
            attachment_invoices=[{"invoice_no": "265300000001"}],
        )
        source_snapshot = self._source_snapshot()
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[],
            admission_records=[
                replace(record, attachment_file_count=0, attachment_invoices=[])
            ],
            payment_statuses={
                "flow-in-progress-1": OAPaymentStatusRecord(
                    flow_id="flow-in-progress-1",
                    pay_status=0,
                )
            },
        )
        watermark_before = self.connection.fetch_one(
            """
            select version, payload
            from app.oa_sync_watermarks
            where sync_key = 'oa_pending_payment_source:default:2026-05'
            """
        )

        result = source_snapshot.commit_targeted_attachment_refresh(records=[record])
        counts = self.connection.fetch_one(
            """
            select
                (select count(*) from app.oa_applications where row_id = %s) as completed_count,
                (select count(*) from app.oa_pending_payment_admissions where oa_id = %s) as pending_count,
                (select count(*) from app.oa_sync_watermarks) as watermark_count
            """,
            (record.id, record.id),
        )

        self.assertEqual(result.upserted_completed_count, 0)
        self.assertEqual(result.upserted_pending_count, 1)
        self.assertEqual(result.pending_admission_changed_scopes, ("2026-05",))
        self.assertEqual(int(counts["completed_count"]), 0)
        self.assertEqual(int(counts["pending_count"]), 1)
        self.assertEqual(int(counts["watermark_count"]), 1)
        self.assertEqual(
            self.connection.fetch_one(
                """
                select version, payload
                from app.oa_sync_watermarks
                where sync_key = 'oa_pending_payment_source:default:2026-05'
                """
            ),
            watermark_before,
        )

    def test_targeted_attachment_refresh_moves_pending_to_completed_atomically(self) -> None:
        pending = replace(_in_progress_record(), apply_type="日常报销")
        completed = replace(pending, workflow_status="completed")
        source_snapshot = self._source_snapshot()
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[],
            admission_records=[pending],
            payment_statuses={
                "flow-in-progress-1": OAPaymentStatusRecord(
                    flow_id="flow-in-progress-1",
                    pay_status=0,
                )
            },
        )

        result = source_snapshot.commit_targeted_attachment_refresh(records=[completed])
        counts = self.connection.fetch_one(
            """
            select
                (select count(*) from app.oa_applications where row_id = %s) as completed_count,
                (select count(*) from app.oa_pending_payment_admissions where oa_id = %s) as pending_count
            """,
            (pending.id, pending.id),
        )

        self.assertEqual(result.upserted_completed_count, 1)
        self.assertEqual(int(counts["completed_count"]), 1)
        self.assertEqual(int(counts["pending_count"]), 0)

    def test_targeted_attachment_refresh_refuses_completed_owner_regression(self) -> None:
        completed = replace(_in_progress_record(), apply_type="日常报销", workflow_status="completed")
        in_progress = replace(completed, workflow_status="in_progress")
        source_snapshot = self._source_snapshot()
        source_snapshot.commit_targeted_attachment_refresh(records=[completed])

        with self.assertRaisesRegex(RuntimeError, "owner regression"):
            source_snapshot.commit_targeted_attachment_refresh(records=[in_progress])

        counts = self.connection.fetch_one(
            """
            select
                (select count(*) from app.oa_applications where row_id = %s) as completed_count,
                (select count(*) from app.oa_pending_payment_admissions where oa_id = %s) as pending_count
            """,
            (completed.id, completed.id),
        )
        self.assertEqual(int(counts["completed_count"]), 1)
        self.assertEqual(int(counts["pending_count"]), 0)

    def test_targeted_refresh_cannot_revive_owner_removed_by_full_snapshot(self) -> None:
        pending = replace(_in_progress_record(), apply_type="日常报销")
        source_snapshot = self._source_snapshot()
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[],
            admission_records=[pending],
            payment_statuses={
                "flow-in-progress-1": OAPaymentStatusRecord(
                    flow_id="flow-in-progress-1",
                    pay_status=0,
                )
            },
        )
        source_snapshot.commit_authoritative_snapshot(
            scope_key="2026-05",
            tenant_id="default",
            projection_records=[],
            admission_records=[],
            payment_statuses={},
        )

        with self.assertRaisesRegex(RuntimeError, "exactly one existing pending owner"):
            source_snapshot.commit_targeted_attachment_refresh(records=[pending])

        counts = self.connection.fetch_one(
            """
            select
                (select count(*) from app.oa_applications where row_id = %s) as completed_count,
                (select count(*) from app.oa_pending_payment_admissions where oa_id = %s) as pending_count
            """,
            (pending.id, pending.id),
        )
        self.assertEqual(int(counts["completed_count"]), 0)
        self.assertEqual(int(counts["pending_count"]), 0)

    def test_full_snapshot_completed_owner_wins_same_id_in_progress_duplicate(self) -> None:
        in_progress = _in_progress_record()
        completed = replace(in_progress, workflow_status="completed")

        self._source_snapshot().commit_authoritative_snapshot(
            scope_key="2026-05",
            projection_records=[completed],
            admission_records=[in_progress, completed],
            payment_statuses={
                "flow-in-progress-1": OAPaymentStatusRecord(
                    flow_id="flow-in-progress-1",
                    pay_status=0,
                )
            },
        )
        counts = self.connection.fetch_one(
            """
            select
                (select count(*) from app.oa_applications where row_id = %s) as completed_count,
                (select count(*) from app.oa_pending_payment_admissions where oa_id = %s) as pending_count
            """,
            (completed.id, completed.id),
        )

        self.assertEqual(int(counts["completed_count"]), 1)
        self.assertEqual(int(counts["pending_count"]), 0)



def _record() -> OAApplicationRecord:
    return OAApplicationRecord(
        id="oa-integration-1",
        month="2026-05",
        section="unpaired",
        case_id=None,
        applicant="集成测试申请人",
        project_name="集成测试项目",
        apply_type="支付申请",
        amount="100.00",
        counterparty_name="集成测试供应商",
        reason="真实 PostgreSQL 闭环",
        relation_code="pending_match",
        relation_label="待找流水与发票",
        relation_tone="warn",
        workflow_status="completed",
        detail_fields={
            "Mongo文档ID": "flow-integration-1",
            "paymentFlowId": "flow-integration-1",
        },
    )


def _record_with_attachment_file() -> OAApplicationRecord:
    record = _record()
    record.expense_items = [
        {
            "expense_item_id": "oa-integration-1:item:0",
            "row_index": "0",
            "settlement_amount": "100.00",
            "attachment_files": [
                {
                    "source_attachment_key": "actual-oa-attachment-key",
                    "source_attachment_name": "invoice.pdf",
                    "fileName": "invoice.pdf",
                }
            ],
        }
    ]
    return record


def _attachment_cache_payload(cache_key: str) -> dict[str, object]:
    invoice = {
        "source_attachment_key": cache_key,
        "source_expense_item_id": "oa-integration-1:item:0",
        "source_expense_row_index": "0",
        "source_attachment_name": "invoice.pdf",
        "evidence_type": "tax_invoice",
        "digital_invoice_no": "26534000000060092011",
        "total_with_tax": "100.00",
    }
    return {
        "parser_version": "integration-v1",
        "cache_schema_version": "integration-v1",
        "invoices": [invoice],
        "evidences": [],
        "artifacts": [],
    }


def _in_progress_record() -> OAApplicationRecord:
    return OAApplicationRecord(
        id="oa-in-progress-1",
        month="2026-05",
        section="unpaired",
        case_id=None,
        applicant="集成测试进行中申请人",
        project_name="集成测试项目",
        apply_type="支付申请",
        amount="88.00",
        counterparty_name="集成测试供应商",
        reason="进行中 OA admission 隔离",
        relation_code="pending_match",
        relation_label="待找流水与发票",
        relation_tone="warn",
        workflow_status="in_progress",
        detail_fields={
            "Mongo文档ID": "flow-in-progress-1",
            "paymentFlowId": "flow-in-progress-1",
        },
    )


if __name__ == "__main__":
    unittest.main()
