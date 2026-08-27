from __future__ import annotations

from contextlib import contextmanager
from copy import deepcopy
from decimal import Decimal
from http import HTTPStatus
from io import BytesIO
import json
from pathlib import Path
import tempfile
from typing import Any, Callable
import unittest
from urllib.parse import quote

from openpyxl import load_workbook

from fin_ops_platform.app.routes_oa_pending_payments import OaPendingPaymentApiRoutes
from fin_ops_platform.app.server import Application, Response
from tests.app_test_support import (
    build_local_state_application as build_application,
    configure_access_control,
)
from fin_ops_platform.domain.enums import TransactionDirection
from fin_ops_platform.domain.models import BankTransaction
from fin_ops_platform.services.oa_adapter import OAApplicationRecord
from fin_ops_platform.services.oa_identity_service import OAUserIdentity
from fin_ops_platform.services.oa_pending_payment_canonical_rows import build_oa_pending_payment_rows
from fin_ops_platform.services.oa_pending_payment_export import (
    build_oa_pending_payment_export_workbook,
)
from fin_ops_platform.services.oa_pending_payment_query_contract import OaPendingPaymentError
from fin_ops_platform.services.workbench_pair_relation_service import WorkbenchPairRelationService





class FakeCommandService:
    def __init__(self) -> None:
        self.link_calls: list[tuple[dict[str, Any], str]] = []

    def link_bank_transactions(self, payload: dict[str, Any], *, actor_id: str) -> dict[str, Any]:
        self.link_calls.append((dict(payload), actor_id))
        return {
            "success": True,
            "action": "oa_pending_payment_link_bank_transactions",
            "oaRowIds": payload.get("oa_row_ids") or payload.get("oaRowIds") or [],
            "bankTransactionIds": payload.get("bank_transaction_ids") or payload.get("bankTransactionIds") or [],
            "relation": {"status": "confirmed"},
            "paymentStatusSync": {"code": "queued", "label": "已进入自动同步"},
        }


class FakeQueryService:
    def __init__(self) -> None:
        self.candidate_queries: list[tuple[dict[str, list[str]], str]] = []
        self.export_queries: list[tuple[dict[str, list[str]], str]] = []

    def bank_transaction_candidates(
        self,
        query: dict[str, list[str]],
        *,
        tenant_id: str,
    ) -> dict[str, Any]:
        self.candidate_queries.append((dict(query), tenant_id))
        return {
            "rows": [
                {
                    "id": "bank-api",
                    "counterpartyName": "API供应商",
                    "tradeTime": "2026-05-21 10:00:00",
                    "amount": "100.00",
                    "relationStatus": query.get("relation_status", ["all"])[0],
                    "relationStatusLabel": "未配对",
                    "linkedOaRowIds": query.get("oa_row_ids", []),
                }
            ],
            "pagination": {"page": 1, "pageSize": 100, "total": 1},
        }

    def export_sources(
        self,
        query: dict[str, list[str]],
        *,
        tenant_id: str,
    ) -> dict[str, Any]:
        self.export_queries.append((dict(query), tenant_id))
        sources = ("completed", "in_progress")
        content = build_oa_pending_payment_export_workbook(
            [
                {
                    "source_kind": "completed",
                    "oa_id": "oa-completed-api",
                    "workflow_no": "OA-2026-001",
                    "workflow_status": "已完成",
                    "applicant": "完成申请人",
                },
                {
                    "source_kind": "in_progress",
                    "oa_id": "oa-progress-api",
                    "workflow_no": "OA-2026-002",
                    "workflow_status": "进行中",
                    "applicant": "进行中申请人",
                },
            ],
            sources=sources,
        )
        return {
            "filename": "OA事实源_2026-08-19.xlsx",
            "content": content,
            "sources": list(sources),
            "counts": {"completed": 1, "in_progress": 1},
            "row_count": 2,
        }


class FakeRelationCommandService:
    def __init__(self) -> None:
        self.active_relations: list[dict[str, object]] = []
        self.confirm_calls: list[dict[str, object]] = []

    def active_relations_for_row_ids(self, row_ids: list[str]) -> list[dict[str, object]]:
        wanted = {str(row_id) for row_id in row_ids}
        return [
            relation
            for relation in self.active_relations
            if wanted & {str(row_id) for row_id in list(relation.get("row_ids") or [])}
        ]

    def confirm_relation(self, **kwargs: object) -> dict[str, object]:
        self.confirm_calls.append(dict(kwargs))
        relation = {
            "case_id": kwargs["case_id"],
            "row_ids": list(kwargs["row_ids"]),  # type: ignore[arg-type]
            "row_types": list(kwargs["row_types"]),  # type: ignore[arg-type]
            "relation_mode": kwargs["relation_mode"],
            "amount_check": dict(kwargs["amount_check"]),  # type: ignore[arg-type]
            "month_scope": kwargs["month_scope"],
        }
        self.active_relations.append(relation)
        return {
            "status": "confirmed",
            "relation": relation,
            "changed_case_ids": [kwargs["case_id"]],
            "affected_months": [kwargs["month_scope"]],
        }


class OaPendingPaymentApiTests(unittest.TestCase):





    def test_production_service_wires_page_query_repository_from_postgres_connection(self) -> None:
        connection = object()
        app = object.__new__(Application)
        app._state_store = type("StateStore", (), {"_connection": connection})()

        service = Application._oa_pending_payment_query_service(app)

        self.assertIs(service, Application._oa_pending_payment_query_service(app))
        self.assertIs(service._repository._connection, connection)  # type: ignore[union-attr]




    def test_confirm_paid_route_is_removed_from_route_owner(self) -> None:
        routes = OaPendingPaymentApiRoutes(command_service=FakeCommandService())

        response = routes.route(
            "POST",
            "/api/oa-pending-payments/confirm-paid",
            {},
            json.dumps({"oa_row_id": "oa-api", "bank_transaction_id": "bank-api"}),
            {},
        )

        self.assertIsNone(response)
        self.assertFalse(hasattr(routes, "confirm_paid"))

        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            app_response = app.handle_request(
                "POST",
                "/api/oa-pending-payments/confirm-paid",
                body=json.dumps({"oa_row_id": "oa-api", "bank_transaction_id": "bank-api"}),
            )

        self.assertEqual(app_response.status_code, 404)

    def test_link_bank_transactions_route_delegates_to_command_service_with_write_actor(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            command_service = FakeCommandService()
            app._oa_pending_payment_api_routes = OaPendingPaymentApiRoutes(command_service=command_service)
            app._workbench_write_auth_context = lambda _headers: ("tester", "default")  # type: ignore[method-assign]

            response = app.handle_request(
                "POST",
                "/api/oa-pending-payments/link-bank-transactions",
                body=json.dumps({"oa_row_ids": ["oa-api"], "bank_transaction_ids": ["bank-api"]}),
            )

        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.body)
        self.assertTrue(payload["success"])
        self.assertEqual(payload["action"], "oa_pending_payment_link_bank_transactions")
        self.assertEqual(payload["paymentStatusSync"]["code"], "queued")
        self.assertNotIn("readModelRefresh", payload)
        self.assertEqual(command_service.link_calls, [({"oa_row_ids": ["oa-api"], "bank_transaction_ids": ["bank-api"]}, "tester")])

    def test_writeback_paid_route_is_removed(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            response = app.handle_request(
                "POST",
                "/api/oa-pending-payments/writeback-paid",
                body=json.dumps({"oa_row_ids": ["oa-api"]}),
            )

        self.assertEqual(response.status_code, HTTPStatus.NOT_FOUND)

    def test_bank_transaction_candidates_route_delegates_to_tenant_query_service(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            query_service = FakeQueryService()
            app._oa_pending_payment_api_routes = OaPendingPaymentApiRoutes(query_service=query_service)  # type: ignore[arg-type]

            response = app.handle_request(
                "GET",
                "/api/oa-pending-payments/bank-transaction-candidates?relation_status=unmatched&oa_row_ids=oa-api&oa_row_ids=oa-extra",
            )

        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.body)
        self.assertEqual(payload["rows"][0]["id"], "bank-api")
        self.assertEqual(payload["rows"][0]["relationStatus"], "unmatched")
        self.assertEqual(payload["rows"][0]["linkedOaRowIds"], ["oa-api", "oa-extra"])
        query, tenant_id = query_service.candidate_queries[0]
        self.assertEqual(query["oa_row_ids"], ["oa-api", "oa-extra"])
        self.assertEqual(tenant_id, "default")

    def test_export_route_returns_xlsx_and_records_metadata_only_audit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            query_service = FakeQueryService()
            app._oa_pending_payment_api_routes = OaPendingPaymentApiRoutes(  # noqa: SLF001
                query_service=query_service  # type: ignore[arg-type]
            )

            response = app.handle_request(
                "GET",
                "/api/oa-pending-payments/export?sources=completed,in_progress",
            )
            audit_entries = app._audit_service.as_dicts()  # noqa: SLF001

        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertEqual(
            response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("OA%E4%BA%8B%E5%AE%9E%E6%BA%90_2026-08-19.xlsx", response.headers["Content-Disposition"])
        self.assertEqual(response.headers["Cache-Control"], "no-store")
        self.assertEqual(response.headers["Access-Control-Expose-Headers"], "Content-Disposition")
        workbook = load_workbook(BytesIO(response.body), read_only=True, data_only=False)
        self.assertEqual(workbook.sheetnames, ["已完成OA", "进行中OA"])
        self.assertEqual(workbook["已完成OA"]["A2"].value, "oa-completed-api")
        self.assertEqual(workbook["进行中OA"]["A2"].value, "oa-progress-api")
        workbook.close()
        self.assertEqual(
            query_service.export_queries,
            [({"sources": ["completed,in_progress"]}, "default")],
        )
        audit = audit_entries[-1]
        self.assertEqual(audit["action"], "oa_pending_payment_source_export_downloaded")
        self.assertEqual(audit["entity_type"], "oa_pending_payment_source_export")
        self.assertEqual(audit["metadata"]["event_type"], "operation.completed")
        self.assertEqual(audit["metadata"]["outcome"], "success")
        self.assertEqual(audit["metadata"]["page_key"], "oa-pending-payments")
        self.assertEqual(audit["metadata"]["operation_location"], "/api/oa-pending-payments/export")
        self.assertEqual(audit["metadata"]["sources"], ["completed", "in_progress"])
        self.assertEqual(audit["metadata"]["counts"], {"completed": 1, "in_progress": 1})
        self.assertEqual(audit["metadata"]["row_count"], 2)
        self.assertNotIn("amount", audit["metadata"])
        self.assertNotIn("reason", audit["metadata"])

    def test_canonical_query_runtime_failure_returns_service_unavailable(self) -> None:
        class FailingQueryService:
            def rows(self, _query: dict[str, list[str]], *, tenant_id: str) -> dict[str, Any]:
                raise RuntimeError(f"canonical PostgreSQL unavailable for {tenant_id}")

        routes = OaPendingPaymentApiRoutes(query_service=FailingQueryService())  # type: ignore[arg-type]
        routes.configure_platform_ports(
            resolve_read_session=lambda _headers: (object(), None),
            resolve_read_tenant=lambda _session: "tenant-a",
            write_auth_context=lambda _headers: ("tester", "tenant-a"),
            json_response=_json_test_response,
            load_json_body=lambda _body: ({}, None),
            error_response=_oa_pending_payment_error_test_response,
        )

        response = routes.route(
            "GET",
            "/api/oa-pending-payments/rows",
            {"page": ["1"]},
            None,
            {},
        )

        self.assertEqual(response.status_code, HTTPStatus.SERVICE_UNAVAILABLE)
        self.assertEqual(
            json.loads(response.body)["error"]["code"],
            "oa_pending_payment_service_unavailable",
        )

    def test_candidate_bank_relation_is_not_visible_or_marked_paid(self) -> None:
        bank = BankTransaction(
            id="bank-candidate",
            account_no="622200001234",
            txn_direction=TransactionDirection.OUTFLOW,
            counterparty_name_raw="候选供应商",
            amount=Decimal("100.00"),
            signed_amount=Decimal("-100.00"),
            txn_date="2026-05-21",
            trade_time="2026-05-21 10:00:00",
        )
        rows = build_oa_pending_payment_rows(
            records=[self._oa("oa-candidate", "候选申请人", "100.00", detail_fields={"申请日期": "2026-05-20"})],
            relations=[
                {
                    "case_id": "candidate-oa-bank",
                    "row_ids": ["oa-candidate", "bank-candidate"],
                    "row_types": ["oa", "bank"],
                    "relation_mode": "unlinked_evidence",
                    "relation_status": "unlinked",
                    "amount_check": {"matched": True},
                }
            ],
            bank_transactions=[bank],
            invoices=[],
            payment_statuses_by_flow_id={},
            flow_id_resolver=lambda _record: None,
            scope_key="2026-05",
        )
        row = rows[0]

        self.assertEqual(row["bankTransaction"]["relationCount"], 0)
        self.assertEqual(row["bankTransaction"]["summaries"], [])
        self.assertEqual(row["bankTransaction"]["paidTotal"], "0.00")
        self.assertNotEqual(row["paymentStatus"]["code"], "paid")






    def test_legacy_application_rebuild_helpers_are_removed(self) -> None:
        removed_helpers = [
            "list_oa_pending_payment_scope_shards",
            "mark_oa_pending_payment_scope_empty",
            "rebuild_oa_pending_payment_read_model_scope",
            "_oa_pending_payment_live_rows",
            "_oa_pending_payment_live_rows_for_view",
        ]

        for helper_name in removed_helpers:
            with self.subTest(helper_name=helper_name):
                self.assertFalse(hasattr(Application, helper_name))

    def test_read_endpoints_require_fin_ops_access(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            app._oa_identity_service.resolve_identity = lambda _token: OAUserIdentity(
                user_id="006",
                username="YNSYLP006",
                nickname="未授权用户",
                display_name="未授权用户",
                roles=["finance", "finops_full_access"],
                permissions=["finops:app:view"],
            )
            endpoints = [
                "/api/oa-pending-payments/rows",
                "/api/oa-pending-payments/export?sources=completed",
                "/api/oa-pending-payments/bank-transaction-candidates",
                "/api/oa-pending-payments/oa/oa-api/detail",
                "/api/oa-pending-payments/bank-transactions/bank-api/detail",
                "/api/oa-pending-payments/invoices/inv-api/detail",
                "/api/oa-pending-payments/rows/row-api/relation-details?kind=bank",
                "/api/oa-pending-payments/rows/row-api/relation-details?kind=oa",
            ]

            responses = [
                app.handle_request("GET", endpoint, headers={"Authorization": "Bearer blocked-token"})
                for endpoint in endpoints
            ]

        self.assertTrue(all(response.status_code == 403 for response in responses))
        self.assertTrue(all(json.loads(response.body)["error"] in {"forbidden", "permission_denied"} for response in responses))


    def test_all_module_endpoints_require_module_owned_authentication(self) -> None:
        requests = [
            ("GET", "/api/oa-pending-payments/rows", None),
            ("GET", "/api/oa-pending-payments/export?sources=completed", None),
            ("GET", "/api/oa-pending-payments/bank-transaction-candidates", None),
            ("GET", "/api/oa-pending-payments/oa/oa-api/detail", None),
            ("GET", "/api/oa-pending-payments/bank-transactions/bank-api/detail", None),
            ("GET", "/api/oa-pending-payments/invoices/inv-api/detail", None),
            ("GET", "/api/oa-pending-payments/rows/row-api/relation-details?kind=bank", None),
            (
                "POST",
                "/api/oa-pending-payments/link-bank-transactions",
                {"oa_row_ids": ["oa-api"], "bank_transaction_ids": ["bank-api"]},
            ),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir), install_test_session=False)
            responses = [
                app.handle_request(method, route, body=json.dumps(body) if body is not None else None)
                for method, route, body in requests
            ]

        self.assertTrue(all(response.status_code == HTTPStatus.UNAUTHORIZED for response in responses))
        self.assertTrue(all(json.loads(response.body)["error"] == "invalid_oa_session" for response in responses))

    def test_module_owned_write_auth_rejects_readonly_user(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            configure_access_control(app, read_export_only=["OA_READONLY"])
            app._oa_identity_service.resolve_identity = lambda _token: OAUserIdentity(
                user_id="oa-readonly-id",
                username="OA_READONLY",
                nickname="OA只读用户",
                display_name="OA只读用户",
                roles=[],
                permissions=[],
            )
            headers = {"Authorization": "Bearer oa-readonly-token"}
            responses = [app.handle_request(
                "POST",
                "/api/oa-pending-payments/link-bank-transactions",
                headers=headers,
                body=json.dumps({"oa_row_ids": ["oa-api"], "bank_transaction_ids": ["bank-api"]}),
            )]

        self.assertTrue(all(response.status_code == HTTPStatus.FORBIDDEN for response in responses))
        self.assertTrue(all(json.loads(response.body)["error"] == "permission_denied" for response in responses))

    def test_export_route_allows_read_export_only_user(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            configure_access_control(app, read_export_only=["OA_READONLY"])
            app._oa_identity_service.resolve_identity = lambda _token: OAUserIdentity(  # noqa: SLF001
                user_id="oa-readonly-id",
                username="OA_READONLY",
                nickname="OA只读用户",
                display_name="OA只读用户",
                roles=[],
                permissions=[],
            )
            query_service = FakeQueryService()
            app._oa_pending_payment_api_routes = OaPendingPaymentApiRoutes(  # noqa: SLF001
                query_service=query_service  # type: ignore[arg-type]
            )

            response = app.handle_request(
                "GET",
                "/api/oa-pending-payments/export?sources=completed,in_progress",
                headers={"Authorization": "Bearer oa-readonly-token"},
            )

        self.assertEqual(response.status_code, HTTPStatus.OK)
        self.assertEqual(query_service.export_queries[0][1], "default")



    @staticmethod
    def _oa(
        oa_id: str,
        applicant: str,
        amount: str,
        *,
        detail_fields: dict[str, object] | None = None,
        workflow_status: str | None = "completed",
    ) -> OAApplicationRecord:
        return OAApplicationRecord(
            id=oa_id,
            month="2026-05",
            section="审批通过",
            case_id=None,
            applicant=applicant,
            project_name="API项目",
            apply_type="报销",
            amount=amount,
            counterparty_name="API供应商",
            reason="API测试",
            relation_code="",
            relation_label="",
            relation_tone="",
            workflow_status=workflow_status,
            detail_fields=detail_fields or {},
            project_name_display="API项目",
        )






























def _json_test_response(
    status_code: HTTPStatus,
    payload: dict[str, Any],
    response_headers: dict[str, str] | None = None,
) -> Response:
    response = Response(
        status_code=int(status_code),
        body="" if status_code == HTTPStatus.NOT_MODIFIED else json.dumps(payload, ensure_ascii=False),
    )
    response.headers.update(response_headers or {})
    return response




def _oa_pending_payment_error_test_response(exc: Exception) -> Response:
    if not isinstance(exc, OaPendingPaymentError):
        raise exc
    return _json_test_response(
        exc.status_code,
        {
            "error": {
                "code": exc.error_code,
                "message": str(exc),
                "details": exc.details,
            }
        },
    )






if __name__ == "__main__":
    unittest.main()
