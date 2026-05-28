from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import load_workbook

from fin_ops_platform.services.bank_details_export_service import (
    BANK_DETAIL_EXPORT_COLUMNS,
    BANK_DETAIL_EXPORT_ROW_LIMIT,
    BankDetailsExportError,
    BankDetailsExportService,
)


class _PagedRowsLoader:
    def __init__(self, rows: list[dict[str, object]]) -> None:
        self.rows = list(rows)
        self.calls: list[dict[str, object]] = []

    def __call__(self, **kwargs: object) -> dict[str, object]:
        self.calls.append(dict(kwargs))
        page = int(kwargs.get("page") or 1)
        page_size = int(kwargs.get("page_size") or 100)
        start = (page - 1) * page_size
        end = start + page_size
        account_key = str(kwargs.get("account_key") or "").strip()
        keyword = str(kwargs.get("keyword") or "").strip()
        rows = [
            row for row in self.rows
            if (not account_key or row.get("account_key") == account_key)
            and (not keyword or keyword in str(row))
        ]
        return {
            "rows": rows[start:end],
            "pagination": {"page": page, "page_size": page_size, "total": len(rows)},
        }


def _accounts() -> dict[str, object]:
    return {
        "accounts": [
            {"account_key": "工商银行:6386", "bank_name": "工商银行", "account_last4": "6386", "display_name": "工商银行 6386"},
            {"account_key": "民生银行:9486", "bank_name": "民生银行", "account_last4": "9486", "display_name": "民生银行 9486"},
            {"account_key": "交通银行:3847", "bank_name": "交通银行", "account_last4": "3847", "display_name": "交通银行 3847"},
            {"account_key": "建设银行:8106", "bank_name": "建设银行", "account_last4": "8106", "display_name": "建设银行 8106"},
            {"account_key": "光大银行:8826", "bank_name": "光大银行", "account_last4": "8826", "display_name": "光大银行 8826"},
            {"account_key": "平安银行:0093", "bank_name": "平安银行", "account_last4": "0093", "display_name": "平安银行 0093"},
        ]
    }


def _row(
    row_id: str,
    *,
    bank_name: str,
    account_last4: str,
    account_key: str,
    direction: str = "expense",
    amount: str = "100.00",
    purpose_text: str = "",
    summary_text: str = "",
    note_text: str = "",
    category: str | None = None,
    category_primary_label: str | None = None,
    category_sub_label: str | None = None,
    oa_tag: str = "无oa",
    invoice_tag: str = "无发票",
) -> dict[str, object]:
    return {
        "id": row_id,
        "account_key": account_key,
        "trade_time": "2026-04-16 11:09:14+08:00",
        "counterparty_name": "云南溯源科技有限公司",
        "direction": direction,
        "direction_label": "收" if direction == "income" else "支",
        "amount": amount,
        "balance": "276.63",
        "bank_name": bank_name,
        "account_last4": account_last4,
        "auto_category_label": category,
        "auto_category_primary_label": category_primary_label,
        "auto_category_sub_label": category_sub_label,
        "effective_category_label": category,
        "effective_category_primary_label": category_primary_label,
        "effective_category_sub_label": category_sub_label,
        "category_primary_label": category_primary_label,
        "category_sub_label": category_sub_label,
        "oa_relation_tag": oa_tag,
        "invoice_relation_tag": invoice_tag,
        "purpose_text": purpose_text,
        "summary_text": summary_text,
        "note_text": note_text,
    }


class BankDetailsExportServiceTests(unittest.TestCase):
    def test_all_bank_export_builds_summary_and_bank_sheets_with_professional_columns(self) -> None:
        rows = [
            _row("icbc", bank_name="工商银行", account_last4="6386", account_key="工商银行:6386", purpose_text="工行用途", summary_text="工行摘要", note_text="工行附言", category="手续费", category_primary_label="费用", category_sub_label="手续费"),
            _row("internal", bank_name="建设银行", account_last4="1410", account_key="建设银行:1410", direction="income", amount="13000.00", summary_text="内部往来收入", category="内部往来款"),
            _row("bocom", bank_name="交通银行", account_last4="3847", account_key="交通银行:3847", summary_text="交行摘要"),
            _row("ccb", bank_name="建设银行", account_last4="8106", account_key="建设银行:8106", summary_text="建行摘要", note_text="建行备注"),
            _row("cmbc", bank_name="民生银行", account_last4="9486", account_key="民生银行:9486", note_text="客户附言", oa_tag="有oa", invoice_tag="有发票"),
            _row("ceb", bank_name="光大银行", account_last4="8826", account_key="光大银行:8826", summary_text="光大摘要"),
            _row("pingan", bank_name="平安银行", account_last4="0093", account_key="平安银行:0093", purpose_text="平安交易用途", summary_text="平安摘要"),
        ]
        service = BankDetailsExportService(transaction_page_loader=_PagedRowsLoader(rows), account_loader=lambda **_kwargs: _accounts())

        result = service.export(mode="all", date_from="2026-04-01", date_to="2026-05-18", keyword=None)
        workbook = load_workbook(BytesIO(result.content))

        self.assertEqual(result.row_count, 7)
        self.assertEqual(workbook.sheetnames, ["全部流水", "工商银行", "建设银行", "交通银行", "民生银行", "光大银行", "平安银行"])
        sheet = workbook["全部流水"]
        self.assertEqual([cell.value for cell in sheet[1]], BANK_DETAIL_EXPORT_COLUMNS)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, f"A1:Q{sheet.max_row}")
        self.assertEqual(sheet["A2"].value, "2026-04-16 11:09:14")
        self.assertIsInstance(sheet["G2"].value, (int, float))
        self.assertIsNone(sheet["F2"].value)
        self.assertEqual(sheet["I2"].value, "手续费")
        self.assertEqual(sheet["J2"].value, "费用")
        self.assertEqual(sheet["K2"].value, "手续费")
        self.assertEqual(sheet["I3"].value, "内部往来款")
        self.assertEqual(sheet["L6"].value, "有oa")
        self.assertEqual(sheet["M6"].value, "有发票")
        self.assertEqual(sheet["N2"].value, "工行用途")
        self.assertEqual(sheet["O4"].value, "交行摘要")
        self.assertEqual(sheet["P5"].value, "建行备注")
        self.assertEqual(sheet["P6"].value, "客户附言")
        self.assertEqual(sheet["N8"].value, "平安交易用途")

    def test_export_forwards_category_label_filters_to_transaction_loader(self) -> None:
        loader = _PagedRowsLoader(
            [
                _row(
                    "icbc",
                    bank_name="工商银行",
                    account_last4="6386",
                    account_key="工商银行:6386",
                    category="手续费",
                    category_primary_label="费用",
                    category_sub_label="手续费",
                )
            ]
        )
        service = BankDetailsExportService(transaction_page_loader=loader, account_loader=lambda **_kwargs: _accounts())

        service.export(
            mode="all",
            date_from="2026-04-01",
            date_to="2026-05-18",
            keyword=None,
            category_primary_label="费用",
            category_sub_label="手续费",
        )

        self.assertEqual(loader.calls[0]["category_primary_label"], "费用")
        self.assertEqual(loader.calls[0]["category_sub_label"], "手续费")

    def test_account_export_validates_account_metadata_and_allows_empty_filtered_result(self) -> None:
        loader = _PagedRowsLoader([])
        service = BankDetailsExportService(transaction_page_loader=loader, account_loader=lambda **_kwargs: _accounts())

        result = service.export(
            mode="account",
            account_key="民生银行:9486",
            date_from="2026-04-01",
            date_to="2026-05-18",
            keyword="不存在",
        )
        workbook = load_workbook(BytesIO(result.content))

        self.assertEqual(result.row_count, 0)
        self.assertEqual(workbook.sheetnames, ["民生银行"])
        self.assertEqual([cell.value for cell in workbook["民生银行"][1]], BANK_DETAIL_EXPORT_COLUMNS)
        self.assertIn("民生银行9486", result.filename)

    def test_export_pages_through_more_than_existing_page_cap(self) -> None:
        rows = [
            _row(f"row-{index}", bank_name="民生银行", account_last4="9486", account_key="民生银行:9486")
            for index in range(1200)
        ]
        loader = _PagedRowsLoader(rows)
        service = BankDetailsExportService(transaction_page_loader=loader, account_loader=lambda **_kwargs: _accounts())

        result = service.export(mode="all", date_from="2026-04-01", date_to="2026-05-18", keyword=None)
        workbook = load_workbook(BytesIO(result.content))

        self.assertEqual(result.row_count, 1200)
        self.assertEqual(workbook["全部流水"].max_row, 1201)
        self.assertEqual([call["page"] for call in loader.calls], [1, 2, 3])
        self.assertTrue(all(call["page_size"] == 500 for call in loader.calls))

    def test_export_escapes_formula_like_text_cells(self) -> None:
        rows = [
            {
                **_row(
                    "formula-row",
                    bank_name="工商银行",
                    account_last4="6386",
                    account_key="工商银行:6386",
                    purpose_text="+cmd",
                    summary_text="@SUM(A1:A2)",
                    note_text="-1+2",
                ),
                "counterparty_name": "=HYPERLINK(\"http://example.com\")",
            }
        ]
        service = BankDetailsExportService(transaction_page_loader=_PagedRowsLoader(rows), account_loader=lambda **_kwargs: _accounts())

        result = service.export(mode="all", date_from="2026-04-01", date_to="2026-05-18", keyword=None)
        sheet = load_workbook(BytesIO(result.content))["全部流水"]

        self.assertEqual(sheet["D2"].value, "'=HYPERLINK(\"http://example.com\")")
        self.assertEqual(sheet["N2"].value, "'+cmd")
        self.assertEqual(sheet["O2"].value, "'@SUM(A1:A2)")
        self.assertEqual(sheet["P2"].value, "'-1+2")

    def test_export_rebuilds_text_columns_from_bank_text_fields_or_legacy_fields(self) -> None:
        rows = [
            {
                **_row("fields", bank_name="工商银行", account_last4="6386", account_key="工商银行:6386"),
                "purpose_text": "",
                "summary_text": "",
                "note_text": "",
                "bank_text_fields": [
                    {"label": "用途", "value": "工行用途"},
                    {"label": "摘要", "value": "工行摘要"},
                    {"label": "附言", "value": "工行附言"},
                ],
            },
            {
                **_row("legacy", bank_name="建设银行", account_last4="8106", account_key="建设银行:8106"),
                "purpose_text": "",
                "summary_text": "",
                "note_text": "",
                "summary": "旧摘要",
                "purpose": "旧用途",
                "remark": "旧备注",
            },
        ]
        service = BankDetailsExportService(transaction_page_loader=_PagedRowsLoader(rows), account_loader=lambda **_kwargs: _accounts())

        result = service.export(mode="all", date_from="2026-04-01", date_to="2026-05-18", keyword=None)
        sheet = load_workbook(BytesIO(result.content))["全部流水"]

        self.assertEqual(sheet["N2"].value, "工行用途")
        self.assertEqual(sheet["O2"].value, "工行摘要")
        self.assertEqual(sheet["P2"].value, "工行附言")
        self.assertIn(sheet["N3"].value, (None, ""))
        self.assertEqual(sheet["O3"].value, "旧摘要")
        self.assertEqual(sheet["P3"].value, "旧备注")

    def test_export_maps_legacy_minsheng_text_to_note_only(self) -> None:
        rows = [
            {
                **_row("cmbc-legacy", bank_name="民生银行", account_last4="9486", account_key="民生银行:9486"),
                "purpose_text": "",
                "summary_text": "",
                "note_text": "",
                "purpose": "客户附言内容",
                "summary": "客户附言内容",
            }
        ]
        service = BankDetailsExportService(transaction_page_loader=_PagedRowsLoader(rows), account_loader=lambda **_kwargs: _accounts())

        result = service.export(mode="all", date_from="2026-04-01", date_to="2026-05-18", keyword=None)
        sheet = load_workbook(BytesIO(result.content))["全部流水"]

        self.assertIn(sheet["N2"].value, (None, ""))
        self.assertIn(sheet["O2"].value, (None, ""))
        self.assertEqual(sheet["P2"].value, "客户附言内容")

    def test_row_limit_is_enforced_before_collecting_all_pages(self) -> None:
        rows = [
            _row(f"row-{index}", bank_name="民生银行", account_last4="9486", account_key="民生银行:9486")
            for index in range(BANK_DETAIL_EXPORT_ROW_LIMIT + 1)
        ]
        service = BankDetailsExportService(transaction_page_loader=_PagedRowsLoader(rows), account_loader=lambda **_kwargs: _accounts())

        with self.assertRaises(BankDetailsExportError) as context:
            service.export(mode="all", date_from="2026-04-01", date_to="2026-05-18", keyword=None)

        self.assertEqual(context.exception.error_code, "bank_detail_export_row_limit_exceeded")

    def test_invalid_account_key_is_rejected_before_row_collection(self) -> None:
        loader = _PagedRowsLoader([])
        service = BankDetailsExportService(transaction_page_loader=loader, account_loader=lambda **_kwargs: _accounts())

        with self.assertRaises(BankDetailsExportError) as context:
            service.export(mode="account", account_key="不存在:0000", date_from="2026-04-01", date_to="2026-05-18", keyword=None)

        self.assertEqual(context.exception.error_code, "bank_detail_export_account_not_found")
        self.assertEqual(loader.calls, [])


if __name__ == "__main__":
    unittest.main()
