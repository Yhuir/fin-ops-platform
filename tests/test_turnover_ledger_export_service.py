from __future__ import annotations

from datetime import date
from io import BytesIO
import unittest

from openpyxl import load_workbook

from fin_ops_platform.services.turnover_ledger_export_service import TurnoverLedgerExportService


class TurnoverLedgerExportServiceTests(unittest.TestCase):
    def _grouped_payload(self) -> dict[str, object]:
        return {
            "summary": {
                "pending_repayment_amount": "100000.00",
                "pending_collection_amount": "5000.00",
            },
            "family_summaries": [],
            "filters": {"family": "all"},
            "pagination": {"page": 1, "page_size": 100, "total": 2},
            "groups": [
                {
                    "group_id": "counterparty:company:梁希涛",
                    "counterparty_name": "梁希涛",
                    "family": "company",
                    "family_label": "公司往来",
                    "pending_direction": "repayment",
                    "pending_amount": "100000.00",
                    "summary_row": {
                        "relation_id": "turnover_rel_001",
                        "row_kind": "summary",
                        "display_level": "group_summary",
                        "status": "suggested",
                        "status_label": "待人工确认",
                        "borrow_amount": "200000.00",
                        "borrow_date": "2026-02-04",
                        "repayment_amount": "100000.00",
                        "repayment_date": "2026-03-05",
                        "counterparty_bank_name": "建行 8106",
                        "repayment_remark": "还款",
                        "interest_rate_type": "annual",
                        "interest_rate_value": "0.060000",
                        "interest_paid_amount": "120.50",
                        "loan_days": None,
                        "accrued_interest": "953.42",
                        "interest_paid_date": "2026-04-01",
                        "interest_payment_method": "银行转账",
                        "note": "页面备注",
                        "bank_row_ids": ["bank_001", "bank_002"],
                        "row_tone": "warning",
                    },
                    "lot_rows": [
                        {
                            "relation_id": "turnover_rel_001",
                            "row_kind": "lot",
                            "lot_id": "lot_001",
                            "status": "suggested",
                            "status_label": "待人工确认",
                            "borrow_amount": "120000.00",
                            "borrow_date": "2026-02-04",
                            "repayment_amount": "100000.00",
                            "repayment_date": "2026-03-05",
                            "balance_amount": "20000.00",
                            "counterparty_bank_name": "建行 8106",
                            "repayment_remark": "还款",
                            "interest_rate_type": "annual",
                            "interest_rate_value": "0.060000",
                            "interest_paid_amount": "120.50",
                            "loan_days": 29,
                            "accrued_interest": "572.05",
                            "interest_paid_date": "2026-04-01",
                            "interest_payment_method": "银行转账",
                            "note": "页面备注",
                            "bank_row_ids": ["bank_001", "bank_002"],
                            "row_tone": "info",
                        },
                        {
                            "relation_id": "turnover_rel_001",
                            "row_kind": "lot",
                            "lot_id": "lot_002",
                            "status": "suggested",
                            "status_label": "待人工确认",
                            "borrow_amount": "80000.00",
                            "borrow_date": "2026-02-04",
                            "repayment_amount": "0.00",
                            "repayment_date": None,
                            "balance_amount": "80000.00",
                            "counterparty_bank_name": "建行 8106",
                            "repayment_remark": "",
                            "interest_rate_type": "annual",
                            "interest_rate_value": "0.060000",
                            "interest_paid_amount": "0.00",
                            "loan_days": 30,
                            "accrued_interest": "381.37",
                            "interest_paid_date": None,
                            "interest_payment_method": "",
                            "note": "",
                            "bank_row_ids": ["bank_003"],
                            "row_tone": "info",
                        },
                    ],
                    "allocation_lots": [
                        {
                            "relation_id": "turnover_rel_001",
                            "row_kind": "allocation_lot",
                            "lot_id": "lot_001",
                            "borrow_amount": "120000.00",
                            "allocated_repayment_amount": "100000.00",
                            "balance_amount": "20000.00",
                            "loan_days": 29,
                            "accrued_interest": "572.05",
                        },
                        {
                            "relation_id": "turnover_rel_001",
                            "row_kind": "allocation_lot",
                            "lot_id": "lot_002",
                            "borrow_amount": "80000.00",
                            "allocated_repayment_amount": "0.00",
                            "balance_amount": "80000.00",
                            "loan_days": 30,
                            "accrued_interest": "381.37",
                        },
                    ],
                    "flow_rows": [
                        {
                            "relation_id": "turnover_rel_001",
                            "row_kind": "flow",
                            "flow_id": "bank:bank_001",
                            "source_bank_row_id": "bank_001",
                            "flow_direction": "income",
                            "flow_amount": "120000.00",
                            "borrow_amount": "120000.00",
                            "borrow_date": "2026-02-04",
                            "repayment_amount": "0.00",
                            "repayment_date": None,
                            "counterparty_bank_name": "建行 8106",
                            "summary_text": "暂借款",
                            "status": "suggested",
                            "status_label": "待人工确认",
                            "interest_rate_type": "annual",
                            "interest_rate_value": "0.060000",
                            "interest_paid_amount": "0.00",
                            "loan_days": None,
                            "accrued_interest": "0.00",
                            "interest_paid_date": None,
                            "interest_payment_method": "",
                            "note": "",
                            "allocation_status": "allocated",
                            "allocated_lot_ids": ["lot_001"],
                            "bank_row_ids": ["bank_001"],
                        },
                        {
                            "relation_id": "turnover_rel_001",
                            "row_kind": "flow",
                            "flow_id": "bank:bank_002",
                            "source_bank_row_id": "bank_002",
                            "flow_direction": "income",
                            "flow_amount": "80000.00",
                            "borrow_amount": "80000.00",
                            "borrow_date": "2026-02-04",
                            "repayment_amount": "0.00",
                            "repayment_date": None,
                            "counterparty_bank_name": "建行 8106",
                            "summary_text": "暂借款",
                            "status": "suggested",
                            "status_label": "待人工确认",
                            "interest_rate_type": "annual",
                            "interest_rate_value": "0.060000",
                            "interest_paid_amount": "0.00",
                            "loan_days": None,
                            "accrued_interest": "0.00",
                            "interest_paid_date": None,
                            "interest_payment_method": "",
                            "note": "",
                            "allocation_status": "unallocated",
                            "allocated_lot_ids": ["lot_002"],
                            "bank_row_ids": ["bank_002"],
                        },
                        {
                            "relation_id": "turnover_rel_001",
                            "row_kind": "flow",
                            "flow_id": "bank:bank_003",
                            "source_bank_row_id": "bank_003",
                            "flow_direction": "expense",
                            "flow_amount": "100000.00",
                            "borrow_amount": "0.00",
                            "borrow_date": None,
                            "repayment_amount": "100000.00",
                            "repayment_date": "2026-03-05",
                            "counterparty_bank_name": "建行 8106",
                            "summary_text": "还款",
                            "status": "suggested",
                            "status_label": "待人工确认",
                            "interest_rate_type": "annual",
                            "interest_rate_value": "0.060000",
                            "interest_paid_amount": "0.00",
                            "loan_days": None,
                            "accrued_interest": "0.00",
                            "interest_paid_date": None,
                            "interest_payment_method": "",
                            "note": "",
                            "allocation_status": "allocated",
                            "allocated_lot_ids": ["lot_001"],
                            "bank_row_ids": ["bank_003"],
                        },
                    ],
                    "rows": [
                        {
                            "relation_id": "turnover_rel_001",
                            "status": "suggested",
                            "status_label": "待人工确认",
                            "borrow_amount": "200000.00",
                            "borrow_date": "2026-02-04",
                            "repayment_amount": "100000.00",
                            "repayment_date": "2026-03-05",
                            "counterparty_bank_name": "建行 8106",
                            "repayment_remark": "还款",
                            "interest_rate_type": "annual",
                            "interest_rate_value": "0.060000",
                            "interest_paid_amount": "120.50",
                            "loan_days": 29,
                            "accrued_interest": "953.42",
                            "interest_paid_date": "2026-04-01",
                            "interest_payment_method": "银行转账",
                            "note": "页面备注",
                            "bank_row_ids": ["bank_001", "bank_002"],
                            "row_tone": "warning",
                        }
                    ],
                },
                {
                    "group_id": "counterparty:business:昆明建设集团",
                    "counterparty_name": "昆明建设集团",
                    "family": "business",
                    "family_label": "业务往来",
                    "pending_direction": "collection",
                    "pending_amount": "5000.00",
                    "rows": [
                        {
                            "relation_id": "turnover_rel_002",
                            "status": "suggested",
                            "status_label": "待人工确认",
                            "borrow_amount": "5000.00",
                            "borrow_date": "2026-03-06",
                            "repayment_amount": "0.00",
                            "repayment_date": None,
                            "counterparty_bank_name": "交行 3847",
                            "repayment_remark": "质保金",
                            "interest_rate_type": "none",
                            "interest_rate_value": "0.000000",
                            "interest_paid_amount": "0.00",
                            "loan_days": 0,
                            "accrued_interest": "0.00",
                            "interest_paid_date": None,
                            "interest_payment_method": "",
                            "note": "",
                        }
                    ],
                },
            ],
        }

    def test_preview_flattens_grouped_payload_to_summary_and_real_flow_rows(self) -> None:
        service = TurnoverLedgerExportService(lambda **_: self._grouped_payload())

        payload = service.preview(family="company")

        self.assertEqual(payload["filters"]["family"], "company")
        self.assertEqual(payload["columns"][:8], ["序号", "行类型", "源银行流水ID", "流水方向", "流水金额", "往来大类", "对方户名", "待还款金额"])
        self.assertEqual(len(payload["rows"]), 4)
        row = payload["rows"][0]
        flow_row = payload["rows"][1]
        self.assertEqual([item["row_type"] for item in payload["rows"]], ["summary", "flow", "flow", "flow"])
        self.assertEqual(row["行类型"], "合计")
        self.assertEqual(row["row_type"], "summary")
        self.assertEqual(row["source_bank_row_id"], "")
        self.assertEqual(row["flow_direction"], "")
        self.assertEqual(row["flow_amount"], "0.00")
        self.assertEqual(row["lot_id"], "")
        self.assertEqual(row["balance_amount"], "100000.00")
        self.assertEqual(flow_row["行类型"], "真实流水")
        self.assertEqual(flow_row["源银行流水ID"], "bank_001")
        self.assertEqual(flow_row["流水方向"], "income")
        self.assertEqual(flow_row["流水金额"], "120000.00")
        self.assertEqual(flow_row["source_bank_row_id"], "bank_001")
        self.assertEqual(flow_row["flow_direction"], "income")
        self.assertEqual(flow_row["flow_amount"], "120000.00")
        self.assertEqual(flow_row["lot_id"], "")
        self.assertEqual(flow_row["余额"], "100000.00")
        self.assertEqual(flow_row["balance_amount"], "100000.00")
        self.assertEqual(row["往来大类"], "公司往来")
        self.assertEqual(row["对方户名"], "梁希涛")
        self.assertEqual(row["待还款金额"], "100000.00")
        self.assertEqual(row["待收款金额"], "0.00")
        self.assertEqual(row["关系状态"], "待人工确认")
        self.assertEqual(
            [item["source_bank_row_id"] for item in payload["rows"] if item["row_type"] == "flow"],
            ["bank_001", "bank_002", "bank_003"],
        )
        self.assertNotIn("row_tone", row)
        self.assertNotIn("bank_row_ids", row)
        self.assertNotIn("allocation_status", flow_row)
        self.assertNotIn("allocated_lot_ids", flow_row)

    def test_preview_does_not_fallback_to_lot_rows_as_flow_rows(self) -> None:
        payload = self._grouped_payload()
        company_group = payload["groups"][0]
        company_group.pop("flow_rows")
        service = TurnoverLedgerExportService(lambda **_: payload)

        preview = service.preview(family="company")

        self.assertEqual([row["row_type"] for row in preview["rows"]], ["summary"])

    def test_preview_applies_limit_after_summary_and_flow_flattening_and_reports_totals(self) -> None:
        service = TurnoverLedgerExportService(lambda **_: self._grouped_payload())

        preview = service.preview(family="company", limit=2)

        self.assertEqual([row["row_type"] for row in preview["rows"]], ["summary", "flow"])
        self.assertEqual(preview["pagination"], {"preview_count": 2, "total": 4, "limit": 2})
        self.assertEqual(preview["totals"]["row_count"], 4)
        self.assertEqual(preview["totals"]["pending_repayment_amount"], "100000.00")
        self.assertEqual(preview["totals"]["pending_collection_amount"], "0.00")

    def test_preview_empty_grouped_payload_keeps_current_empty_shape(self) -> None:
        service = TurnoverLedgerExportService(
            lambda **_: {"summary": {}, "family_summaries": [], "filters": {}, "pagination": {"total": 0}, "groups": []}
        )

        preview = service.preview(family="all", limit=20)

        self.assertEqual(preview["rows"], [])
        self.assertEqual(preview["totals"]["row_count"], 0)
        self.assertEqual(preview["pagination"], {"preview_count": 0, "total": 0, "limit": 20})
        self.assertEqual(preview["filters"], {"family": "all"})

    def test_export_builds_xlsx_and_filename_for_family_scope(self) -> None:
        service = TurnoverLedgerExportService(lambda **_: self._grouped_payload())

        filename, content = service.export(family="business", today=date(2026, 5, 12))
        workbook = load_workbook(BytesIO(content))
        sheet = workbook.active

        self.assertEqual(filename, "往来款台账-业务往来-2026-05-12.xlsx")
        self.assertEqual(sheet.cell(row=1, column=1).value, "序号")
        self.assertEqual(sheet.cell(row=1, column=2).value, "行类型")
        self.assertEqual(sheet.cell(row=1, column=3).value, "源银行流水ID")
        self.assertEqual(sheet.cell(row=1, column=10).value, "余额")
        self.assertEqual(sheet.cell(row=2, column=2).value, "合计")
        self.assertEqual(sheet.cell(row=2, column=6).value, "业务往来")
        self.assertEqual(sheet.cell(row=2, column=7).value, "昆明建设集团")
        self.assertEqual(sheet.max_row, 2)


if __name__ == "__main__":
    unittest.main()
