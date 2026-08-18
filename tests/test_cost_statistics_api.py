from __future__ import annotations

import json
import unittest
from dataclasses import replace
from time import monotonic

from openpyxl import load_workbook

from fin_ops_platform.domain.enums import BatchType
from fin_ops_platform.services.oa_adapter import (
    InMemoryOAAdapter,
    OAApplicationRecord,
)
from fin_ops_platform.services.workbench_query_service import WorkbenchQueryService
from tests.app_test_support import build_local_state_application


class CostStatisticsApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.app = build_local_state_application()
        oa = OAApplicationRecord(
            id="oa-cost-api-001",
            month="2026-03",
            section="unpaired",
            case_id=None,
            applicant="刘际涛",
            project_name="云南溯源科技",
            apply_type="支付申请",
            amount="1250.00",
            counterparty_name="昆明设备供应商",
            reason="PLC 模块采购",
            relation_code="pending_match",
            relation_label="待关联",
            relation_tone="warn",
            workflow_status="completed",
            completed_at="2026-03-12 09:30:00",
            expense_type="设备货款及材料费",
            expense_content="PLC 模块采购",
            detail_fields={
                "项目编号": "P-001",
                "费用类型": "设备货款及材料费",
                "费用内容": "PLC 模块采购",
            },
        )
        self.app._workbench_query_service = WorkbenchQueryService(  # noqa: SLF001
            oa_adapter=InMemoryOAAdapter({"2026-03": [oa]})
        )
        self.oa = oa
        preview = self.app._import_service.preview_import(  # noqa: SLF001
            batch_type=BatchType.BANK_TRANSACTION,
            source_name="cost-statistics-direct-canonical.json",
            imported_by="cost-statistics-test",
            rows=[
                {
                    "account_no": "62228888",
                    "txn_date": "2026-03-10",
                    "trade_time": "2026-03-10 21:27:55",
                    "counterparty_name": "昆明设备供应商",
                    "debit_amount": "1250.00",
                    "credit_amount": "",
                    "bank_serial_no": "COST-DIRECT-001",
                    "summary": "PLC 模块采购",
                    "remark": "设备采购款",
                }
            ],
        )
        self.app._import_service.confirm_import(preview.id)  # noqa: SLF001
        self.bank_id = self.app._import_service.list_transactions()[-1].id  # noqa: SLF001
        self.app._workbench_pair_relation_service.create_active_relation(  # noqa: SLF001
            case_id="CASE-COST-DIRECT-001",
            row_ids=["oa-cost-api-001", self.bank_id],
            row_types=["oa", "bank"],
            relation_mode="manual_confirmed",
            created_by="cost-statistics-test",
            month_scope="2026-03",
            note="direct canonical Cost fixture",
            special_metadata={},
        )

    def _get(self, path: str):
        return self.app.handle_request("GET", path)

    def _json(self, path: str) -> tuple[int, dict[str, object]]:
        response = self._get(path)
        return response.status_code, json.loads(response.body)

    def test_all_explorer_views_read_current_canonical_facts(self) -> None:
        status, time_payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=time"
        )
        self.assertEqual(status, 200)
        time_row = time_payload["rows"][0]
        self.assertEqual(time_row["counterparty_name"], "昆明设备供应商")
        self.assertEqual(time_row["oa_applicant"], "")
        paths = {
            "time": "/api/cost-statistics/explorer?scope=2026-03&view=time",
            "project": (
                "/api/cost-statistics/explorer?scope=2026-03&view=project"
                "&project_name=云南溯源科技"
                "&expense_type=设备货款及材料费"
            ),
            "bank": (
                "/api/cost-statistics/explorer?scope=2026-03&view=bank"
                f"&payment_account_label={time_row['payment_account_label']}"
                "&project_name=云南溯源科技"
            ),
            "expense_type": (
                "/api/cost-statistics/explorer?scope=2026-03&view=expense_type"
                "&expense_type=设备货款及材料费"
            ),
            "bank_tag": (
                "/api/cost-statistics/explorer?scope=2026-03&view=bank_tag"
                f"&bank_tag_primary_label={time_row['bank_tag_primary_label']}"
                f"&bank_tag_sub_label={time_row['bank_tag_sub_label']}"
            ),
        }
        for view, path in paths.items():
            with self.subTest(view=view):
                status, payload = self._json(path)
                self.assertEqual(status, 200)
                self.assertGreaterEqual(payload["row_count"], 1)
                self.assertEqual(payload["summary"]["total_amount"], "1250.00")
                self.assertNotIn("read_model_status", payload)
                self.assertNotIn("read_model_version", payload)
                self.assertNotIn("refresh_scope_keys", payload)

    def test_project_view_and_detail_use_the_same_active_relation(self) -> None:
        status, page = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=云南溯源科技"
            "&expense_type=设备货款及材料费"
        )
        self.assertEqual(status, 200)
        self.assertEqual(page["rows"][0]["project_name"], "云南溯源科技")
        self.assertEqual(page["rows"][0]["oa_applicant"], "刘际涛")
        self.assertEqual(page["rows"][0]["counterparty_name"], "昆明设备供应商")
        self.assertEqual(
            page["rows"][0]["expense_type"],
            "设备货款及材料费",
        )

        allocation_id = page["rows"][0]["allocation_id"]
        status, detail = self._json(
            f"/api/cost-statistics/allocations/{allocation_id}"
            "?scope=2026-03&view=project"
        )
        self.assertEqual(status, 200)
        self.assertEqual(detail["kind"], "oa_allocation")
        self.assertEqual(detail["allocation"]["project_name"], "云南溯源科技")
        self.assertEqual(detail["allocation"]["amount"], "1250.00")
        self.assertEqual(detail["allocation"]["oa_original_amount"], "1250.00")
        self.assertEqual(detail["allocation"]["oa_allocation_weight"], "100.00%")
        self.assertEqual(detail["allocation"]["bank_event_amount"], "1250.00")
        self.assertEqual(detail["allocation"]["oa_applicant"], "刘际涛")
        self.assertEqual(detail["payment_evidence"][0]["transaction_id"], self.bank_id)
        self.assertEqual(detail["payment_evidence"][0]["direction"], "支出")
        self.assertEqual(detail["reconciliation"]["relation_case_id"], "CASE-COST-DIRECT-001")
        self.assertEqual(detail["reconciliation"]["net_cash_cost"], "1250.00")
        self.assertEqual(detail["reconciliation"]["cash_payment_ratio"], "100.00%")

    def test_transaction_detail_loads_only_the_requested_scope(self) -> None:
        repository = self.app._cost_statistics_canonical_repository  # noqa: SLF001
        original_load_snapshot = repository.load_snapshot
        calls: list[dict[str, object]] = []

        def load_snapshot(**kwargs):
            calls.append(dict(kwargs))
            return original_load_snapshot(**kwargs)

        repository.load_snapshot = load_snapshot

        status, _detail = self._json(
            f"/api/cost-statistics/bank-transactions/{self.bank_id}"
            "?scope=2026-03&view=time"
        )

        self.assertEqual(status, 200)
        self.assertEqual(
            calls,
            [
                {
                    "scope_kind": "month",
                    "scope_value": "2026-03",
                    "view": "time",
                    "include_statistics": False,
                }
            ],
        )

    def test_detail_endpoints_are_explicit_and_old_route_stays_removed(self) -> None:
        old_response = self._get(
            f"/api/cost-statistics/transactions/{self.bank_id}"
            "?scope=2026-03&view=time"
        )
        self.assertEqual(old_response.status_code, 404)

        status, payload = self._json(
            f"/api/cost-statistics/bank-transactions/{self.bank_id}"
            "?scope=2026-03&view=project"
        )
        self.assertEqual(status, 200)
        self.assertEqual(payload["kind"], "bank_transaction")

    def test_duplicate_oa_allocation_across_active_relations_returns_conflict(
        self,
    ) -> None:
        repository = self.app._cost_statistics_canonical_repository  # noqa: SLF001
        original_relations = repository._relations_provider()  # noqa: SLF001
        duplicate = {
            **original_relations[0],
            "case_id": "CASE-COST-DIRECT-DUPLICATE",
        }
        repository._relations_provider = (  # noqa: SLF001
            lambda: [*original_relations, duplicate]
        )

        status, payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
        )

        self.assertEqual(status, 409)
        self.assertEqual(
            payload["error"],
            "cost_statistics_allocation_integrity_conflict",
        )

    def test_oa_cost_views_exclude_in_progress_oa(self) -> None:
        in_progress_oa = replace(self.oa, workflow_status="in_progress")
        self.app._cost_statistics_canonical_repository._oa_rows_by_ids_provider = (  # noqa: SLF001
            lambda _row_ids: [in_progress_oa]
        )

        for view in ("project", "bank", "expense_type"):
            with self.subTest(view=view):
                status, payload = self._json(
                    "/api/cost-statistics/explorer"
                    f"?scope=2026-03&view={view}"
                )
                self.assertEqual(status, 200)
                self.assertEqual(payload["summary"]["total_amount"], "0.00")
                self.assertEqual(payload["rows"], [])
                self.assertEqual(payload["row_count"], 0)

        time_status, time_payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=time"
        )
        self.assertEqual(time_status, 200)
        self.assertEqual(time_payload["row_count"], 1)
        self.assertEqual(time_payload["rows"][0]["transaction_id"], self.bank_id)

    def test_daily_reimbursement_items_drive_views_detail_and_export(self) -> None:
        daily_oa = replace(
            self.oa,
            applicant="报销成员甲",
            apply_type="日常报销",
            project_name="项目A；项目B",
            expense_type="交通费；住宿费",
            expense_content="市内交通；出差住宿",
            expense_items=[
                {
                    "expense_item_id": "api-item-1",
                    "project_id": "P-A",
                    "project_name": "项目A",
                    "expense_type": "交通费",
                    "expense_content": "市内交通",
                    "amount": "500.00",
                },
                {
                    "expense_item_id": "api-item-2",
                    "project_id": "P-B",
                    "project_name": "项目B",
                    "expense_type": "住宿费",
                    "expense_content": "出差住宿",
                    "amount": "750.00",
                },
            ],
        )
        self.app._cost_statistics_canonical_repository._oa_rows_by_ids_provider = (  # noqa: SLF001
            lambda _row_ids: [daily_oa]
        )

        status, project_page = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=项目A&expense_type=交通费"
        )
        self.assertEqual(status, 200)
        self.assertEqual(project_page["row_count"], 1)
        self.assertEqual(project_page["rows"][0]["amount"], "500.00")
        self.assertEqual(project_page["rows"][0]["oa_applicant"], "报销成员甲")
        self.assertNotIn(
            "多项目",
            {
                facet["project_name"]
                for facet in project_page["facets"]["projects"]
            },
        )

        allocation_id = project_page["rows"][0]["allocation_id"]
        status, detail = self._json(
            f"/api/cost-statistics/allocations/{allocation_id}"
            "?scope=2026-03&view=project"
        )
        self.assertEqual(status, 200)
        self.assertEqual(detail["allocation"]["project_name"], "项目A")
        self.assertEqual(detail["allocation"]["amount"], "500.00")
        self.assertEqual(detail["allocation"]["oa_applicant"], "报销成员甲")
        self.assertEqual(detail["payment_evidence"][0]["amount"], "1250.00")

        status, preview = self._json(
            "/api/cost-statistics/export-preview"
            "?month=2026-03&view=project"
            "&project_name=项目B"
        )
        self.assertEqual(status, 200)
        self.assertEqual(preview["summary"]["total_amount"], "750.00")
        applicant_column = preview["columns"].index("申请/报销人")
        self.assertEqual(preview["rows"][0][applicant_column], "报销成员甲")

    def test_missing_oa_applicant_stays_empty_without_fake_fallback(self) -> None:
        oa_without_applicant = replace(
            self.oa,
            applicant="",
            detail_fields={
                "项目编号": "P-001",
                "费用类型": "设备货款及材料费",
                "费用内容": "PLC 模块采购",
            },
        )
        self.app._cost_statistics_canonical_repository._oa_rows_by_ids_provider = (  # noqa: SLF001
            lambda _row_ids: [oa_without_applicant]
        )

        status, payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=云南溯源科技"
            "&expense_type=设备货款及材料费"
        )

        self.assertEqual(status, 200)
        self.assertEqual(payload["rows"][0]["oa_applicant"], "")

    def test_next_request_observes_relation_withdrawal_without_refresh_job(self) -> None:
        before_status, before = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=云南溯源科技"
            "&expense_type=设备货款及材料费"
        )
        self.assertEqual(before_status, 200)
        self.assertEqual(before["row_count"], 1)

        self.app._workbench_pair_relation_service.cancel_active_relations_for_row_ids(  # noqa: SLF001
            [self.bank_id],
            created_by="cost-statistics-test",
        )

        after_status, after = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=云南溯源科技"
            "&expense_type=设备货款及材料费"
        )
        self.assertEqual(after_status, 200)
        self.assertEqual(after["row_count"], 0)
        self.assertEqual(after["rows"], [])

        time_status, time_payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=time"
        )
        self.assertEqual(time_status, 200)
        self.assertEqual(time_payload["row_count"], 1)
        self.assertEqual(time_payload["rows"][0]["transaction_id"], self.bank_id)

    def test_export_preview_and_workbook_use_canonical_snapshot(self) -> None:
        status, preview = self._json(
            "/api/cost-statistics/export-preview"
            "?month=2026-03&view=project"
            "&project_name=云南溯源科技"
        )
        self.assertEqual(status, 200)
        self.assertEqual(preview["summary"]["row_count"], 1)
        applicant_column = preview["columns"].index("申请/报销人")
        self.assertEqual(preview["rows"][0][applicant_column], "刘际涛")

        response = self._get(
            "/api/cost-statistics/export"
            "?month=2026-03&view=project"
            "&project_name=云南溯源科技"
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["Content-Type"],
        )
        workbook = load_workbook(filename=__import__("io").BytesIO(response.body))
        self.assertIn("成本明细", workbook.sheetnames)
        detail_sheet = workbook["成本明细"]
        self.assertEqual(detail_sheet["D1"].value, "申请/报销人")
        self.assertEqual(detail_sheet["D2"].value, "刘际涛")

    def test_each_api_request_loads_exactly_one_snapshot(self) -> None:
        repository = self.app._cost_statistics_canonical_repository  # noqa: SLF001
        original = repository.load_snapshot
        calls = 0

        def counted_snapshot(*args, **kwargs):
            nonlocal calls
            calls += 1
            return original(*args, **kwargs)

        repository.load_snapshot = counted_snapshot
        status, _payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=云南溯源科技"
            "&expense_type=设备货款及材料费"
        )
        self.assertEqual(status, 200)
        self.assertEqual(calls, 1)

    def test_follow_up_request_can_skip_rebuilding_global_statistics(self) -> None:
        status, payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=云南溯源科技"
            "&expense_type=设备货款及材料费&include_statistics=false"
        )

        self.assertEqual(status, 200)
        self.assertIsNone(payload["statistics"])
        self.assertEqual(payload["row_count"], 1)

    def test_search_uses_current_view_canonical_fields_before_aggregation(
        self,
    ) -> None:
        status, time_payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=time"
            "&query=昆明设备"
        )
        self.assertEqual(status, 200)
        self.assertEqual(time_payload["row_count"], 1)
        self.assertEqual(time_payload["summary"]["total_amount"], "1250.00")
        self.assertEqual(time_payload["rows"][0]["project_name"], "")
        self.assertEqual(time_payload["rows"][0]["expense_type"], "")

        status, expense_payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=expense_type"
            "&query=PLC"
        )
        self.assertEqual(status, 200)
        self.assertEqual(
            [
                row["expense_type"]
                for row in expense_payload["facets"]["expense_types"]
            ],
            ["设备货款及材料费"],
        )

        status, empty_payload = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=expense_type"
            "&query=不存在的内容"
        )
        self.assertEqual(status, 200)
        self.assertEqual(empty_payload["summary"]["total_amount"], "0.00")
        self.assertEqual(empty_payload["facets"]["expense_types"], [])

    def test_invalid_query_contracts_fail_closed(self) -> None:
        cases = (
            (
                "/api/cost-statistics/explorer?scope=bad&view=time",
                "invalid_cost_statistics_query",
            ),
            (
                "/api/cost-statistics/explorer?scope=2026-03&view=bad",
                "invalid_cost_statistics_query",
            ),
            (
                "/api/cost-statistics/explorer"
                "?scope=2026-03&view=time&include_statistics=maybe",
                "invalid_cost_statistics_query",
            ),
            (
                "/api/cost-statistics/explorer?scope=2026-03&view=time"
                f"&query={'x' * 201}",
                "invalid_cost_statistics_query",
            ),
        )
        for path, error_code in cases:
            with self.subTest(path=path):
                status, payload = self._json(path)
                self.assertEqual(status, 400)
                self.assertEqual(payload["error"], error_code)

    def test_time_tag_rules_default_to_all_and_custom_empty_filters_raw_bank_views(self) -> None:
        status, rules = self._json("/api/cost-statistics/time-tag-rules")
        self.assertEqual(status, 200)
        self.assertEqual(rules["mode"], "all")
        self.assertEqual(rules["selected_tag_codes"], [])
        self.assertTrue(rules["available_tags"])

        response = self.app.handle_request(
            "PUT",
            "/api/cost-statistics/time-tag-rules",
            body=json.dumps(
                {
                    "expected_version": rules["version"],
                    "mode": "custom",
                    "selected_tag_codes": [],
                }
            ),
        )
        self.assertEqual(response.status_code, 200)
        saved = json.loads(response.body)
        self.assertEqual(saved["mode"], "custom")

        status, time_page = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=time"
        )
        self.assertEqual(status, 200)
        self.assertEqual(time_page["row_count"], 0)

        legacy = self.app.handle_request("GET", "/api/cost-statistics/tag-rules")
        self.assertEqual(legacy.status_code, 404)

    def test_no_oa_scope_defaults_empty_and_includes_only_unpaired_rows_after_save(self) -> None:
        preview = self.app._import_service.preview_import(  # noqa: SLF001
            batch_type=BatchType.BANK_TRANSACTION,
            source_name="cost-statistics-no-oa.json",
            imported_by="cost-statistics-test",
            rows=[
                {
                    "account_no": "62228888",
                    "txn_date": "2026-03-11",
                    "trade_time": "2026-03-11 08:00:00",
                    "counterparty_name": "银行",
                    "debit_amount": "8.00",
                    "credit_amount": "",
                    "bank_serial_no": "COST-NO-OA-001",
                    "summary": "手续费",
                    "remark": "手续费",
                }
            ],
        )
        self.app._import_service.confirm_import(preview.id)  # noqa: SLF001
        unpaired_bank_id = self.app._import_service.list_transactions()[-1].id  # noqa: SLF001

        status, rules = self._json("/api/cost-statistics/no-oa-rules")
        self.assertEqual(status, 200)
        self.assertEqual(rules["projects"], [])
        self.assertEqual(len(rules["available_tags"]), 1)
        candidate_code = rules["available_tags"][0]["code"]
        self.assertEqual(rules["available_tags"][0]["label"], "手续费")

        response = self.app.handle_request(
            "PUT",
            "/api/cost-statistics/no-oa-rules",
            body=json.dumps(
                {
                    "expected_version": rules["version"],
                    "projects": [
                        {
                            "id": "yunnan-no-oa",
                            "display_name": "云南溯源无 OA 分类",
                            "tag_codes": [candidate_code],
                        }
                    ],
                },
                ensure_ascii=False,
            ),
        )
        self.assertEqual(response.status_code, 200)
        saved = json.loads(response.body)
        self.assertEqual(saved["projects"][0]["display_name"], "云南溯源无 OA 分类")

        duplicate_response = self.app.handle_request(
            "PUT",
            "/api/cost-statistics/no-oa-rules",
            body=json.dumps(
                {
                    "expected_version": saved["version"],
                    "projects": [
                        {
                            "id": "project-a",
                            "display_name": "项目 A",
                            "tag_codes": [candidate_code],
                        },
                        {
                            "id": "project-b",
                            "display_name": "项目 B",
                            "tag_codes": [candidate_code],
                        },
                    ],
                },
                ensure_ascii=False,
            ),
        )
        self.assertEqual(duplicate_response.status_code, 400)
        self.assertEqual(
            json.loads(duplicate_response.body)["error"],
            "duplicate_cost_statistics_no_oa_tag_assignment",
        )

        status, project = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=project"
            "&project_name=云南溯源无%20OA%20分类"
            "&expense_type=无%20OA%20分类"
        )
        self.assertEqual(status, 200)
        self.assertEqual(project["row_count"], 1)
        self.assertEqual(project["rows"][0]["transaction_id"], unpaired_bank_id)
        self.assertEqual(project["rows"][0]["row_kind"], "bank_transaction")
        self.assertEqual(project["rows"][0]["amount"], "8.00")
        self.assertEqual(project["summary"]["total_amount"], "1258.00")

        status, time_page = self._json(
            "/api/cost-statistics/explorer?scope=2026-03&view=time"
        )
        self.assertEqual(status, 200)
        self.assertEqual(
            {row["transaction_id"] for row in time_page["rows"]},
            {self.bank_id, unpaired_bank_id},
        )
        self.assertEqual(time_page["summary"]["total_amount"], "1258.00")

    def test_local_direct_read_stays_within_regression_budget(self) -> None:
        samples: list[float] = []
        for _index in range(5):
            started = monotonic()
            status, payload = self._json(
                "/api/cost-statistics/explorer?scope=2026-03&view=project"
                "&project_name=云南溯源科技"
                "&expense_type=设备货款及材料费"
            )
            samples.append(monotonic() - started)
            self.assertEqual(status, 200)
            self.assertEqual(payload["row_count"], 1)
        self.assertLess(max(samples), 0.5)


if __name__ == "__main__":
    unittest.main()
