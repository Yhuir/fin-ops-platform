import json
import pickle
import tempfile
import unittest
from contextlib import contextmanager
from http import HTTPStatus
from io import BytesIO
from unittest.mock import patch
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from pymongo.errors import ServerSelectionTimeoutError
from openpyxl import load_workbook

from fin_ops_platform.app.http_adapter import WsgiHttpAdapter
from fin_ops_platform.app.server import Application
from tests.app_test_support import build_local_state_application as build_application
from fin_ops_platform.services.bank_details_export_service import BANK_DETAIL_EXPORT_ROW_LIMIT
from fin_ops_platform.domain.enums import BatchType
from fin_ops_platform.services.oa_identity_service import OAUserIdentity
from fin_ops_platform.services.mongo_oa_adapter import MongoOAAdapter, MongoOASettings
from fin_ops_platform.services.object_identity_policy import FinancialObjectIdentityPolicy
from fin_ops_platform.services.oa_adapter import InMemoryOAAdapter, OAApplicationRecord
from fin_ops_platform.services.settings_data_reset_service import RESET_OA_AND_REBUILD_ACTION
from fin_ops_platform.services.etc_service import UploadedEtcZipFile
from fin_ops_platform.services.workbench_query_service import WorkbenchQueryService
from tests.test_etc_backend import etc_zip
from tests.mock_import_files import INVOICE_JAN

class FailingMongoWorkbenchOAAdapter(MongoOAAdapter):
    def __init__(self) -> None:
        super().__init__(settings=MongoOASettings(host="127.0.0.1", database="form_data_db"))

    def _collection(self):
        raise ServerSelectionTimeoutError("mock mongo unavailable")

class MemoryAttachmentInvoiceCache:
    def __init__(self) -> None:
        self.entries: dict[str, dict[str, object]] = {}

    def load_oa_attachment_invoice_cache_entry(self, cache_key: str) -> dict[str, object] | None:
        return self.entries.get(cache_key)

    def save_oa_attachment_invoice_cache_entry(self, cache_key: str, payload: dict[str, object]) -> None:
        self.entries[cache_key] = dict(payload)

def workbench_operation_targets(scope_key: str) -> list[dict[str, str]]:
    return [
        {"read_model_key": "workbench", "scope_key": scope_key},
        {"read_model_key": "workbench_relation", "scope_key": scope_key},
    ]

class StaticMongoWorkbenchOAAdapter(MongoOAAdapter):
    def __init__(
        self,
        *,
        form_documents: dict[str, list[dict]],
        project_documents: list[dict] | None = None,
        attachment_invoice_cache: MemoryAttachmentInvoiceCache | None = None,
    ) -> None:
        super().__init__(
            settings=MongoOASettings(host="127.0.0.1", database="form_data_db"),
            attachment_invoice_cache=attachment_invoice_cache,
        )
        self._form_documents = form_documents
        self._project_documents = project_documents or []

    def _load_form_documents(self, form_id: str, month: str | None = None) -> list[dict]:
        documents = [self._with_default_completed_status(document) for document in self._form_documents.get(str(form_id), [])]
        if month is None:
            return documents
        filtered: list[dict] = []
        for document in documents:
            data = document.get("data", {})
            application_date = str(data.get("applicationDate") or data.get("ApplicationDate") or "")
            if application_date.startswith(month):
                filtered.append(document)
        return filtered

    def _load_project_documents(self) -> list[dict]:
        return list(self._project_documents)

    def _load_form_month_documents(self, form_id: str) -> list[dict]:
        return [self._with_default_completed_status(document) for document in self._form_documents.get(str(form_id), [])]

    def _load_form_documents_by_external_ids(self, form_id: str, external_ids: set[str]) -> list[dict]:
        documents = [self._with_default_completed_status(document) for document in self._form_documents.get(str(form_id), [])]
        normalized_external_ids = {str(external_id).strip() for external_id in external_ids if str(external_id).strip()}
        return [
            document
            for document in documents
            if self._document_external_id(form_id, document) in normalized_external_ids
        ]

    @staticmethod
    def _with_default_completed_status(document: dict) -> dict:
        normalized = dict(document)
        data = dict(normalized.get("data", {}))
        if "status" not in data or data.get("status") in (None, ""):
            data["status"] = "已完成"
        normalized["data"] = data
        return normalized

class RetentionScopedMongoWorkbenchOAAdapter(StaticMongoWorkbenchOAAdapter):
    def __init__(
        self,
        *,
        form_documents: dict[str, list[dict]],
        project_documents: list[dict] | None = None,
        attachment_invoice_cache: MemoryAttachmentInvoiceCache | None = None,
        row_id_records: dict[str, list[OAApplicationRecord]] | None = None,
    ) -> None:
        super().__init__(
            form_documents=form_documents,
            project_documents=project_documents,
            attachment_invoice_cache=attachment_invoice_cache,
        )
        self.month_calls: list[str] = []
        self.bulk_call_count = 0
        self.row_id_calls: list[list[str]] = []
        self._row_id_records = row_id_records or {}

    def list_available_months(self) -> list[str]:
        months: set[str] = set()
        for documents in self._form_documents.values():
            for document in documents:
                data = self._with_default_completed_status(document).get("data", {})
                application_date = str(data.get("applicationDate") or data.get("ApplicationDate") or "")
                if len(application_date) >= 7:
                    months.add(application_date[:7])
        return sorted(months)

    def list_application_records(self, month: str) -> list[OAApplicationRecord]:
        self.month_calls.append(month)
        return super().list_application_records(month)

    def list_all_application_records(self) -> list[OAApplicationRecord]:
        self.bulk_call_count += 1
        raise AssertionError("should not bulk scan all OA records")

    def list_application_records_by_row_ids(self, row_ids: list[str]) -> list[OAApplicationRecord]:
        normalized = [str(row_id) for row_id in row_ids]
        self.row_id_calls.append(normalized)
        records: list[OAApplicationRecord] = []
        for row_id in normalized:
            records.extend(self._row_id_records.get(row_id, []))
        return records

class ErrorMonthListRetentionMongoWorkbenchOAAdapter(RetentionScopedMongoWorkbenchOAAdapter):
    def list_available_months(self) -> list[str]:
        self._set_read_status("error", "OA 连接失败")
        return []

class MutatingRecordDict(dict):
    def values(self):
        values_iterator = super().values()
        did_mutate = False
        for row in values_iterator:
            if not did_mutate:
                self["oa-mutated-during-iteration"] = {
                    "id": "oa-mutated-during-iteration",
                    "type": "oa",
                    "_month": "2026-01",
                    "_section": "unpaired",
                    "oa_bank_relation": {"code": "pending_match", "label": "待找流水与发票", "tone": "warn"},
                }
                did_mutate = True
            yield row

class RelationDistributionFacadeStub:
    def __init__(self, rows: list[dict[str, object]] | None = None) -> None:
        self.rows = list(rows or [])
        self.calls: list[dict[str, object]] = []

    def set_rows(self, rows: list[dict[str, object]]) -> None:
        self.rows = list(rows)

    def get_by_row_ids(self, row_ids: list[str], **kwargs: object) -> dict[str, object]:
        wanted = {str(row_id) for row_id in row_ids}
        self.calls.append({"row_ids": list(row_ids), "kwargs": dict(kwargs)})
        return {
            "status": "fresh",
            "rows": [dict(row) for row in self.rows if str(row.get("row_id") or "") in wanted],
            "groups": [],
            "source_versions": {"schema_version": "test"},
            "read_model_scope_keys": ["2026-03", "2026-04", "2026-05"],
            "refresh_enqueued": False,
            "stale_reasons": [],
        }

class BankDetailCanonicalQueryFixture:
    def __init__(self, app: Application) -> None:
        self._app = app

    def bank_detail_scope_keys_for_range(self, *, date_from: str | None = None, date_to: str | None = None) -> list[str]:
        months: set[str] = set()
        start_month = str(date_from or "")[:7]
        end_month = str(date_to or "")[:7]
        for transaction in self._app._import_service.list_transactions(month="all"):
            month = str(getattr(transaction, "txn_date", "") or getattr(transaction, "trade_time", "") or "")[:7]
            if len(month) != 7:
                continue
            if start_month and month < start_month:
                continue
            if end_month and month > end_month:
                continue
            months.add(month)
        return sorted(months) or ["all"]

    def list_bank_detail_transactions(self, **kwargs: object) -> dict[str, object]:
        return self._app._bank_details_service.list_transactions(**kwargs)

    def list_bank_detail_accounts(self, *, date_from: str | None = None, date_to: str | None = None) -> dict[str, object]:
        return self._app._bank_details_service.list_accounts(date_from=date_from, date_to=date_to)

    def accounts_payload(
        self,
        *,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> dict[str, object]:
        return self.list_bank_detail_accounts(date_from=date_from, date_to=date_to)

    def transactions_payload(self, **kwargs: object) -> dict[str, object]:
        return self.list_bank_detail_transactions(**kwargs)

    def export_payload(
        self,
        *,
        include_accounts: bool,
        **kwargs: object,
    ) -> dict[str, object]:
        transaction_kwargs = {
            **kwargs,
            "page": 1,
            "page_size": BANK_DETAIL_EXPORT_ROW_LIMIT + 1,
        }
        return {
            "transactions": self.list_bank_detail_transactions(**transaction_kwargs),
            "accounts": (
                self.list_bank_detail_accounts(
                    date_from=kwargs.get("date_from") if isinstance(kwargs.get("date_from"), str) else None,
                    date_to=kwargs.get("date_to") if isinstance(kwargs.get("date_to"), str) else None,
                )
                if include_accounts
                else None
            ),
        }

class WorkbenchV2ApiTests(unittest.TestCase):
    def _install_workbench_query_service(self, app: Application, query_service: WorkbenchQueryService) -> None:
        app._workbench_query_service = query_service

    def _create_imported_bank_transaction(
        self,
        app: Application,
        *,
        trade_time: str = "2026-04-03 09:00:00",
        counterparty_name: str = "供应商A",
        summary: str = "付款",
        remark: str = "货款",
    ) -> str:
        preview = app._import_service.preview_import(
            batch_type=BatchType.BANK_TRANSACTION,
            source_name="bank.xlsx",
            imported_by="YNSYLP005",
            rows=[
                {
                    "account_no": "6222000011116386",
                    "account_name": "云南溯源科技有限公司基本户",
                    "txn_date": trade_time[:10],
                    "trade_time": trade_time,
                    "pay_receive_time": trade_time,
                    "counterparty_name": counterparty_name,
                    "debit_amount": "100.00",
                    "credit_amount": "",
                    "summary": summary,
                    "remark": remark,
                    "selected_bank_name": "工商银行",
                    "selected_bank_last4": "6386",
                }
            ],
        )
        app._import_service.confirm_import(preview.id)
        return app._import_service.list_transactions()[-1].id

    def _install_bank_relation_distribution(
        self,
        app: Application,
        rows: list[dict[str, object]],
    ) -> RelationDistributionFacadeStub:
        facade = RelationDistributionFacadeStub(rows)
        app._bank_details_relation_tag_projection_service._relation_facade = facade
        return facade

    def _install_bank_detail_canonical_query_fixture(self, app: Application) -> BankDetailCanonicalQueryFixture:
        query_service = BankDetailCanonicalQueryFixture(app)
        application_service = app._bank_details_application_service()
        application_service._query_service = query_service
        app._bank_details_application_service = lambda: application_service
        return query_service

    def test_http_server_dispatches_patch_bank_transaction_categories(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            transaction_id = self._create_imported_bank_transaction(app)
            request_body = json.dumps(
                {
                    "updates": [
                        {
                            "transaction_id": transaction_id,
                            "category_code": "borrow_in_company_pending_repayment",
                            "expected_version": 0,
                        }
                    ]
                }
            ).encode("utf-8")
            response_status: list[str] = []
            response_headers: dict[str, str] = {}

            def start_response(status: str, headers: list[tuple[str, str]]) -> None:
                response_status.append(status)
                response_headers.update(headers)

            response_body = b"".join(
                WsgiHttpAdapter(app)(
                    {
                        "REQUEST_METHOD": "PATCH",
                        "PATH_INFO": "/api/bank-details/transactions/categories",
                        "QUERY_STRING": "",
                        "CONTENT_TYPE": "application/json",
                        "CONTENT_LENGTH": str(len(request_body)),
                        "wsgi.input": BytesIO(request_body),
                    },
                    start_response,
                )
            ).decode("utf-8")

        self.assertEqual(response_status, ["410 Gone"])
        self.assertEqual(response_headers["Content-Type"], "application/json; charset=utf-8")
        payload = json.loads(response_body)
        self.assertEqual(payload["error"], "manual_bank_transaction_category_disabled")

    def test_bank_details_api_returns_auto_and_effective_category_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            transaction_id = self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-03 09:15:30",
                summary="网银手续费",
                remark="转账手续费",
            )
            self._install_bank_detail_canonical_query_fixture(app)

            response = app.handle_request(
                "GET",
                "/api/bank-details/transactions?account_key=%E5%B7%A5%E5%95%86%E9%93%B6%E8%A1%8C%3A6386",
            )
            payload = json.loads(response.body)

        self.assertEqual(response.status_code, 200)
        row = next(row for row in payload["rows"] if row["id"] == transaction_id)
        self.assertEqual(row["trade_time"], "2026-04-03 09:15:30")
        self.assertEqual(row["auto_category_code"], "fee")
        self.assertEqual(row["auto_category_label"], "手续费")
        self.assertEqual(row["auto_category_primary_label"], "费用")
        self.assertEqual(row["auto_category_sub_label"], "手续费")
        self.assertEqual(row["auto_category_label_path"], ["费用", "手续费"])
        self.assertEqual(row["auto_category_source"], "auto")
        self.assertEqual(row["effective_category_code"], "fee")
        self.assertEqual(row["effective_category_label"], "手续费")
        self.assertEqual(row["effective_category_primary_label"], "费用")
        self.assertEqual(row["effective_category_sub_label"], "手续费")
        self.assertEqual(row["effective_category_label_path"], ["费用", "手续费"])
        self.assertEqual(row["effective_category_source"], "auto")
        self.assertEqual(row["category_code"], "fee")
        self.assertEqual(row["category_label"], "手续费")
        self.assertEqual(row["category_primary_label"], "费用")
        self.assertEqual(row["category_sub_label"], "手续费")
        self.assertEqual(row["category_label_path"], ["费用", "手续费"])
        self.assertEqual(row["category_source"], "auto")
        self.assertEqual(payload["category_counts"]["fee"], 1)

    def test_bank_details_api_filters_by_auto_category_primary_and_sub_label(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            fee_id = self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-03 09:15:30",
                summary="网银手续费",
                remark="转账手续费",
            )
            self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-02 09:15:30",
                summary="工资发放",
                remark="员工工资",
            )
            self._install_bank_detail_canonical_query_fixture(app)

            response = app.handle_request(
                "GET",
                "/api/bank-details/transactions"
                "?category_primary_label=%E8%B4%B9%E7%94%A8"
                "&category_sub_label=%E6%89%8B%E7%BB%AD%E8%B4%B9",
            )
            payload = json.loads(response.body)

        self.assertEqual(response.status_code, 200, response.body)
        self.assertEqual([row["id"] for row in payload["rows"]], [fee_id])
        self.assertEqual(payload["pagination"]["total"], 1)

    def test_bank_details_api_filters_uncategorized_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            uncategorized_id = self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-03 09:15:30",
                summary="普通付款",
                remark="普通用途",
            )
            self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-02 09:15:30",
                summary="网银手续费",
                remark="转账手续费",
            )
            self._install_bank_detail_canonical_query_fixture(app)

            response = app.handle_request(
                "GET",
                "/api/bank-details/transactions?category_code=uncategorized",
            )
            payload = json.loads(response.body)

        self.assertEqual(response.status_code, 200, response.body)
        self.assertEqual([row["id"] for row in payload["rows"]], [uncategorized_id])
        self.assertEqual(payload["pagination"]["total"], 1)
        self.assertEqual(payload["category_counts"]["uncategorized"], 1)
        self.assertEqual(payload["category_counts"]["fee"], 0)

    def test_bank_details_api_passes_keyword_to_server_side_transaction_search(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-03 09:00:00",
                counterparty_name="普通供应商A",
                summary="普通付款",
                remark="普通用途",
            )
            self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-02 09:00:00",
                counterparty_name="普通供应商B",
                summary="普通付款",
                remark="普通用途",
            )
            target_id = self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-01 09:00:00",
                counterparty_name="跨页目标供应商",
                summary="网银手续费",
                remark="跨页目标用途",
            )
            self._install_bank_detail_canonical_query_fixture(app)

            response = app.handle_request(
                "GET",
                "/api/bank-details/transactions"
                "?account_key=%E5%B7%A5%E5%95%86%E9%93%B6%E8%A1%8C%3A6386"
                "&date_from=2026-04-01"
                "&date_to=2026-04-30"
                "&page=1"
                "&page_size=2"
                "&keyword=%E8%B7%A8%E9%A1%B5%E7%9B%AE%E6%A0%87",
            )
            payload = json.loads(response.body)

        self.assertEqual(response.status_code, 200, response.body)
        self.assertEqual([row["id"] for row in payload["rows"]], [target_id])
        self.assertEqual(payload["pagination"]["total"], 1)
        self.assertEqual(payload["category_counts"]["fee"], 1)
        self.assertEqual(payload["category_counts"]["uncategorized"], 0)

    def test_bank_details_export_api_returns_xlsx_and_records_audit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            transaction_id = self._create_imported_bank_transaction(
                app,
                trade_time="2026-04-16 11:09:14+08:00",
                counterparty_name="云南溯源科技有限公司",
                summary="网银手续费",
                remark="工行附言",
            )
            facade = self._install_bank_relation_distribution(
                app,
                [
                    {
                        "row_id": transaction_id,
                        "row_type": "bank_transaction",
                        "group_ids": ["CASE-BANK-DETAIL-EXPORT"],
                        "linked_oa": [{"id": "oa-export"}],
                        "linked_input_invoices": [],
                        "linked_output_invoices": [],
                    }
                ],
            )
            self._install_bank_detail_canonical_query_fixture(app)

            response = app.handle_request(
                "GET",
                "/api/bank-details/transactions/export"
                "?mode=all"
                "&date_from=2026-04-01"
                "&date_to=2026-05-18"
                "&keyword=%E6%BA%AF%E6%BA%90",
            )
            workbook = load_workbook(BytesIO(response.body))
            audit_entries = app._audit_service.as_dicts()

        self.assertEqual(response.status_code, 200, response.body)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response.headers["Content-Type"])
        self.assertIn("filename*=UTF-8''", response.headers["Content-Disposition"])
        self.assertEqual(workbook.sheetnames, ["全部流水", "工商银行"])
        sheet = workbook["全部流水"]
        self.assertEqual(sheet["A2"].value, "2026-04-16 11:09:14")
        self.assertEqual(sheet["D2"].value, "云南溯源科技有限公司")
        self.assertEqual(sheet["I2"].value, "手续费")
        self.assertEqual(sheet["J2"].value, "费用")
        self.assertEqual(sheet["K2"].value, "手续费")
        self.assertEqual(sheet["M2"].value, "有oa")
        self.assertEqual(sheet["N2"].value, "无发票")
        self.assertEqual(sheet["Q2"].value, "工行附言")
        self.assertEqual(sheet["R2"].value, transaction_id)
        self.assertEqual(facade.calls[0]["kwargs"]["reason"], "bank_details_relation_tag_projection")
        self.assertEqual(audit_entries[-1]["action"], "bank_detail_export_downloaded")
        self.assertEqual(audit_entries[-1]["entity_type"], "bank_detail_export")
        self.assertEqual(audit_entries[-1]["metadata"]["row_count"], 1)
        self.assertEqual(audit_entries[-1]["metadata"]["filters"]["keyword"], "溯源")

    def test_bank_details_export_api_validates_account_mode_and_row_limit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            repository = self._install_bank_detail_canonical_query_fixture(app)
            missing_account_response = app.handle_request(
                "GET",
                "/api/bank-details/transactions/export?mode=account",
            )
            with patch.object(
                repository,
                "list_bank_detail_transactions",
                return_value={
                    "rows": [],
                    "pagination": {"page": 1, "page_size": 100, "total": BANK_DETAIL_EXPORT_ROW_LIMIT + 1},
                    "read_model_status": "fresh",
                    "category_counts": {"uncategorized": 0},
                },
            ):
                row_limit_response = app.handle_request(
                    "GET",
                    "/api/bank-details/transactions/export?mode=all",
                )

        self.assertEqual(missing_account_response.status_code, 400)
        self.assertEqual(json.loads(missing_account_response.body)["error"], "bank_detail_export_account_required")
        self.assertEqual(row_limit_response.status_code, 400)
        self.assertEqual(json.loads(row_limit_response.body)["error"], "bank_detail_export_row_limit_exceeded")

    def test_bank_details_export_api_uses_canonical_query_without_retired_read_model_refresh(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))

            class FakeCanonicalQueryService:
                def __init__(self) -> None:
                    self.called = False

                def export_payload(self, **_kwargs: object) -> dict[str, object]:
                    self.called = True
                    return {
                        "transactions": {
                            "account_key": None,
                            "date_from": "2026-04-01",
                            "date_to": "2026-05-18",
                            "rows": [],
                            "category_counts": {"uncategorized": 0},
                            "pagination": {"page": 1, "page_size": 100, "total": 0},
                        },
                        "accounts": None,
                    }

            canonical_query = FakeCanonicalQueryService()
            application_service = app._bank_details_application_service()
            application_service._query_service = canonical_query
            app._bank_details_application_service = lambda: application_service
            app._bank_detail_sql_read_repository = SimpleNamespace(
                list_bank_detail_transactions=lambda **_kwargs: (_ for _ in ()).throw(
                    AssertionError("retired bank-detail read model must not be queried")
                )
            )
            with patch.object(
                app._bank_details_service,
                "list_transactions",
                side_effect=AssertionError("canonical export must not use the legacy page service"),
            ):
                response = app.handle_request(
                    "GET",
                    "/api/bank-details/transactions/export?mode=all&date_from=2026-04-01&date_to=2026-05-18",
                )
            audit_entries = app._audit_service.as_dicts()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(canonical_query.called)
        self.assertTrue(any(entry["action"] == "bank_detail_export_downloaded" for entry in audit_entries))

    def test_bank_details_api_projects_workbench_relation_tags(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            transaction_id = self._create_imported_bank_transaction(
                app,
                trade_time="2026-05-02 09:00:00",
                summary="付款",
                remark="项目款",
            )
            facade = self._install_bank_relation_distribution(
                app,
                [
                    {
                        "row_id": transaction_id,
                        "row_type": "bank_transaction",
                        "group_ids": ["CASE-BANK-DETAILS"],
                        "linked_oa": [{"id": "oa-bank-details-001"}],
                        "linked_input_invoices": [{"id": "iv-bank-details-001"}],
                        "linked_output_invoices": [],
                    }
                ],
            )
            self._install_bank_detail_canonical_query_fixture(app)
            linked_response = app.handle_request(
                "GET",
                "/api/bank-details/transactions?account_key=%E5%B7%A5%E5%95%86%E9%93%B6%E8%A1%8C%3A6386",
            )
            facade.set_rows(
                [
                    {
                        "row_id": transaction_id,
                        "row_type": "bank_transaction",
                        "group_ids": [],
                        "linked_oa": [],
                        "linked_input_invoices": [],
                        "linked_output_invoices": [],
                    }
                ]
            )
            unlinked_response = app.handle_request(
                "GET",
                "/api/bank-details/transactions?account_key=%E5%B7%A5%E5%95%86%E9%93%B6%E8%A1%8C%3A6386",
            )
            app.close()

        self.assertEqual(linked_response.status_code, 200, linked_response.body)
        linked_payload = json.loads(linked_response.body)
        linked_row = next(row for row in linked_payload["rows"] if row["id"] == transaction_id)
        self.assertEqual(linked_row["oa_relation_tag"], "有oa")
        self.assertEqual(linked_row["invoice_relation_tag"], "有发票")
        self.assertEqual(linked_row["relation_tags"], ["有oa", "有发票"])
        self.assertEqual(linked_row["relation_case_id"], "CASE-BANK-DETAILS")
        self.assertEqual(facade.calls[0]["kwargs"]["reason"], "bank_details_relation_tag_projection")

        self.assertEqual(unlinked_response.status_code, 200, unlinked_response.body)
        unlinked_payload = json.loads(unlinked_response.body)
        unlinked_row = next(row for row in unlinked_payload["rows"] if row["id"] == transaction_id)
        self.assertEqual(unlinked_row["oa_relation_tag"], "无oa")
        self.assertEqual(unlinked_row["invoice_relation_tag"], "无发票")
        self.assertEqual(unlinked_row["relation_tags"], ["无oa", "无发票"])
        self.assertNotIn("relation_case_id", unlinked_row)

    def test_bank_details_api_ignores_raw_workbench_group_relation_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            transaction_id = self._create_imported_bank_transaction(
                app,
                trade_time="2026-03-12 10:16:38",
                summary="电子转账",
                remark="汽油费",
            )
            candidate_case_id = "candidate:fuel-oa-bank"
            raw_payload = {
                "month": "all",
                "summary": {},
                "paired": {"oa": [], "bank": [], "invoice": []},
                "unpaired": {
                    "oa": [
                        {
                            "id": "oa-pay-fuel-001",
                            "type": "oa",
                            "case_id": None,
                            "apply_type": "支付申请",
                            "amount": "100.00",
                            "counterparty_name": "供应商A",
                            "oa_bank_relation": {
                                "code": "pending_match",
                                "label": "待找流水与发票",
                                "tone": "warn",
                            },
                        }
                    ],
                    "bank": [
                        {
                            "id": transaction_id,
                            "type": "bank",
                            "case_id": candidate_case_id,
                            "debit_amount": "100.00",
                            "credit_amount": "",
                            "counterparty_name": "供应商A",
                            "trade_time": "2026-03-12 10:16:38",
                            "invoice_relation": {
                                "code": "suggested_match",
                                "label": "待人工确认",
                                "tone": "warn",
                            },
                        }
                    ],
                    "invoice": [],
                },
            }

            self.assertIn(candidate_case_id, json.dumps(raw_payload, ensure_ascii=False))
            self._install_bank_detail_canonical_query_fixture(app)
            response = app.handle_request(
                "GET",
                "/api/bank-details/transactions?date_from=2026-03-12&date_to=2026-03-12&page_size=500",
            )

        self.assertEqual(response.status_code, 200, response.body)
        payload = json.loads(response.body)
        row = next(row for row in payload["rows"] if row["id"] == transaction_id)
        self.assertEqual(row["oa_relation_tag"], "无oa")
        self.assertEqual(row["invoice_relation_tag"], "无发票")
        self.assertEqual(row["relation_tags"], ["无oa", "无发票"])
        self.assertNotIn("relation_case_id", row)

    def test_disabled_manual_clear_does_not_suppress_auto_in_bank_details_api(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            transaction_id = self._create_imported_bank_transaction(
                app,
                summary="网银手续费",
                remark="转账手续费",
            )
            self._install_bank_detail_canonical_query_fixture(app)

            save_response = app.handle_request(
                "PATCH",
                "/api/bank-details/transactions/categories",
                body=json.dumps(
                    {
                        "updates": [
                            {
                                "transaction_id": transaction_id,
                                "category_code": None,
                                "expected_version": 0,
                            }
                        ]
                    }
                ),
            )
            save_payload = json.loads(save_response.body)
            list_response = app.handle_request(
                "GET",
                "/api/bank-details/transactions?account_key=%E5%B7%A5%E5%95%86%E9%93%B6%E8%A1%8C%3A6386",
            )
            list_payload = json.loads(list_response.body)

        self.assertEqual(save_response.status_code, 410)
        self.assertEqual(save_payload["error"], "manual_bank_transaction_category_disabled")
        self.assertEqual(list_response.status_code, 200)
        row = next(row for row in list_payload["rows"] if row["id"] == transaction_id)
        self.assertEqual(row["auto_category_code"], "fee")
        self.assertEqual(row["effective_category_code"], "fee")
        self.assertEqual(row["category_code"], "fee")
        self.assertEqual(row["category_source"], "auto")
        self.assertEqual(list_payload["category_counts"]["fee"], 1)
        self.assertEqual(list_payload["category_counts"]["uncategorized"], 0)

    def test_import_file_confirm_returns_preview_stale_when_existing_records_change(self) -> None:
        app = build_application()
        boundary = "----finops-import-stale-boundary"
        preview_body = (
            f"--{boundary}\r\n"
            'Content-Disposition: form-data; name="imported_by"\r\n\r\n'
            "user_finance_01\r\n"
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="files"; filename="{INVOICE_JAN.name}"\r\n'
            f"Content-Type: {INVOICE_JAN.content_type}\r\n\r\n"
        ).encode("utf-8") + INVOICE_JAN.content + f"\r\n--{boundary}--\r\n".encode("utf-8")
        preview_response = app.handle_request(
            "POST",
            "/imports/files/preview",
            body=preview_body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        )
        self.assertEqual(preview_response.status_code, 200)
        preview_payload = json.loads(preview_response.body)
        session = app._file_import_service.get_session(preview_payload["session"]["id"])
        competing_preview = app._import_service.preview_import(
            batch_type=session.files[0].batch_type,
            source_name="competing.json",
            imported_by="user_finance_02",
            rows=[session.files[0].row_results[0].raw_payload],
        )
        app._import_service.confirm_import(competing_preview.id)

        confirm_response = app.handle_request(
            "POST",
            "/imports/files/confirm",
            json.dumps(
                {
                    "session_id": preview_payload["session"]["id"],
                    "selected_file_ids": [preview_payload["files"][0]["id"]],
                }
            ),
        )

        self.assertEqual(confirm_response.status_code, 409)
        self.assertEqual(json.loads(confirm_response.body)["error"], "preview_stale")

    def test_enqueued_workbench_auto_matching_only_marks_durable_scopes(self) -> None:
        app = build_application()

        with (
            patch.object(
                app,
                "_mark_workbench_matching_dirty_scopes",
                return_value=["2026-05"],
            ) as mark_dirty,
            patch.object(app._matching_service, "run", return_value=SimpleNamespace(result_count=99)) as run_legacy,
        ):
            job = app._enqueue_workbench_auto_matching_for_scopes(
                ["2026-05"],
                reason="unit",
                owner_user_id="system",
            )

        self.assertIsNotNone(job)
        mark_dirty.assert_called_once_with(["2026-05"], reason="unit")
        run_legacy.assert_not_called()
        job_payload = app._background_job_service.get_job(job.job_id, "system").to_payload()
        self.assertEqual(job_payload["result_summary"]["queued_months"], ["2026-05"])
        self.assertEqual(job_payload["result_summary"]["affected_months"], ["2026-05"])
        self.assertNotIn("matching_results", job_payload["result_summary"])

    def test_workbench_auto_matching_surfaces_durable_queue_failure(self) -> None:
        app = build_application()
        with (
            patch.object(
                app,
                "_mark_workbench_matching_dirty_scopes",
                side_effect=RuntimeError("durable queue unavailable"),
            ),
            self.assertRaisesRegex(RuntimeError, "durable queue unavailable"),
        ):
            app._run_workbench_auto_matching_for_scopes(["2026-05"], reason="unit_failure")

    def test_current_oa_attachment_invoice_parser_version_includes_source_schema_version(self) -> None:
        app = build_application()
        app._workbench_query_service._oa_adapter = MongoOAAdapter(
            settings=MongoOASettings(host="127.0.0.1", database="form_data_db")
        )

        self.assertEqual(
            app._current_oa_attachment_invoice_parser_version(),
            MongoOAAdapter._attachment_invoice_cache_parser_version(),
        )

    def test_get_api_workbench_ignored_uses_sql_read_model_without_rebuild(self) -> None:
        class SqlReadRepository:
            def list_workbench_ignored_rows(self, *, scope_key: str) -> list[dict[str, object]]:
                self.scope_key = scope_key
                return [{"id": "bk-sql-ignored-001", "type": "bank"}]

        app = build_application()
        repository = SqlReadRepository()
        app._workbench_sql_read_repository = repository

        response = app.handle_request("GET", "/api/workbench/ignored?month=all")

        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.body)
        self.assertEqual(repository.scope_key, "all")
        self.assertEqual(payload["rows"], [{"id": "bk-sql-ignored-001", "type": "bank"}])

    def test_bank_policy_metadata_uses_selected_row_category_before_resolver(self) -> None:
        app = build_application()
        rules_payload = app._app_settings_service.get_bank_flow_rule_batch_tag_rules_payload()
        app._app_settings_service.update_bank_flow_rule_batch_tag_rules(
            {
                "expected_version": rules_payload["version"],
                "rules": [
                    {
                        **rule,
                        "requires_oa": True if rule["tag_code"] == "external_turnover" else rule["requires_oa"],
                        "requires_invoice": False if rule["tag_code"] == "external_turnover" else rule["requires_invoice"],
                    }
                    for rule in rules_payload["rules"]
                ],
            },
            actor_id="tester",
        )

        with patch.object(
            app,
            "_bank_transaction_category_codes_for_workbench_row_ids",
            side_effect=AssertionError("selected row category should avoid resolver"),
        ):
            metadata = app._workbench_write_facade()._bank_transaction_paired_policy_metadata(
                row_ids=["bank-fast-policy-1"],
                row_types=["bank"],
                selected_rows=[
                    {
                        "id": "bank-fast-policy-1",
                        "type": "bank",
                        "category_code": "external_turnover",
                    }
                ],
            )

        self.assertEqual(metadata["paired_requirement_tag_codes"], ["external_turnover"])
        self.assertTrue(metadata["requires_oa"])
        self.assertFalse(metadata["requires_invoice"])

    def test_confirm_link_ignores_empty_row_ids_in_minimal_hot_path(self) -> None:
        app = build_application()

        response = app._handle_live_workbench_confirm_link(
            {
                "month": "2026-03",
                "row_ids": ["oa-o-202603-001", None, "  ", "bk-o-202603-001"],
                "case_id": "CASE-NORMALIZE-ROWIDS-001",
            }
        )

        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.body)
        self.assertEqual(
            payload["affected_row_ids"],
            ["oa-o-202603-001", "bk-o-202603-001"],
        )

    def test_cancel_link_does_not_resolve_source_rows_in_hot_path(self) -> None:
        app = build_application()
        app._workbench_pair_relation_service.create_active_relation(
            case_id="CASE-MINIMAL-CANCEL-001",
            row_ids=["oa-o-202603-001", "bk-o-202603-001", "iv-o-202603-001"],
            row_types=["oa", "bank", "invoice"],
            relation_mode="manual_confirmed",
            created_by="YNSYLP005",
            month_scope="2026-03",
        )

        with patch.object(app, "_resolve_live_rows_direct", side_effect=AssertionError("should not resolve source rows")):
            cancel_response = app._handle_live_workbench_cancel_link(
                {
                    "month": "2026-03",
                    "row_id": "bk-o-202603-001",
                    "comment": "reopen",
                }
            )

        self.assertEqual(cancel_response.status_code, 200)
        cancel_payload = json.loads(cancel_response.body)
        self.assertCountEqual(
            cancel_payload["affected_row_ids"],
            ["oa-o-202603-001", "bk-o-202603-001", "iv-o-202603-001"],
        )
        self.assertNotIn("updated_rows", cancel_payload)

    def test_cancel_exception_resolves_all_scope_bank_rows_without_oa_query_service(self) -> None:
        app = build_application()
        preview = app._import_service.preview_import(
            batch_type=BatchType.BANK_TRANSACTION,
            source_name="cancel-exception-bank-fast-path.xlsx",
            imported_by="user_finance_01",
            rows=[
                {
                    "account_no": "62220031",
                    "account_name": "云南溯源科技有限公司建设银行基本户",
                    "txn_date": "2026-03-10",
                    "trade_time": "2026-03-10 09:00:00",
                    "pay_receive_time": "2026-03-10 09:00:00",
                    "counterparty_name": "测试取消异常流水",
                    "debit_amount": "100.00",
                    "credit_amount": "",
                    "summary": "测试取消异常流水",
                },
            ],
        )
        app._import_service.confirm_import(preview.id)
        bank_row_id = next(
            transaction.id
            for transaction in app._import_service.list_transactions()
            if transaction.source_batch_id == preview.id
        )
        app.handle_request("GET", "/api/workbench?month=all")

        exception_response = app._handle_live_workbench_oa_bank_exception(
            {
                "month": "all",
                "row_ids": [bank_row_id],
                "exception_code": "oa_bank_amount_mismatch",
                "exception_label": "金额不一致，继续异常",
                "comment": "测试异常处理",
            }
        )
        self.assertEqual(exception_response.status_code, 200)

        with patch.object(
            app._workbench_query_service,
            "get_row_record",
            side_effect=AssertionError("bank rows should resolve from live detail without OA query service"),
        ):
            cancel_response = app._handle_live_workbench_cancel_exception(
                {
                    "month": "all",
                    "row_ids": [bank_row_id],
                    "comment": "撤回异常处理",
                }
            )

        self.assertEqual(cancel_response.status_code, 200)
        cancel_payload = json.loads(cancel_response.body)
        self.assertTrue(cancel_payload["success"])
        self.assertEqual(cancel_payload["action"], "cancel_exception")
        self.assertEqual(cancel_payload["affected_row_ids"], [bank_row_id])

    def test_confirm_link_accepts_unbalanced_same_type_rows_and_rebuilds_live_cache_once(self) -> None:
        app = build_application()
        preview = app._import_service.preview_import(
            batch_type=BatchType.BANK_TRANSACTION,
            source_name="multi-bank.xlsx",
            imported_by="user_finance_01",
            rows=[
                {
                    "account_no": "62220031",
                    "account_name": "云南溯源科技有限公司建设银行基本户",
                    "txn_date": "2026-03-10",
                    "trade_time": "2026-03-10 09:00:00",
                    "pay_receive_time": "2026-03-10 09:00:00",
                    "counterparty_name": "测试对方A",
                    "debit_amount": "100.00",
                    "credit_amount": "",
                    "summary": "测试流水A",
                },
                {
                    "account_no": "62220032",
                    "account_name": "云南溯源科技有限公司建设银行基本户",
                    "txn_date": "2026-03-10",
                    "trade_time": "2026-03-10 09:05:00",
                    "pay_receive_time": "2026-03-10 09:05:00",
                    "counterparty_name": "测试对方B",
                    "debit_amount": "250.00",
                    "credit_amount": "",
                    "summary": "测试流水B",
                },
            ],
        )
        app._import_service.confirm_import(preview.id)
        row_ids = [transaction.id for transaction in app._import_service.list_transactions() if transaction.source_batch_id == preview.id]

        with patch.object(
            app._live_workbench_service,
            "get_rows_detail",
            wraps=app._live_workbench_service.get_rows_detail,
        ) as get_rows_detail:
            confirm_response = app._handle_live_workbench_confirm_link(
                {
                    "month": "2026-03",
                    "row_ids": row_ids,
                    "case_id": "CASE-LIVE-BULK-001",
                }
            )

        self.assertEqual(confirm_response.status_code, 200)
        confirm_payload = json.loads(confirm_response.body)
        self.assertEqual(confirm_payload["action"], "confirm_link")
        self.assertCountEqual(confirm_payload["affected_row_ids"], row_ids)
        get_rows_detail.assert_called_once_with(row_ids)

    def test_confirm_link_resolves_underlying_live_row_services_without_group_payload_dependency(self) -> None:
        app = build_application()
        app._live_workbench_service = _StubLiveWorkbenchService()
        app._workbench_row_detail_api_routes = app._build_workbench_row_detail_api_routes()

        confirm_response = app._handle_live_workbench_confirm_link(
            {
                "month": "2026-03",
                "row_ids": ["oa-o-202603-001", "txn-live-202603-001"],
            }
        )

        self.assertEqual(confirm_response.status_code, 200)
        confirm_payload = json.loads(confirm_response.body)
        self.assertEqual(confirm_payload["action"], "confirm_link")
        self.assertEqual(confirm_payload["affected_row_ids"], ["oa-o-202603-001", "txn-live-202603-001"])


    def test_oa_sync_status_endpoint_reads_durable_queue_status(self) -> None:
        app = build_application()
        app._app_status_runtime_statuses = lambda: {
            "read_model_statuses": None,
            "outbox_statuses": {
                "oa.sync": {
                    "status": "pending",
                    "count": 1,
                    "scopes": [
                        {
                            "event_type": "oa.sync",
                            "scope_type": "oa",
                            "scope_key": "2026-03",
                            "status": "pending",
                            "count": 1,
                        }
                    ],
                }
            },
            "worker_statuses": {"oa-sync": {"status": "ready"}},
        }
        app._postgres_oa_projection_latest_sync_run = lambda: {
            "id": "oa-run-1",
            "status": "succeeded",
            "finished_at": "2026-07-03T08:10:00+00:00",
            "upserted_count": 334,
            "scanned_count": 334,
        }

        response = app.handle_request("GET", "/api/oa-sync/status")
        payload = json.loads(response.body)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["status"], "refreshing")
        self.assertCountEqual(payload["dirty_scopes"], ["2026-03", "all"])
        self.assertEqual(payload["outbox_status"], "pending")
        self.assertEqual(payload["worker_status"], "ready")
        self.assertEqual(payload["last_synced_at"], "2026-07-03T08:10:00+00:00")
        self.assertEqual(payload["last_upserted_count"], 334)

    def test_oa_sync_status_endpoint_does_not_treat_failed_run_as_synced(self) -> None:
        app = build_application()
        app._app_status_runtime_statuses = lambda: {
            "read_model_statuses": None,
            "outbox_statuses": {},
            "worker_statuses": {"oa-sync": {"status": "ready"}},
        }
        app._postgres_oa_projection_latest_sync_run = lambda: {
            "id": "oa-run-failed",
            "status": "failed",
            "finished_at": "2026-07-03T08:20:00+00:00",
            "last_error": "mongo timeout",
        }

        response = app.handle_request("GET", "/api/oa-sync/status")
        payload = json.loads(response.body)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["status"], "error")
        self.assertEqual(payload["message"], "mongo timeout")
        self.assertIsNone(payload["last_synced_at"])

    def test_in_process_oa_polling_and_hot_rebuild_entrypoints_are_removed(self) -> None:
        app = build_application()

        for method_name in [
            "start_oa_sync_polling_worker",
            "_poll_oa_source_once",
            "_handle_oa_source_changed",
            "_schedule_oa_sync_dirty_scope_rebuild",
            "_rebuild_oa_sync_dirty_scopes_once",
            "_hot_rebuild_workbench_read_model_scopes",
        ]:
            self.assertFalse(hasattr(app, method_name), method_name)

    def test_oa_sync_events_endpoint_is_removed_for_polling_mode(self) -> None:
        app = build_application()

        response = app.handle_request("GET", "/api/oa-sync/events?once=1")

        self.assertEqual(response.status_code, 404)

    def _submit_batch_accounting_mismatch_with_note(
        self,
        app: Application,
        *,
        note: str,
    ) -> dict[str, object]:
        response = app.handle_request(
            "POST",
            "/api/batch-accounting/submit",
            json.dumps(
                {
                    "year": "2026",
                    "bank_row_id": "txn_imported_202601_batch_001",
                    "oa_row_ids": ["oa-exp-ba-001"],
                    "note": note,
                    "actor": "finance-user",
                }
            ),
        )
        self.assertEqual(response.status_code, 200, response.body)
        return json.loads(response.body)

    @staticmethod
    def _find_group_by_row_id(groups: list[dict[str, object]], row_id: str) -> dict[str, object]:
        for group in groups:
            for key in ("oa_rows", "bank_rows", "invoice_rows"):
                for row in list(group.get(key) or []):
                    if isinstance(row, dict) and row.get("id") == row_id:
                        return group
        raise AssertionError(f"group not found for row {row_id}")

if __name__ == "__main__":
    unittest.main()

def flatten_groups(groups: list[dict[str, object]], record_type: str) -> list[dict[str, object]]:
    key = f"{record_type}_rows"
    flattened: list[dict[str, object]] = []
    for group in groups:
        flattened.extend(group[key])
    return flattened

def all_groups(payload: dict[str, object]) -> list[dict[str, object]]:
    return [*payload["paired"]["groups"], *payload["unpaired"]["groups"]]

def assert_auto_linked_group(
    testcase: unittest.TestCase,
    payload: dict[str, object],
    expected_row_ids: set[str],
) -> dict[str, object]:
    paired = payload["paired"]
    assert isinstance(paired, dict)
    groups = paired["groups"]
    assert isinstance(groups, list)
    testcase.assertEqual(len(groups), 1)
    group = groups[0]
    assert isinstance(group, dict)
    testcase.assertEqual(group["group_type"], "relation")
    testcase.assertEqual(group.get("relation_note"), "系统自动配对")
    amount_check = group.get("amount_check")
    testcase.assertIsInstance(amount_check, dict)
    assert isinstance(amount_check, dict)
    testcase.assertEqual(amount_check.get("status"), "matched")
    actual_row_ids = {
        str(row.get("id"))
        for key in ("oa_rows", "bank_rows", "invoice_rows")
        for row in list(group.get(key) or [])
        if isinstance(row, dict)
    }
    testcase.assertEqual(actual_row_ids, expected_row_ids)
    for key in ("oa_rows", "bank_rows", "invoice_rows"):
        for row in list(group.get(key) or []):
            assert isinstance(row, dict)
            testcase.assertTrue(str(row.get("case_id") or "").startswith("CASE-AUTO-"))
            relation_amount_check = row.get("relation_amount_check")
            testcase.assertIsInstance(relation_amount_check, dict)
            assert isinstance(relation_amount_check, dict)
            testcase.assertEqual(relation_amount_check.get("status"), "matched")
    return group

def build_personal_advance_repayment_raw_payload(
    *,
    oa_amount: str = "300000.00",
    bank_debit_amount: str = "300000.00",
    bank_credit_amounts: list[str] | None = None,
    include_bank_debit: bool = True,
    include_invoice: bool = False,
) -> dict[str, object]:
    credit_amounts = ["200000.00", "100000.00"] if bank_credit_amounts is None else list(bank_credit_amounts)
    oa_rows = [
        {
            "id": "oa-personal-advance-001",
            "type": "oa",
            "case_id": None,
            "applicant": "测试员工",
            "project_name": "个人暂借款",
            "apply_type": "支付申请",
            "amount": oa_amount,
            "counterparty_name": "测试员工",
            "reason": "个人暂借款",
            "oa_bank_relation": {"code": "pending_match", "label": "待找流水与发票", "tone": "warn"},
            "available_actions": ["detail", "confirm_link", "mark_exception"],
            "summary_fields": {"申请人": "测试员工"},
            "detail_fields": {"申请日期": "2026-03-01"},
        }
    ]
    bank_rows: list[dict[str, object]] = []
    if include_bank_debit:
        bank_rows.append(
            {
                "id": "bank-personal-advance-out-001",
                "type": "bank",
                "case_id": None,
                "trade_time": "2026-03-02 09:00:00",
                "pay_receive_time": "2026-03-02 09:00:00",
                "debit_amount": bank_debit_amount,
                "credit_amount": "",
                "counterparty_name": "测试员工",
                "invoice_relation": {"code": "pending_invoice_match", "label": "待关联发票", "tone": "warn"},
                "available_actions": ["detail", "confirm_link", "mark_exception"],
                "summary_fields": {"交易时间": "2026-03-02 09:00:00"},
                "detail_fields": {"摘要": "个人暂借款付款"},
            }
        )
    for index, amount in enumerate(credit_amounts, start=1):
        bank_rows.append(
            {
                "id": f"bank-personal-advance-in-{index:03d}",
                "type": "bank",
                "case_id": None,
                "trade_time": f"2026-03-{index + 2:02d} 09:00:00",
                "pay_receive_time": f"2026-03-{index + 2:02d} 09:00:00",
                "debit_amount": "",
                "credit_amount": amount,
                "counterparty_name": "测试员工",
                "invoice_relation": {"code": "pending_invoice_match", "label": "待关联发票", "tone": "warn"},
                "available_actions": ["detail", "confirm_link", "mark_exception"],
                "summary_fields": {"交易时间": f"2026-03-{index + 2:02d} 09:00:00"},
                "detail_fields": {"摘要": "个人暂借款还款"},
            }
        )
    invoice_rows = (
        [
            {
                "id": "invoice-personal-advance-001",
                "type": "invoice",
                "case_id": None,
                "amount": "300000.00",
                "total_with_tax": "300000.00",
                "seller_name": "测试供应商",
                "invoice_type": "进项发票",
                "invoice_bank_relation": {"code": "pending_collection", "label": "待匹配流水", "tone": "warn"},
                "available_actions": ["detail", "confirm_link", "mark_exception"],
            }
        ]
        if include_invoice
        else []
    )
    return {
        "month": "2026-03",
        "summary": {
            "oa_count": len(oa_rows),
            "bank_count": len(bank_rows),
            "invoice_count": len(invoice_rows),
            "paired_count": 0,
            "unpaired_count": len(oa_rows) + len(bank_rows) + len(invoice_rows),
            "exception_count": 0,
        },
        "paired": {"oa": [], "bank": [], "invoice": []},
        "unpaired": {"oa": oa_rows, "bank": bank_rows, "invoice": invoice_rows},
    }

def build_oa_retention_raw_payload(
    *,
    oa_rows: list[dict[str, object]],
    bank_rows: list[dict[str, object]],
    invoice_rows: list[dict[str, object]],
) -> dict[str, object]:
    return {
        "month": "all",
        "summary": {
            "oa_count": len(oa_rows),
            "bank_count": len(bank_rows),
            "invoice_count": len(invoice_rows),
            "paired_count": 0,
            "unpaired_count": len(oa_rows) + len(bank_rows) + len(invoice_rows),
            "exception_count": 0,
        },
        "paired": {"oa": [], "bank": [], "invoice": []},
        "unpaired": {"oa": oa_rows, "bank": bank_rows, "invoice": invoice_rows},
    }

def build_oa_retention_oa_row(row_id: str, case_id: str, application_date: str) -> dict[str, object]:
    return {
        "id": row_id,
        "type": "oa",
        "case_id": case_id,
        "applicant": "测试申请人",
        "project_name": "测试项目",
        "apply_type": "支付申请",
        "amount": "100.00",
        "counterparty_name": "测试供应商",
        "reason": "测试保OA",
        "oa_bank_relation": {"code": "pending_match", "label": "待找流水与发票", "tone": "warn"},
        "available_actions": ["detail"],
        "summary_fields": {"申请人": "测试申请人"},
        "detail_fields": {"申请日期": application_date},
    }

def build_oa_retention_bank_row(row_id: str, case_id: str, trade_time: str) -> dict[str, object]:
    return {
        "id": row_id,
        "type": "bank",
        "case_id": case_id,
        "trade_time": trade_time,
        "debit_amount": "100.00",
        "credit_amount": "",
        "counterparty_name": "测试供应商",
        "invoice_relation": {"code": "pending_match", "label": "待匹配", "tone": "warn"},
        "available_actions": ["detail"],
    }

def build_ccb_bank_row(row_id: str, trade_time: str, amount: str) -> dict[str, object]:
    return {
        "id": row_id,
        "type": "bank",
        "case_id": None,
        "trade_time": trade_time,
        "pay_receive_time": trade_time,
        "direction": "支出",
        "debit_amount": amount,
        "credit_amount": "",
        "counterparty_name": "建设银行可见性测试供应商",
        "payment_account_label": "建设银行 8106",
        "invoice_relation": {"code": "pending_invoice_match", "label": "待关联发票", "tone": "warn"},
        "available_actions": ["detail", "confirm_link", "mark_exception"],
        "summary_fields": {"交易时间": trade_time, "账号": "建设银行 8106"},
        "detail_fields": {"交易时间": trade_time, "账号": "建设银行 8106"},
    }

def build_oa_retention_invoice_row(row_id: str, case_id: str, issue_date: str) -> dict[str, object]:
    return {
        "id": row_id,
        "type": "invoice",
        "case_id": case_id,
        "seller_name": "测试供应商",
        "buyer_name": "云南溯源科技有限公司",
        "issue_date": issue_date,
        "amount": "100.00",
        "invoice_bank_relation": {"code": "pending_collection", "label": "待匹配流水", "tone": "warn"},
        "available_actions": ["detail"],
    }

def build_relation_amount_raw_payload(*, invoice_amount: str) -> dict[str, object]:
    return {
        "month": "2026-05",
        "summary": {"oa_count": 1, "bank_count": 1, "invoice_count": 1, "paired_count": 0, "unpaired_count": 3, "exception_count": 0},
        "paired": {"oa": [], "bank": [], "invoice": []},
        "unpaired": {
            "oa": [
                {
                    "id": "oa-o-202605-001",
                    "type": "oa",
                    "case_id": "",
                    "apply_type": "支付申请",
                    "amount": "100.00",
                    "counterparty_name": "测试供应商",
                    "oa_bank_relation": {"code": "pending_match", "label": "待找流水与发票", "tone": "warn"},
                    "available_actions": ["detail"],
                }
            ],
            "bank": [
                {
                    "id": "bk-o-202605-001",
                    "type": "bank",
                    "case_id": "",
                    "trade_time": "2026-05-02 09:00:00",
                    "debit_amount": "100.00",
                    "credit_amount": "",
                    "counterparty_name": "测试供应商",
                    "invoice_relation": {"code": "pending_match", "label": "待匹配", "tone": "warn"},
                    "available_actions": ["detail"],
                }
            ],
            "invoice": [
                {
                    "id": "iv-o-202605-001",
                    "type": "invoice",
                    "case_id": "",
                    "seller_name": "测试供应商",
                    "buyer_name": "云南溯源科技有限公司",
                    "issue_date": "2026-05-02",
                    "amount": invoice_amount,
                    "total_with_tax": invoice_amount,
                    "invoice_type": "进项专票",
                    "invoice_bank_relation": {"code": "pending_collection", "label": "待匹配流水", "tone": "warn"},
                    "available_actions": ["detail"],
                }
            ],
        },
    }

def row_detail_side_effect_for_raw_payload(raw_payload: dict[str, object]):
    rows_by_id = {
        row["id"]: dict(row)
        for pane in ("oa", "bank", "invoice")
        for row in raw_payload["unpaired"][pane]
    }

    def row_detail(row_id: str, **_kwargs: object) -> dict[str, object]:
        return {"row": dict(rows_by_id[row_id])}

    return row_detail

def build_batch_accounting_raw_payload() -> dict[str, object]:
    return {
        "month": "all",
        "summary": {"oa_count": 1, "bank_count": 1, "invoice_count": 0, "paired_count": 0, "unpaired_count": 2, "exception_count": 0},
        "paired": {"oa": [], "bank": [], "invoice": []},
        "unpaired": {
            "oa": [
                {
                    "id": "oa-exp-ba-001",
                    "type": "oa",
                    "case_id": "",
                    "applicant": "刘晨",
                    "apply_time": "2026-01-06",
                    "project_name": "品牌广告投放",
                    "amount": "700.00",
                    "reason": "1月日常报销",
                    "apply_type": "日常报销",
                    "expense_type": "交通费",
                    "oa_bank_relation": {"code": "pending_match", "label": "待找流水与发票", "tone": "warn"},
                    "available_actions": ["detail"],
                    "summary_fields": {"申请日期": "2026-01-06"},
                }
            ],
            "bank": [
                {
                    "id": "txn_imported_202601_batch_001",
                    "type": "bank",
                    "case_id": "",
                    "trade_time": "2026-01-07 15:54:00",
                    "pay_receive_time": "2026-01-07 15:54:00",
                    "counterparty_name": "批量账务集中处理",
                    "debit_amount": "1200.00",
                    "credit_amount": "",
                    "payment_account_label": "建行基本户 8106",
                    "bank_name": "建行",
                    "account_last4": "8106",
                    "invoice_relation": {"code": "pending_match", "label": "待匹配", "tone": "warn"},
                    "available_actions": ["detail"],
                    "summary_fields": {"交易时间": "2026-01-07 15:54:00"},
                }
            ],
            "invoice": [],
        },
    }

def build_etc_batch_raw_payload(*, bank_amount: str | None) -> dict[str, object]:
    bank_rows: list[dict[str, object]] = []
    if bank_amount is not None:
        bank_rows.append(
            {
                "id": "bk-etc-202606-001",
                "type": "bank",
                "case_id": "",
                "trade_time": "2026-06-03 09:00:00",
                "debit_amount": bank_amount,
                "credit_amount": "",
                "counterparty_name": "云南高速通行费",
                "invoice_relation": {"code": "pending_match", "label": "待匹配", "tone": "warn"},
                "available_actions": ["detail"],
            }
        )
    return {
        "month": "2026-06",
        "summary": {
            "oa_count": 1,
            "bank_count": len(bank_rows),
            "invoice_count": 0,
            "paired_count": 0,
            "unpaired_count": 1 + len(bank_rows),
            "exception_count": 0,
        },
        "paired": {"oa": [], "bank": [], "invoice": []},
        "unpaired": {
            "oa": [
                {
                    "id": "oa-etc-202606-001",
                    "type": "oa",
                    "source": "etc_batch",
                    "etc_batch_id": "etc_20260503_001",
                    "etcBatchId": "etc_20260503_001",
                    "tags": ["ETC批量提交"],
                    "case_id": "",
                    "apply_type": "支付申请",
                    "amount": "100.00",
                    "counterparty_name": "云南高速通行费",
                    "reason": "ETC批量提交\netc_batch_id=etc_20260503_001",
                    "oa_bank_relation": {"code": "pending_match", "label": "待找流水", "tone": "warn"},
                    "available_actions": ["detail"],
                }
            ],
            "bank": bank_rows,
            "invoice": [],
        },
    }

class _StubLiveWorkbenchService:
    def has_rows_for_month(self, month: str) -> bool:
        return month == "2026-03"

    def get_workbench(self, month: str) -> dict[str, object]:
        if month != "2026-03":
            return {
                "month": month,
                "summary": {"oa_count": 0, "bank_count": 0, "invoice_count": 0, "paired_count": 0, "unpaired_count": 0, "exception_count": 0},
                "paired": {"oa": [], "bank": [], "invoice": []},
                "unpaired": {"oa": [], "bank": [], "invoice": []},
            }
        return {
            "month": "2026-03",
            "summary": {"oa_count": 0, "bank_count": 1, "invoice_count": 0, "paired_count": 0, "unpaired_count": 1, "exception_count": 0},
            "paired": {"oa": [], "bank": [], "invoice": []},
            "unpaired": {
                "oa": [],
                "bank": [
                    {
                        "id": "txn-live-202603-001",
                        "type": "bank",
                        "case_id": "CASE-LIVE-202603-001",
                        "trade_time": "2026-03-28 11:20:00",
                        "debit_amount": "58,000.00",
                        "credit_amount": "",
                        "counterparty_name": "智能工厂设备商",
                        "payment_account_label": "工商银行 账户 8888",
                        "invoice_relation": {"code": "suggested_match", "label": "待人工确认", "tone": "warn"},
                        "pay_receive_time": "2026-03-28 11:20:00",
                        "remark": "设备尾款待支付",
                        "repayment_date": "",
                        "available_actions": ["detail"],
                    }
                ],
                "invoice": [],
            },
        }

    def get_row_detail(self, row_id: str) -> dict[str, object]:
        if row_id != "txn-live-202603-001":
            raise KeyError(row_id)
        return {
            "id": "txn-live-202603-001",
            "type": "bank",
            "case_id": "CASE-LIVE-202603-001",
            "trade_time": "2026-03-28 11:20:00",
            "debit_amount": "58,000.00",
            "credit_amount": "",
            "counterparty_name": "智能工厂设备商",
            "payment_account_label": "工商银行 账户 8888",
            "invoice_relation": {"code": "suggested_match", "label": "待人工确认", "tone": "warn"},
            "pay_receive_time": "2026-03-28 11:20:00",
            "remark": "设备尾款待支付",
            "repayment_date": "",
            "available_actions": ["detail"],
            "summary_fields": {"和发票关联情况": "待人工确认", "备注": "设备尾款待支付"},
            "detail_fields": {"备注": "设备尾款待支付"},
        }

class _AttachmentRecord:
    def __init__(self) -> None:
        self.id = "oa-attach-202603-001"
        self.month = "2026-03"
        self.section = "unpaired"
        self.case_id = None
        self.applicant = "刘际涛"
        self.project_name = "玉烟维护项目"
        self.apply_type = "日常报销"
        self.amount = "58,000.00"
        self.counterparty_name = "智能工厂设备商"
        self.reason = "设备尾款报销"
        self.relation_code = "pending_match"
        self.relation_label = "待找流水与发票"
        self.relation_tone = "warn"
        self.expense_type = "设备货款及材料费"
        self.expense_content = "设备尾款报销"
        self.detail_fields = {
            "OA单号": "OA-ATT-001",
            "申请日期": "2026-03-28",
            "明细行号": "0",
        }
        self.attachment_invoices = [
            {
                "invoice_code": "053002200111",
                "invoice_no": "40512344",
                "seller_tax_no": "91530100678728169X",
                "seller_name": "智能工厂设备商",
                "buyer_tax_no": "915300007194052520",
                "buyer_name": "云南溯源科技有限公司",
                "issue_date": "2026-03-28",
                "amount": "58,000.00",
                "tax_rate": "13%",
                "tax_amount": "0.00",
                "total_with_tax": "58,000.00",
                "invoice_type": "进项发票",
                "source_expense_row_index": "0",
                "source_expense_item_id": "oa-attach-202603-001:item:0:equipment",
                "source_attachment_key": "oa-attach-202603-001:item:0:att:equipment",
                "source_attachment_name": "设备发票.pdf",
                "attachment_name": "设备发票.pdf",
                "invoice_kind": "增值税电子专用发票",
            }
        ]

class AttachmentAwareOAAdapter:
    def list_application_records(self, month: str) -> list[object]:
        if month != "2026-03":
            return []
        return [_AttachmentRecord()]

class _SourceBoundAttachmentRecord:
    def __init__(self) -> None:
        self.id = "oa-exp-hurong-248"
        self.month = "2026-03"
        self.section = "unpaired"
        self.case_id = None
        self.applicant = "胡瑢"
        self.project_name = "2024-2026年度红塔集团工作证管理系统维护项目"
        self.apply_type = "日常报销"
        self.amount = "248.00"
        self.counterparty_name = ""
        self.reason = "工作证管理系统维护项目报销"
        self.relation_code = "pending_match"
        self.relation_label = "待找流水与发票"
        self.relation_tone = "warn"
        self.expense_type = "项目费用"
        self.expense_content = "工作证维护费用"
        self.detail_fields = {"OA单号": "OA-HR-248", "申请日期": "2026-03-04"}
        self.expense_items = [
            {
                "row_index": "0",
                "expense_item_id": "oa-exp-hurong-248:item:0:maint",
                "amount": "196.00",
                "expense_content": "付款项1",
            },
            {
                "row_index": "1",
                "expense_item_id": "oa-exp-hurong-248:item:1:service",
                "amount": "52.00",
                "expense_content": "付款项2",
            },
        ]
        self.attachment_invoices = [
            {
                "invoice_no": "24800001",
                "seller_name": "红塔供应商A",
                "buyer_name": "云南溯源科技有限公司",
                "issue_date": "2026-03-04",
                "amount": "100.00",
                "total_with_tax": "100.00",
                "invoice_type": "进项发票",
                "source_expense_row_index": "0",
                "source_expense_item_id": "oa-exp-hurong-248:item:0:maint",
                "source_attachment_key": "oa-exp-hurong-248:item:0:att:a",
                "source_attachment_name": "付款项1-发票A.pdf",
            },
            {
                "invoice_no": "24800002",
                "seller_name": "红塔供应商B",
                "buyer_name": "云南溯源科技有限公司",
                "issue_date": "2026-03-04",
                "amount": "96.00",
                "total_with_tax": "96.00",
                "invoice_type": "进项发票",
                "source_expense_row_index": "0",
                "source_expense_item_id": "oa-exp-hurong-248:item:0:maint",
                "source_attachment_key": "oa-exp-hurong-248:item:0:att:b",
                "source_attachment_name": "付款项1-发票B.pdf",
            },
            {
                "invoice_no": "24800003",
                "seller_name": "红塔供应商C",
                "buyer_name": "云南溯源科技有限公司",
                "issue_date": "2026-03-04",
                "amount": "52.00",
                "total_with_tax": "52.00",
                "invoice_type": "进项发票",
                "source_expense_row_index": "1",
                "source_expense_item_id": "oa-exp-hurong-248:item:1:service",
                "source_attachment_key": "oa-exp-hurong-248:item:1:att:c",
                "source_attachment_name": "付款项2-发票C.pdf",
            },
        ]
        self.attachment_file_count = 3

class _SingleSourceAttachmentRecord:
    def __init__(self) -> None:
        self.id = "oa-exp-hurong-292"
        self.month = "2026-03"
        self.section = "unpaired"
        self.case_id = None
        self.applicant = "胡瑢"
        self.project_name = "红云红河烟草能源管理运维项目"
        self.apply_type = "日常报销"
        self.amount = "292.00"
        self.counterparty_name = ""
        self.reason = "能源管理运维项目报销"
        self.relation_code = "pending_match"
        self.relation_label = "待找流水与发票"
        self.relation_tone = "warn"
        self.expense_type = "项目费用"
        self.expense_content = "能源管理运维费用"
        self.detail_fields = {"OA单号": "OA-HR-292", "申请日期": "2026-03-24", "明细行号": "0"}
        source_invoice = {
            "invoice_no": "29200001",
            "seller_name": "能源运维供应商",
            "buyer_name": "云南溯源科技有限公司",
            "issue_date": "2026-03-24",
            "amount": "292.00",
            "total_with_tax": "292.00",
            "invoice_type": "进项发票",
            "source_expense_row_index": "0",
            "source_expense_item_id": "oa-exp-hurong-292:item:0:energy",
            "source_attachment_key": "oa-exp-hurong-292:item:0:att:only",
            "source_attachment_name": "能源管理运维发票.pdf",
        }
        self.expense_items = [
            {
                "row_index": "0",
                "expense_item_id": "oa-exp-hurong-292:item:0:energy",
                "amount": "292.00",
                "attachment_invoices": [dict(source_invoice)],
            }
        ]
        self.attachment_invoices = [source_invoice]
        self.attachment_file_count = 1

class SourceBoundAttachmentOAAdapter:
    def list_application_records(self, month: str) -> list[object]:
        if month != "2026-03":
            return []
        return [_SourceBoundAttachmentRecord(), _SingleSourceAttachmentRecord()]

def oa_2035_attachment_evidences() -> list[dict[str, str]]:
    base = {
        "issue_date": "2026-03-04",
        "paid_at": "2026-03-04 09:00:00",
        "buyer_name": "云南溯源科技有限公司",
    }
    return [
        {
            **base,
            "evidence_id": "oa2035-inv-25",
            "evidence_type": "machine_invoice",
            "document_kind": "云南通用机打发票",
            "invoice_no": "20350025",
            "seller_name": "云南高速公路联网收费有限公司",
            "amount": "25.00",
            "total_with_tax": "25.00",
            "tax_amount": "0.00",
            "source_expense_row_index": "0",
            "source_expense_item_id": "oa-2035:item:0:toll",
            "source_attachment_key": "oa-2035:item:0:att:invoice-a",
            "source_attachment_name": "过路费机打发票合图.jpg",
        },
        {
            **base,
            "evidence_id": "oa2035-inv-23",
            "evidence_type": "machine_invoice",
            "document_kind": "云南通用机打发票",
            "invoice_no": "20350023",
            "seller_name": "云南高速公路联网收费有限公司",
            "amount": "23.00",
            "total_with_tax": "23.00",
            "tax_amount": "0.00",
            "source_expense_row_index": "0",
            "source_expense_item_id": "oa-2035:item:0:toll",
            "source_attachment_key": "oa-2035:item:0:att:invoice-b",
            "source_attachment_name": "过路费机打发票合图.jpg",
        },
        {
            **base,
            "evidence_id": "oa2035-inv-200",
            "evidence_type": "tax_invoice",
            "document_kind": "增值税电子普通发票",
            "invoice_no": "20350200",
            "seller_name": "云南中油严家山交通服务有限公司",
            "amount": "176.99",
            "total_with_tax": "200.00",
            "tax_amount": "23.01",
            "source_expense_row_index": "1",
            "source_expense_item_id": "oa-2035:item:1:fuel",
            "source_attachment_key": "oa-2035:item:1:att:invoice",
            "source_attachment_name": "加油费电子发票.pdf",
        },
        {
            **base,
            "evidence_id": "oa2035-pay-25",
            "evidence_type": "payment_receipt",
            "document_kind": "微信支付凭证",
            "merchant_name": "云南高速公路联网收费有限公司",
            "transaction_no": "wx-toll-25",
            "payment_method": "微信",
            "amount": "25.00",
            "source_expense_row_index": "0",
            "source_expense_item_id": "oa-2035:item:0:toll",
            "source_attachment_key": "oa-2035:item:0:att:payment-25",
            "source_attachment_name": "微信付款凭证25.jpg",
        },
        {
            **base,
            "evidence_id": "oa2035-pay-23",
            "evidence_type": "payment_receipt",
            "document_kind": "微信支付凭证",
            "merchant_name": "云南高速公路联网收费有限公司",
            "transaction_no": "wx-toll-23",
            "payment_method": "微信",
            "amount": "23.00",
            "source_expense_row_index": "0",
            "source_expense_item_id": "oa-2035:item:0:toll",
            "source_attachment_key": "oa-2035:item:0:att:payment-23",
            "source_attachment_name": "微信付款凭证23.jpg",
        },
        {
            **base,
            "evidence_id": "oa2035-pay-200",
            "evidence_type": "payment_receipt",
            "document_kind": "微信支付凭证",
            "merchant_name": "云南中油严家山交通服务有限公司",
            "transaction_no": "wx-fuel-200",
            "payment_method": "微信",
            "amount": "200.00",
            "source_expense_row_index": "1",
            "source_expense_item_id": "oa-2035:item:1:fuel",
            "source_attachment_key": "oa-2035:item:1:att:payment",
            "source_attachment_name": "微信加油付款凭证200.jpg",
        },
    ]

class _FailingOverrideStateStore:
    def save_workbench_overrides(self, snapshot: dict[str, object], *, changed_row_ids: list[str] | None = None) -> None:
        raise TimeoutError("mock override persistence timeout")

    def save_workbench_read_models(
        self,
        snapshot: dict[str, object],
        *,
        changed_scope_keys: list[str] | None = None,
    ) -> None:
        return None

class _FailingPairRelationStateStore:
    def save_workbench_pair_relations(
        self,
        snapshot: dict[str, object],
        *,
        changed_case_ids: list[str] | None = None,
    ) -> None:
        raise TimeoutError("mock pair relation persistence timeout")

    def save_workbench_read_models(
        self,
        snapshot: dict[str, object],
        *,
        changed_scope_keys: list[str] | None = None,
    ) -> None:
        return None

class _FailingReadModelStateStore:
    def save_workbench_read_models(
        self,
        snapshot: dict[str, object],
        *,
        changed_scope_keys: list[str] | None = None,
    ) -> None:
        raise TimeoutError("mock read model persistence timeout")

class _BankCategoryPersistenceSpyStateStore:
    def __init__(self) -> None:
        self.full_save_calls = 0
        self.category_save_calls = 0
        self.read_model_save_calls = 0
        self.turnover_relation_save_calls = 0

    def save(self, payload: dict[str, object]) -> None:
        self.full_save_calls += 1
        raise AssertionError("bank category saves must not persist full application state")

    def save_bank_transaction_categories(self, snapshot: dict[str, object]) -> None:
        self.category_save_calls += 1

    def save_workbench_read_models(
        self,
        snapshot: dict[str, object],
        *,
        changed_scope_keys: list[str] | None = None,
    ) -> None:
        self.read_model_save_calls += 1

    def save_turnover_relations(self, snapshot: dict[str, object]) -> None:
        self.turnover_relation_save_calls += 1
