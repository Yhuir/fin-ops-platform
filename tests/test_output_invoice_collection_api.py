from __future__ import annotations

from copy import deepcopy
from decimal import Decimal
from http import HTTPStatus
from io import BytesIO
import json
from pathlib import Path
import tempfile
from typing import Any
import unittest
from urllib.parse import quote

from openpyxl import load_workbook

from fin_ops_platform.app.server import build_application
from fin_ops_platform.domain.enums import InvoiceType, TransactionDirection
from fin_ops_platform.domain.models import BankTransaction, Counterparty, Invoice
from fin_ops_platform.services.imports import ImportNormalizationService
from fin_ops_platform.services.invoice_usage_collection_source_versions import output_invoice_collection_source_versions
from fin_ops_platform.services.output_invoice_collection_service import OutputInvoiceCollectionQueryService
from fin_ops_platform.services.workbench_pair_relation_service import WorkbenchPairRelationService


class FakeOutputRelationFacade:
    def __init__(self, relations: list[dict[str, Any]]) -> None:
        self.relations = [dict(relation) for relation in relations]

    def get_by_row_ids(self, row_ids: list[str], **_kwargs: Any) -> dict[str, Any]:
        wanted = {str(row_id) for row_id in row_ids}
        groups = [self._group(relation) for relation in self.relations if wanted & set(relation.get("row_ids") or [])]
        return self._payload(groups)

    def list_by_month(self, _month: str, **_kwargs: Any) -> dict[str, Any]:
        return self._payload([self._group(relation) for relation in self.relations])

    def _payload(self, groups: list[dict[str, Any]]) -> dict[str, Any]:
        rows: list[dict[str, Any]] = []
        for group in groups:
            group_id = str(group["group_id"])
            payload = group["payload"]
            for row_id, row_type in zip(payload["row_ids"], payload["row_types"]):
                rows.append({"row_id": row_id, "row_type": row_type, "relation_status": "linked", "group_ids": [group_id]})
        return {"status": "fresh", "rows": rows, "groups": groups, "source_versions": {}, "read_model_scope_keys": []}

    @staticmethod
    def _group(relation: dict[str, Any]) -> dict[str, Any]:
        case_id = str(relation.get("case_id") or "")
        row_ids = [str(row_id) for row_id in list(relation.get("row_ids") or [])]
        row_types = [str(row_type) for row_type in list(relation.get("row_types") or [])]
        return {
            "group_id": case_id,
            "scope_month": relation.get("month_scope") or "2026-05",
            "oa_row_ids": [row_id for row_id, row_type in zip(row_ids, row_types) if row_type == "oa"],
            "bank_transaction_ids": [row_id for row_id, row_type in zip(row_ids, row_types) if row_type == "bank"],
            "input_invoice_ids": [],
            "output_invoice_ids": [row_id for row_id, row_type in zip(row_ids, row_types) if row_type == "invoice"],
            "payload": {
                "case_id": case_id,
                "row_ids": row_ids,
                "row_types": row_types,
                "relation_mode": relation.get("relation_mode") or "",
                "amount_check": dict(relation.get("amount_check") or {}),
                "special_metadata": dict(relation.get("special_metadata") or {}),
            },
        }


class OutputInvoiceCollectionApiTests(unittest.TestCase):
    def test_rows_route_returns_output_invoice_collection_read_model(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            self._install_service(
                app,
                invoices=[
                    self._invoice("out-api-1", "1001", "甲客户", total_with_tax="30.00"),
                    self._invoice("out-api-2", "1002", "乙客户", total_with_tax="10.00"),
                    self._invoice("out-api-3", "1003", "甲客户", total_with_tax="20.00"),
                ],
            )
            filters = quote(json.dumps([{"field": "buyer_name", "operator": "in", "values": ["甲客户"]}]))

            response = app.handle_request(
                "GET",
                f"/api/output-invoice-collections/rows?page=1&page_size=1&filters={filters}&sort_field=total_with_tax&sort_direction=desc",
            )

        payload = json.loads(response.body)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["readModelStatus"], "live_query")
        self.assertEqual(payload["pagination"], {"page": 1, "pageSize": 1, "total": 2})
        self.assertEqual(payload["rows"][0]["invoiceId"], "out-api-1")
        self.assertEqual(payload["rows"][0]["invoice"]["buyerName"], "甲客户")

    def test_export_preview_and_download_use_current_filter_without_pagination(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            self._install_service(
                app,
                invoices=[
                    self._invoice("out-export-1", "9001", "甲客户", total_with_tax="300.00"),
                    self._invoice("out-export-2", "9002", "乙客户", total_with_tax="100.00"),
                    self._invoice("out-export-3", "9003", "甲客户", total_with_tax="200.00"),
                ],
            )
            filters = quote(json.dumps([{"field": "buyer_name", "operator": "in", "values": ["甲客户"]}]))

            preview_response = app.handle_request(
                "GET",
                f"/api/output-invoice-collections/export-preview?filters={filters}&sort_field=total_with_tax&sort_direction=desc",
            )
            download_response = app.handle_request(
                "GET",
                f"/api/output-invoice-collections/export?filters={filters}&sort_field=total_with_tax&sort_direction=desc",
            )

        preview = json.loads(preview_response.body)
        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(preview["row_count"], 2)
        self.assertIn("发票号码", preview["columns"])
        self.assertIn("红蓝票依据", preview["columns"])
        self.assertEqual(preview["sample_rows"][0]["发票号码"], "9001")
        self.assertEqual(preview["sample_rows"][0]["购方"], "甲客户")

        self.assertEqual(download_response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_response.headers["Content-Type"])
        self.assertIn("filename*=", download_response.headers["Content-Disposition"])
        workbook = load_workbook(BytesIO(download_response.body))
        sheet = workbook.active
        self.assertEqual(sheet.title, "销项收款")
        self.assertEqual(sheet.cell(row=1, column=2).value, "发票号码")
        self.assertEqual(sheet.cell(row=2, column=2).value, "9001")
        self.assertEqual(sheet.cell(row=2, column=4).value, "甲客户")
        self.assertEqual(sheet.cell(row=3, column=2).value, "9003")
        self.assertIsNone(sheet.cell(row=4, column=2).value)

    def test_export_rejects_row_count_over_contract_limit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            self._install_service(app, invoices=[self._invoice("out-limit", "9101", "超量客户")])
            row = json.loads(app.handle_request("GET", "/api/output-invoice-collections/rows").body)["rows"][0]

            def too_many_rows(_query: dict[str, list[str]]) -> dict[str, object]:
                return {
                    "rows": [deepcopy(row) for _index in range(20001)],
                    "pagination": {"page": 1, "pageSize": 200, "total": 20001},
                    "summary": {},
                    "read_model_status": "fresh",
                }

            app._get_output_invoice_collection_all_rows_from_sql_read_model = too_many_rows
            preview_response = app.handle_request("GET", "/api/output-invoice-collections/export-preview")
            download_response = app.handle_request("GET", "/api/output-invoice-collections/export")

        self.assertEqual(preview_response.status_code, 400)
        self.assertEqual(
            json.loads(preview_response.body)["error"]["code"],
            "output_invoice_collection_export_row_limit_exceeded",
        )
        self.assertEqual(download_response.status_code, 400)
        self.assertEqual(
            json.loads(download_response.body)["error"]["details"]["limit"],
            20000,
        )

    def test_detail_rules_preview_history_and_relation_routes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            invoice = self._invoice("out-detail", "2001", "详情客户")
            bank = self._bank("bank-detail", "100.00", TransactionDirection.INFLOW)
            pair_service = WorkbenchPairRelationService()
            pair_service.create_active_relation(
                case_id="case-detail",
                row_ids=[invoice.id, bank.id],
                row_types=["invoice", "bank"],
                relation_mode="manual_confirmed",
                created_by="tester",
                amount_check={"matched": True},
            )
            self._install_service(app, invoices=[invoice], transactions=[bank], pair_service=pair_service)

            rows_response = app.handle_request("GET", "/api/output-invoice-collections/rows")
            row = json.loads(rows_response.body)["rows"][0]
            filter_response = app.handle_request("GET", "/api/output-invoice-collections/filter-options?month=2026-05")
            rules_response = app.handle_request("GET", "/api/output-invoice-collections/status-rules")
            invoice_response = app.handle_request("GET", "/api/output-invoice-collections/invoices/out-detail/detail")
            bank_response = app.handle_request("GET", "/api/output-invoice-collections/bank-transactions/bank-detail/detail")
            relation_response = app.handle_request(
                "GET",
                f"/api/output-invoice-collections/rows/{row['id']}/relation-details?kind=bank",
            )
            preview_response = app.handle_request(
                "POST",
                "/api/output-invoice-collections/receipt-preview",
                body=json.dumps({"rowId": row["id"]}),
            )
            history_response = app.handle_request(
                "GET",
                "/api/output-invoice-collections/receipts/history?invoice_id=out-detail",
            )

        self.assertEqual(filter_response.status_code, 200)
        self.assertEqual(rules_response.status_code, 200)
        self.assertEqual(invoice_response.status_code, 200)
        self.assertEqual(bank_response.status_code, 200)
        self.assertEqual(relation_response.status_code, 200)
        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(history_response.status_code, 200)
        self.assertIn("collection_status", [field["field"] for field in json.loads(filter_response.body)["fields"]])
        self.assertEqual(json.loads(rules_response.body)["rules"][0]["label"], "开票已收款，冲红并退款")
        self.assertEqual(json.loads(invoice_response.body)["id"], "out-detail")
        self.assertEqual(json.loads(bank_response.body)["id"], "bank-detail")
        self.assertEqual(json.loads(relation_response.body)["kind"], "bank")
        self.assertTrue(json.loads(preview_response.body)["canPreview"])
        self.assertTrue(json.loads(history_response.body)["sourceAvailable"])
        self.assertEqual(json.loads(history_response.body)["receipts"], [])

    def test_invoice_relation_details_returns_all_related_output_invoices(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            invoices = [
                self._invoice("out-relation-a", "3001", "多发票客户", total_with_tax="300.00"),
                self._invoice("out-relation-b", "3002", "多发票客户", total_with_tax="100.00"),
                self._invoice("out-relation-c", "3003", "多发票客户", total_with_tax="200.00"),
            ]
            pair_service = WorkbenchPairRelationService()
            pair_service.create_active_relation(
                case_id="case-output-invoices",
                row_ids=[invoice.id for invoice in invoices],
                row_types=["invoice", "invoice", "invoice"],
                relation_mode="manual_confirmed",
                created_by="tester",
                amount_check={"matched": True},
            )
            self._install_service(app, invoices=invoices, pair_service=pair_service)

            rows_response = app.handle_request("GET", "/api/output-invoice-collections/rows")
            row = json.loads(rows_response.body)["rows"][0]
            relation_response = app.handle_request(
                "GET",
                f"/api/output-invoice-collections/rows/{row['id']}/relation-details?kind=invoice",
            )

        payload = json.loads(relation_response.body)
        self.assertEqual(relation_response.status_code, 200)
        self.assertEqual(row["invoiceRelations"]["relationCount"], 3)
        self.assertEqual(row["invoiceRelations"]["totalWithTax"], "600.00")
        self.assertEqual(payload["kind"], "invoice")
        self.assertEqual(payload["relationCount"], 3)
        self.assertEqual(
            {summary["invoiceId"] for summary in payload["summaries"]},
            {"out-relation-a", "out-relation-b", "out-relation-c"},
        )

    def test_detail_routes_require_output_collection_read_session(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            invoice = self._invoice("out-auth-detail", "2101", "权限客户")
            bank = self._bank("bank-auth-detail", "100.00", TransactionDirection.INFLOW)
            self._install_service(app, invoices=[invoice], transactions=[bank])
            row = json.loads(app.handle_request("GET", "/api/output-invoice-collections/rows").body)["rows"][0]

            def deny_read_session(headers: dict[str, str] | None = None) -> tuple[None, object]:
                return None, app._json_response(
                    HTTPStatus.UNAUTHORIZED,
                    {"error": {"code": "oa_session_required", "message": "需要登录 OA。"}},
                )

            app._resolve_output_invoice_collection_read_session = deny_read_session
            invoice_response = app.handle_request("GET", "/api/output-invoice-collections/invoices/out-auth-detail/detail")
            bank_response = app.handle_request("GET", "/api/output-invoice-collections/bank-transactions/bank-auth-detail/detail")
            relation_response = app.handle_request(
                "GET",
                f"/api/output-invoice-collections/rows/{row['id']}/relation-details?kind=bank",
            )

        self.assertEqual(invoice_response.status_code, 401)
        self.assertEqual(bank_response.status_code, 401)
        self.assertEqual(relation_response.status_code, 401)

    def test_routes_return_structured_validation_and_not_found_errors(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            self._install_service(app, invoices=[])

            invalid_page = app.handle_request("GET", "/api/output-invoice-collections/rows?page=0")
            invalid_sort = app.handle_request("GET", "/api/output-invoice-collections/rows?sort_field=unknown")
            invalid_filters = quote(json.dumps([{"field": "bad", "operator": "equals", "value": "x"}]))
            invalid_filter = app.handle_request(
                "GET",
                f"/api/output-invoice-collections/rows?filters={invalid_filters}",
            )
            missing_detail = app.handle_request("GET", "/api/output-invoice-collections/invoices/missing/detail")

        self.assertEqual(invalid_page.status_code, 400)
        self.assertEqual(json.loads(invalid_page.body)["error"]["code"], "invalid_paging")
        self.assertEqual(invalid_sort.status_code, 400)
        self.assertEqual(json.loads(invalid_sort.body)["error"]["code"], "invalid_sort_field")
        self.assertEqual(invalid_filter.status_code, 400)
        self.assertEqual(json.loads(invalid_filter.body)["error"]["code"], "invalid_filter_field")
        self.assertEqual(missing_detail.status_code, 404)
        self.assertEqual(json.loads(missing_detail.body)["error"]["code"], "invoice_not_found")

    def test_lifecycle_write_routes_overlay_rows_and_create_real_receipt_history(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            invoice = self._invoice("out-lifecycle", "3001", "生命周期客户", total_with_tax="100.00")
            red_invoice = self._invoice("out-red", "3002", "生命周期客户", total_with_tax="-100.00")
            bank = self._bank("bank-lifecycle", "100.00", TransactionDirection.INFLOW)
            pair_service = WorkbenchPairRelationService()
            pair_service.create_active_relation(
                case_id="case-lifecycle",
                row_ids=[invoice.id, bank.id],
                row_types=["invoice", "bank"],
                relation_mode="manual_confirmed",
                created_by="tester",
                amount_check={"matched": True},
            )
            self._install_service(app, invoices=[invoice, red_invoice], transactions=[bank], pair_service=pair_service)
            rows = json.loads(app.handle_request("GET", "/api/output-invoice-collections/rows").body)["rows"]
            row = next(item for item in rows if item["invoiceId"] == "out-lifecycle")

            status_response = app.handle_request(
                "PUT",
                f"/api/output-invoice-collections/rows/{row['id']}/collection-status",
                body=json.dumps(
                    {
                        "statusCode": "pending_red_invoice",
                        "expectedCollectionDate": "2026-06-20",
                        "note": "待冲红",
                        "expectedVersion": 0,
                    }
                ),
            )
            reminder_response = app.handle_request(
                "PUT",
                f"/api/output-invoice-collections/rows/{row['id']}/collection-reminder",
                body=json.dumps({"remindAt": "2026-06-15T09:00:00+08:00", "channel": "oa", "note": "提醒"}),
            )
            reminder_id = json.loads(reminder_response.body)["reminder"]["id"]
            red_relation_response = app.handle_request(
                "POST",
                f"/api/output-invoice-collections/rows/{row['id']}/red-invoice-relations",
                body=json.dumps(
                    {
                        "relatedInvoiceIdentityKey": "id:out-red",
                        "relatedInvoiceId": "out-red",
                        "relationType": "red_invoice",
                        "evidence": "客户邮件确认",
                    }
                ),
            )
            relation_id = json.loads(red_relation_response.body)["relation"]["id"]
            receipt_response = app.handle_request(
                "POST",
                f"/api/output-invoice-collections/rows/{row['id']}/receipts",
                body=json.dumps({"bankTransactionId": "bank-lifecycle", "idempotencyKey": "receipt-api-1"}),
            )
            receipt_id = json.loads(receipt_response.body)["receipt"]["id"]
            reminder_delete_response = app.handle_request(
                "DELETE",
                f"/api/output-invoice-collections/rows/{row['id']}/collection-reminder/{reminder_id}",
            )
            red_relation_delete_response = app.handle_request(
                "DELETE",
                f"/api/output-invoice-collections/red-invoice-relations/{relation_id}",
            )
            void_response = app.handle_request(
                "POST",
                f"/api/output-invoice-collections/receipts/{receipt_id}/void",
                body=json.dumps({"reason": "API 作废"}),
            )
            reissue_response = app.handle_request(
                "POST",
                f"/api/output-invoice-collections/receipts/{receipt_id}/reissue",
                body=json.dumps({"reason": "API 重开"}),
            )
            duplicate_reissue_response = app.handle_request(
                "POST",
                f"/api/output-invoice-collections/receipts/{receipt_id}/reissue",
                body=json.dumps({"reason": "API 重复重开"}),
            )
            missing_void_response = app.handle_request(
                "POST",
                "/api/output-invoice-collections/receipts/not-found/void",
                body=json.dumps({"reason": "missing"}),
            )
            history_response = app.handle_request(
                "GET",
                "/api/output-invoice-collections/receipts/history?invoice_id=out-lifecycle",
            )
            refreshed_row = json.loads(app.handle_request("GET", "/api/output-invoice-collections/rows").body)["rows"][0]

        self.assertEqual(status_response.status_code, 200)
        self.assertEqual(reminder_response.status_code, 200)
        self.assertEqual(red_relation_response.status_code, 200)
        self.assertEqual(receipt_response.status_code, 200)
        self.assertEqual(reminder_delete_response.status_code, 200)
        self.assertEqual(red_relation_delete_response.status_code, 200)
        self.assertEqual(void_response.status_code, 200)
        self.assertEqual(reissue_response.status_code, 200)
        self.assertEqual(duplicate_reissue_response.status_code, 409)
        self.assertEqual(missing_void_response.status_code, 404)
        self.assertEqual(refreshed_row["collectionStatus"]["code"], "pending_red_invoice")
        self.assertIsNone(refreshed_row["collectionStatus"]["reminder"])
        self.assertFalse(any(item["source"] == "manual" for item in refreshed_row["redInvoiceRelation"]["summaries"]))
        self.assertEqual(refreshed_row["receipt"]["status"], "issued")
        self.assertTrue(json.loads(history_response.body)["sourceAvailable"])
        self.assertEqual([item["status"] for item in json.loads(history_response.body)["receipts"]], ["issued", "voided"])

    def test_sql_fresh_rows_route_applies_lifecycle_overlay_before_response(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            app = build_application(data_dir=Path(temp_dir))
            invoice = self._invoice("out-sql-overlay", "4001", "SQL 覆盖客户", total_with_tax="100.00")
            red_invoice = self._invoice("out-sql-red", "4002", "SQL 覆盖客户", total_with_tax="-100.00")
            self._install_service(app, invoices=[invoice, red_invoice])
            live_rows = json.loads(app.handle_request("GET", "/api/output-invoice-collections/rows").body)["rows"]
            row = next(item for item in live_rows if item["invoiceId"] == "out-sql-overlay")
            stale_sql_row = deepcopy(row)
            stale_sql_row["invoiceIdentityKey"] = f"id:{row['invoiceId']}"

            status_response = app.handle_request(
                "PUT",
                f"/api/output-invoice-collections/rows/{row['id']}/collection-status",
                body=json.dumps(
                    {
                        "statusCode": "pending_red_invoice",
                        "expectedCollectionDate": "2026-06-20",
                        "note": "SQL fresh 路径也必须展示人工状态",
                        "expectedVersion": 0,
                    }
                ),
            )
            relation_response = app.handle_request(
                "POST",
                f"/api/output-invoice-collections/rows/{row['id']}/red-invoice-relations",
                body=json.dumps(
                    {
                        "relatedInvoiceIdentityKey": "id:out-sql-red",
                        "relatedInvoiceId": "out-sql-red",
                        "relationType": "red_invoice",
                        "evidence": "SQL overlay 验证",
                    }
                ),
            )
            app._output_invoice_collection_sql_read_repository = _FreshOutputInvoiceCollectionSqlRepository([stale_sql_row])

            response = app.handle_request("GET", "/api/output-invoice-collections/rows?month=2026-05")

        payload = json.loads(response.body)
        self.assertEqual(status_response.status_code, 200)
        self.assertEqual(relation_response.status_code, 200)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["readModelStatus"], "fresh")
        self.assertEqual(payload["rows"][0]["collectionStatus"]["manualOverride"]["note"], "SQL fresh 路径也必须展示人工状态")
        red_relation_summaries = payload["rows"][0]["redInvoiceRelation"]["summaries"]
        self.assertTrue(
            any(
                item.get("source") == "manual" and item.get("relatedInvoiceId") == "out-sql-red"
                for item in red_relation_summaries
            )
        )

    @staticmethod
    def _install_service(
        app: object,
        *,
        invoices: list[Invoice],
        transactions: list[BankTransaction] | None = None,
        pair_service: WorkbenchPairRelationService | None = None,
    ) -> None:
        import_service = ImportNormalizationService(
            existing_invoices=invoices,
            existing_transactions=transactions or [],
        )
        relation_service = pair_service or WorkbenchPairRelationService()
        app._import_service = import_service
        app._workbench_pair_relation_service = relation_service
        app._output_invoice_collection_query_service = OutputInvoiceCollectionQueryService(
            import_service=import_service,
            relation_facade=FakeOutputRelationFacade(relation_service.list_active_relations()),
            lifecycle_repository=getattr(app, "_output_invoice_collection_lifecycle_repository", None),
        )

    @staticmethod
    def _invoice(invoice_id: str, invoice_no: str, buyer_name: str, *, total_with_tax: str = "100.00") -> Invoice:
        counterparty = Counterparty(
            id=f"cp-{invoice_id}",
            name=buyer_name,
            normalized_name=buyer_name,
            counterparty_type="customer",
            tax_no="91530000BUYER",
        )
        return Invoice(
            id=invoice_id,
            invoice_type=InvoiceType.OUTPUT,
            invoice_no=invoice_no,
            counterparty=counterparty,
            amount=Decimal(total_with_tax),
            signed_amount=Decimal(total_with_tax),
            invoice_date="2026-05-20",
            seller_name="云南溯源科技有限公司",
            buyer_name=buyer_name,
            seller_tax_no="91530000SELLER",
            buyer_tax_no="91530000BUYER",
            tax_rate="6%",
            tax_amount=Decimal("0.00"),
            total_with_tax=Decimal(total_with_tax),
            taxable_item_name="服务费",
            is_positive_invoice="是",
        )

    @staticmethod
    def _bank(transaction_id: str, amount: str, direction: TransactionDirection) -> BankTransaction:
        return BankTransaction(
            id=transaction_id,
            account_no="622200001234",
            txn_direction=direction,
            counterparty_name_raw="详情客户",
            amount=Decimal(amount),
            signed_amount=Decimal(amount) if direction == TransactionDirection.INFLOW else -Decimal(amount),
            txn_date="2026-05-21",
            trade_time="2026-05-21 10:00:00",
            imported_bank_name="中国银行",
            imported_bank_last4="1234",
            summary="服务费",
        )


class _FreshOutputInvoiceCollectionSqlRepository:
    def __init__(self, rows: list[dict[str, object]]) -> None:
        self._rows = rows

    def list_output_invoice_collection_rows(self, **kwargs: object) -> dict[str, object]:
        page = int(kwargs.get("page") or 1)
        page_size = int(kwargs.get("page_size") or 50)
        return {
            "rows": deepcopy(self._rows),
            "pagination": {"page": page, "pageSize": page_size, "total": len(self._rows)},
            "summary": {
                "invoiceCount": len(self._rows),
                "totalWithTax": "100.00",
                "collectedAmount": "0.00",
                "pendingAmount": "100.00",
                "pendingCollectionCount": 1,
                "partialCollectionCount": 0,
                "receiptPendingCount": 0,
            },
            "filterConfig": [],
            "read_model_status": "fresh",
            "source_versions": output_invoice_collection_source_versions(),
        }
