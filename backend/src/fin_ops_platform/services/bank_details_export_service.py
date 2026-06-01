from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
import re
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BANK_DETAIL_EXPORT_ROW_LIMIT = 20000
BANK_DETAIL_EXPORT_PAGE_SIZE = 500
BANK_DETAIL_EXPORT_COLUMNS = [
    "交易时间",
    "银行",
    "账号尾号",
    "对方户名",
    "收支方向",
    "收入金额",
    "支出金额",
    "余额",
    "自动分类",
    "自动分类主标签",
    "自动分类子标签",
    "自动分类第三级业务",
    "OA 关系",
    "发票关系",
    "用途/交易用途",
    "摘要",
    "备注/附言/客户附言",
    "流水 ID",
]
MONEY_COLUMNS = {"收入金额", "支出金额", "余额"}
PURPOSE_TEXT_LABELS = ("用途", "交易用途")
SUMMARY_TEXT_LABELS = ("摘要",)
NOTE_TEXT_LABELS = ("备注", "附言", "客户附言")


@dataclass(slots=True)
class BankDetailsExportResult:
    filename: str
    content: bytes
    row_count: int
    sheet_names: list[str]


class BankDetailsExportError(ValueError):
    def __init__(self, error_code: str, message: str) -> None:
        super().__init__(message)
        self.error_code = error_code


class BankDetailsExportService:
    def __init__(
        self,
        *,
        transaction_page_loader: Callable[..., dict[str, Any]],
        account_loader: Callable[..., dict[str, Any]],
    ) -> None:
        self._transaction_page_loader = transaction_page_loader
        self._account_loader = account_loader

    def export(
        self,
        *,
        mode: str,
        account_key: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        keyword: str | None = None,
        category_code: str | None = None,
        category_primary_label: str | None = None,
        category_sub_label: str | None = None,
        category_third_label: str | None = None,
        today: date | None = None,
    ) -> BankDetailsExportResult:
        normalized_mode = str(mode or "").strip().lower()
        if normalized_mode not in {"all", "account"}:
            raise BankDetailsExportError("bank_detail_export_invalid_mode", "导出模式无效。")
        normalized_account_key = str(account_key or "").strip()
        account_payload = self._account_payload(
            normalized_account_key,
            date_from=date_from,
            date_to=date_to,
        ) if normalized_mode == "account" else None
        rows = self._collect_rows(
            account_key=normalized_account_key if normalized_mode == "account" else None,
            date_from=date_from,
            date_to=date_to,
            keyword=keyword,
            category_code=category_code,
            category_primary_label=category_primary_label,
            category_sub_label=category_sub_label,
            category_third_label=category_third_label,
        )
        formal_rows = [self._formal_row(row) for row in rows]
        workbook = self._build_workbook(formal_rows, mode=normalized_mode, account_payload=account_payload)
        sheet_names = list(workbook.sheetnames)
        filename = self._filename(
            mode=normalized_mode,
            account_payload=account_payload,
            date_from=date_from,
            date_to=date_to,
            today=today,
        )
        return BankDetailsExportResult(
            filename=filename,
            content=self._serialize_workbook(workbook),
            row_count=len(formal_rows),
            sheet_names=sheet_names,
        )

    def _account_payload(self, account_key: str, *, date_from: str | None, date_to: str | None) -> dict[str, Any]:
        if not account_key:
            raise BankDetailsExportError("bank_detail_export_account_required", "请选择具体银行账户后再导出当前账户。")
        payload = self._account_loader(date_from=date_from, date_to=date_to)
        for account in list(payload.get("accounts") or []):
            if isinstance(account, dict) and str(account.get("account_key") or "").strip() == account_key:
                return dict(account)
        raise BankDetailsExportError("bank_detail_export_account_not_found", "当前银行账户不存在或不在当前筛选范围内。")

    def _collect_rows(
        self,
        *,
        account_key: str | None,
        date_from: str | None,
        date_to: str | None,
        keyword: str | None,
        category_code: str | None,
        category_primary_label: str | None,
        category_sub_label: str | None,
        category_third_label: str | None,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        page = 1
        total: int | None = None
        while total is None or len(rows) < total:
            payload = self._transaction_page_loader(
                account_key=account_key,
                date_from=date_from,
                date_to=date_to,
                keyword=keyword,
                category_code=category_code,
                category_primary_label=category_primary_label,
                category_sub_label=category_sub_label,
                category_third_label=category_third_label,
                page=page,
                page_size=BANK_DETAIL_EXPORT_PAGE_SIZE,
            )
            pagination = payload.get("pagination") if isinstance(payload.get("pagination"), dict) else {}
            total = int(pagination.get("total") or len(payload.get("rows") or []))
            if total > BANK_DETAIL_EXPORT_ROW_LIMIT:
                raise BankDetailsExportError(
                    "bank_detail_export_row_limit_exceeded",
                    f"当前筛选命中 {total} 行，超过 {BANK_DETAIL_EXPORT_ROW_LIMIT} 行导出上限，请缩小筛选范围。",
                )
            page_rows = [dict(row) for row in list(payload.get("rows") or []) if isinstance(row, dict)]
            rows.extend(page_rows)
            if not page_rows:
                break
            page += 1
        return rows[:total or len(rows)]

    @classmethod
    def _formal_row(cls, row: dict[str, Any]) -> dict[str, Any]:
        direction = str(row.get("direction") or "").strip()
        amount = cls._money(row.get("amount"))
        text_fields = cls._bank_text_display_fields(row)
        return {
            "交易时间": cls._trade_time_text(row.get("trade_time")),
            "银行": cls._text(row.get("bank_name")),
            "账号尾号": cls._text(row.get("account_last4")),
            "对方户名": cls._text(row.get("counterparty_name")),
            "收支方向": "收入" if direction == "income" else "支出",
            "收入金额": amount if direction == "income" else None,
            "支出金额": amount if direction != "income" else None,
            "余额": cls._money(row.get("balance")),
            "自动分类": cls._text(row.get("auto_category_label") or row.get("effective_category_label")) or "-",
            "自动分类主标签": cls._text(
                row.get("auto_category_primary_label")
                or row.get("effective_category_primary_label")
                or row.get("category_primary_label")
            ) or "-",
            "自动分类子标签": cls._text(
                row.get("auto_category_sub_label")
                or row.get("effective_category_sub_label")
                or row.get("category_sub_label")
            ) or "-",
            "自动分类第三级业务": cls._text(
                row.get("auto_category_third_label")
                or row.get("effective_category_third_label")
                or row.get("category_third_label")
            ) or "-",
            "OA 关系": cls._text(row.get("oa_relation_tag")) or "无oa",
            "发票关系": cls._text(row.get("invoice_relation_tag")) or "无发票",
            "用途/交易用途": text_fields["purpose_text"],
            "摘要": text_fields["summary_text"],
            "备注/附言/客户附言": text_fields["note_text"],
            "流水 ID": cls._text(row.get("id")),
        }

    @classmethod
    def _build_workbook(
        cls,
        rows: list[dict[str, Any]],
        *,
        mode: str,
        account_payload: dict[str, Any] | None,
    ) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)
        if mode == "all":
            cls._append_sheet(workbook, "全部流水", rows)
            for bank_name in cls._bank_order(rows):
                cls._append_sheet(workbook, bank_name, [row for row in rows if row.get("银行") == bank_name])
        else:
            sheet_name = cls._sheet_name(cls._text((account_payload or {}).get("bank_name")) or "银行流水", set())
            cls._append_sheet(workbook, sheet_name, rows)
        return workbook

    @classmethod
    def _append_sheet(cls, workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet(cls._sheet_name(title, set(workbook.sheetnames)))
        sheet.append(BANK_DETAIL_EXPORT_COLUMNS)
        for row in rows:
            sheet.append([cls._excel_cell_value(row.get(column)) for column in BANK_DETAIL_EXPORT_COLUMNS])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(BANK_DETAIL_EXPORT_COLUMNS))}{max(sheet.max_row, 1)}"
        header_fill = PatternFill("solid", fgColor="E8EEF6")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="1F2937")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index, column_name in enumerate(BANK_DETAIL_EXPORT_COLUMNS, start=1):
            width = cls._column_width(column_name)
            sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = width
            for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=sheet.max_row):
                for item in cell:
                    item.alignment = Alignment(vertical="top", wrap_text=column_name not in MONEY_COLUMNS)
                    if column_name in MONEY_COLUMNS:
                        item.number_format = '#,##0.00'

    @staticmethod
    def _bank_order(rows: list[dict[str, Any]]) -> list[str]:
        seen: dict[str, None] = {}
        for row in rows:
            bank_name = str(row.get("银行") or "").strip() or "未知银行"
            seen.setdefault(bank_name, None)
        return list(seen)

    @staticmethod
    def _column_width(column_name: str) -> int:
        widths = {
            "交易时间": 20,
            "银行": 14,
            "账号尾号": 10,
            "对方户名": 30,
            "收支方向": 10,
            "收入金额": 14,
            "支出金额": 14,
            "余额": 14,
            "自动分类": 18,
            "自动分类主标签": 16,
            "自动分类子标签": 16,
            "自动分类第三级业务": 16,
            "OA 关系": 10,
            "发票关系": 12,
            "用途/交易用途": 24,
            "摘要": 24,
            "备注/附言/客户附言": 30,
            "流水 ID": 28,
        }
        return widths.get(column_name, 16)

    @classmethod
    def _filename(
        cls,
        *,
        mode: str,
        account_payload: dict[str, Any] | None,
        date_from: str | None,
        date_to: str | None,
        today: date | None,
    ) -> str:
        start = cls._compact_date(date_from)
        end = cls._compact_date(date_to)
        date_segment = f"{start}-{end}" if start and end else (today or date.today()).isoformat().replace("-", "")
        if mode == "account":
            bank = cls._sanitize_filename_part(cls._text((account_payload or {}).get("bank_name")) or "银行")
            last4 = cls._sanitize_filename_part(cls._text((account_payload or {}).get("account_last4")) or "")
            return f"银行明细_{bank}{last4}_{date_segment}.xlsx"
        return f"银行明细_当前筛选_全部银行_{date_segment}.xlsx"

    @staticmethod
    def _serialize_workbook(workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _sheet_name(value: str, existing: set[str]) -> str:
        base = re.sub(r"[:\\/?*\\[\\]]", "-", str(value or "Sheet").strip())[:31] or "Sheet"
        candidate = base
        index = 2
        while candidate in existing:
            suffix = f"_{index}"
            candidate = f"{base[:31 - len(suffix)]}{suffix}"
            index += 1
        return candidate

    @staticmethod
    def _compact_date(value: str | None) -> str:
        text = str(value or "").strip()
        return text[:10].replace("-", "") if text else ""

    @staticmethod
    def _sanitize_filename_part(value: str) -> str:
        return str(value or "").strip().replace("/", "-").replace("\\", "-").replace(":", "：")[:80]

    @staticmethod
    def _trade_time_text(value: Any) -> str:
        text = str(value or "").strip().replace("T", " ")
        if len(text) >= 25 and text[19] in {"+", "-"} and text[20:22].isdigit() and text[23:25].isdigit():
            return text[:19]
        if text.endswith("Z") and len(text) >= 20:
            return text[:19]
        return text

    @staticmethod
    def _money(value: Any) -> float | None:
        if value in (None, ""):
            return None
        try:
            return float(Decimal(str(value).replace(",", "")))
        except Exception:
            return None

    @staticmethod
    def _text(value: Any) -> str:
        return str(value or "").strip()

    @classmethod
    def _bank_text_display_fields(cls, row: dict[str, Any]) -> dict[str, str]:
        fields_by_label = cls._bank_text_fields_by_label(row.get("bank_text_fields"))
        summary_text = cls._first_field_value(fields_by_label, SUMMARY_TEXT_LABELS)
        purpose_text = cls._first_field_value(fields_by_label, PURPOSE_TEXT_LABELS)
        note_text = cls._first_field_value(fields_by_label, NOTE_TEXT_LABELS)
        if not fields_by_label:
            return cls._legacy_bank_text_display_fields(row)
        return {
            "purpose_text": purpose_text.strip(),
            "summary_text": summary_text.strip(),
            "note_text": note_text.strip(),
        }

    @classmethod
    def _legacy_bank_text_display_fields(cls, row: dict[str, Any]) -> dict[str, str]:
        bank_name = cls._text(row.get("bank_name"))
        summary = cls._text(row.get("summary_text") or row.get("summary"))
        purpose = cls._text(row.get("purpose_text") or row.get("purpose"))
        note = cls._text(row.get("note_text") or row.get("note") or row.get("remark"))
        if "民生" in bank_name:
            return {"purpose_text": "", "summary_text": "", "note_text": note or purpose or summary}
        if "交通" in bank_name or "光大" in bank_name:
            return {"purpose_text": "", "summary_text": summary or purpose or note, "note_text": ""}
        if "建设" in bank_name:
            return {"purpose_text": "", "summary_text": summary, "note_text": note or purpose}
        if "平安" in bank_name:
            return {"purpose_text": purpose or note, "summary_text": summary, "note_text": ""}
        if "工商" in bank_name:
            return {"purpose_text": purpose if purpose != note else "", "summary_text": summary, "note_text": note if note else ""}
        return {"purpose_text": purpose, "summary_text": summary, "note_text": note}

    @classmethod
    def _bank_text_fields_by_label(cls, value: Any) -> dict[str, str]:
        fields: dict[str, str] = {}
        if isinstance(value, dict):
            iterable = [{"label": label, "value": field_value} for label, field_value in value.items()]
        else:
            iterable = list(value or []) if isinstance(value, list) else []
        for item in iterable:
            if not isinstance(item, dict):
                continue
            label = cls._text(item.get("label"))
            field_value = cls._text(item.get("value"))
            if label and field_value and label not in fields:
                fields[label] = field_value
        return fields

    @staticmethod
    def _first_field_value(fields_by_label: dict[str, str], labels: tuple[str, ...]) -> str:
        for label in labels:
            value = fields_by_label.get(label)
            if value:
                return value
        return ""

    @staticmethod
    def _excel_cell_value(value: Any) -> Any:
        if not isinstance(value, str):
            return value
        text = value.strip()
        if len(text) > 1 and text[0] in {"=", "+", "-", "@"}:
            return f"'{text}"
        return text
