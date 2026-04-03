from __future__ import annotations

from collections import defaultdict
from decimal import Decimal, InvalidOperation
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from fin_ops_platform.domain.enums import TransactionDirection
from fin_ops_platform.services.imports import ImportNormalizationService
from fin_ops_platform.services.live_workbench_service import format_decimal


ZERO = Decimal("0.00")
EXCLUDED_COST_EXPENSE_TYPES = {"借款", "还款"}


class ProjectDetailExportService:
    def __init__(
        self,
        import_service: ImportNormalizationService,
        *,
        grouped_workbench_loader: Callable[[str], dict[str, Any]],
        raw_workbench_loader: Callable[[str], dict[str, Any]] | None = None,
    ) -> None:
        self._import_service = import_service
        self._grouped_workbench_loader = grouped_workbench_loader
        self._raw_workbench_loader = raw_workbench_loader

    def build_export_payload(self, *, month: str, project_name: str) -> dict[str, Any]:
        return self.build_export_payload_with_options(
            month=month,
            project_name=project_name,
            include_oa_details=True,
            include_invoice_details=True,
            include_exception_rows=True,
            include_ignored_rows=True,
            include_expense_content_summary=True,
            sort_by="time",
        )

    def build_export_payload_with_options(
        self,
        *,
        month: str,
        project_name: str,
        expense_types: list[str] | None = None,
        start_month: str | None = None,
        end_month: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        include_oa_details: bool,
        include_invoice_details: bool,
        include_exception_rows: bool,
        include_ignored_rows: bool,
        include_expense_content_summary: bool,
        sort_by: str,
    ) -> dict[str, Any]:
        normal_entries: list[dict[str, Any]] = []
        oa_rows: list[dict[str, Any]] = []
        invoice_rows: list[dict[str, Any]] = []
        exception_rows: list[dict[str, Any]] = []
        ignored_rows: list[dict[str, Any]] = []
        target_months = self._resolve_target_months(
            month,
            start_month=start_month or (start_date[:7] if start_date else None),
            end_month=end_month or (end_date[:7] if end_date else None),
        )
        resolved_sort_by = self._normalize_sort_by(sort_by)
        selected_expense_types = self._normalize_expense_types(expense_types)

        for scoped_month in target_months:
            grouped_payload = self._grouped_workbench_loader(scoped_month)
            for section in ("paired", "open"):
                section_payload = grouped_payload.get(section) or {}
                for group in section_payload.get("groups") or []:
                    selected_oa_rows = [
                        row
                        for row in (group.get("oa_rows") or [])
                        if self._clean_text(row.get("project_name")) == project_name
                    ]
                    if not selected_oa_rows:
                        continue

                    bank_rows = list(group.get("bank_rows") or [])
                    grouped_invoice_rows = list(group.get("invoice_rows") or [])
                    group_id = str(group.get("group_id") or "")
                    context = self._resolve_group_cost_context(selected_oa_rows)

                    if selected_expense_types:
                        if context is None or context["expense_type"] not in selected_expense_types:
                            continue

                    if self._group_has_exception(selected_oa_rows, bank_rows, grouped_invoice_rows):
                        exception_rows.extend(
                            self._build_exception_rows(
                                month=scoped_month,
                                project_name=project_name,
                                oa_rows=selected_oa_rows,
                                bank_rows=bank_rows,
                                invoice_rows=grouped_invoice_rows,
                            )
                        )
                        continue

                    if context is None:
                        exception_rows.extend(
                            self._build_ambiguous_context_rows(
                                month=scoped_month,
                                project_name=project_name,
                                oa_rows=selected_oa_rows,
                            )
                        )
                        continue

                    normal_entries.extend(
                        self._build_normal_entries(
                            month=scoped_month,
                            group_id=group_id,
                            bank_rows=bank_rows,
                            context=context,
                            oa_rows=selected_oa_rows,
                        )
                    )
                    if include_oa_details:
                        oa_rows.extend(self._build_oa_detail_rows(selected_oa_rows, group_id=group_id))
                    if include_invoice_details:
                        invoice_rows.extend(
                            self._build_invoice_detail_rows(
                                grouped_invoice_rows,
                                group_id=group_id,
                                project_name=project_name,
                            )
                        )

            if include_ignored_rows:
                ignored_rows.extend(
                    self._extract_ignored_rows(
                        project_name=project_name,
                        month=scoped_month,
                        expense_types=selected_expense_types,
                    )
                )

        normal_entries = self._sort_transaction_rows(normal_entries, resolved_sort_by)
        scope_label = self._build_scope_label(
            month=month,
            target_months=target_months,
            start_month=start_month,
            end_month=end_month,
        )

        return {
            "month": month,
            "project_name": project_name,
            "months": target_months,
            "options": {
                "include_oa_details": include_oa_details,
                "include_invoice_details": include_invoice_details,
                "include_exception_rows": include_exception_rows,
                "include_ignored_rows": include_ignored_rows,
                "include_expense_content_summary": include_expense_content_summary,
                "sort_by": resolved_sort_by,
                "scope_label": scope_label,
                "expense_types": sorted(selected_expense_types),
            },
            "summary": self._build_summary(
                project_name=project_name,
                normal_entries=normal_entries,
                oa_rows=oa_rows,
                invoice_rows=invoice_rows,
                exception_rows=exception_rows if include_exception_rows else [],
                ignored_rows=ignored_rows if include_ignored_rows else [],
                scope_label=scope_label,
            ),
            "expense_type_rows": self._build_expense_type_rows(normal_entries),
            "expense_content_rows": self._build_expense_content_rows(normal_entries) if include_expense_content_summary else [],
            "transaction_rows": normal_entries,
            "oa_rows": oa_rows,
            "invoice_rows": invoice_rows,
            "exception_rows": exception_rows if include_exception_rows else [],
            "ignored_rows": ignored_rows if include_ignored_rows else [],
        }

    def build_workbook(self, payload: dict[str, Any]) -> Workbook:
        options = dict(payload.get("options") or {})
        include_oa_details = bool(options.get("include_oa_details", True))
        include_invoice_details = bool(options.get("include_invoice_details", True))
        include_exception_rows = bool(options.get("include_exception_rows", True))
        include_ignored_rows = bool(options.get("include_ignored_rows", True))
        include_expense_content_summary = bool(options.get("include_expense_content_summary", True))
        workbook = Workbook()
        intro_sheet = workbook.active
        intro_sheet.title = "导出说明"
        self._fill_key_value_sheet(
            intro_sheet,
            [
                ("项目名称", payload["project_name"]),
                ("统计范围", payload["summary"]["scope_label"]),
                ("月份列表", "、".join(payload["months"])),
                ("数据口径", "仅统计可稳定取得 OA 项目名称、费用类型、费用内容的支出流水"),
                ("导出结构", self._build_export_structure_label(options)),
            ],
            widths=[18, 84],
        )

        summary_sheet = workbook.create_sheet("项目汇总")
        self._fill_key_value_sheet(
            summary_sheet,
            [
                ("项目名称", payload["summary"]["project_name"]),
                ("统计期间", payload["summary"]["scope_label"]),
                ("总支出金额", payload["summary"]["total_amount"]),
                ("支出流水笔数", payload["summary"]["transaction_count"]),
                ("费用类型数", payload["summary"]["expense_type_count"]),
                ("已关联OA笔数", payload["summary"]["oa_count"]),
                ("已关联发票笔数", payload["summary"]["invoice_count"]),
                ("已处理异常笔数", payload["summary"]["exception_count"]),
                ("已忽略笔数", payload["summary"]["ignored_count"]),
            ],
            widths=[18, 28],
        )

        expense_type_sheet = workbook.create_sheet("按费用类型汇总")
        self._fill_table_sheet(
            expense_type_sheet,
            headers=["费用类型", "金额", "占比", "笔数", "费用内容数"],
            rows=[
                [
                    row["expense_type"],
                    row["total_amount"],
                    row["percentage"],
                    row["transaction_count"],
                    row["expense_content_count"],
                ]
                for row in payload["expense_type_rows"]
            ],
            widths=[24, 14, 12, 10, 12],
        )

        if include_expense_content_summary:
            expense_content_sheet = workbook.create_sheet("按费用内容汇总")
            self._fill_table_sheet(
                expense_content_sheet,
                headers=["费用类型", "费用内容", "金额", "笔数"],
                rows=[
                    [
                        row["expense_type"],
                        row["expense_content"],
                        row["total_amount"],
                        row["transaction_count"],
                    ]
                    for row in payload["expense_content_rows"]
                ],
                widths=[24, 42, 14, 10],
            )

        transaction_sheet = workbook.create_sheet("流水明细")
        self._fill_table_sheet(
            transaction_sheet,
            headers=["时间", "交易流水ID", "资金方向", "对方户名", "支付账户", "金额", "备注", "项目名称", "费用类型", "费用内容", "OA单号", "关联组ID"],
            rows=[
                [
                    row["trade_time"],
                    row["transaction_id"],
                    row["direction"],
                    row["counterparty_name"],
                    row["payment_account_label"],
                    row["amount"],
                    row["remark"],
                    row["project_name"],
                    row["expense_type"],
                    row["expense_content"],
                    row["oa_form_nos"],
                    row["group_id"],
                ]
                for row in payload["transaction_rows"]
            ],
            widths=[22, 18, 12, 24, 22, 14, 28, 28, 24, 34, 18, 18],
        )

        if include_oa_details:
            oa_sheet = workbook.create_sheet("OA关联明细")
            self._fill_table_sheet(
                oa_sheet,
                headers=["OA单号", "申请人", "项目名称", "费用类型", "费用内容", "OA金额", "关联组ID"],
                rows=[
                    [
                        row["oa_form_no"],
                        row["applicant"],
                        row["project_name"],
                        row["expense_type"],
                        row["expense_content"],
                        row["amount"],
                        row["group_id"],
                    ]
                    for row in payload["oa_rows"]
                ],
                widths=[18, 12, 28, 24, 34, 14, 18],
            )

        if include_invoice_details:
            invoice_sheet = workbook.create_sheet("发票关联明细")
            self._fill_table_sheet(
                invoice_sheet,
                headers=["发票号码", "销方名称", "购方名称", "发票金额", "税额", "项目名称", "关联状态", "关联组ID"],
                rows=[
                    [
                        row["invoice_no"],
                        row["seller_name"],
                        row["buyer_name"],
                        row["amount"],
                        row["tax_amount"],
                        row["project_name"],
                        row["status_label"],
                        row["group_id"],
                    ]
                    for row in payload["invoice_rows"]
                ],
                widths=[18, 28, 28, 14, 14, 28, 18, 18],
            )

        if include_exception_rows or include_ignored_rows:
            exception_sheet = workbook.create_sheet("异常与未闭环")
            all_exception_rows = [
                *[
                    [
                        row["record_type"],
                        row["record_id"],
                        row["project_name"],
                        row["expense_type"],
                        row["amount"],
                        row["status_label"],
                        row["remark"],
                    ]
                    for row in payload["exception_rows"]
                ],
                *[
                    [
                        "ignored",
                        row["record_id"],
                        row["project_name"],
                        row["expense_type"],
                        row["amount"],
                        "已忽略",
                        row["remark"],
                    ]
                    for row in payload["ignored_rows"]
                ],
            ]
            self._fill_table_sheet(
                exception_sheet,
                headers=["记录类型", "记录ID", "项目名称", "费用类型", "金额", "状态", "备注"],
                rows=all_exception_rows,
                widths=[14, 20, 28, 24, 14, 22, 32],
            )

        return workbook

    def _build_summary(
        self,
        *,
        project_name: str,
        normal_entries: list[dict[str, Any]],
        oa_rows: list[dict[str, Any]],
        invoice_rows: list[dict[str, Any]],
        exception_rows: list[dict[str, Any]],
        ignored_rows: list[dict[str, Any]],
        scope_label: str,
    ) -> dict[str, Any]:
        total_amount = sum((row["amount_decimal"] for row in normal_entries), start=ZERO)
        return {
            "project_name": project_name,
            "scope_label": scope_label,
            "total_amount": format_decimal(total_amount),
            "transaction_count": len(normal_entries),
            "expense_type_count": len({row["expense_type"] for row in normal_entries}),
            "oa_count": len(oa_rows),
            "invoice_count": len(invoice_rows),
            "exception_count": len(exception_rows),
            "ignored_count": len(ignored_rows),
        }

    def _build_expense_type_rows(self, entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
        total_amount = sum((row["amount_decimal"] for row in entries), start=ZERO)
        buckets: dict[str, dict[str, Any]] = {}
        for row in entries:
            bucket = buckets.setdefault(
                row["expense_type"],
                {
                    "expense_type": row["expense_type"],
                    "amount_decimal": ZERO,
                    "transaction_count": 0,
                    "expense_contents": set(),
                },
            )
            bucket["amount_decimal"] += row["amount_decimal"]
            bucket["transaction_count"] += 1
            bucket["expense_contents"].add(row["expense_content"])
        return [
            {
                "expense_type": bucket["expense_type"],
                "total_amount": format_decimal(bucket["amount_decimal"]),
                "percentage": self._format_percentage(bucket["amount_decimal"], total_amount),
                "transaction_count": bucket["transaction_count"],
                "expense_content_count": len(bucket["expense_contents"]),
            }
            for bucket in sorted(
                buckets.values(),
                key=lambda item: (-item["amount_decimal"], item["expense_type"]),
            )
        ]

    def _build_expense_content_rows(self, entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
        buckets: dict[tuple[str, str], dict[str, Any]] = {}
        for row in entries:
            key = (row["expense_type"], row["expense_content"])
            bucket = buckets.setdefault(
                key,
                {
                    "expense_type": row["expense_type"],
                    "expense_content": row["expense_content"],
                    "amount_decimal": ZERO,
                    "transaction_count": 0,
                },
            )
            bucket["amount_decimal"] += row["amount_decimal"]
            bucket["transaction_count"] += 1
        return [
            {
                "expense_type": bucket["expense_type"],
                "expense_content": bucket["expense_content"],
                "total_amount": format_decimal(bucket["amount_decimal"]),
                "transaction_count": bucket["transaction_count"],
            }
            for bucket in sorted(
                buckets.values(),
                key=lambda item: (-item["amount_decimal"], item["expense_type"], item["expense_content"]),
            )
        ]

    def _build_normal_entries(
        self,
        *,
        month: str,
        group_id: str,
        bank_rows: list[dict[str, Any]],
        context: dict[str, str],
        oa_rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        oa_form_nos = "、".join(
            sorted(
                {
                    self._clean_text((row.get("detail_fields") or {}).get("OA单号"))
                    for row in oa_rows
                    if self._clean_text((row.get("detail_fields") or {}).get("OA单号"))
                }
            )
        )
        rows: list[dict[str, Any]] = []
        for bank_row in bank_rows:
            amount = self._extract_outflow_amount(bank_row)
            if amount is None:
                continue
            rows.append(
                {
                    "month": month,
                    "group_id": group_id,
                    "transaction_id": str(bank_row.get("id") or ""),
                    "trade_time": str(bank_row.get("trade_time") or bank_row.get("pay_receive_time") or ""),
                    "direction": str(bank_row.get("direction") or "支出"),
                    "counterparty_name": str(bank_row.get("counterparty_name") or ""),
                    "payment_account_label": str(bank_row.get("payment_account_label") or ""),
                    "remark": str(bank_row.get("remark") or ""),
                    "project_name": context["project_name"],
                    "expense_type": context["expense_type"],
                    "expense_content": context["expense_content"],
                    "amount_decimal": amount,
                    "amount": format_decimal(amount),
                    "oa_form_nos": oa_form_nos or "—",
                }
            )
        return rows

    def _build_oa_detail_rows(self, oa_rows: list[dict[str, Any]], *, group_id: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row in oa_rows:
            detail_fields = row.get("detail_fields") or {}
            rows.append(
                {
                    "oa_form_no": self._clean_text(detail_fields.get("OA单号")) or "—",
                    "applicant": self._clean_text(row.get("applicant")) or "—",
                    "project_name": self._clean_text(row.get("project_name")) or "—",
                    "expense_type": self._clean_text(row.get("expense_type")) or self._clean_text(detail_fields.get("费用类型")) or "—",
                    "expense_content": self._clean_text(row.get("expense_content")) or self._clean_text(detail_fields.get("费用内容")) or "—",
                    "amount": self._clean_text(row.get("amount")) or "—",
                    "group_id": group_id,
                }
            )
        return rows

    def _build_invoice_detail_rows(
        self,
        invoice_rows: list[dict[str, Any]],
        *,
        group_id: str,
        project_name: str,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row in invoice_rows:
            detail_fields = row.get("detail_fields") or {}
            relation = row.get("invoice_bank_relation") or row.get("invoice_relation") or {}
            rows.append(
                {
                    "invoice_no": self._clean_text(detail_fields.get("发票号码")) or self._clean_text(row.get("invoice_no")) or "—",
                    "seller_name": self._clean_text(row.get("seller_name")) or "—",
                    "buyer_name": self._clean_text(row.get("buyer_name")) or "—",
                    "amount": self._clean_text(row.get("amount")) or "—",
                    "tax_amount": self._clean_text(row.get("tax_amount")) or "—",
                    "project_name": project_name,
                    "status_label": self._clean_text(relation.get("label")) or "—",
                    "group_id": group_id,
                }
            )
        return rows

    def _build_exception_rows(
        self,
        *,
        month: str,
        project_name: str,
        oa_rows: list[dict[str, Any]],
        bank_rows: list[dict[str, Any]],
        invoice_rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        rows.extend(self._exception_rows_for_type(month=month, project_name=project_name, rows=oa_rows, row_type="oa"))
        rows.extend(self._exception_rows_for_type(month=month, project_name=project_name, rows=bank_rows, row_type="bank"))
        rows.extend(self._exception_rows_for_type(month=month, project_name=project_name, rows=invoice_rows, row_type="invoice"))
        return rows

    def _exception_rows_for_type(
        self,
        *,
        month: str,
        project_name: str,
        rows: list[dict[str, Any]],
        row_type: str,
    ) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for row in rows:
            relation = (
                row.get("oa_bank_relation")
                or row.get("invoice_relation")
                or row.get("invoice_bank_relation")
                or {}
            )
            amount = (
                self._clean_text(row.get("amount"))
                or self._clean_text(row.get("debit_amount"))
                or self._clean_text(row.get("credit_amount"))
                or "—"
            )
            records.append(
                {
                    "month": month,
                    "record_type": row_type,
                    "record_id": self._clean_text(row.get("id")) or "—",
                    "project_name": project_name,
                    "expense_type": self._clean_text(row.get("expense_type")) or self._clean_text((row.get("detail_fields") or {}).get("费用类型")) or "—",
                    "amount": amount,
                    "status_label": self._clean_text(relation.get("label")) or "待人工处理",
                    "remark": self._clean_text(row.get("expense_content")) or self._clean_text(row.get("remark")) or self._clean_text((row.get("detail_fields") or {}).get("费用内容")) or "—",
                }
            )
        return [record for record in records if record["expense_type"] not in EXCLUDED_COST_EXPENSE_TYPES]

    def _build_ambiguous_context_rows(self, *, month: str, project_name: str, oa_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row in oa_rows:
            rows.append(
                {
                    "month": month,
                    "record_type": "oa",
                    "record_id": self._clean_text(row.get("id")) or "—",
                    "project_name": project_name,
                    "expense_type": self._clean_text(row.get("expense_type")) or "—",
                    "amount": self._clean_text(row.get("amount")) or "—",
                    "status_label": "项目成本上下文不唯一",
                    "remark": self._clean_text(row.get("expense_content")) or self._clean_text(row.get("reason")) or "—",
                }
            )
        return rows

    def _extract_ignored_rows(
        self,
        *,
        project_name: str,
        month: str,
        expense_types: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        if self._raw_workbench_loader is None:
            return []
        payload = self._raw_workbench_loader(month)
        rows: list[dict[str, Any]] = []
        for section in ("paired", "open"):
            section_payload = payload.get(section) or {}
            for row_type in ("oa", "bank", "invoice"):
                for row in section_payload.get(row_type) or []:
                    if not row.get("ignored"):
                        continue
                    if self._clean_text(row.get("project_name")) != project_name:
                        continue
                    expense_type = self._clean_text(row.get("expense_type")) or "—"
                    if expense_type in EXCLUDED_COST_EXPENSE_TYPES:
                        continue
                    if expense_types and expense_type not in expense_types:
                        continue
                    rows.append(
                        {
                            "record_type": row_type,
                            "record_id": self._clean_text(row.get("id")) or "—",
                            "project_name": project_name,
                            "expense_type": expense_type,
                            "amount": self._clean_text(row.get("amount")) or self._clean_text(row.get("debit_amount")) or self._clean_text(row.get("credit_amount")) or "—",
                            "remark": self._clean_text(row.get("expense_content")) or self._clean_text(row.get("reason")) or self._clean_text(row.get("remark")) or "已忽略",
                        }
                    )
        return rows

    @staticmethod
    def _group_has_exception(oa_rows: list[dict[str, Any]], bank_rows: list[dict[str, Any]], invoice_rows: list[dict[str, Any]]) -> bool:
        return any(
            bool(row.get("handled_exception"))
            for row in [*oa_rows, *bank_rows, *invoice_rows]
        )

    @staticmethod
    def _normalize_expense_types(expense_types: list[str] | None) -> set[str]:
        return {item.strip() for item in (expense_types or []) if item and item.strip()}

    def _resolve_target_months(
        self,
        month: str,
        *,
        start_month: str | None = None,
        end_month: str | None = None,
    ) -> list[str]:
        normalized = (month or "").strip()
        if start_month and end_month and start_month > end_month:
            start_month, end_month = end_month, start_month
        if normalized and normalized.lower() != "all":
            months = [normalized]
        else:
            months = sorted(
                {
                    (transaction.txn_date or "")[:7]
                    for transaction in self._import_service.list_transactions()
                    if (transaction.txn_date or "")[:7]
                },
                reverse=True,
            )
        if start_month:
            months = [item for item in months if item >= start_month]
        if end_month:
            months = [item for item in months if item <= end_month]
        return months

    @staticmethod
    def _normalize_sort_by(sort_by: str) -> str:
        return sort_by if sort_by in {"time", "expense_type", "amount_desc"} else "time"

    def _sort_transaction_rows(self, rows: list[dict[str, Any]], sort_by: str) -> list[dict[str, Any]]:
        if sort_by == "amount_desc":
            return sorted(rows, key=lambda item: (-item["amount_decimal"], item["trade_time"], item["transaction_id"]))
        if sort_by == "expense_type":
            return sorted(rows, key=lambda item: (item["expense_type"], item["trade_time"], item["transaction_id"]))
        return sorted(rows, key=lambda item: (item["trade_time"], item["transaction_id"]))

    @staticmethod
    def _build_scope_label(
        *,
        month: str,
        target_months: list[str],
        start_month: str | None,
        end_month: str | None,
    ) -> str:
        if start_month and end_month:
            return f"{start_month}至{end_month}"
        if str(month).strip().lower() == "all":
            return "全部期间"
        if target_months:
            return target_months[0]
        return month or "—"

    @staticmethod
    def _build_export_structure_label(options: dict[str, Any]) -> str:
        parts = ["项目汇总", "费用类型汇总"]
        if options.get("include_expense_content_summary", True):
            parts.append("费用内容汇总")
        parts.append("流水明细")
        if options.get("include_oa_details", True):
            parts.append("OA关联")
        if options.get("include_invoice_details", True):
            parts.append("发票关联")
        if options.get("include_exception_rows", True) or options.get("include_ignored_rows", True):
            parts.append("异常与未闭环")
        return "、".join(parts)

    @staticmethod
    def _format_percentage(amount: Decimal, total: Decimal) -> str:
        if total == ZERO:
            return "0.0%"
        return f"{((amount / total) * Decimal('100')).quantize(Decimal('0.1'))}%"

    @staticmethod
    def _fill_key_value_sheet(sheet: Any, rows: list[tuple[str, Any]], *, widths: list[int]) -> None:
        sheet.append(["字段", "值"])
        for key, value in rows:
            sheet.append([key, value])
        ProjectDetailExportService._apply_sheet_layout(sheet, widths)

    @staticmethod
    def _fill_table_sheet(sheet: Any, *, headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        ProjectDetailExportService._apply_sheet_layout(sheet, widths)

    @staticmethod
    def _apply_sheet_layout(sheet: Any, widths: list[int]) -> None:
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", horizontal="left")
        sheet.freeze_panes = "A2"
        if sheet.max_column and sheet.max_row:
            sheet.auto_filter.ref = sheet.dimensions
        for index, width in enumerate(widths, start=1):
            column = sheet.cell(row=1, column=index).column_letter
            sheet.column_dimensions[column].width = width

    def _resolve_group_cost_context(self, oa_rows: list[dict[str, Any]]) -> dict[str, str] | None:
        contexts: set[tuple[str, str, str]] = set()
        for row in oa_rows:
            project_name = self._clean_text(row.get("project_name"))
            expense_type = self._clean_text(row.get("expense_type"))
            expense_content = self._clean_text(row.get("expense_content")) or self._clean_text(row.get("reason"))
            detail_fields = row.get("detail_fields")
            if isinstance(detail_fields, dict):
                if not expense_type:
                    expense_type = self._clean_text(detail_fields.get("费用类型"))
                if not expense_content:
                    expense_content = self._clean_text(detail_fields.get("费用内容"))
            if expense_type in EXCLUDED_COST_EXPENSE_TYPES:
                continue
            if not (project_name and expense_type and expense_content):
                continue
            contexts.add((project_name, expense_type, expense_content))
        if len(contexts) != 1:
            return None
        project_name, expense_type, expense_content = next(iter(contexts))
        return {
            "project_name": project_name,
            "expense_type": expense_type,
            "expense_content": expense_content,
        }

    def _extract_outflow_amount(self, bank_row: dict[str, Any]) -> Decimal | None:
        transaction_id = str(bank_row.get("id", ""))
        try:
            transaction = self._import_service.get_transaction(transaction_id)
        except KeyError:
            transaction = None

        if transaction is not None:
            if transaction.txn_direction != TransactionDirection.OUTFLOW:
                return None
            return transaction.amount

        debit_amount = self._parse_decimal(bank_row.get("debit_amount"))
        credit_amount = self._parse_decimal(bank_row.get("credit_amount"))
        if credit_amount not in (None, ZERO):
            return None
        if debit_amount in (None, ZERO):
            return None
        return debit_amount

    @staticmethod
    def _parse_decimal(value: Any) -> Decimal | None:
        if value in (None, "", "—", "--"):
            return None
        text = str(value).replace(",", "").strip()
        if not text:
            return None
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _clean_text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text in {"-", "--", "—", "——"}:
            return ""
        return text
