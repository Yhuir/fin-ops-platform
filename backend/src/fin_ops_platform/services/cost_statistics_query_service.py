from __future__ import annotations

import base64
from decimal import Decimal
from io import BytesIO
import json
import re
from typing import Any

from openpyxl import Workbook

from fin_ops_platform.services.cost_statistics_policy import CostStatisticsPolicy


COST_STATISTICS_EXPORT_ROW_LIMIT = 20000
COST_STATISTICS_EXPORT_PREVIEW_SIZE = 8


class CostStatisticsExportLimitError(ValueError):
    def __init__(
        self,
        *,
        view: str,
        total: int,
        limit: int = COST_STATISTICS_EXPORT_ROW_LIMIT,
    ) -> None:
        super().__init__(
            f"当前筛选命中 {total} 行，超过 {limit} 行导出上限，请缩小筛选范围。"
        )
        self.error_code = "cost_statistics_export_row_limit_exceeded"
        self.details = {"view": view, "total": total, "limit": limit}


class CostStatisticsQueryService:
    """Serve Cost pages from one canonical snapshot per API request."""

    def __init__(self, *, canonical_repository: Any) -> None:
        if not callable(getattr(canonical_repository, "load_snapshot", None)):
            raise ValueError(
                "Cost statistics query service requires a canonical snapshot repository."
            )
        self._canonical_repository = canonical_repository

    def get_explorer_page(
        self,
        *,
        scope: str,
        view: str,
        project_scope: str,
        filters: dict[str, str | None],
        cursor: str | None,
        page_size: int,
        include_statistics: bool = True,
        if_none_match: str | None = None,
    ) -> tuple[dict[str, Any], bool, str, bool]:
        del if_none_match
        normalized_project_scope = self._normalize_project_scope(project_scope)
        scope_kind, scope_value, normalized_scope = self._normalize_page_scope(
            scope
        )
        normalized_view, normalized_filters = self._normalize_page_query(
            view,
            filters,
        )
        normalized_page_size = self._normalize_page_size(page_size)
        query_binding = self._page_query_binding(
            scope=normalized_scope,
            view=normalized_view,
            filters=normalized_filters,
            page_size=normalized_page_size,
        )
        cursor_values = self._decode_page_cursor(
            cursor,
            query_binding=query_binding,
        )
        policy = CostStatisticsPolicy(
            self._canonical_repository.load_snapshot(
                scope_kind=scope_kind,
                scope_value=scope_value,
                view=normalized_view,
                include_statistics=include_statistics,
            ),
            project_scope=normalized_project_scope,
        )
        raw_page = policy.explorer_page(
            scope_kind=scope_kind,
            scope_value=scope_value,
            view=normalized_view,
            filters=normalized_filters,
            cursor_values=cursor_values,
            page_size=normalized_page_size,
            include_statistics=include_statistics,
        )
        facets = {
            "projects": [],
            "expense_types": [],
            "bank_accounts": [],
            "bank_tag_primary": [],
            "bank_tag_sub": [],
        }
        primary = list(raw_page.get("primary_facets") or [])
        secondary = list(raw_page.get("secondary_facets") or [])
        if normalized_view == "project":
            facets["projects"] = primary
            facets["expense_types"] = secondary
        elif normalized_view == "bank":
            facets["bank_accounts"] = self._merge_bank_account_facets(
                primary,
                list(raw_page.get("bank_accounts") or []),
            )
            facets["projects"] = secondary
        elif normalized_view == "expense_type":
            facets["expense_types"] = primary
        elif normalized_view == "bank_tag":
            facets["bank_tag_primary"] = primary
            facets["bank_tag_sub"] = secondary
        next_cursor_values = raw_page.get("next_cursor_values")
        payload = {
            "scope": normalized_scope,
            "view": normalized_view,
            "summary": dict(raw_page.get("summary") or {}),
            "statistics": (
                dict(raw_page["statistics"])
                if isinstance(raw_page.get("statistics"), dict)
                else None
            ),
            "available_years": list(raw_page.get("available_years") or []),
            "facets": facets,
            "rows": [
                dict(row)
                for row in list(raw_page.get("rows") or [])
                if isinstance(row, dict)
            ],
            "row_count": int(raw_page.get("row_count") or 0),
            "next_cursor": (
                self._encode_page_cursor(
                    tuple(str(value) for value in next_cursor_values),
                    query_binding=query_binding,
                )
                if isinstance(next_cursor_values, (list, tuple))
                and len(next_cursor_values) == 4
                else None
            ),
            "cost_statistics_tag_selection_version": int(
                raw_page.get("tag_selection_version") or 1
            ),
        }
        return payload, False, "", False

    def get_transaction_detail(
        self,
        transaction_id: str,
        *,
        project_scope: str,
        view: str,
        scope: str,
    ) -> dict[str, Any]:
        normalized_project_scope = self._normalize_project_scope(project_scope)
        scope_kind, scope_value, _normalized_scope = self._normalize_page_scope(
            scope
        )
        normalized_view, _filters = self._normalize_page_query(view, {})
        normalized_transaction_id = str(transaction_id or "").strip()
        if not normalized_transaction_id:
            raise KeyError(transaction_id)
        row = self._policy(project_scope=normalized_project_scope).transaction(
            transaction_id=normalized_transaction_id,
            bank_flow_view=normalized_view in {"time", "bank_tag"},
            scope_kind=scope_kind,
            scope_value=scope_value,
        )
        if not isinstance(row, dict):
            raise KeyError(transaction_id)
        cost_allocations = [
            {
                "row_key": str(item.get("row_key") or ""),
                "project_name": str(item.get("project_name") or "未归集项目"),
                "project_id": str(item.get("project_id") or ""),
                "expense_type": str(item.get("expense_type") or "未分类"),
                "expense_content": str(item.get("expense_content") or ""),
                "oa_applicant": str(item.get("oa_applicant") or "—"),
                "amount": _plain_money(
                    _decimal_from_value(item.get("amount"))
                    or Decimal("0.00")
                ),
            }
            for item in list(row.get("cost_allocations") or [])
            if isinstance(item, dict)
        ]
        project_names = sorted(
            {item["project_name"] for item in cost_allocations}
        )
        expense_types = sorted(
            {item["expense_type"] for item in cost_allocations}
        )
        allocation_amount = sum(
            (
                _decimal_from_value(item["amount"]) or Decimal("0.00")
                for item in cost_allocations
            ),
            start=Decimal("0.00"),
        )
        label_path = row.get("bank_tag_label_path")
        month = str(row.get("month") or row.get("trade_time") or "")[:7] or "all"
        return {
            "month": month,
            "transaction": {
                "id": normalized_transaction_id,
                "project_name": (
                    "、".join(project_names)
                    if project_names
                    else str(row.get("project_name") or "")
                ),
                "expense_type": (
                    "、".join(expense_types)
                    if expense_types
                    else str(row.get("expense_type") or "")
                ),
                "expense_content": (
                    "、".join(
                        sorted(
                            {
                                item["expense_content"]
                                for item in cost_allocations
                                if item["expense_content"]
                            }
                        )
                    )
                    or str(row.get("expense_content") or "")
                ),
                "trade_time": str(row.get("trade_time") or ""),
                "direction": str(row.get("direction") or ""),
                "amount": _plain_money(
                    allocation_amount
                    if cost_allocations
                    else _decimal_from_value(row.get("amount"))
                    or Decimal("0.00")
                ),
                "counterparty_name": str(
                    row.get("counterparty_name") or ""
                ),
                "payment_account_label": str(
                    row.get("payment_account_label") or ""
                ),
                "remark": str(row.get("remark") or ""),
                "oa_applicant": (
                    "、".join(
                        sorted(
                            {
                                item["oa_applicant"]
                                for item in cost_allocations
                                if item["oa_applicant"]
                            }
                        )
                    )
                    or str(row.get("oa_applicant") or "")
                ),
                "cost_allocations": cost_allocations,
                "summary_fields": {},
                "detail_fields": {},
                "relation_status": "canonical",
                "relation_case_ids": [
                    str(row.get("group_id"))
                ]
                if str(row.get("group_id") or "")
                else [],
                "linked_oa_count": int(row.get("linked_oa_count") or 0),
                "linked_invoice_count": 0,
                "bank_tag_code": str(row.get("bank_tag_code") or ""),
                "bank_tag_label": str(row.get("bank_tag_label") or ""),
                "bank_tag_primary_label": str(
                    row.get("bank_tag_primary_label") or ""
                ),
                "bank_tag_sub_label": str(
                    row.get("bank_tag_sub_label") or ""
                ),
                "bank_tag_label_path": (
                    list(label_path) if isinstance(label_path, list) else []
                ),
            },
            "relation_context": {
                "row_id": normalized_transaction_id,
                "row_type": "bank_transaction",
                "relation_status": "canonical",
                "group_ids": [
                    str(row.get("group_id"))
                ]
                if str(row.get("group_id") or "")
                else [],
                "linked_oa": [],
                "linked_bank_transactions": [],
                "linked_input_invoices": [],
                "linked_output_invoices": [],
            },
        }

    def get_export_preview(self, **kwargs: Any) -> dict[str, Any]:
        project_scope = self._normalize_project_scope(
            str(kwargs.get("project_scope") or "active")
        )
        view = str(kwargs.get("view") or "").strip()
        month = str(kwargs.get("month") or "all").strip() or "all"
        project_names = self._normalize_text_set(
            kwargs.get("project_names")
            or (
                [kwargs.get("project_name")]
                if kwargs.get("project_name")
                else []
            )
        )
        expense_types = self._normalize_text_set(kwargs.get("expense_types"))
        range_kwargs = self._range_kwargs(kwargs)
        if view not in {"time", "bank_tag", "project", "expense_type"}:
            raise ValueError(
                "view must be time, bank_tag, project, or expense_type."
            )
        if view == "project" and not project_names:
            raise ValueError("project_name is required for project export preview")
        if view == "expense_type" and not expense_types:
            raise ValueError(
                "expense_type is required for expense_type export preview"
            )
        row_shape = "raw_bank" if view in {"time", "bank_tag"} else "raw_cost"
        page = self._policy(project_scope=project_scope).export_page(
            month=month,
            project_names=sorted(project_names),
            expense_types=sorted(expense_types),
            row_shape=row_shape,
            offset=0,
            page_size=COST_STATISTICS_EXPORT_ROW_LIMIT + 1,
            include_summary=True,
            **range_kwargs,
        )
        summary = self._export_page_summary(page)
        total = int(summary.get("source_row_count") or 0)
        self._ensure_export_row_limit(view=view, total=total)
        entries = [
            self._export_entry_from_row(row)
            for row in list(page.get("rows") or [])[:COST_STATISTICS_EXPORT_PREVIEW_SIZE]
        ]
        scope_label = self._build_scope_label(month=month, **range_kwargs)
        if view == "time":
            columns = [
                "时间",
                "项目名称",
                "费用类型",
                "金额",
                "费用内容",
                "资金方向",
                "对方户名",
                "支付账户",
            ]
            rows = [self._time_row_from_entry(entry) for entry in entries]
            sheet_names = ["按时间统计"]
            file_name = self._build_filename(
                month=scope_label,
                view=view,
            )
            extra = self._directional_summary_from_export_summary(summary)
        elif view == "bank_tag":
            columns = [
                "时间",
                "主标签",
                "子标签",
                "资金方向",
                "金额",
                "费用内容",
                "对方户名",
                "支付账户",
            ]
            rows = [self._bank_tag_row_from_entry(entry) for entry in entries]
            sheet_names = ["按标签统计"]
            file_name = self._build_filename(
                month=scope_label,
                view=view,
            )
            extra = self._directional_summary_from_export_summary(summary)
        elif view == "project":
            project_name = sorted(project_names)[0]
            columns = [
                "时间",
                "资金方向",
                "费用类型",
                "金额",
                "费用内容",
                "对方户名",
                "支付账户",
            ]
            rows = [
                [
                    entry["trade_time"],
                    entry["direction"],
                    entry["expense_type"],
                    _plain_money(entry["amount_decimal"]),
                    entry["expense_content"],
                    entry["counterparty_name"],
                    entry["payment_account_label"],
                ]
                for entry in entries
            ]
            sheet_names = self._project_sheet_names(
                include_oa_details=True,
                include_invoice_details=True,
                include_exception_rows=True,
                include_ignored_rows=True,
                include_expense_content_summary=True,
            )
            file_name = self._build_filename(
                month=scope_label,
                view=view,
                project_name=project_name,
            )
            extra = {}
        else:
            columns = [
                "时间",
                "项目名称",
                "资金方向",
                "金额",
                "费用内容",
                "对方户名",
                "支付账户",
            ]
            rows = [
                [
                    entry["trade_time"],
                    entry["project_name"],
                    entry["direction"],
                    _plain_money(entry["amount_decimal"]),
                    entry["expense_content"],
                    entry["counterparty_name"],
                    entry["payment_account_label"],
                ]
                for entry in entries
            ]
            sheet_names = ["按费用类型统计"]
            file_name = self._build_filename(
                month=scope_label,
                view=view,
                expense_type=self._build_expense_type_label(expense_types),
            )
            extra = {}
        return self._preview_payload(
            view=view,
            file_name=file_name,
            scope_label=scope_label,
            sheet_names=sheet_names,
            columns=columns,
            rows=rows,
            total_count=total,
            total_amount=_plain_money(
                _decimal_from_value(summary.get("total_amount"))
                or Decimal("0.00")
            ),
            summary_extra=extra,
        )

    def export_view(self, **kwargs: Any) -> tuple[str, bytes]:
        project_scope = self._normalize_project_scope(
            str(kwargs.get("project_scope") or "active")
        )
        view = str(kwargs.get("view") or "").strip()
        month = str(kwargs.get("month") or "all").strip() or "all"
        project_names = self._normalize_text_set(
            kwargs.get("project_names")
            or (
                [kwargs.get("project_name")]
                if kwargs.get("project_name")
                else []
            )
        )
        expense_types = self._normalize_text_set(kwargs.get("expense_types"))
        aggregate_by = self._normalize_project_aggregate_by(
            kwargs.get("aggregate_by")
        )
        range_kwargs = self._range_kwargs(kwargs)
        if view == "transaction":
            transaction_id = str(kwargs.get("transaction_id") or "").strip()
            if not transaction_id:
                raise ValueError(
                    "transaction_id is required for transaction export"
                )
            payload = self.get_transaction_detail(
                transaction_id,
                project_scope=project_scope,
                view="project",
                scope=month,
            )
            workbook = self._transaction_workbook(payload)
            return (
                self._build_filename(
                    month=payload["month"],
                    view=view,
                    project_name=payload["transaction"]["project_name"],
                    transaction_id=transaction_id,
                ),
                self._serialize_workbook(workbook),
            )
        if view not in {
            "month",
            "time",
            "bank_tag",
            "project",
            "expense_type",
        }:
            raise ValueError(f"unsupported export view: {view}")
        if view == "project" and not project_names:
            raise ValueError("project_name is required for project export")
        if view == "expense_type" and not expense_types:
            raise ValueError(
                "expense_type is required for expense_type export"
            )
        row_shape = "raw_cost"
        export_month = month
        if view in {"time", "bank_tag"}:
            row_shape = "raw_bank"
        elif view == "month":
            row_shape = "month_summary"
        elif view == "project" and (
            aggregate_by is not None or len(project_names) > 1
        ):
            export_month = "all"
            row_shape = (
                "project_month"
                if (aggregate_by or "month") == "month"
                else "project_year"
            )
        policy = self._policy(project_scope=project_scope)
        page = policy.export_page(
            month=export_month,
            project_names=sorted(project_names),
            expense_types=sorted(expense_types),
            row_shape=row_shape,
            offset=0,
            page_size=COST_STATISTICS_EXPORT_ROW_LIMIT + 1,
            include_summary=True,
            **range_kwargs,
        )
        summary = self._export_page_summary(page)
        total = int(
            summary.get("row_count")
            if view == "month"
            else summary.get("source_row_count")
            or 0
        )
        self._ensure_export_row_limit(view=view, total=total)
        entries = [
            self._export_entry_from_row(row)
            for row in list(page.get("rows") or [])
        ]
        scope_label = self._build_scope_label(month=month, **range_kwargs)
        if view == "time":
            workbook = self._table_workbook(
                "按时间统计",
                [
                    "时间",
                    "项目名称",
                    "费用类型",
                    "金额",
                    "费用内容",
                    "资金方向",
                    "对方户名",
                    "支付账户",
                ],
                (self._time_row_from_entry(entry) for entry in entries),
            )
            filename = self._build_filename(
                month=scope_label,
                view=view,
            )
        elif view == "bank_tag":
            workbook = self._table_workbook(
                "按标签统计",
                [
                    "时间",
                    "主标签",
                    "子标签",
                    "资金方向",
                    "金额",
                    "费用内容",
                    "对方户名",
                    "支付账户",
                ],
                (self._bank_tag_row_from_entry(entry) for entry in entries),
            )
            filename = self._build_filename(
                month=scope_label,
                view=view,
            )
        elif view == "month":
            workbook = self._table_workbook(
                "月份汇总",
                [
                    "项目名称",
                    "费用类型",
                    "金额",
                    "费用内容",
                    "支出笔数",
                ],
                (
                    [
                        entry["project_name"],
                        entry["expense_type"],
                        entry["amount"],
                        entry["expense_content"],
                        entry["transaction_count"],
                    ]
                    for entry in entries
                ),
            )
            filename = self._build_filename(month=month, view=view)
        elif view == "project":
            if aggregate_by is not None or len(project_names) > 1:
                workbook = self._table_workbook(
                    "按项目统计",
                    [
                        "统计周期",
                        "项目名称",
                        "费用类型",
                        "金额",
                        "费用内容",
                        "支出笔数",
                    ],
                    (
                        [
                            entry["period_label"],
                            entry["project_name"],
                            entry["expense_type"],
                            entry["amount"],
                            entry["expense_content"],
                            entry["transaction_count"],
                        ]
                        for entry in entries
                    ),
                )
                filename = self._build_filename(
                    month=self._build_scope_label(
                        month="all",
                        **range_kwargs,
                    ),
                    view=view,
                    project_names=sorted(project_names),
                    aggregate_by=aggregate_by or "month",
                )
            else:
                project_name = sorted(project_names)[0]
                workbook = self._project_detail_workbook(
                    month=month,
                    project_name=project_name,
                    entries=entries,
                    include_oa_details=bool(
                        kwargs.get("include_oa_details", True)
                    ),
                    include_invoice_details=bool(
                        kwargs.get("include_invoice_details", True)
                    ),
                    include_exception_rows=bool(
                        kwargs.get("include_exception_rows", True)
                    ),
                    include_ignored_rows=bool(
                        kwargs.get("include_ignored_rows", True)
                    ),
                    include_expense_content_summary=bool(
                        kwargs.get("include_expense_content_summary", True)
                    ),
                    scope_label=scope_label,
                )
                filename = self._build_filename(
                    month=scope_label,
                    view=view,
                    project_name=project_name,
                )
        else:
            workbook = self._table_workbook(
                "按费用类型统计",
                [
                    "时间",
                    "项目名称",
                    "金额",
                    "费用内容",
                    "资金方向",
                    "对方户名",
                    "支付账户",
                ],
                (
                    [
                        entry["trade_time"],
                        entry["project_name"],
                        _plain_money(entry["amount_decimal"]),
                        entry["expense_content"],
                        entry["direction"],
                        entry["counterparty_name"],
                        entry["payment_account_label"],
                    ]
                    for entry in entries
                ),
            )
            filename = self._build_filename(
                month=scope_label,
                view=view,
                expense_type=self._build_expense_type_label(expense_types),
            )
        return filename, self._serialize_workbook(workbook)

    def _policy(self, *, project_scope: str) -> CostStatisticsPolicy:
        return CostStatisticsPolicy(
            self._canonical_repository.load_snapshot(),
            project_scope=project_scope,
        )

    @staticmethod
    def _normalize_page_scope(
        scope: str,
    ) -> tuple[str, str | None, str]:
        normalized = str(scope or "").strip().lower()
        if normalized == "all":
            return "all", None, "all"
        if re.fullmatch(r"year:\d{4}", normalized):
            return "year", normalized.split(":", 1)[1], normalized
        if re.fullmatch(r"\d{4}-\d{2}", normalized):
            return "month", normalized, normalized
        raise ValueError("scope must be all, year:YYYY, or YYYY-MM")

    @staticmethod
    def _normalize_page_size(page_size: int) -> int:
        normalized = int(page_size)
        if normalized < 1 or normalized > 200:
            raise ValueError("page_size must be between 1 and 200")
        return normalized

    @staticmethod
    def _normalize_page_query(
        view: str,
        filters: dict[str, str | None],
    ) -> tuple[str, dict[str, str]]:
        normalized_view = str(view or "time").strip().lower() or "time"
        if normalized_view not in {
            "time",
            "project",
            "bank",
            "expense_type",
            "bank_tag",
        }:
            raise ValueError(
                "view must be time, project, bank, expense_type, or bank_tag"
            )
        return (
            normalized_view,
            {
                key: str(filters.get(key) or "").strip()
                for key in (
                    "project_name",
                    "expense_type",
                    "payment_account_label",
                    "bank_tag_primary_label",
                    "bank_tag_sub_label",
                )
            },
        )

    @staticmethod
    def _page_query_binding(
        *,
        scope: str,
        view: str,
        filters: dict[str, str],
        page_size: int,
    ) -> str:
        return json.dumps(
            {
                "scope": scope,
                "view": view,
                "filters": filters,
                "page_size": page_size,
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )

    @staticmethod
    def _encode_page_cursor(
        values: tuple[str, ...],
        *,
        query_binding: str,
    ) -> str:
        encoded = json.dumps(
            {"query": query_binding, "values": list(values)},
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
        return base64.urlsafe_b64encode(encoded).decode("ascii").rstrip("=")

    @staticmethod
    def _decode_page_cursor(
        cursor: str | None,
        *,
        query_binding: str,
    ) -> tuple[str, str, str, str] | None:
        if not cursor:
            return None
        try:
            padding = "=" * (-len(cursor) % 4)
            payload = json.loads(
                base64.urlsafe_b64decode(cursor + padding).decode("utf-8")
            )
            values = payload.get("values")
        except Exception as error:
            raise ValueError("invalid cost statistics cursor") from error
        if (
            not isinstance(payload, dict)
            or payload.get("query") != query_binding
            or not isinstance(values, list)
            or len(values) != 4
        ):
            raise ValueError("cost statistics cursor does not match query")
        return tuple(str(value) for value in values)  # type: ignore[return-value]

    @staticmethod
    def _merge_bank_account_facets(
        observed: list[dict[str, Any]],
        configured: list[Any],
    ) -> list[dict[str, Any]]:
        merged = [
            dict(item) for item in observed if isinstance(item, dict)
        ]
        labels = {
            str(item.get("payment_account_label") or "") for item in merged
        }
        for item in configured:
            if not isinstance(item, dict):
                continue
            label = str(item.get("payment_account_label") or "").strip()
            if not label or label in labels:
                continue
            labels.add(label)
            merged.append(
                {
                    "payment_account_label": label,
                    "total_amount": "0.00",
                    "transaction_count": 0,
                    "project_count": 0,
                    "percentage_label": "0.0%",
                }
            )
        return merged

    @staticmethod
    def empty_explorer_page_payload(
        scope: str,
        view: str,
    ) -> dict[str, Any]:
        return {
            "scope": scope,
            "view": view,
            "summary": {
                "row_count": 0,
                "transaction_count": 0,
                "total_amount": "0.00",
                "expense_amount": "0.00",
                "income_amount": "0.00",
                "expense_transaction_count": 0,
                "income_transaction_count": 0,
            },
            "statistics": {},
            "available_years": [],
            "facets": {
                "projects": [],
                "expense_types": [],
                "bank_accounts": [],
                "bank_tag_primary": [],
                "bank_tag_sub": [],
            },
            "rows": [],
            "row_count": 0,
            "next_cursor": None,
        }

    @staticmethod
    def explorer_entry_count(payload: dict[str, Any]) -> int:
        return int(payload.get("row_count") or 0)

    @staticmethod
    def _normalize_project_scope(project_scope: str) -> str:
        normalized = str(project_scope or "active").strip().lower()
        if normalized not in {"active", "all"}:
            raise ValueError("project_scope must be active or all")
        return normalized

    @staticmethod
    def _export_page_summary(page: dict[str, Any]) -> dict[str, Any]:
        summary = page.get("summary")
        return dict(summary) if isinstance(summary, dict) else {}

    @staticmethod
    def _export_entry_from_row(
        raw_row: dict[str, Any],
    ) -> dict[str, Any]:
        trade_time = str(
            raw_row.get("trade_time")
            or raw_row.get("trade_time_text")
            or ""
        )
        amount = _decimal_from_value(raw_row.get("amount")) or Decimal(
            "0.00"
        )
        primary = (
            str(
                raw_row.get("bank_tag_primary_label")
                or raw_row.get("bank_tag_label")
                or ""
            ).strip()
            or "未标记"
        )
        sub = (
            str(
                raw_row.get("bank_tag_sub_label")
                or raw_row.get("bank_tag_label")
                or ""
            ).strip()
            or primary
        )
        label_path = [
            str(item).strip()
            for item in list(raw_row.get("bank_tag_label_path") or [])
            if str(item).strip()
        ]
        if not label_path:
            label_path = (
                [primary] if primary == sub else [primary, sub]
            )
        return {
            "transaction_id": str(
                raw_row.get("transaction_id") or ""
            ).strip(),
            "month": str(
                raw_row.get("month") or raw_row.get("scope_month") or ""
            )[:7]
            or trade_time[:7],
            "trade_time": trade_time,
            "direction": str(raw_row.get("direction") or "支出"),
            "project_name": str(
                raw_row.get("project_name") or ""
            ).strip(),
            "expense_type": str(
                raw_row.get("expense_type") or ""
            ).strip(),
            "expense_content": str(
                raw_row.get("expense_content") or ""
            ).strip(),
            "amount_decimal": amount,
            "amount": _plain_money(amount),
            "counterparty_name": str(
                raw_row.get("counterparty_name") or ""
            ).strip(),
            "payment_account_label": str(
                raw_row.get("payment_account_label") or ""
            ).strip(),
            "remark": str(raw_row.get("remark") or "").strip(),
            "oa_applicant": str(
                raw_row.get("oa_applicant") or "—"
            ).strip()
            or "—",
            "bank_tag_code": str(
                raw_row.get("bank_tag_code") or ""
            ).strip(),
            "bank_tag_label": str(
                raw_row.get("bank_tag_label") or sub
            ).strip()
            or sub,
            "bank_tag_primary_label": primary,
            "bank_tag_sub_label": sub,
            "bank_tag_label_path": label_path,
            "period_label": str(raw_row.get("period_label") or "—"),
            "transaction_count": int(
                raw_row.get("transaction_count") or 0
            ),
        }

    @staticmethod
    def _range_kwargs(
        kwargs: dict[str, Any],
    ) -> dict[str, str | None]:
        return {
            "start_month": str(kwargs.get("start_month") or "").strip()
            or None,
            "end_month": str(kwargs.get("end_month") or "").strip()
            or None,
            "start_date": str(kwargs.get("start_date") or "").strip()
            or None,
            "end_date": str(kwargs.get("end_date") or "").strip() or None,
        }

    @staticmethod
    def _normalize_text_set(values: object) -> set[str]:
        if values is None:
            return set()
        if isinstance(values, str):
            iterable: list[object] = [values]
        else:
            try:
                iterable = list(values)  # type: ignore[arg-type]
            except TypeError:
                iterable = [values]
        return {
            str(value).strip()
            for value in iterable
            if str(value or "").strip()
        }

    @staticmethod
    def _normalize_project_aggregate_by(value: object) -> str | None:
        normalized = str(value or "").strip().lower()
        if not normalized:
            return None
        if normalized not in {"month", "year"}:
            raise ValueError("aggregate_by must be month or year")
        return normalized

    @staticmethod
    def _build_scope_label(
        *,
        month: str,
        start_month: str | None = None,
        end_month: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
    ) -> str:
        if start_date and end_date:
            return f"{start_date}至{end_date}"
        if start_month and end_month:
            return f"{start_month}至{end_month}"
        if str(month or "").strip().lower() == "all":
            return "全部期间"
        return month or "—"

    @staticmethod
    def _ensure_export_row_limit(*, view: str, total: int) -> None:
        if total > COST_STATISTICS_EXPORT_ROW_LIMIT:
            raise CostStatisticsExportLimitError(view=view, total=total)

    @staticmethod
    def _preview_payload(
        *,
        view: str,
        file_name: str,
        scope_label: str,
        sheet_names: list[str],
        columns: list[str],
        rows: list[list[Any]],
        total_count: int,
        total_amount: str,
        summary_extra: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        return {
            "view": view,
            "file_name": file_name,
            "scope_label": scope_label,
            "summary": {
                "row_count": total_count,
                "transaction_count": total_count,
                "total_amount": total_amount,
                "sheet_count": len(sheet_names),
                **dict(summary_extra or {}),
            },
            "sheet_names": sheet_names,
            "columns": columns,
            "rows": rows[:COST_STATISTICS_EXPORT_PREVIEW_SIZE],
        }

    @staticmethod
    def _time_row_from_entry(entry: dict[str, Any]) -> list[Any]:
        return [
            entry["trade_time"],
            entry["project_name"],
            entry["expense_type"],
            _plain_money(entry["amount_decimal"]),
            entry["expense_content"],
            entry["direction"],
            entry["counterparty_name"],
            entry["payment_account_label"],
        ]

    @staticmethod
    def _bank_tag_row_from_entry(
        entry: dict[str, Any],
    ) -> list[Any]:
        return [
            entry["trade_time"],
            entry["bank_tag_primary_label"],
            entry["bank_tag_sub_label"],
            entry["direction"],
            _plain_money(entry["amount_decimal"]),
            entry["expense_content"],
            entry["counterparty_name"],
            entry["payment_account_label"],
        ]

    @staticmethod
    def _directional_summary_from_export_summary(
        summary: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "expense_amount": _plain_money(
                _decimal_from_value(summary.get("expense_amount"))
                or Decimal("0.00")
            ),
            "income_amount": _plain_money(
                _decimal_from_value(summary.get("income_amount"))
                or Decimal("0.00")
            ),
            "expense_transaction_count": int(
                summary.get("expense_transaction_count") or 0
            ),
            "income_transaction_count": int(
                summary.get("income_transaction_count") or 0
            ),
        }

    @staticmethod
    def _build_expense_type_label(expense_types: set[str]) -> str:
        ordered = sorted(expense_types)
        if not ordered:
            return "未命名费用类型"
        if len(ordered) == 1:
            return ordered[0]
        return f"{ordered[0]}等{len(ordered)}类"

    @staticmethod
    def _project_sheet_names(
        *,
        include_oa_details: bool,
        include_invoice_details: bool,
        include_exception_rows: bool,
        include_ignored_rows: bool,
        include_expense_content_summary: bool,
    ) -> list[str]:
        sheet_names = ["导出说明", "项目汇总", "按费用类型汇总"]
        if include_expense_content_summary:
            sheet_names.append("按费用内容汇总")
        sheet_names.append("流水明细")
        if include_oa_details:
            sheet_names.append("OA关联明细")
        if include_invoice_details:
            sheet_names.append("发票关联明细")
        if include_exception_rows or include_ignored_rows:
            sheet_names.append("异常与未闭环")
        return sheet_names

    @staticmethod
    def _table_workbook(
        title: str,
        headers: list[str],
        rows: Any,
    ) -> Workbook:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        for index in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + index)].width = 18
        return workbook

    def _project_detail_workbook(
        self,
        *,
        month: str,
        project_name: str,
        entries: list[dict[str, Any]],
        include_oa_details: bool,
        include_invoice_details: bool,
        include_exception_rows: bool,
        include_ignored_rows: bool,
        include_expense_content_summary: bool,
        scope_label: str,
    ) -> Workbook:
        workbook = Workbook(write_only=True)
        intro_sheet = workbook.create_sheet("导出说明")
        self._fill_key_value_sheet(
            intro_sheet,
            [
                ("项目名称", project_name),
                ("统计范围", scope_label),
                ("月份列表", month),
                ("数据口径", "统一事实源只读一致性快照"),
                ("导出结构", "项目汇总、费用类型汇总、流水明细"),
            ],
        )
        summary_sheet = workbook.create_sheet("项目汇总")
        expense_type_sheet = workbook.create_sheet("按费用类型汇总")
        expense_content_sheet = (
            workbook.create_sheet("按费用内容汇总")
            if include_expense_content_summary
            else None
        )
        detail_sheet = workbook.create_sheet("流水明细")
        detail_headers = [
            "时间",
            "交易流水ID",
            "资金方向",
            "对方户名",
            "支付账户",
            "金额",
            "备注",
            "项目名称",
            "费用类型",
            "费用内容",
            "OA单号",
            "关联组ID",
        ]
        detail_sheet.append(detail_headers)
        for index in range(1, len(detail_headers) + 1):
            detail_sheet.column_dimensions[chr(64 + index)].width = 18
        if include_oa_details:
            self._append_table_sheet(
                workbook.create_sheet("OA关联明细"),
                [
                    "OA单号",
                    "申请人",
                    "项目名称",
                    "费用类型",
                    "费用内容",
                    "OA金额",
                    "关联组ID",
                ],
                [],
            )
        if include_invoice_details:
            self._append_table_sheet(
                workbook.create_sheet("发票关联明细"),
                [
                    "发票号码",
                    "销方名称",
                    "购方名称",
                    "发票金额",
                    "税额",
                    "项目名称",
                    "关联状态",
                    "关联组ID",
                ],
                [],
            )
        if include_exception_rows or include_ignored_rows:
            self._append_table_sheet(
                workbook.create_sheet("异常与未闭环"),
                [
                    "记录类型",
                    "记录ID",
                    "项目名称",
                    "费用类型",
                    "金额",
                    "状态",
                    "备注",
                ],
                [],
            )
        total_amount = Decimal("0.00")
        type_buckets: dict[str, dict[str, Any]] = {}
        content_buckets: dict[tuple[str, str], dict[str, Any]] = {}
        for entry in entries:
            total_amount += entry["amount_decimal"]
            type_bucket = type_buckets.setdefault(
                entry["expense_type"],
                {
                    "amount_decimal": Decimal("0.00"),
                    "transaction_count": 0,
                    "expense_contents": set(),
                },
            )
            type_bucket["amount_decimal"] += entry["amount_decimal"]
            type_bucket["transaction_count"] += 1
            type_bucket["expense_contents"].add(
                entry["expense_content"]
            )
            content_key = (
                entry["expense_type"],
                entry["expense_content"],
            )
            content_bucket = content_buckets.setdefault(
                content_key,
                {
                    "amount_decimal": Decimal("0.00"),
                    "transaction_count": 0,
                },
            )
            content_bucket["amount_decimal"] += entry["amount_decimal"]
            content_bucket["transaction_count"] += 1
            detail_sheet.append(
                [
                    entry["trade_time"],
                    entry["transaction_id"],
                    entry["direction"],
                    entry["counterparty_name"],
                    entry["payment_account_label"],
                    _plain_money(entry["amount_decimal"]),
                    entry["remark"],
                    entry["project_name"],
                    entry["expense_type"],
                    entry["expense_content"],
                    "—",
                    "—",
                ]
            )
        self._fill_key_value_sheet(
            summary_sheet,
            [
                ("项目名称", project_name),
                ("统计期间", scope_label),
                ("总支出金额", _plain_money(total_amount)),
                ("支出流水笔数", len(entries)),
                ("费用类型数", len(type_buckets)),
                ("已关联OA笔数", 0),
                ("已关联发票笔数", 0),
                ("已处理异常笔数", 0),
                ("已忽略笔数", 0),
            ],
        )
        self._append_table_sheet(
            expense_type_sheet,
            ["费用类型", "金额", "占比", "笔数", "费用内容数"],
            (
                [
                    expense_type,
                    _plain_money(bucket["amount_decimal"]),
                    _percentage(bucket["amount_decimal"], total_amount),
                    bucket["transaction_count"],
                    len(bucket["expense_contents"]),
                ]
                for expense_type, bucket in sorted(
                    type_buckets.items(),
                    key=lambda item: (
                        -item[1]["amount_decimal"],
                        item[0],
                    ),
                )
            ),
        )
        if include_expense_content_summary:
            assert expense_content_sheet is not None
            self._append_table_sheet(
                expense_content_sheet,
                ["费用类型", "费用内容", "金额", "笔数"],
                [
                    [
                        expense_type,
                        expense_content,
                        _plain_money(bucket["amount_decimal"]),
                        bucket["transaction_count"],
                    ]
                    for (
                        expense_type,
                        expense_content,
                    ), bucket in sorted(
                        content_buckets.items(),
                        key=lambda item: (
                            -item[1]["amount_decimal"],
                            item[0][0],
                            item[0][1],
                        ),
                    )
                ],
            )
        return workbook

    @staticmethod
    def _transaction_workbook(payload: dict[str, Any]) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "流水详情"
        transaction = payload["transaction"]
        sheet.append(["字段", "值"])
        for key, value in [
            ("交易ID", transaction["id"]),
            ("月份", payload["month"]),
            ("项目名称", transaction["project_name"]),
            ("费用类型", transaction["expense_type"]),
            ("费用内容", transaction["expense_content"]),
            ("交易时间", transaction["trade_time"]),
            ("资金方向", transaction["direction"]),
            ("金额", transaction["amount"]),
            ("对方户名", transaction["counterparty_name"]),
            ("OA提交人", transaction["oa_applicant"]),
            ("支付账户", transaction["payment_account_label"]),
            ("备注", transaction["remark"]),
        ]:
            sheet.append([key, value])
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 52
        return workbook

    @staticmethod
    def _fill_key_value_sheet(
        sheet: Any,
        rows: list[tuple[str, Any]],
    ) -> None:
        sheet.append(["字段", "值"])
        for key, value in rows:
            sheet.append([key, value])
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 52

    @staticmethod
    def _append_table_sheet(
        sheet: Any,
        headers: list[str],
        rows: Any,
    ) -> None:
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        for index in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + index)].width = 18

    @staticmethod
    def _serialize_workbook(workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _build_filename(
        *,
        month: str,
        view: str,
        project_name: str | None = None,
        project_names: list[str] | None = None,
        aggregate_by: str | None = None,
        expense_type: str | None = None,
        transaction_id: str | None = None,
    ) -> str:
        month_segment = (
            "全部期间"
            if (month or "").strip().lower() == "all"
            else month
        )
        if view == "time":
            return f"成本统计_{month_segment}_按时间统计.xlsx"
        if view == "bank_tag":
            return f"成本统计_{month_segment}_按标签统计.xlsx"
        if view == "month":
            return f"成本统计_{month_segment}_月份汇总.xlsx"
        if view == "project":
            if aggregate_by is not None:
                project_label = (
                    "、".join(
                        project_names
                        or ([project_name] if project_name else [])
                    )
                    or "未命名项目"
                )
                period = "月" if aggregate_by == "month" else "年"
                return (
                    f"成本统计_{month_segment}_按项目统计_按{period}_"
                    f"{_sanitize_filename(project_label)}.xlsx"
                )
            return (
                f"成本统计_{month_segment}_项目明细_"
                f"{_sanitize_filename(project_name or '未命名项目')}.xlsx"
            )
        if view == "expense_type":
            return (
                f"成本统计_{month_segment}_按费用类型统计_"
                f"{_sanitize_filename(expense_type or '未命名费用类型')}.xlsx"
            )
        return (
            f"成本统计_{month_segment}_流水详情_"
            f"{_sanitize_filename(project_name or '未命名项目')}_"
            f"{_sanitize_filename(transaction_id or 'unknown')}.xlsx"
        )


def _plain_money(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01')):.2f}"


def _decimal_from_value(value: object) -> Decimal | None:
    if value in (None, "", "--", "—"):
        return None
    try:
        return Decimal(str(value).replace(",", ""))
    except Exception:
        return None


def _percentage(value: Decimal, total: Decimal) -> str:
    if total == Decimal("0.00"):
        return "0.00%"
    return (
        f"{(value / total * Decimal('100')).quantize(Decimal('0.01'))}%"
    )


def _sanitize_filename(value: str) -> str:
    sanitized = (
        str(value or "")
        .strip()
        .replace("/", "-")
        .replace("\\", "-")
        .replace(":", "：")
    )
    return sanitized[:80] if len(sanitized) > 80 else sanitized
