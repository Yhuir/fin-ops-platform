from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from fin_ops_platform.services.oa_pending_payment_query_contract import OaPendingPaymentError


OA_PENDING_PAYMENT_EXPORT_ROW_LIMIT = 20_000
OA_PENDING_PAYMENT_EXPORT_SOURCES = ("completed", "in_progress")
OA_PENDING_PAYMENT_EXPORT_COLUMNS = (
    ("oa_id", "OA ID", 26),
    ("workflow_no", "OA单号", 18),
    ("workflow_status", "流程状态", 12),
    ("month", "归属月份", 12),
    ("applicant", "申请人", 14),
    ("apply_type", "申请类型", 16),
    ("application_time", "申请时间", 20),
    ("completed_at", "完成时间", 20),
    ("project_name", "项目名称", 32),
    ("amount", "申请金额", 16),
    ("counterparty_name", "往来单位", 28),
    ("reason", "申请事由", 40),
    ("expense_type", "费用类型", 24),
    ("expense_content", "费用内容", 40),
)
OA_PENDING_PAYMENT_EXPORT_SHEETS = {
    "completed": "已完成OA",
    "in_progress": "进行中OA",
}


def parse_oa_pending_payment_export_sources(query: dict[str, list[str]]) -> tuple[str, ...]:
    requested = {
        source.strip()
        for value in query.get("sources", [])
        for source in str(value or "").split(",")
        if source.strip()
    }
    if not requested:
        raise OaPendingPaymentError(
            "oa_pending_payment_export_sources_required",
            "请至少选择一种 OA 来源。",
        )
    invalid = sorted(requested.difference(OA_PENDING_PAYMENT_EXPORT_SOURCES))
    if invalid:
        raise OaPendingPaymentError(
            "invalid_oa_pending_payment_export_source",
            "OA 导出来源只能是 completed 或 in_progress。",
            details={"sources": invalid},
        )
    return tuple(source for source in OA_PENDING_PAYMENT_EXPORT_SOURCES if source in requested)


def build_oa_pending_payment_export_workbook(
    rows: Iterable[dict[str, Any]],
    *,
    sources: tuple[str, ...],
) -> bytes:
    workbook = Workbook(write_only=True)
    sheets = {
        source: _create_sheet(workbook, OA_PENDING_PAYMENT_EXPORT_SHEETS[source])
        for source in sources
    }
    for row in rows:
        source = str(row.get("source_kind") or "").strip()
        sheet = sheets.get(source)
        if sheet is None:
            continue
        sheet.append([_xlsx_text(row.get(key)) for key, _label, _width in OA_PENDING_PAYMENT_EXPORT_COLUMNS])

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def oa_pending_payment_export_filename(*, today: date | None = None) -> str:
    return f"OA事实源_{(today or date.today()).isoformat()}.xlsx"


def _create_sheet(workbook: Workbook, title: str) -> Any:
    sheet = workbook.create_sheet(title=title)
    sheet.freeze_panes = "A2"
    for index, (_key, _label, width) in enumerate(OA_PENDING_PAYMENT_EXPORT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    header: list[WriteOnlyCell] = []
    for _key, label, _width in OA_PENDING_PAYMENT_EXPORT_COLUMNS:
        cell = WriteOnlyCell(sheet, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        header.append(cell)
    sheet.append(header)
    return sheet


def _xlsx_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        normalized = value.isoformat(sep=" ", timespec="seconds")
    elif isinstance(value, date):
        normalized = value.isoformat()
    else:
        normalized = str(value)
    normalized = ILLEGAL_CHARACTERS_RE.sub("", normalized).strip()
    if normalized in {"—", "--", "None"}:
        return ""
    if normalized.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{normalized}"
    return normalized[:32_767]
