from __future__ import annotations

import hashlib
import re
import unicodedata
import warnings
from copy import deepcopy
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Callable
from uuid import uuid4
from zipfile import BadZipFile, ZipFile

import xlrd
from openpyxl import load_workbook

from fin_ops_platform.domain.enums import BatchType, ImportDecision
from fin_ops_platform.domain.models import ImportedBatchRowResult
from fin_ops_platform.services.import_preview_audit import (
    BankTransactionIdentityStrategy,
    ImportPreviewAuditCounts,
    ImportPreviewAuditRow,
    ImportPreviewDuplicateGroup,
    ImportPreviewFileAudit,
    ImportPreviewSessionAudit,
    ImportPreviewStaleError,
    ImportRecordIdentity,
    InvoiceIdentityStrategy,
    build_import_preview_session_audit,
)
from fin_ops_platform.services.imports import ImportNormalizationService

DATE_ONLY_RE = re.compile(r"^(\d{4})[-/](\d{2})[-/](\d{2})$")
DATE_TIME_RE = re.compile(r"^(\d{4})[-/](\d{2})[-/](\d{2})[ T](\d{2}):(\d{2}):(\d{2})$")
COMPACT_DATE_RE = re.compile(r"^(\d{4})(\d{2})(\d{2})$")
COMPACT_DATE_TIME_RE = re.compile(r"^(\d{4})(\d{2})(\d{2})[ T]?(\d{2})(\d{2})(\d{2})$")
COMPANY_TAX_NOS = {"91330106589876543T", "915300007194052520"}
COMPANY_NAME_KEYWORDS = ("云南溯源科技有限公司", "溯源科技有限公司")
ACCOUNT_METADATA_KEYWORDS = ("账号", "账户", "卡号")
BANK_TEXT_FIELD_LABELS = ("摘要", "备注", "用途", "交易用途", "客户附言", "附言")
INVOICE_REQUIRED_HEADERS = {"发票代码", "发票号码", "销方识别号", "购买方名称", "开票日期", "金额", "税额"}
INVOICE_HEADER_ALIASES = {
    "数电号码": "数电发票号码",
    "销方税号": "销方识别号",
    "销方企业名称": "销方名称",
    "购方税号": "购方识别号",
    "购方企业名称": "购买方名称",
    "商品名称": "货物或应税劳务名称",
    "规格": "规格型号",
    "发票类型": "发票票种",
}
TEMPLATE_DEFINITIONS: list[dict[str, Any]] = [
    {
        "template_code": "invoice_export",
        "label": "发票导出",
        "file_extensions": [".xlsx"],
        "record_type": "invoice",
        "allowed_batch_types": [BatchType.INPUT_INVOICE.value, BatchType.OUTPUT_INVOICE.value],
        "required_headers": sorted(INVOICE_REQUIRED_HEADERS),
    },
    {
        "template_code": "bank_statement",
        "label": "银行流水",
        "file_extensions": [".xls", ".xlsx"],
        "record_type": "bank_transaction",
        "allowed_batch_types": [BatchType.BANK_TRANSACTION.value],
        "required_headers": ["交易日期或时间", "借方和贷方金额，或金额和收支方向"],
    },
]

BANK_FIELD_LABELS = {
    "account_no": "本方账号",
    "account_name": "本方户名",
    "trade_time": "交易时间",
    "txn_date": "交易日期",
    "txn_clock": "交易时刻",
    "debit_amount": "支出金额",
    "credit_amount": "收入金额",
    "amount": "交易金额",
    "direction": "收支方向",
    "balance": "账户余额",
    "counterparty_name": "对方名称",
    "counterparty_account_no": "对方账号",
    "counterparty_bank_name": "对方开户行",
    "summary": "摘要",
    "remark": "备注/用途",
    "bank_serial_no": "银行流水号",
    "account_detail_no": "账户明细编号",
    "enterprise_serial_no": "企业流水号",
    "voucher_kind": "凭证种类",
    "voucher_no": "凭证号",
    "currency": "币种",
    "booked_date": "记账日期",
}
BANK_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "account_no": ("账号", "帐号", "客户账号", "查询账号", "本方账号", "账户号", "卡号"),
    "account_name": ("账户名称", "帐户名称", "户名", "本方户名", "客户名称"),
    "trade_time": ("交易时间", "交易日期时间", "交易日期及时间", "发生时间", "入账时间", "记账时间"),
    "txn_date": ("交易日期", "交易日", "日期"),
    "txn_clock": ("交易时刻", "交易时间点", "时间"),
    "debit_amount": (
        "借方发生额", "借方发生额支取", "借方发生额支出", "借方金额", "借方金额支出", "支出", "支出金额", "转出金额", "付出金额",
    ),
    "credit_amount": (
        "贷方发生额", "贷方发生额收入", "贷方金额", "贷方金额收入", "收入", "收入金额", "转入金额", "收款金额",
    ),
    "amount": ("交易金额", "发生额", "金额", "本次金额"),
    "direction": ("收支方向", "交易方向", "借贷标志", "收付标志", "收支标志", "借贷方向"),
    "balance": ("余额", "账户余额", "帐户余额", "交易后余额", "可用余额"),
    "counterparty_name": (
        "对方户名", "对方名称", "对方单位", "对方账号名称", "对手方名称", "对手方户名", "收付款人名称",
    ),
    "counterparty_account_no": ("对方账号", "对方账户", "对手方账号", "收付款人账号"),
    "counterparty_bank_name": ("对方开户行", "对方开户机构", "对方银行", "对方账号开户行", "对手方开户行", "对方行名"),
    "summary": ("摘要", "交易摘要"),
    "remark": ("备注", "用途", "交易用途", "客户附言", "附言"),
    "bank_serial_no": ("银行流水号", "交易流水号", "流水号", "核心唯一流水号", "银行交易流水号"),
    "account_detail_no": ("账户明细编号交易流水号", "账户明细编号", "明细编号"),
    "enterprise_serial_no": ("企业流水号", "业务流水号", "客户流水号"),
    "voucher_kind": ("凭证种类", "凭证类型"),
    "voucher_no": ("凭证号", "凭证号码"),
    "currency": ("币种", "货币种类"),
    "booked_date": ("记账日期", "入账日期", "账务日期"),
}
BANK_NAME_MARKERS = {
    "工商银行": ("中国工商银行", "工商银行", "[historydetail]"),
    "建设银行": ("中国建设银行", "建设银行"),
    "光大银行": ("中国光大银行", "光大银行"),
    "民生银行": ("中国民生银行", "民生银行"),
    "平安银行": ("平安银行", "核心唯一流水号"),
    "交通银行": ("交通银行",),
}
XLS_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
MAX_WORKBOOK_SHEETS = 32
MAX_WORKBOOK_ROWS = 200_000
MAX_WORKBOOK_COLUMNS = 256
MAX_WORKBOOK_CELLS = 2_000_000
MAX_XLSX_MEMBERS = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200


@dataclass(slots=True)
class UploadedImportFile:
    file_name: str
    content: bytes
    template_code_override: str | None = None
    batch_type_override: str | None = None
    selected_bank_mapping_id: str | None = None
    selected_bank_name: str | None = None
    selected_bank_short_name: str | None = None
    selected_bank_last4: str | None = None
    field_mapping: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True, slots=True)
class SourceControlEvidence:
    status: str
    computed_row_count: int
    declared_row_count: int | None = None
    computed_debit_total: str | None = None
    declared_debit_total: str | None = None
    computed_credit_total: str | None = None
    declared_credit_total: str | None = None
    mismatch_fields: tuple[str, ...] = ()


@dataclass(slots=True)
class FileImportPreviewItem:
    id: str
    file_name: str
    template_code: str | None
    batch_type: BatchType | None
    status: str
    message: str
    row_count: int
    success_count: int = 0
    error_count: int = 0
    duplicate_count: int = 0
    suspected_duplicate_count: int = 0
    updated_count: int = 0
    preview_batch_id: str | None = None
    batch_id: str | None = None
    stored_file_path: str | None = None
    override_template_code: str | None = None
    override_batch_type: BatchType | None = None
    selected_bank_mapping_id: str | None = None
    selected_bank_name: str | None = None
    selected_bank_short_name: str | None = None
    selected_bank_last4: str | None = None
    detected_bank_name: str | None = None
    detected_last4: str | None = None
    bank_selection_conflict: bool = False
    conflict_message: str | None = None
    header_signature: str | None = None
    mapping_candidates: list[dict[str, str]] = field(default_factory=list)
    mapping_fields: list[dict[str, Any]] = field(default_factory=list)
    field_mapping: dict[str, str] = field(default_factory=dict)
    mapping_source: str | None = None
    row_results: list[ImportedBatchRowResult] = field(default_factory=list)
    normalized_rows: list[dict[str, Any]] = field(default_factory=list)
    audit: ImportPreviewAuditCounts = field(default_factory=ImportPreviewAuditCounts)
    content_sha256: str | None = None
    duplicate_file_name: str | None = None
    source_control: SourceControlEvidence | None = None


@dataclass(slots=True)
class FileImportSession:
    id: str
    imported_by: str
    file_count: int
    status: str
    files: list[FileImportPreviewItem]
    created_at: datetime = field(default_factory=lambda: datetime.now(UTC))
    audit: ImportPreviewAuditCounts = field(default_factory=ImportPreviewAuditCounts)
    duplicate_groups: list[ImportPreviewDuplicateGroup] = field(default_factory=list)


@dataclass(slots=True)
class ParsedImportFile:
    template_code: str
    batch_type: BatchType
    rows: list[dict[str, Any]]
    detected_bank_name: str | None = None
    header_signature: str | None = None
    mapping_candidates: list[dict[str, str]] = field(default_factory=list)
    mapping_fields: list[dict[str, Any]] = field(default_factory=list)
    field_mapping: dict[str, str] = field(default_factory=dict)
    mapping_source: str | None = None
    source_control: SourceControlEvidence | None = None


class BankStatementMappingRequired(ValueError):
    def __init__(
        self,
        message: str,
        *,
        header_signature: str | None = None,
        candidates: list[dict[str, str]] | None = None,
        mapping_fields: list[dict[str, Any]] | None = None,
        field_mapping: dict[str, str] | None = None,
        detected_bank_name: str | None = None,
    ) -> None:
        super().__init__(message)
        self.header_signature = header_signature
        self.candidates = list(candidates or [])
        self.mapping_fields = list(mapping_fields or [])
        self.field_mapping = dict(field_mapping or {})
        self.detected_bank_name = detected_bank_name


class FileImportService:
    def __init__(self, import_service: ImportNormalizationService, *, file_store: Any | None = None) -> None:
        self._import_service = import_service
        self._session_counter = 0
        self._file_counter = 0
        self._sessions: dict[str, FileImportSession] = {}
        self._file_store = file_store

    @classmethod
    def from_snapshot(
        cls,
        import_service: ImportNormalizationService,
        snapshot: dict[str, Any] | None,
        *,
        file_store: Any | None = None,
    ) -> FileImportService:
        service = cls(import_service, file_store=file_store)
        if not snapshot:
            return service
        service._session_counter = int(snapshot.get("session_counter", 0))
        service._file_counter = int(snapshot.get("file_counter", 0))
        service._sessions = dict(snapshot.get("sessions", {}))
        return service

    def snapshot(self) -> dict[str, Any]:
        return {
            "session_counter": self._session_counter,
            "file_counter": self._file_counter,
            "sessions": self._sessions,
        }

    def preview_session_persistence_payload(self, session_id: str) -> dict[str, Any]:
        session = self._sessions[session_id]
        batch_ids = [
            str(item.preview_batch_id).strip()
            for item in session.files
            if str(item.preview_batch_id or "").strip()
        ]
        return {
            "imports": self._import_service.persistence_snapshot_for_batches(
                batch_ids,
                include_facts=False,
            ),
            "file_imports": {
                "session_counter": self._session_counter,
                "file_counter": self._file_counter,
                "sessions": {session.id: session},
            },
        }

    def confirmed_session_persistence_payload(
        self,
        *,
        session_id: str,
        selected_file_ids: list[str],
    ) -> dict[str, Any]:
        session = self._sessions[session_id]
        selected = {str(file_id).strip() for file_id in selected_file_ids if str(file_id).strip()}
        batch_ids = [
            str(item.batch_id).strip()
            for item in session.files
            if item.id in selected and item.status == "confirmed" and str(item.batch_id or "").strip()
        ]
        return {
            "imports": self._import_service.persistence_snapshot_for_batches(batch_ids),
            "file_imports": {
                "session_counter": self._session_counter,
                "file_counter": self._file_counter,
                "sessions": {session.id: session},
            },
        }

    def list_templates(self) -> list[dict[str, Any]]:
        return [dict(template) for template in TEMPLATE_DEFINITIONS]

    def preview_files(self, *, imported_by: str, uploads: list[UploadedImportFile]) -> FileImportSession:
        session = FileImportSession(
            id=self._next_session_id(),
            imported_by=imported_by,
            file_count=len(uploads),
            status="preview_ready",
            files=[],
        )

        seen_hashes: dict[str, str] = {}
        for upload in uploads:
            file_id = self._next_file_id()
            stored_file_path = self._store_upload_file(
                session.id,
                file_id,
                upload,
                imported_by=imported_by,
            )
            content_sha256 = hashlib.sha256(upload.content).hexdigest()
            duplicate_name = seen_hashes.get(content_sha256)
            if duplicate_name:
                file_item = self._build_preview_error_item(
                    file_id=file_id,
                    upload=upload,
                    stored_file_path=stored_file_path,
                    message=f"文件内容与本次上传的“{duplicate_name}”完全相同。",
                    status="duplicate_file",
                    content_sha256=content_sha256,
                    duplicate_file_name=duplicate_name,
                    template_code_override=upload.template_code_override,
                    batch_type_override=upload.batch_type_override,
                    selected_bank_mapping_id=upload.selected_bank_mapping_id,
                    selected_bank_name=upload.selected_bank_name,
                    selected_bank_short_name=upload.selected_bank_short_name,
                    selected_bank_last4=upload.selected_bank_last4,
                )
            else:
                file_item = self._preview_single_file(
                    imported_by=imported_by,
                    upload=upload,
                    file_id=file_id,
                    stored_file_path=stored_file_path,
                    template_code_override=upload.template_code_override,
                    batch_type_override=upload.batch_type_override,
                    selected_bank_mapping_id=upload.selected_bank_mapping_id,
                    selected_bank_name=upload.selected_bank_name,
                    selected_bank_short_name=upload.selected_bank_short_name,
                    selected_bank_last4=upload.selected_bank_last4,
                    field_mapping=upload.field_mapping,
                )
            seen_hashes.setdefault(content_sha256, upload.file_name)
            session.files.append(file_item)

        if any(file.status != "preview_ready" for file in session.files):
            session.status = "preview_ready_with_errors"

        self._refresh_session_audit(session)
        self._sessions[session.id] = session
        return session

    def get_session(self, session_id: str) -> FileImportSession:
        return self._sessions[session_id]

    def assert_session_owner(self, *, session_id: str, imported_by: str) -> FileImportSession:
        session = self._sessions[session_id]
        if str(session.imported_by) != str(imported_by):
            raise PermissionError("import session belongs to another user")
        return session

    def list_active_sessions(self, *, imported_by: str, mode: str | None = None) -> list[FileImportSession]:
        def matches_mode(session: FileImportSession) -> bool:
            batch_types = {item.batch_type for item in session.files if item.batch_type is not None}
            if mode == "bank_transaction":
                return BatchType.BANK_TRANSACTION in batch_types
            if mode == "invoice":
                return bool(batch_types & {BatchType.INPUT_INVOICE, BatchType.OUTPUT_INVOICE})
            return True

        return sorted(
            (
                session
                for session in self._sessions.values()
                if session.imported_by == imported_by
                and session.status in {"preview_ready", "preview_ready_with_errors"}
                and matches_mode(session)
            ),
            key=lambda session: (session.created_at, session.id),
            reverse=True,
        )

    def discard_session(self, *, session_id: str, imported_by: str) -> FileImportSession:
        session = self.assert_session_owner(session_id=session_id, imported_by=imported_by)
        if session.status == "reverted":
            return session
        if any(item.status == "confirmed" for item in session.files):
            raise ValueError("confirmed import sessions cannot be discarded")
        for item in session.files:
            if item.preview_batch_id:
                self._import_service.discard_preview(item.preview_batch_id)
            item.status = "reverted"
            item.batch_id = None
        session.status = "reverted"
        self._sessions[session.id] = session
        return session

    def review_rows(
        self,
        *,
        session_id: str,
        kind: str,
        offset: int,
        limit: int,
    ) -> dict[str, Any]:
        session = self._sessions[session_id]
        if kind == "duplicates":
            rows = [
                {
                    **dict(row),
                    "record_type": group.record_type,
                    "duplicate_type": group.duplicate_type,
                }
                for group in session.duplicate_groups
                for row in group.rows
            ]
        elif kind == "unimported":
            rows = []
            for item in session.files:
                for row_result, normalized in zip(item.row_results, item.normalized_rows, strict=True):
                    decision = (
                        row_result.decision.value
                        if isinstance(row_result.decision, ImportDecision)
                        else str(row_result.decision)
                    )
                    if decision not in {"duplicate_skipped", "suspected_duplicate", "error"}:
                        continue
                    rows.append(
                        {
                            "file_id": item.id,
                            "file_name": item.file_name,
                            "row_no": row_result.row_no,
                            "record_type": row_result.source_record_type,
                            "decision": decision,
                            "decision_reason": row_result.decision_reason,
                            "identity_kind": row_result.identity_kind,
                            **self._audit_row_display_fields(row_result.source_record_type, normalized),
                        }
                    )
        else:
            raise ValueError("kind must be duplicates or unimported")
        page_rows = rows[offset : offset + limit]
        return {
            "rows": page_rows,
            "total": len(rows),
            "offset": offset,
            "limit": limit,
            "has_more": offset + len(page_rows) < len(rows),
        }

    def confirm_session(
        self,
        *,
        session_id: str,
        selected_file_ids: list[str],
        progress_callback: Callable[[FileImportSession, int, int], None] | None = None,
    ) -> FileImportSession:
        session = self._sessions[session_id]
        selected = set(selected_file_ids)
        if not selected:
            raise ValueError("at least one selected file is required")
        known_ids = {item.id for item in session.files}
        unknown_ids = sorted(selected - known_ids)
        if unknown_ids:
            raise KeyError(f"Unknown selected file ids: {', '.join(unknown_ids)}")

        confirmed_any = False
        selected_items = [item for item in session.files if item.id in selected]
        invalid_items = [item.id for item in selected_items if item.status not in {"preview_ready", "confirmed"}]
        if invalid_items:
            raise ValueError(f"selected files are not confirmable: {', '.join(sorted(invalid_items))}")
        if any(item.status == "preview_ready" for item in selected_items):
            self.assert_session_preview_current(session_id=session_id)
        progress_total = len(selected_items)
        progress_current = 0
        rollback_session = deepcopy(session)
        try:
            for item in session.files:
                if item.id not in selected:
                    if item.status == "preview_ready":
                        item.status = "skipped"
                        item.batch_id = None
                    continue
                progress_current += 1
                if item.status == "confirmed":
                    confirmed_any = True
                    if progress_callback is not None:
                        progress_callback(session, progress_current, progress_total)
                    continue
                if not item.preview_batch_id:
                    if progress_callback is not None:
                        progress_callback(session, progress_current, progress_total)
                    continue
                batch = self._import_service.confirm_import(item.preview_batch_id)
                item.batch_id = batch.id
                item.status = "confirmed"
                confirmed_any = True
                if progress_callback is not None:
                    progress_callback(session, progress_current, progress_total)

            session.status = "confirmed" if confirmed_any else "skipped"
            self._refresh_session_audit(session)
            self._sessions[session.id] = session
            return session
        except Exception:
            self._sessions[session.id] = rollback_session
            raise

    def assert_session_preview_current(self, *, session_id: str) -> None:
        session = self._sessions[session_id]
        if session.status == "reverted":
            raise ValueError("discarded import sessions cannot be confirmed")
        current_audit = self._build_session_audit(session, refresh_existing=True)
        if current_audit.audit.stale_projection() != session.audit.stale_projection():
            raise ImportPreviewStaleError(preview=session.audit, current=current_audit.audit)

    def retry_session_files(
        self,
        *,
        session_id: str,
        selected_file_ids: list[str],
        overrides: dict[str, dict[str, Any]] | None = None,
    ) -> FileImportSession:
        session = self._sessions[session_id]
        if session.status == "reverted":
            raise ValueError("discarded import sessions cannot be retried")
        override_map = overrides or {}
        selected = set(selected_file_ids)
        for item in session.files:
            if item.id not in selected:
                continue
            if item.status == "confirmed":
                raise ValueError("confirmed files cannot be retried directly")
            if not item.stored_file_path:
                raise ValueError("original upload file is missing")
            if self._file_store is None:
                raise ValueError("import file storage is not configured")
            upload = UploadedImportFile(
                file_name=item.file_name,
                content=self._file_store.read_import_file(item.stored_file_path),
                selected_bank_mapping_id=item.selected_bank_mapping_id,
                selected_bank_name=item.selected_bank_name,
                selected_bank_short_name=item.selected_bank_short_name,
                selected_bank_last4=item.selected_bank_last4,
                field_mapping=item.field_mapping,
            )
            override_payload = override_map.get(item.id, {})
            field_mapping = override_payload.get("field_mapping")
            if not isinstance(field_mapping, dict):
                field_mapping = item.field_mapping
            refreshed = self._preview_single_file(
                imported_by=session.imported_by,
                upload=upload,
                file_id=item.id,
                stored_file_path=item.stored_file_path,
                template_code_override=override_payload.get("template_code"),
                batch_type_override=override_payload.get("batch_type"),
                selected_bank_mapping_id=override_payload.get("bank_mapping_id") or item.selected_bank_mapping_id,
                selected_bank_name=override_payload.get("bank_name") or item.selected_bank_name,
                selected_bank_short_name=override_payload.get("bank_short_name") or item.selected_bank_short_name,
                selected_bank_last4=override_payload.get("last4") or item.selected_bank_last4,
                field_mapping={str(key): str(value) for key, value in field_mapping.items()},
            )
            item.template_code = refreshed.template_code
            item.batch_type = refreshed.batch_type
            item.status = refreshed.status
            item.message = refreshed.message
            item.row_count = refreshed.row_count
            item.success_count = refreshed.success_count
            item.error_count = refreshed.error_count
            item.duplicate_count = refreshed.duplicate_count
            item.suspected_duplicate_count = refreshed.suspected_duplicate_count
            item.updated_count = refreshed.updated_count
            item.preview_batch_id = refreshed.preview_batch_id
            item.row_results = refreshed.row_results
            item.normalized_rows = refreshed.normalized_rows
            item.override_template_code = refreshed.override_template_code
            item.override_batch_type = refreshed.override_batch_type
            item.selected_bank_mapping_id = refreshed.selected_bank_mapping_id
            item.selected_bank_name = refreshed.selected_bank_name
            item.selected_bank_short_name = refreshed.selected_bank_short_name
            item.selected_bank_last4 = refreshed.selected_bank_last4
            item.detected_bank_name = refreshed.detected_bank_name
            item.detected_last4 = refreshed.detected_last4
            item.bank_selection_conflict = refreshed.bank_selection_conflict
            item.conflict_message = refreshed.conflict_message
            item.header_signature = refreshed.header_signature
            item.mapping_candidates = refreshed.mapping_candidates
            item.mapping_fields = refreshed.mapping_fields
            item.field_mapping = refreshed.field_mapping
            item.mapping_source = refreshed.mapping_source
            item.content_sha256 = refreshed.content_sha256
            item.duplicate_file_name = refreshed.duplicate_file_name
            item.source_control = refreshed.source_control

        session.status = "preview_ready_with_errors" if any(
            file.status != "preview_ready" for file in session.files
        ) else "preview_ready"
        self._refresh_session_audit(session)
        self._sessions[session.id] = session
        return session

    def replay_confirmed_session_files(
        self,
        *,
        source_session_id: str,
        selected_file_ids: list[str],
        imported_by: str,
        expected_repaired_duplicate_count: int = 0,
        repaired_duplicate_decision_reason: str | None = None,
        repaired_duplicate_evidence: list[dict[str, Any]] | None = None,
        expected_canonical_owner_count: int = 0,
        canonical_owner_evidence: list[dict[str, Any]] | None = None,
    ) -> FileImportSession:
        source_session = self._sessions[source_session_id]
        selected = {str(file_id).strip() for file_id in selected_file_ids if str(file_id).strip()}
        if not selected:
            raise ValueError("at least one selected source file is required")
        source_items = [item for item in source_session.files if item.id in selected]
        if {item.id for item in source_items} != selected:
            raise KeyError("one or more replay source files do not belong to the source session")
        if any(item.status != "confirmed" for item in source_items):
            raise ValueError("only confirmed source files can be replayed")
        if self._file_store is None:
            raise ValueError("import file storage is not configured")

        replay_session = FileImportSession(
            id=self._next_session_id(),
            imported_by=str(imported_by or "").strip() or "system",
            file_count=len(source_items),
            status="preview_ready",
            files=[],
        )
        selected_file_ids_set = {item.id for item in source_items}
        evidence_by_file: dict[str, list[dict[str, Any]]] = {
            file_id: [] for file_id in selected_file_ids_set
        }
        canonical_owner_evidence_by_file: dict[str, list[dict[str, Any]]] = {
            file_id: [] for file_id in selected_file_ids_set
        }
        for evidence in repaired_duplicate_evidence or []:
            file_id = str(evidence.get("file_id") or "").strip()
            if file_id not in selected_file_ids_set:
                raise ValueError(
                    "controlled replay repaired evidence references an unselected source file"
                )
            evidence_by_file[file_id].append(dict(evidence))
        for evidence in canonical_owner_evidence or []:
            file_id = str(evidence.get("file_id") or "").strip()
            if file_id not in selected_file_ids_set:
                raise ValueError(
                    "controlled replay canonical owner evidence references an unselected source file"
                )
            canonical_owner_evidence_by_file[file_id].append(dict(evidence))
        if sum(len(items) for items in canonical_owner_evidence_by_file.values()) != int(
            expected_canonical_owner_count
        ):
            raise ValueError("controlled replay canonical owner evidence count changed")
        repaired_duplicate_count = 0
        for source_item in source_items:
            if not source_item.stored_file_path:
                raise ValueError(f"replay source file {source_item.id} is missing stored content")
            content = self._file_store.read_import_file(source_item.stored_file_path)
            content_sha256 = hashlib.sha256(content).hexdigest()
            if source_item.content_sha256 and content_sha256 != source_item.content_sha256:
                raise ValueError(f"replay source file {source_item.id} checksum changed")
            upload = UploadedImportFile(
                file_name=source_item.file_name,
                content=content,
                selected_bank_mapping_id=source_item.selected_bank_mapping_id,
                selected_bank_name=source_item.selected_bank_name,
                selected_bank_short_name=source_item.selected_bank_short_name,
                selected_bank_last4=source_item.selected_bank_last4,
                field_mapping=dict(source_item.field_mapping),
            )
            replay_item = self._preview_single_file(
                imported_by=replay_session.imported_by,
                upload=upload,
                file_id=self._next_file_id(),
                stored_file_path=source_item.stored_file_path,
                template_code_override=source_item.template_code,
                batch_type_override=(source_item.batch_type.value if source_item.batch_type else None),
                selected_bank_mapping_id=source_item.selected_bank_mapping_id,
                selected_bank_name=source_item.selected_bank_name,
                selected_bank_short_name=source_item.selected_bank_short_name,
                selected_bank_last4=source_item.selected_bank_last4,
                field_mapping=dict(source_item.field_mapping),
                skip_duplicate_file_guard=True,
            )
            repaired_duplicate_count += self._resolve_repaired_replay_duplicates(
                replay_item=replay_item,
                repaired_duplicate_decision_reason=repaired_duplicate_decision_reason,
                repaired_duplicate_evidence=evidence_by_file[source_item.id],
                canonical_owner_evidence=canonical_owner_evidence_by_file[source_item.id],
                source_file_id=source_item.id,
            )
            replay_session.files.append(replay_item)

        if repaired_duplicate_count != int(expected_repaired_duplicate_count):
            raise ValueError(
                "controlled replay repaired duplicate count changed: "
                f"expected {int(expected_repaired_duplicate_count)}, "
                f"resolved {repaired_duplicate_count}"
            )

        if any(file.status != "preview_ready" for file in replay_session.files):
            replay_session.status = "preview_ready_with_errors"
        self._refresh_session_audit(replay_session)
        self._sessions[replay_session.id] = replay_session
        return replay_session

    def _resolve_repaired_replay_duplicates(
        self,
        *,
        replay_item: FileImportPreviewItem,
        repaired_duplicate_decision_reason: str | None,
        repaired_duplicate_evidence: list[dict[str, Any]],
        canonical_owner_evidence: list[dict[str, Any]] | None = None,
        source_file_id: str | None = None,
    ) -> int:
        if replay_item.batch_type != BatchType.BANK_TRANSACTION:
            return 0
        normalized_repair_reason = str(
            repaired_duplicate_decision_reason or ""
        ).strip()
        if not normalized_repair_reason:
            return 0
        evidence_by_row_no: dict[int, tuple[str, dict[str, Any]]] = {}
        for evidence in repaired_duplicate_evidence:
            row_no = int(evidence.get("row_no") or 0)
            if row_no <= 0 or row_no in evidence_by_row_no:
                raise ValueError("controlled replay repaired evidence has invalid row numbers")
            if (
                str(evidence.get("decision_reason") or "").strip()
                != normalized_repair_reason
                or str(evidence.get("source_record_type") or "").strip()
                != "bank_transaction"
                or not str(evidence.get("data_fingerprint") or "").strip()
                or str(evidence.get("linked_object_type") or "").strip()
                != "bank_transaction"
                or not str(evidence.get("linked_object_id") or "").strip()
            ):
                raise ValueError(
                    "controlled replay repaired evidence lacks canonical row identity"
                )
            evidence_by_row_no[row_no] = ("repaired_duplicate", evidence)
        for evidence in canonical_owner_evidence or []:
            row_no = int(evidence.get("row_no") or 0)
            if row_no <= 0 or row_no in evidence_by_row_no:
                raise ValueError("controlled replay canonical owner evidence has invalid row numbers")
            if (
                str(evidence.get("source_record_type") or "").strip()
                != "bank_transaction"
                or not str(evidence.get("data_fingerprint") or "").strip()
                or str(evidence.get("linked_object_type") or "").strip()
                != "bank_transaction"
                or not str(evidence.get("linked_object_id") or "").strip()
            ):
                raise ValueError(
                    "controlled replay canonical owner evidence lacks canonical row identity"
                )
            evidence_by_row_no[row_no] = ("canonical_owner", evidence)

        resolved_count = 0
        for replay_result in replay_item.row_results:
            evidence_entry = evidence_by_row_no.pop(replay_result.row_no, None)
            if evidence_entry is None:
                if replay_result.decision == ImportDecision.SUSPECTED_DUPLICATE:
                    raise ValueError(
                        "controlled replay suspected duplicate lacks authoritative row evidence: "
                        f"source_file_id={source_file_id or ''}; "
                        f"row_no={replay_result.row_no}; "
                        f"fingerprint={replay_result.data_fingerprint}; "
                        f"linked_object_id={replay_result.linked_object_id or ''}"
                    )
                continue
            evidence_kind, evidence = evidence_entry
            mismatches = [
                field_name
                for field_name, matches in (
                    (
                        "source_record_type",
                        replay_result.source_record_type
                        == str(evidence["source_record_type"]),
                    ),
                    (
                        "data_fingerprint",
                        replay_result.data_fingerprint
                        == str(evidence["data_fingerprint"]),
                    ),
                    (
                        "decision",
                        replay_result.decision
                        in {
                            ImportDecision.CREATED,
                            ImportDecision.SUSPECTED_DUPLICATE,
                            ImportDecision.DUPLICATE_SKIPPED,
                        },
                    ),
                    (
                        "linked_object_shape",
                        (
                            replay_result.decision == ImportDecision.CREATED
                            and replay_result.linked_object_type is None
                            and replay_result.linked_object_id is None
                        )
                        or (
                            replay_result.decision
                            in {
                                ImportDecision.SUSPECTED_DUPLICATE,
                                ImportDecision.DUPLICATE_SKIPPED,
                            }
                            and replay_result.linked_object_type
                            == "bank_transaction"
                            and bool(str(replay_result.linked_object_id or "").strip())
                        ),
                    ),
                )
                if not matches
            ]
            if mismatches:
                raise ValueError(
                    "controlled replay current preview changed from repaired row evidence: "
                    f"row_no={replay_result.row_no}; fields={','.join(mismatches)}"
                )
            replay_result.decision = ImportDecision.DUPLICATE_SKIPPED
            replay_result.decision_reason = (
                "Controlled replay matched an explicitly repaired canonical row owner."
            )
            replay_result.linked_object_type = "bank_transaction"
            replay_result.linked_object_id = str(evidence["linked_object_id"])
            if evidence_kind == "repaired_duplicate":
                resolved_count += 1
        if evidence_by_row_no:
            raise ValueError(
                "controlled replay repaired evidence rows are missing from the current preview"
            )

        if replay_item.preview_batch_id:
            preview = self._import_service.get_batch(replay_item.preview_batch_id)
            preview.batch.success_count = sum(
                row.decision == ImportDecision.CREATED
                for row in preview.row_results
            )
            preview.batch.duplicate_count = sum(
                row.decision == ImportDecision.DUPLICATE_SKIPPED
                for row in preview.row_results
            )
            preview.batch.suspected_duplicate_count = sum(
                row.decision == ImportDecision.SUSPECTED_DUPLICATE
                for row in preview.row_results
            )
            replay_item.success_count = preview.batch.success_count
            replay_item.duplicate_count = preview.batch.duplicate_count
            replay_item.suspected_duplicate_count = preview.batch.suspected_duplicate_count
        return resolved_count

    def _preview_single_file(
        self,
        *,
        imported_by: str,
        upload: UploadedImportFile,
        file_id: str,
        stored_file_path: str | None,
        template_code_override: str | None = None,
        batch_type_override: str | None = None,
        selected_bank_mapping_id: str | None = None,
        selected_bank_name: str | None = None,
        selected_bank_short_name: str | None = None,
        selected_bank_last4: str | None = None,
        field_mapping: dict[str, str] | None = None,
        skip_duplicate_file_guard: bool = False,
    ) -> FileImportPreviewItem:
        content_sha256 = hashlib.sha256(upload.content).hexdigest()
        duplicate_file_name = (
            None
            if skip_duplicate_file_guard
            else self._find_confirmed_duplicate_file(
                content_sha256=content_sha256,
                exclude_file_id=file_id,
            )
        )
        if duplicate_file_name:
            return self._build_preview_error_item(
                file_id=file_id,
                upload=upload,
                stored_file_path=stored_file_path,
                message=f"该文件内容已通过“{duplicate_file_name}”确认导入。",
                status="duplicate_file",
                content_sha256=content_sha256,
                duplicate_file_name=duplicate_file_name,
                template_code_override=template_code_override,
                batch_type_override=batch_type_override,
                selected_bank_mapping_id=selected_bank_mapping_id,
                selected_bank_name=selected_bank_name,
                selected_bank_short_name=selected_bank_short_name,
                selected_bank_last4=selected_bank_last4,
            )
        try:
            rows = self._read_rows(upload)
            try:
                parsed = self._parse_rows(
                    rows=rows,
                    template_code_override=template_code_override,
                    batch_type_override=batch_type_override,
                    field_mapping=field_mapping,
                )
                if field_mapping and parsed.batch_type == BatchType.BANK_TRANSACTION:
                    parsed.mapping_source = "manual"
            except BankStatementMappingRequired as mapping_error:
                saved_mapping = self._saved_field_mapping(mapping_error.header_signature)
                if field_mapping or not saved_mapping:
                    return self._build_preview_error_item(
                        file_id=file_id,
                        upload=upload,
                        stored_file_path=stored_file_path,
                        message=str(mapping_error),
                        template_code_override=template_code_override,
                        batch_type_override=batch_type_override,
                        selected_bank_mapping_id=selected_bank_mapping_id,
                        selected_bank_name=selected_bank_name,
                        selected_bank_short_name=selected_bank_short_name,
                        selected_bank_last4=selected_bank_last4,
                        template_code="bank_statement",
                        batch_type=BatchType.BANK_TRANSACTION,
                        header_signature=mapping_error.header_signature,
                        mapping_candidates=mapping_error.candidates,
                        mapping_fields=mapping_error.mapping_fields,
                        field_mapping=mapping_error.field_mapping,
                        detected_bank_name=mapping_error.detected_bank_name,
                    )
                parsed = self._parse_rows(
                    rows=rows,
                    template_code_override="bank_statement",
                    batch_type_override=BatchType.BANK_TRANSACTION.value,
                    field_mapping=saved_mapping,
                )
                parsed.mapping_source = "saved"
        except ValueError as exc:
            return self._build_preview_error_item(
                file_id=file_id,
                upload=upload,
                stored_file_path=stored_file_path,
                message=str(exc),
                template_code_override=template_code_override,
                batch_type_override=batch_type_override,
                selected_bank_mapping_id=selected_bank_mapping_id,
                selected_bank_name=selected_bank_name,
                selected_bank_short_name=selected_bank_short_name,
                selected_bank_last4=selected_bank_last4,
            )
        except Exception:
            return self._build_preview_error_item(
                file_id=file_id,
                upload=upload,
                stored_file_path=stored_file_path,
                message="文件读取失败，请确认文件未损坏且为受支持的 Excel 模板。",
                template_code_override=template_code_override,
                batch_type_override=batch_type_override,
                selected_bank_mapping_id=selected_bank_mapping_id,
                selected_bank_name=selected_bank_name,
                selected_bank_short_name=selected_bank_short_name,
                selected_bank_last4=selected_bank_last4,
            )

        detected_bank_name, detected_last4 = self._detect_bank_selection(parsed)
        if parsed.source_control and parsed.source_control.status == "mismatch":
            return self._build_preview_error_item(
                file_id=file_id,
                upload=upload,
                stored_file_path=stored_file_path,
                message="文件控制合计与解析明细不一致，已阻止导入。",
                status="source_control_mismatch",
                content_sha256=content_sha256,
                source_control=parsed.source_control,
                template_code_override=template_code_override,
                batch_type_override=batch_type_override,
                selected_bank_mapping_id=selected_bank_mapping_id,
                selected_bank_name=selected_bank_name,
                selected_bank_short_name=selected_bank_short_name,
                selected_bank_last4=selected_bank_last4,
                template_code=parsed.template_code,
                batch_type=parsed.batch_type,
                detected_bank_name=detected_bank_name,
            )
        conflict_message = self._build_bank_selection_conflict_message(
            selected_bank_name=selected_bank_name,
            selected_bank_short_name=selected_bank_short_name,
            selected_bank_last4=selected_bank_last4,
            detected_bank_name=detected_bank_name,
            detected_last4=detected_last4,
        )
        bank_selection_conflict = bool(conflict_message)
        if parsed.batch_type == BatchType.BANK_TRANSACTION:
            for row in parsed.rows:
                if not clean(row.get("account_no")) and (selected_bank_mapping_id or selected_bank_last4):
                    row["account_no"] = selected_bank_mapping_id or selected_bank_last4
                row["selected_bank_mapping_id"] = selected_bank_mapping_id
                row["selected_bank_name"] = selected_bank_name
                row["selected_bank_short_name"] = selected_bank_short_name
                row["selected_bank_last4"] = selected_bank_last4
                row["detected_bank_name"] = detected_bank_name
                row["detected_last4"] = detected_last4

        try:
            preview = self._import_service.preview_import(
                batch_type=parsed.batch_type,
                source_name=upload.file_name,
                imported_by=imported_by,
                rows=parsed.rows,
            )
        except Exception:
            return self._build_preview_error_item(
                file_id=file_id,
                upload=upload,
                stored_file_path=stored_file_path,
                message="文件预览失败，请检查字段格式后重试。",
                template_code_override=template_code_override,
                batch_type_override=batch_type_override,
                selected_bank_mapping_id=selected_bank_mapping_id,
                selected_bank_name=selected_bank_name,
                selected_bank_short_name=selected_bank_short_name,
                selected_bank_last4=selected_bank_last4,
            )
        return FileImportPreviewItem(
            id=file_id,
            file_name=upload.file_name,
            template_code=parsed.template_code,
            batch_type=parsed.batch_type,
            status="preview_ready",
            message="模板识别成功。",
            row_count=len(parsed.rows),
            success_count=preview.success_count,
            error_count=preview.error_count,
            duplicate_count=preview.duplicate_count,
            suspected_duplicate_count=preview.suspected_duplicate_count,
            updated_count=preview.updated_count,
            preview_batch_id=preview.id,
            stored_file_path=stored_file_path,
            override_template_code=template_code_override,
            override_batch_type=BatchType(batch_type_override) if batch_type_override else None,
            selected_bank_mapping_id=selected_bank_mapping_id,
            selected_bank_name=selected_bank_name,
            selected_bank_short_name=selected_bank_short_name,
            selected_bank_last4=selected_bank_last4,
            detected_bank_name=detected_bank_name,
            detected_last4=detected_last4,
            bank_selection_conflict=bank_selection_conflict,
            conflict_message=conflict_message,
            header_signature=parsed.header_signature,
            mapping_candidates=parsed.mapping_candidates,
            mapping_fields=parsed.mapping_fields,
            field_mapping=parsed.field_mapping,
            mapping_source=parsed.mapping_source,
            row_results=preview.row_results,
            normalized_rows=preview.normalized_rows,
            content_sha256=content_sha256,
            source_control=parsed.source_control,
        )

    def _refresh_session_audit(self, session: FileImportSession) -> None:
        audit = self._build_session_audit(session, refresh_existing=False)
        session.audit = audit.audit
        session.duplicate_groups = audit.duplicate_groups
        file_audits = {file_audit.file_id: file_audit.audit for file_audit in audit.files}
        for item in session.files:
            item.audit = file_audits.get(item.id, ImportPreviewAuditCounts())

    def _build_session_audit(self, session: FileImportSession, *, refresh_existing: bool) -> ImportPreviewSessionAudit:
        rows: list[ImportPreviewAuditRow] = []
        for item in session.files:
            if item.batch_type is None:
                continue
            for row_result, normalized in zip(item.row_results, item.normalized_rows, strict=True):
                decision = row_result.decision
                linked_object_type = row_result.linked_object_type
                linked_object_id = row_result.linked_object_id
                if refresh_existing and row_result.decision != ImportDecision.ERROR:
                    decision, linked_object_type, linked_object_id = self._import_service.current_import_decision_for_normalized_row(
                        batch_type=item.batch_type,
                        normalized=normalized,
                    )
                identity = self._identity_for_row(row_result.source_record_type, normalized)
                rows.append(
                    ImportPreviewAuditRow(
                        file_id=item.id,
                        file_name=item.file_name,
                        row_no=row_result.row_no,
                        record_type=row_result.source_record_type,
                        identity_key=identity.identity_key,
                        identity_kind=identity.identity_kind,
                        decision=decision,
                        decision_reason=row_result.decision_reason,
                        linked_object_type=linked_object_type,
                        linked_object_id=linked_object_id,
                        **self._audit_row_display_fields(row_result.source_record_type, normalized),
                    )
                )
        audit = build_import_preview_session_audit(rows)
        existing_file_ids = {file_audit.file_id for file_audit in audit.files}
        for item in session.files:
            if item.id not in existing_file_ids:
                audit.files.append(self._empty_file_audit(item))
        return audit

    @staticmethod
    def _empty_file_audit(item: FileImportPreviewItem) -> ImportPreviewFileAudit:
        return ImportPreviewFileAudit(
            file_id=item.id,
            file_name=item.file_name,
            audit=ImportPreviewAuditCounts(error_count=item.error_count),
        )

    @staticmethod
    def _identity_for_row(record_type: str, normalized: dict[str, Any]) -> ImportRecordIdentity:
        if record_type == "bank_transaction":
            return BankTransactionIdentityStrategy().identify(normalized)
        return InvoiceIdentityStrategy().identify(normalized)

    @staticmethod
    def _audit_row_display_fields(record_type: str, normalized: dict[str, Any]) -> dict[str, str | None]:
        def text(value: Any) -> str | None:
            return None if value in (None, "") else str(value)

        if record_type == "bank_transaction":
            return {
                "account_no": text(normalized.get("account_no")),
                "trade_time": text(
                    normalized.get("trade_time") or normalized.get("pay_receive_time") or normalized.get("txn_date")
                ),
                "direction": text(normalized.get("txn_direction") or normalized.get("direction")),
                "amount": text(normalized.get("amount")),
                "counterparty_name": text(
                    normalized.get("counterparty_name_raw") or normalized.get("counterparty_name")
                ),
            }
        invoice_code = text(normalized.get("invoice_code"))
        invoice_no = text(normalized.get("digital_invoice_no") or normalized.get("invoice_no"))
        if invoice_code and invoice_no and not normalized.get("digital_invoice_no"):
            invoice_no = f"{invoice_code}-{invoice_no}"
        return {
            "invoice_no": invoice_no,
            "invoice_date": text(normalized.get("invoice_date")),
            "seller_name": text(normalized.get("seller_name")),
            "buyer_name": text(normalized.get("buyer_name")),
            "amount": text(normalized.get("amount")),
            "tax_amount": text(normalized.get("tax_amount")),
            "total_with_tax": text(normalized.get("total_with_tax")),
        }

    @staticmethod
    def _build_preview_error_item(
        *,
        file_id: str,
        upload: UploadedImportFile,
        stored_file_path: str | None,
        message: str,
        template_code_override: str | None,
        batch_type_override: str | None,
        selected_bank_mapping_id: str | None,
        selected_bank_name: str | None,
        selected_bank_short_name: str | None,
        selected_bank_last4: str | None,
        template_code: str | None = None,
        batch_type: BatchType | None = None,
        header_signature: str | None = None,
        mapping_candidates: list[dict[str, str]] | None = None,
        mapping_fields: list[dict[str, Any]] | None = None,
        field_mapping: dict[str, str] | None = None,
        detected_bank_name: str | None = None,
        status: str = "unrecognized_template",
        content_sha256: str | None = None,
        duplicate_file_name: str | None = None,
        source_control: SourceControlEvidence | None = None,
    ) -> FileImportPreviewItem:
        return FileImportPreviewItem(
            id=file_id,
            file_name=upload.file_name,
            template_code=template_code,
            batch_type=batch_type,
            status=status,
            message=message,
            row_count=0,
            stored_file_path=stored_file_path,
            override_template_code=template_code_override,
            override_batch_type=BatchType(batch_type_override) if batch_type_override else None,
            selected_bank_mapping_id=selected_bank_mapping_id,
            selected_bank_name=selected_bank_name,
            selected_bank_short_name=selected_bank_short_name,
            selected_bank_last4=selected_bank_last4,
            detected_bank_name=detected_bank_name,
            header_signature=header_signature,
            mapping_candidates=list(mapping_candidates or []),
            mapping_fields=list(mapping_fields or []),
            field_mapping=dict(field_mapping or {}),
            content_sha256=content_sha256 or hashlib.sha256(upload.content).hexdigest(),
            duplicate_file_name=duplicate_file_name,
            source_control=source_control,
        )

    def _parse_rows(
        self,
        *,
        rows: list[list[str]],
        template_code_override: str | None = None,
        batch_type_override: str | None = None,
        field_mapping: dict[str, str] | None = None,
    ) -> ParsedImportFile:
        template_code = template_code_override
        if not template_code:
            try:
                template_code = detect_invoice_template(rows)
            except ValueError:
                template_code = "bank_statement"
        if template_code == "invoice_export":
            parsed_rows = parse_invoice_rows(rows)
            resolved_batch_type = self._resolve_invoice_batch_type(parsed_rows, batch_type_override)
            for parsed_row in parsed_rows:
                parsed_row["counterparty_name"] = (
                    parsed_row.get("buyer_name") if resolved_batch_type == BatchType.OUTPUT_INVOICE else parsed_row.get("seller_name")
                )
            return ParsedImportFile(
                template_code=template_code,
                batch_type=resolved_batch_type,
                rows=parsed_rows,
                source_control=SourceControlEvidence(
                    status="not_applicable",
                    computed_row_count=len(parsed_rows),
                ),
            )
        if template_code == "bank_statement" or batch_type_override == BatchType.BANK_TRANSACTION.value:
            return parse_bank_statement_rows(rows, field_mapping=field_mapping)
        raise ValueError("无法识别文件模板。")

    def _saved_field_mapping(self, header_signature: str | None) -> dict[str, str]:
        if not header_signature:
            return {}
        for session in reversed(list(self._sessions.values())):
            for item in reversed(session.files):
                if item.header_signature == header_signature and item.field_mapping and item.mapping_source in {"manual", "saved"}:
                    return dict(item.field_mapping)
        return {}

    @staticmethod
    def _read_rows(upload: UploadedImportFile) -> list[list[str]]:
        suffix = upload.file_name.lower().rsplit(".", 1)[-1] if "." in upload.file_name else ""
        if suffix == "xlsx":
            return read_xlsx_rows(upload.content)
        if suffix == "xls":
            return read_xls_rows(upload.content)
        raise ValueError("无法识别文件模板。")

    def _next_session_id(self) -> str:
        self._session_counter += 1
        return f"import_session_{uuid4().hex}"

    def _next_file_id(self) -> str:
        self._file_counter += 1
        return f"import_file_{uuid4().hex}"

    def _resolve_invoice_batch_type(self, rows: list[dict[str, Any]], override: str | None) -> BatchType:
        if override:
            return BatchType(override)
        if not rows:
            return BatchType.INPUT_INVOICE
        input_votes = 0
        output_votes = 0
        for row in rows:
            if is_company_identity(row.get("buyer_tax_no"), row.get("buyer_name")) and not is_company_identity(
                row.get("seller_tax_no"),
                row.get("seller_name"),
            ):
                input_votes += 1
            elif is_company_identity(row.get("seller_tax_no"), row.get("seller_name")) and not is_company_identity(
                row.get("buyer_tax_no"),
                row.get("buyer_name"),
            ):
                output_votes += 1
        return BatchType.OUTPUT_INVOICE if output_votes > input_votes else BatchType.INPUT_INVOICE

    def _store_upload_file(
        self,
        session_id: str,
        file_id: str,
        upload: UploadedImportFile,
        *,
        imported_by: str,
    ) -> str | None:
        if self._file_store is None:
            return None
        return self._file_store.store_import_file(
            session_id=session_id,
            file_id=file_id,
            file_name=upload.file_name,
            content=upload.content,
            imported_by=imported_by,
        )

    def _find_confirmed_duplicate_file(self, *, content_sha256: str, exclude_file_id: str) -> str | None:
        for session in self._sessions.values():
            for item in session.files:
                if item.id != exclude_file_id and item.status == "confirmed" and item.content_sha256 == content_sha256:
                    return item.file_name
        finder = getattr(self._file_store, "find_confirmed_import_file_by_sha256", None)
        if not callable(finder):
            return None
        match = finder(content_sha256=content_sha256, exclude_file_id=exclude_file_id)
        if isinstance(match, dict):
            return clean(match.get("file_name")) or None
        return None

    @staticmethod
    def _detect_bank_selection(parsed: ParsedImportFile) -> tuple[str | None, str | None]:
        if parsed.batch_type != BatchType.BANK_TRANSACTION:
            return None, None
        detected_bank_name = parsed.detected_bank_name
        detected_last4 = None
        for row in parsed.rows:
            account_no = clean(row.get("account_no"))
            if len(account_no) >= 4:
                detected_last4 = account_no[-4:]
                break
        return detected_bank_name, detected_last4

    @staticmethod
    def _build_bank_selection_conflict_message(
        *,
        selected_bank_name: str | None,
        selected_bank_short_name: str | None,
        selected_bank_last4: str | None,
        detected_bank_name: str | None,
        detected_last4: str | None,
    ) -> str | None:
        mismatches: list[str] = []
        selected_bank_aliases = {
            normalized
            for raw_name in (selected_bank_name, selected_bank_short_name)
            if (normalized := FileImportService._normalize_bank_name_for_conflict(raw_name or ""))
        }
        detected_bank_alias = FileImportService._normalize_bank_name_for_conflict(detected_bank_name or "")
        if (
            selected_bank_aliases
            and detected_bank_alias
            and not any(
                FileImportService._bank_name_alias_matches(selected_alias, detected_bank_alias)
                for selected_alias in selected_bank_aliases
            )
        ):
            mismatches.append(f"银行选择为{selected_bank_name}，系统识别为{detected_bank_name}")
        if selected_bank_last4 and detected_last4 and selected_bank_last4 != detected_last4:
            mismatches.append(f"后四位选择为{selected_bank_last4}，系统识别为{detected_last4}")
        if not mismatches:
            return None
        return "；".join(mismatches)

    @staticmethod
    def _normalize_bank_name_for_conflict(bank_name: str) -> str:
        normalized = re.sub(r"\s+", "", str(bank_name or "").strip())
        return normalized.removesuffix("银行")

    @staticmethod
    def _bank_name_alias_matches(selected_alias: str, detected_alias: str) -> bool:
        return selected_alias == detected_alias or selected_alias in detected_alias or detected_alias in selected_alias


def detect_invoice_template(rows: list[list[str]]) -> str:
    find_invoice_header_index(rows)
    return "invoice_export"


def read_xlsx_rows(content: bytes) -> list[list[str]]:
    if not content.startswith(b"PK"):
        raise ValueError("文件扩展名为 .xlsx，但文件内容不是有效的 Excel 工作簿。")
    _validate_xlsx_archive(content)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        if not workbook.sheetnames or len(workbook.sheetnames) > MAX_WORKBOOK_SHEETS:
            raise ValueError(f"Excel 工作表数量必须在 1 到 {MAX_WORKBOOK_SHEETS} 之间。")
        first_rows: list[list[str]] | None = None
        for sheet in workbook.worksheets:
            # Some bank and invoice exporters write an invalid/underreported
            # worksheet dimension (for example ``A1``) even though the sheet
            # contains hundreds of rows.  Read-only openpyxl trusts that value
            # unless the dimensions are reset, which otherwise makes a valid
            # file look like it only contains one cell.
            sheet.reset_dimensions()
            rows = _worksheet_rows(sheet)
            if first_rows is None:
                first_rows = rows
            try:
                detect_invoice_template(rows)
                return rows
            except ValueError:
                if find_bank_header_candidate(rows) is not None:
                    return rows
        return first_rows or []
    finally:
        workbook.close()


def _worksheet_rows(sheet: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    cell_count = 0
    for row_index, excel_row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_index > MAX_WORKBOOK_ROWS:
            raise ValueError(f"Excel 明细超过 {MAX_WORKBOOK_ROWS} 行限制。")
        if len(excel_row) > MAX_WORKBOOK_COLUMNS:
            raise ValueError(f"Excel 明细超过 {MAX_WORKBOOK_COLUMNS} 列限制。")
        cell_count += len(excel_row)
        if cell_count > MAX_WORKBOOK_CELLS:
            raise ValueError(f"Excel 明细超过 {MAX_WORKBOOK_CELLS} 个单元格限制。")
        rows.append([stringify_cell(value) for value in excel_row])
    return rows


def read_xls_rows(content: bytes) -> list[list[str]]:
    if not content.startswith(XLS_SIGNATURE):
        raise ValueError("文件扩展名为 .xls，但文件内容不是有效的 Excel 工作簿。")
    workbook = xlrd.open_workbook(file_contents=content)
    if workbook.nsheets < 1 or workbook.nsheets > MAX_WORKBOOK_SHEETS:
        raise ValueError(f"Excel 工作表数量必须在 1 到 {MAX_WORKBOOK_SHEETS} 之间。")
    first_rows: list[list[str]] = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        if sheet.nrows > MAX_WORKBOOK_ROWS:
            raise ValueError(f"Excel 明细超过 {MAX_WORKBOOK_ROWS} 行限制。")
        if sheet.ncols > MAX_WORKBOOK_COLUMNS:
            raise ValueError(f"Excel 明细超过 {MAX_WORKBOOK_COLUMNS} 列限制。")
        if sheet.nrows * sheet.ncols > MAX_WORKBOOK_CELLS:
            raise ValueError(f"Excel 明细超过 {MAX_WORKBOOK_CELLS} 个单元格限制。")
        rows = [
            [stringify_cell(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
        if sheet_index == 0:
            first_rows = rows
        try:
            detect_invoice_template(rows)
            return rows
        except ValueError:
            if find_bank_header_candidate(rows) is not None:
                return rows
    return first_rows


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise ValueError("Excel 压缩包包含过多内部文件。")
            names = {member.filename for member in members}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ValueError("Excel 工作簿结构不完整。")
            total_uncompressed = 0
            for member in members:
                if member.flag_bits & 0x1:
                    raise ValueError("不支持加密的 Excel 文件。")
                if ".." in member.filename.split("/"):
                    raise ValueError("Excel 工作簿包含非法内部路径。")
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("Excel 解压后内容过大。")
                if member.file_size and member.file_size / max(member.compress_size, 1) > MAX_XLSX_COMPRESSION_RATIO:
                    raise ValueError("Excel 压缩比异常，已停止解析。")
    except BadZipFile as exc:
        raise ValueError("Excel 工作簿压缩结构已损坏。") from exc


def parse_invoice_rows(rows: list[list[str]]) -> list[dict[str, Any]]:
    header_index = find_invoice_header_index(rows)
    header = [canonical_invoice_header(cell) for cell in rows[header_index]]
    data_rows = []
    for row in rows[header_index + 1 :]:
        mapped = row_to_dict(header, row)
        if not any(mapped.values()):
            continue
        if is_invoice_summary_footer(mapped):
            continue
        data_rows.append(
            {
                "invoice_code": mapped.get("发票代码"),
                "invoice_no": mapped.get("发票号码"),
                "digital_invoice_no": mapped.get("数电发票号码"),
                "seller_tax_no": mapped.get("销方识别号"),
                "seller_name": mapped.get("销方名称"),
                "buyer_tax_no": mapped.get("购方识别号"),
                "buyer_name": mapped.get("购买方名称"),
                "counterparty_name": mapped.get("销方名称"),
                "invoice_date": to_date_string(mapped.get("开票日期")),
                "tax_classification_code": mapped.get("税收分类编码"),
                "specific_business_type": mapped.get("特定业务类型"),
                "taxable_item_name": mapped.get("货物或应税劳务名称"),
                "specification_model": mapped.get("规格型号"),
                "unit": mapped.get("单位"),
                "quantity": mapped.get("数量"),
                "unit_price": mapped.get("单价"),
                "amount": mapped.get("金额"),
                "tax_rate": mapped.get("税率"),
                "tax_amount": mapped.get("税额"),
                "total_with_tax": mapped.get("价税合计"),
                "invoice_source": mapped.get("发票来源"),
                "invoice_kind": mapped.get("发票票种"),
                "invoice_status_from_source": mapped.get("发票状态"),
                "is_positive_invoice": mapped.get("是否正数发票"),
                "risk_level": mapped.get("发票风险等级"),
                "issuer": mapped.get("开票人"),
                "remark": mapped.get("备注"),
            }
        )
    return aggregate_invoice_line_rows(data_rows)


def aggregate_invoice_line_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = {}
    order: list[tuple[str, ...]] = []
    for index, row in enumerate(rows):
        digital_no = clean(row.get("digital_invoice_no"))
        invoice_code = clean(row.get("invoice_code"))
        invoice_no = clean(row.get("invoice_no"))
        identity = (
            ("digital", digital_no)
            if digital_no
            else ("code_no", invoice_code, invoice_no)
            if invoice_code and invoice_no
            else ("row", str(index))
        )
        if identity not in grouped:
            grouped[identity] = []
            order.append(identity)
        grouped[identity].append(row)

    aggregated: list[dict[str, Any]] = []
    for identity in order:
        line_rows = grouped[identity]
        if len(line_rows) == 1:
            aggregated.append(line_rows[0])
            continue
        line_signatures = [_invoice_line_signature(row) for row in line_rows]
        if len(set(line_signatures)) == 1:
            aggregated.extend(line_rows)
            continue
        if len(set(line_signatures)) != len(line_signatures):
            raise ValueError("同一发票同时包含重复行和不同明细行，无法安全判断合计金额。")
        for field_name in (
            "digital_invoice_no",
            "invoice_code",
            "invoice_no",
            "seller_tax_no",
            "seller_name",
            "buyer_tax_no",
            "buyer_name",
            "invoice_date",
            "invoice_status_from_source",
        ):
            values = {clean(row.get(field_name)) for row in line_rows if clean(row.get(field_name))}
            if len(values) > 1:
                raise ValueError(f"同一发票的 {field_name} 不一致，无法安全合并明细行。")
        try:
            amount = sum((_invoice_line_decimal(row.get("amount")) for row in line_rows), Decimal("0"))
            tax_amount = sum((_invoice_line_decimal(row.get("tax_amount")) for row in line_rows), Decimal("0"))
            total_with_tax = sum(
                (_invoice_line_decimal(row.get("total_with_tax")) for row in line_rows),
                Decimal("0"),
            )
        except InvalidOperation:
            aggregated.extend(line_rows)
            continue
        merged = dict(line_rows[0])
        merged.update(
            {
                "amount": _invoice_line_decimal_text(amount),
                "tax_amount": _invoice_line_decimal_text(tax_amount),
                "total_with_tax": _invoice_line_decimal_text(total_with_tax),
                "source_line_count": len(line_rows),
                "source_line_items": [dict(row) for row in line_rows],
            }
        )
        tax_rates = {clean(row.get("tax_rate")) for row in line_rows if clean(row.get("tax_rate"))}
        merged["tax_rate"] = next(iter(tax_rates)) if len(tax_rates) == 1 else "mixed"
        aggregated.append(merged)
    return aggregated


def _invoice_line_decimal(value: Any) -> Decimal:
    text = clean(value).replace(",", "")
    if not text:
        return Decimal("0")
    return Decimal(text)


def _invoice_line_decimal_text(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.01")), "f")


def _invoice_line_signature(row: dict[str, Any]) -> tuple[str, ...]:
    return tuple(
        clean(row.get(field_name))
        for field_name in (
            "tax_classification_code",
            "taxable_item_name",
            "specification_model",
            "unit",
            "quantity",
            "unit_price",
            "amount",
            "tax_rate",
            "tax_amount",
            "total_with_tax",
        )
    )


def is_invoice_summary_footer(mapped: dict[str, str]) -> bool:
    invoice_kind = clean(mapped.get("发票票种"))
    return invoice_kind.startswith("份数：") and not any(
        clean(mapped.get(key)) for key in ("数电发票号码", "发票代码", "发票号码", "开票日期", "金额", "税额")
    )


@dataclass(slots=True)
class BankHeaderCandidate:
    index: int
    headers: list[str]
    candidates: list[dict[str, str]]
    automatic_mapping: dict[str, str]
    signature: str


def normalize_bank_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value).strip() if value is not None else "").lower()
    text = text.replace("帐号", "账号").replace("帐户", "账户")
    text = re.sub(r"人民币|rmb|cny|/元|\(元\)|元", "", text)
    return re.sub(r"[\s/\\()\[\]{}<>:：_\-—]+", "", text)


_BANK_ALIAS_LOOKUP = {
    normalize_bank_header(alias): field_name
    for field_name, aliases in BANK_FIELD_ALIASES.items()
    for alias in aliases
}


def find_bank_header_candidate(rows: list[list[str]]) -> BankHeaderCandidate | None:
    best: tuple[int, BankHeaderCandidate] | None = None
    for index, row in enumerate(rows[:60]):
        headers = [clean(cell) for cell in row]
        populated = [(column, header) for column, header in enumerate(headers) if header]
        if len(populated) < 3:
            continue
        resolved: dict[str, list[int]] = {}
        keyword_count = 0
        for column, header in populated:
            normalized = normalize_bank_header(header)
            field_name = _BANK_ALIAS_LOOKUP.get(normalized)
            if field_name:
                resolved.setdefault(field_name, []).append(column)
            if re.search(r"时间|日期|金额|余额|对方|账号|户名|摘要|备注|流水|凭证|收支|借方|贷方", header):
                keyword_count += 1
        automatic_mapping = {
            field_name: str(columns[0])
            for field_name, columns in resolved.items()
            if len(columns) == 1
        }
        next_row = rows[index + 1] if index + 1 < len(rows) else []
        data_score = sum(
            1
            for value in next_row
            if re.search(r"\d{4}[-/]?\d{2}[-/]?\d{2}|^-?[\d,]+(?:\.\d+)?$", clean(value))
        )
        score = len(automatic_mapping) * 10 + keyword_count + min(data_score, 5)
        if score < 5:
            continue
        candidates = [
            {"key": str(column), "label": f"第{column + 1}列 · {header}"}
            for column, header in populated
        ]
        signature_source = "|".join(normalize_bank_header(header) for header in headers)
        candidate = BankHeaderCandidate(
            index=index,
            headers=headers,
            candidates=candidates,
            automatic_mapping=automatic_mapping,
            signature=hashlib.sha256(signature_source.encode("utf-8")).hexdigest()[:24],
        )
        if best is None or score > best[0]:
            best = (score, candidate)
    return best[1] if best else None


def parse_bank_statement_rows(
    rows: list[list[str]],
    *,
    field_mapping: dict[str, str] | None = None,
) -> ParsedImportFile:
    candidate = find_bank_header_candidate(rows)
    if candidate is None:
        raise ValueError("无法识别银行流水表头，请确认文件包含交易明细。")
    mapping = dict(candidate.automatic_mapping)
    for field_name, raw_column in dict(field_mapping or {}).items():
        if field_name not in BANK_FIELD_LABELS or not str(raw_column).isdigit():
            raise ValueError("字段映射无效，请重新选择源列。")
        column = int(raw_column)
        if column < 0 or column >= len(candidate.headers) or not candidate.headers[column]:
            raise ValueError("字段映射中的源列不存在，请重新选择。")
        mapping[field_name] = str(column)

    mapping_fields = _bank_mapping_fields(mapping)
    missing = _missing_bank_core_fields(mapping)
    detected_bank_name = detect_bank_name(rows, header_index=candidate.index)
    if missing:
        raise BankStatementMappingRequired(
            f"需要补充字段映射：{'、'.join(missing)}。",
            header_signature=candidate.signature,
            candidates=candidate.candidates,
            mapping_fields=mapping_fields,
            field_mapping=mapping,
            detected_bank_name=detected_bank_name,
        )

    header = candidate.headers
    metadata = extract_key_value_metadata(rows[: candidate.index])
    metadata_account_no = extract_account_no_from_metadata(rows[: candidate.index])
    metadata_account_name = first_metadata_value(metadata, "账户名称", "帐户名称", "户名")
    metadata_currency = first_metadata_value(metadata, "币种", "货币种类") or "CNY"
    parsed_rows: list[dict[str, Any]] = []
    for row in rows[candidate.index + 1 :]:
        if not any(clean(cell) for cell in row):
            continue

        def cell(field_name: str) -> str | None:
            raw_index = mapping.get(field_name)
            if raw_index is None:
                return None
            column = int(raw_index)
            return clean(row[column] if column < len(row) else "") or None

        raw_trade_time = cell("trade_time")
        raw_txn_date = cell("txn_date") or cell("booked_date")
        raw_txn_clock = cell("txn_clock")
        if raw_trade_time and to_date_string(raw_trade_time) is None and raw_txn_date:
            raw_trade_time = f"{raw_txn_date} {raw_trade_time}"
        elif not raw_trade_time and raw_txn_date:
            raw_trade_time = f"{raw_txn_date} {raw_txn_clock or '00:00:00'}"
        txn_date = to_date_string(raw_txn_date or raw_trade_time)
        if not txn_date:
            continue
        trade_time = normalize_datetime_string(raw_trade_time)
        debit_amount, credit_amount = normalize_signed_debit_credit_columns(
            cell("debit_amount"),
            cell("credit_amount"),
        )
        if not debit_amount and not credit_amount:
            debit_amount, credit_amount = split_amount_by_direction(cell("amount"), cell("direction"))
        account_detail_no = cell("account_detail_no")
        voucher_no = cell("voucher_no")
        remark = cell("remark")
        parsed_rows.append(
            {
                "account_no": cell("account_no") or metadata_account_no,
                "account_name": cell("account_name") or metadata_account_name,
                "trade_time": trade_time,
                "pay_receive_time": trade_time,
                "txn_date": txn_date,
                "booked_date": to_date_string(cell("booked_date")),
                "counterparty_name": cell("counterparty_name") or "未知对手方",
                "counterparty_account_no": cell("counterparty_account_no"),
                "counterparty_bank_name": cell("counterparty_bank_name"),
                "credit_amount": credit_amount,
                "debit_amount": debit_amount,
                "balance": cell("balance"),
                "summary": cell("summary"),
                "remark": remark,
                "bank_text_fields": extract_bank_text_fields_from_row(header, row),
                "bank_serial_no": cell("bank_serial_no"),
                "account_detail_no": account_detail_no,
                "enterprise_serial_no": cell("enterprise_serial_no"),
                "voucher_kind": cell("voucher_kind"),
                "voucher_no": voucher_no,
                "currency": cell("currency") or metadata_currency,
            }
        )
    if not parsed_rows:
        raise ValueError("未找到可导入的银行交易明细。")
    return ParsedImportFile(
        template_code="bank_statement",
        batch_type=BatchType.BANK_TRANSACTION,
        rows=parsed_rows,
        detected_bank_name=detected_bank_name,
        header_signature=candidate.signature,
        mapping_candidates=candidate.candidates,
        mapping_fields=mapping_fields,
        field_mapping=mapping,
        mapping_source="manual" if field_mapping else "auto",
        source_control=build_bank_source_control_evidence(rows, parsed_rows),
    )


def build_bank_source_control_evidence(
    source_rows: list[list[str]],
    parsed_rows: list[dict[str, Any]],
) -> SourceControlEvidence:
    metadata = extract_key_value_metadata(source_rows)
    debit_count_raw = _control_metadata_value(
        metadata,
        "借方交易笔数",
        "借方累计笔数",
        "借方笔数",
        "支出笔数",
    )
    credit_count_raw = _control_metadata_value(
        metadata,
        "贷方交易笔数",
        "贷方累计笔数",
        "贷方笔数",
        "收入笔数",
    )
    total_count_raw = _control_metadata_value(metadata, "交易总笔数", "交易笔数", "明细笔数", "合计笔数")
    debit_total_raw = _control_metadata_value(
        metadata,
        "借方交易金额",
        "借方累计发生额",
        "借方金额合计",
        "支出金额合计",
    )
    credit_total_raw = _control_metadata_value(
        metadata,
        "贷方交易金额",
        "贷方累计发生额",
        "贷方金额合计",
        "收入金额合计",
    )
    has_declared_controls = any(
        value is not None
        for value in (debit_count_raw, credit_count_raw, total_count_raw, debit_total_raw, credit_total_raw)
    )
    computed_debit = sum((_control_decimal(row.get("debit_amount")) or Decimal("0") for row in parsed_rows), Decimal("0"))
    computed_credit = sum((_control_decimal(row.get("credit_amount")) or Decimal("0") for row in parsed_rows), Decimal("0"))
    if not has_declared_controls:
        return SourceControlEvidence(
            status="unavailable",
            computed_row_count=len(parsed_rows),
            computed_debit_total=_control_decimal_text(computed_debit),
            computed_credit_total=_control_decimal_text(computed_credit),
        )

    mismatch_fields: list[str] = []
    declared_row_count: int | None = None
    if total_count_raw is not None:
        declared_row_count = _control_integer(total_count_raw)
        if declared_row_count is None or declared_row_count != len(parsed_rows):
            mismatch_fields.append("row_count")
    elif debit_count_raw is not None and credit_count_raw is not None:
        debit_count = _control_integer(debit_count_raw)
        credit_count = _control_integer(credit_count_raw)
        if debit_count is not None and credit_count is not None:
            declared_row_count = debit_count + credit_count
        if declared_row_count is None or declared_row_count != len(parsed_rows):
            mismatch_fields.append("row_count")

    declared_debit = _control_decimal(debit_total_raw)
    declared_credit = _control_decimal(credit_total_raw)
    if debit_total_raw is not None and (declared_debit is None or declared_debit != computed_debit):
        mismatch_fields.append("debit_total")
    if credit_total_raw is not None and (declared_credit is None or declared_credit != computed_credit):
        mismatch_fields.append("credit_total")
    return SourceControlEvidence(
        status="mismatch" if mismatch_fields else "verified",
        computed_row_count=len(parsed_rows),
        declared_row_count=declared_row_count,
        computed_debit_total=_control_decimal_text(computed_debit),
        declared_debit_total=_control_decimal_text(declared_debit) if declared_debit is not None else None,
        computed_credit_total=_control_decimal_text(computed_credit),
        declared_credit_total=_control_decimal_text(declared_credit) if declared_credit is not None else None,
        mismatch_fields=tuple(mismatch_fields),
    )


def _control_metadata_value(metadata: dict[str, str], *aliases: str) -> str | None:
    normalized = {normalize_bank_header(key): value for key, value in metadata.items()}
    for alias in aliases:
        value = normalized.get(normalize_bank_header(alias))
        if clean(value):
            return clean(value)
    return None


def _control_decimal(value: Any) -> Decimal | None:
    text = clean(value).replace(",", "").replace("￥", "").replace("¥", "")
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _control_integer(value: Any) -> int | None:
    amount = _control_decimal(value)
    return int(amount) if amount is not None and amount == amount.to_integral_value() else None


def _control_decimal_text(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.01")), "f")


def _bank_mapping_fields(mapping: dict[str, str]) -> list[dict[str, Any]]:
    return [
        {
            "key": field_name,
            "label": label,
            "selected": mapping.get(field_name),
            "required": field_name in {"trade_time", "txn_date", "debit_amount", "credit_amount", "amount", "direction"},
        }
        for field_name, label in BANK_FIELD_LABELS.items()
    ]


def _missing_bank_core_fields(mapping: dict[str, str]) -> list[str]:
    missing: list[str] = []
    if not ({"trade_time", "txn_date"} & mapping.keys()):
        missing.append("交易日期或时间")
    has_split_amount = {"debit_amount", "credit_amount"}.issubset(mapping)
    has_directed_amount = {"amount", "direction"}.issubset(mapping)
    if not has_split_amount and not has_directed_amount:
        missing.append("借方和贷方金额，或金额和收支方向")
    return missing


def detect_bank_name(rows: list[list[str]], *, header_index: int | None = None) -> str | None:
    limit = min((header_index + 1) if header_index is not None else 20, 20)
    sample = " ".join(clean(cell).lower() for row in rows[:limit] for cell in row if clean(cell))
    for bank_name, markers in BANK_NAME_MARKERS.items():
        if any(marker.lower() in sample for marker in markers):
            return bank_name
    normalized = normalize_bank_header(sample)
    structural_markers = (
        ("工商银行", ("historydetail",)),
        ("平安银行", ("核心唯一流水号",)),
        ("民生银行", ("对方账号名称", "交易流水号")),
        ("建设银行", ("账户明细编号交易流水号",)),
        ("交通银行", ("查询账号", "借方发生额支出", "贷方发生额收入")),
    )
    for bank_name, markers in structural_markers:
        if all(normalize_bank_header(marker) in normalized for marker in markers):
            return bank_name
    return None


def split_amount_by_direction(amount: str | None, direction: str | None) -> tuple[str | None, str | None]:
    amount_text = clean(amount)
    direction_text = clean(direction).lower()
    if not amount_text:
        return None, None
    if any(marker in direction_text for marker in ("支", "出", "借", "付", "debit")):
        return amount_text.lstrip("-"), None
    if any(marker in direction_text for marker in ("收", "入", "贷", "credit")):
        return None, amount_text.lstrip("-")
    return None, None


def first_metadata_value(metadata: dict[str, str], *keys: str) -> str | None:
    normalized = {normalize_bank_header(key): value for key, value in metadata.items()}
    for key in keys:
        value = clean(normalized.get(normalize_bank_header(key)))
        if value:
            return value
    return None


def extract_bank_text_fields_from_row(header: list[str], row: list[str]) -> list[dict[str, str]]:
    fields: list[dict[str, str]] = []
    normalized_header = [normalize_bank_header(label) for label in header]
    for label in BANK_TEXT_FIELD_LABELS:
        normalized_label = normalize_bank_header(label)
        for index, header_label in enumerate(normalized_header):
            if header_label != normalized_label:
                continue
            value = clean(row[index] if index < len(row) else "")
            if value:
                fields.append({"label": clean(header[index]), "value": value})
            break
    return fields


def normalize_row(row: list[str]) -> list[str]:
    return [normalize_header(cell) for cell in row if clean(cell)]


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", clean(value)))


def canonical_invoice_header(value: str) -> str:
    normalized = normalize_header(value)
    return INVOICE_HEADER_ALIASES.get(normalized, normalized)


def clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_invoice_header_index(rows: list[list[str]]) -> int:
    for index, row in enumerate(rows):
        normalized_row = {canonical_invoice_header(cell) for cell in normalize_row(row)}
        if INVOICE_REQUIRED_HEADERS.issubset(normalized_row):
            return index
    raise ValueError("无法识别文件模板。")


def row_to_dict(header: list[str], row: list[str]) -> dict[str, str]:
    payload: dict[str, str] = {}
    width = max(len(header), len(row))
    for index in range(width):
        key = header[index] if index < len(header) else ""
        if not clean(key):
            continue
        payload[clean(key)] = clean(row[index] if index < len(row) else "")
    return payload


def normalize_signed_debit_credit_columns(debit_amount: str | None, credit_amount: str | None) -> tuple[str | None, str | None]:
    debit_text = clean(debit_amount)
    credit_text = clean(credit_amount)
    if debit_text.startswith("-") and not credit_text:
        return None, debit_text[1:].strip()
    if credit_text.startswith("-") and not debit_text:
        return credit_text[1:].strip(), None
    return debit_amount, credit_amount


def extract_key_value_metadata(rows: list[list[str]]) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row in rows:
        for cell in row:
            text = clean(cell)
            if "：" in text:
                key, value = text.split("：", 1)
            elif ":" in text:
                key, value = text.split(":", 1)
            else:
                continue
            normalized_key = clean(key)
            normalized_value = clean(value)
            if normalized_key and normalized_value:
                metadata[normalized_key] = normalized_value
        for index in range(0, len(row) - 1, 2):
            key = clean(row[index]).rstrip(":：")
            value = clean(row[index + 1])
            if key and value:
                metadata[key] = value
    return metadata


def extract_account_no_from_metadata(rows: list[list[str]]) -> str | None:
    metadata = extract_key_value_metadata(rows)
    for key, value in metadata.items():
        if is_account_no_key(key):
            account_no = normalize_account_no(value)
            if account_no:
                return account_no
    return None


def is_account_no_key(key: str) -> bool:
    normalized_key = normalize_header(key)
    if not any(keyword in normalized_key for keyword in ACCOUNT_METADATA_KEYWORDS):
        return False
    return not any(excluded in normalized_key for excluded in ("对方", "户名", "名称", "开户", "余额"))


def normalize_account_no(value: Any) -> str | None:
    digits = re.sub(r"\D+", "", clean(value))
    return digits if len(digits) >= 4 else None


def to_date_string(value: str | None) -> str | None:
    text = clean(value)
    if not text:
        return None
    for pattern in (DATE_ONLY_RE, DATE_TIME_RE, COMPACT_DATE_RE, COMPACT_DATE_TIME_RE):
        match = pattern.match(text)
        if not match:
            continue
        groups = match.groups()
        return f"{groups[0]}-{groups[1]}-{groups[2]}"
    try:
        return datetime.fromisoformat(text.replace("/", "-")).date().isoformat()
    except ValueError:
        return None


def normalize_datetime_string(value: str | None) -> str | None:
    text = clean(value)
    if not text:
        return None
    match = DATE_TIME_RE.match(text)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)} {match.group(4)}:{match.group(5)}:{match.group(6)}"
    match = COMPACT_DATE_TIME_RE.match(text)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)} {match.group(4)}:{match.group(5)}:{match.group(6)}"
    if COMPACT_DATE_RE.match(text):
        return f"{text[:4]}-{text[4:6]}-{text[6:8]} 00:00:00"
    try:
        parsed = datetime.fromisoformat(text.replace("/", "-"))
    except ValueError:
        return text
    return parsed.strftime("%Y-%m-%d %H:%M:%S")


def is_company_identity(tax_no: str | None, company_name: str | None) -> bool:
    normalized_tax_no = clean(tax_no).upper()
    normalized_name = clean(company_name)
    if normalized_tax_no and normalized_tax_no in COMPANY_TAX_NOS:
        return True
    return any(keyword in normalized_name for keyword in COMPANY_NAME_KEYWORDS)


def sanitize_file_name(file_name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", file_name).strip("._")
    return cleaned or "uploaded_file"
