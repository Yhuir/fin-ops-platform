from __future__ import annotations

from copy import deepcopy
from datetime import UTC, datetime
import hashlib
import json
from pathlib import Path
import re
from threading import RLock
from typing import Any, Callable

from fin_ops_platform.services.bank_turnover_tag_semantics import (
    EXTERNAL_TURNOVER_CATEGORY_CODE,
    EXTERNAL_TURNOVER_ROLE,
    EXTERNAL_TURNOVER_THIRD_LABEL_OPTIONS,
    TURNOVER_ACTION_TYPE_OPTIONS,
    infer_turnover_action_type,
    is_external_turnover_definition,
    is_external_turnover_primary_label,
    label_path as turnover_label_path,
    normalize_external_third_label,
    normalize_turnover_action_type,
    turnover_family_for_third_label,
)


BANK_TRANSACTION_CATEGORY_SCHEMA_VERSION = "2026-05-bank-transaction-category-taxonomy"
BANK_TRANSACTION_CATEGORY_TAXONOMY: list[dict[str, Any]] = [
    {
        "root": "借入",
        "groups": [
            {
                "name": "个人往来款",
                "display_name": "个人暂借款",
                "items": [
                    ("borrow_in_personal_pending_repayment", "待还款"),
                    ("borrow_in_personal_repaid", "已还款"),
                ],
            },
            {
                "name": "公司往来款",
                "display_name": "公司暂借款",
                "items": [
                    ("borrow_in_company_pending_repayment", "待还款"),
                    ("borrow_in_company_repaid", "已还款"),
                ],
            },
            {
                "name": "银行往来款",
                "display_name": "银行往来款",
                "items": [
                    ("borrow_in_bank_pending_repayment", "待还款"),
                    ("borrow_in_bank_repaid", "已还款"),
                ],
            },
        ],
    },
    {
        "root": "借出",
        "groups": [
            {
                "name": "个人往来款",
                "display_name": "个人往来款",
                "items": [
                    ("borrow_out_personal_lent", "待收款"),
                    ("borrow_out_personal_pending_collection", "已收款"),
                ],
            },
            {
                "name": "公司往来款",
                "display_name": "公司往来款",
                "items": [
                    ("borrow_out_company_lent", "待收款"),
                    ("borrow_out_company_pending_collection", "已收款"),
                ],
            },
            {
                "name": "银行往来款",
                "display_name": "银行往来款",
                "items": [
                    ("borrow_out_bank_lent", "待收款"),
                    ("borrow_out_bank_pending_collection", "已收款"),
                ],
            },
            {
                "name": "货款往来款",
                "display_name": "货款往来款",
                "items": [
                    ("borrow_out_goods_lent", "待收款"),
                    ("borrow_out_goods_pending_collection", "已收款"),
                ],
            },
        ],
    },
    {
        "root": "业务往来",
        "groups": [
            {
                "name": "质保金",
                "display_name": "质保金",
                "items": [("business_warranty_pending_collection", "待收款")],
            },
            {
                "name": "投标保证金",
                "display_name": "投标保证金",
                "items": [("business_bid_bond_pending_collection", "待收款")],
            },
            {
                "name": "履约保证金",
                "display_name": "履约保证金",
                "items": [("business_performance_bond_pending_collection", "待收款")],
            },
            {
                "name": "已开发票未收款",
                "display_name": "已开发票未收款",
                "items": [("business_invoiced_pending_collection", "待收款")],
            },
        ],
    },
]


def _build_category_definitions() -> dict[str, dict[str, Any]]:
    definitions: dict[str, dict[str, Any]] = {}
    for root_node in BANK_TRANSACTION_CATEGORY_TAXONOMY:
        root = str(root_node["root"])
        for group in list(root_node["groups"]):
            group_name = str(group["name"])
            display_name = str(group["display_name"])
            for code, status in list(group["items"]):
                definitions[str(code)] = {
                    "category_code": str(code),
                    "category_label": f"{display_name}：{status}",
                    "category_path": [root, group_name, str(status)],
                    "category_root": root,
                    "category_group": group_name,
                    "category_status": str(status),
                }
    return definitions


BANK_TRANSACTION_CATEGORY_DEFINITIONS = _build_category_definitions()
BANK_TRANSACTION_AUTO_CATEGORY_DEFINITIONS: dict[str, dict[str, Any]] = {
    "fee": {
        "category_code": "fee",
        "category_label": "手续费",
        "category_path": ["自动识别", "手续费"],
    },
    "salary": {
        "category_code": "salary",
        "category_label": "工资",
        "category_path": ["自动识别", "工资"],
    },
    "holiday_bonus": {
        "category_code": "holiday_bonus",
        "category_label": "过节费",
        "category_path": ["自动识别", "过节费"],
    },
    "bonus": {
        "category_code": "bonus",
        "category_label": "奖金",
        "category_path": ["自动识别", "奖金"],
    },
    "tax_payment": {
        "category_code": "tax_payment",
        "category_label": "税款",
        "category_path": ["自动识别", "税款"],
    },
    "treasury_tax_collection": {
        "category_code": "treasury_tax_collection",
        "category_label": "代理国库税收收缴",
        "category_path": ["自动识别", "代理国库税收收缴"],
    },
    "social_security": {
        "category_code": "social_security",
        "category_label": "社保款",
        "category_path": ["自动识别", "社保款"],
    },
}
BANK_TRANSACTION_LEGACY_CATEGORY_DEFINITIONS: dict[str, dict[str, Any]] = {
    "external_turnover": {
        "category_code": "external_turnover",
        "category_label": "外部往来款",
        "category_path": [],
    },
    "internal_transfer": {
        "category_code": "internal_transfer",
        "category_label": "内部往来款",
        "category_path": ["自动识别", "内部往来款"],
    },
    "offset": {
        "category_code": "offset",
        "category_label": "冲",
        "category_path": [],
    },
    "cash_turnover": {
        "category_code": "cash_turnover",
        "category_label": "现金往来",
        "category_path": [],
    },
}
BANK_TRANSACTION_LEGACY_CATEGORY_LABELS: dict[str, str] = {
    code: str(definition["category_label"])
    for code, definition in BANK_TRANSACTION_LEGACY_CATEGORY_DEFINITIONS.items()
}
BANK_TRANSACTION_CATEGORY_LABELS: dict[str, str] = {
    **{
        code: str(definition["category_label"])
        for code, definition in BANK_TRANSACTION_CATEGORY_DEFINITIONS.items()
    },
    **{
        code: str(definition["category_label"])
        for code, definition in BANK_TRANSACTION_AUTO_CATEGORY_DEFINITIONS.items()
    },
    **BANK_TRANSACTION_LEGACY_CATEGORY_LABELS,
}
BANK_TRANSACTION_CATEGORY_COUNT_KEYS = [
    *BANK_TRANSACTION_CATEGORY_DEFINITIONS.keys(),
    *BANK_TRANSACTION_AUTO_CATEGORY_DEFINITIONS.keys(),
    *BANK_TRANSACTION_LEGACY_CATEGORY_LABELS.keys(),
    "uncategorized",
]
BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION = 1
BANK_TRANSACTION_TAG_CODE_RE = re.compile(r"^[A-Za-z0-9_:-]+$")
BANK_AUTO_TAG_INTERNAL_TRANSFER_CODE = "internal_transfer"
BANK_AUTO_TAG_EDITABLE_CODES = (
    "fee",
    "holiday_bonus",
    "salary",
    "bonus",
    "treasury_tax_collection",
    "social_security",
    "tax_payment",
    "external_turnover",
)
BANK_AUTO_TAG_FIELD_OPTIONS: tuple[dict[str, str], ...] = (
    {"value": "counterparty_name", "label": "对方户名"},
    {"value": "counterparty_account", "label": "对方账号"},
    {"value": "counterparty_bank", "label": "对方开户行"},
    {"value": "purpose_text", "label": "用途/交易用途"},
    {"value": "summary_text", "label": "摘要"},
    {"value": "note_text", "label": "备注/附言/客户附言"},
    {"value": "detail_text", "label": "其他明细"},
    {"value": "all_text", "label": "全部文本"},
)
BANK_AUTO_TAG_FIELD_LABELS = {
    str(option["value"]): str(option["label"])
    for option in BANK_AUTO_TAG_FIELD_OPTIONS
}
BANK_AUTO_TAG_ALLOWED_FIELDS = set(BANK_AUTO_TAG_FIELD_LABELS)
BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS = ("summary_text", "purpose_text", "note_text", "detail_text")
BANK_AUTO_TAG_ALLOWED_DIRECTIONS = {"income", "expense", "any"}
BANK_AUTO_TAG_ALLOWED_ACCOUNT_SCOPE_TYPES = {"any", "bank_account", "account_type", "bank"}
DEFAULT_BANK_AUTO_TAG_DIRECTION = "any"
DEFAULT_BANK_AUTO_TAG_ACCOUNT_SCOPE = {"type": "any", "values": []}
LEGACY_EXTERNAL_TURNOVER_LABEL_REPAIR_SPECS: dict[str, dict[str, Any]] = {
    "收回借出款": {
        "primary_label": "外部往来款收款",
        "direction": "income",
        "contains_any": ["收回借出款", "收回借款", "收回暂借款", "还借款", "还暂借款", "归还借款"],
    },
    "退保证金": {
        "primary_label": "外部往来款收款",
        "direction": "income",
        "contains_any": ["退保证金", "退回保证金", "返还保证金"],
    },
    "退款": {
        "primary_label": "外部往来款收款",
        "direction": "income",
        "contains_any": ["退款", "退回", "返还"],
    },
    "借入款": {
        "primary_label": "外部往来款收款",
        "direction": "income",
        "contains_any": ["借入款", "暂借款", "借款"],
    },
    "归还借款": {
        "primary_label": "外部往来款付款",
        "direction": "expense",
        "contains_any": ["归还借款", "还借款", "还暂借款", "还款"],
    },
    "押金": {
        "primary_label": "外部往来款付款",
        "direction": "expense",
        "contains_any": ["押金"],
    },
    "退押金": {
        "primary_label": "外部往来款收款",
        "direction": "income",
        "contains_any": ["退押金", "退回押金", "返还押金"],
    },
}
BANK_AUTO_TAG_FILE_FIELD_MAPPING_VERSION = "2026-05-29-bank-auto-tag-field-mapping-v1"
BANK_AUTO_TAG_FILE_SCHEMA_VERSION = "2026-05-29-bank-auto-tag-rules-normalized-v1"
BANK_AUTO_TAG_FILE_TEXT_FIELD_LABEL = "用途/交易用途、摘要、备注/附言/客户附言"
BANK_AUTO_TAG_FILE_FIELD_MAPPINGS: dict[str, list[str]] = {
    BANK_AUTO_TAG_FILE_TEXT_FIELD_LABEL: ["purpose_text", "summary_text", "note_text", "detail_text"],
    "对方户": ["counterparty_name"],
}
BANK_AUTO_TAG_FILE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "flow_type": ("流水类型", "flow_type"),
    "primary_label": ("分类（一级）", "主标签", "primary_label"),
    "sub_label": ("银行流水标签（贰级）", "银行流水标签（二级）", "子标签", "sub_label"),
    "query_fields": ("选择查询的项", "选择查询的项（可全选/清空）", "query_fields"),
    "contains": ("包含", "contains"),
    "contains_all": ("必须同时包含", "contains_all"),
    "exact": ("精准命重", "精准命中", "exact"),
    "none_of": ("不包含字样", "none_of"),
    "priority": ("优先级", "priority"),
    "source_row": ("source_row", "源行号"),
}
_BANK_AUTO_TAG_FILE_TERM_SPLIT_RE = re.compile(r"[\r\n、，,；;]+")
DEFAULT_BANK_AUTO_TAG_RULES: dict[str, dict[str, Any]] = {
    "fee": {
        "priority": 2,
        "rule_code": "fee_text_keyword",
        "output_primary_label": "费用",
        "output_sub_label": "手续费",
        "rules": {
            "match_fields": ["counterparty_name", "summary_text", "note_text"],
            "exact": [],
            "contains": ["手续费", "短信服务费"],
            "excludes": [],
        },
    },
    "holiday_bonus": {
        "priority": 2,
        "rule_code": "holiday_bonus_text_keyword",
        "rules": {
            "match_fields": list(BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS),
            "exact": [],
            "contains": ["过节费"],
            "excludes": [],
        },
    },
    "salary": {
        "priority": 2,
        "rule_code": "salary_text_keyword",
        "rules": {
            "match_fields": list(BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS),
            "exact": [],
            "contains": ["工资"],
            "excludes": [],
        },
    },
    "bonus": {
        "priority": 2,
        "rule_code": "bonus_text_keyword",
        "rules": {
            "match_fields": list(BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS),
            "exact": [],
            "contains": ["奖金", "绩效奖", "年终奖"],
            "excludes": [],
        },
    },
    "treasury_tax_collection": {
        "priority": 2,
        "rule_code": "treasury_tax_collection_text_keyword",
        "rules": {
            "match_fields": list(BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS),
            "exact": [],
            "contains": ["代理国库税收收缴", "国库税收收缴"],
            "excludes": [],
        },
    },
    "social_security": {
        "priority": 2,
        "rule_code": "social_security_text_keyword",
        "rules": {
            "match_fields": list(BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS),
            "exact": [],
            "contains": ["社保款", "社保费", "社会保险费", "缴纳社保"],
            "excludes": [],
        },
    },
    "tax_payment": {
        "priority": 2,
        "rule_code": "tax_payment_text_keyword",
        "direction": "expense",
        "rules": {
            "match_fields": list(BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS),
            "exact": [],
            "contains": ["税款", "缴纳税款", "电子缴税", "税库银", "税务局", "完税"],
            "excludes": ["社保及税款", "社保和税款", "社保税款", "社保、税款"],
        },
    },
    "external_turnover": {
        "priority": 2,
        "rule_code": "external_turnover_candidate_text_keyword",
        "direction": "any",
        "account_scope": DEFAULT_BANK_AUTO_TAG_ACCOUNT_SCOPE,
        "stop_on_match": True,
        "review_required": True,
        "route_to": "turnover_ledger_pending",
        "rules": {
            "match_fields": ["purpose_text", "summary_text", "note_text"],
            "exact_any": [],
            "contains_any": [
                "往来款",
                "暂借款",
                "借款",
                "还暂借款",
                "归还借款",
                "还借款",
                "保证金",
                "投标保证金",
                "履约保证金",
                "质量保证金",
                "押金",
                "退款",
                "退回",
                "返还",
                "代垫",
                "垫付",
                "代付",
                "代购",
                "贷款扣款",
                "批量还款",
                "借据号",
                "还息",
                "贷款本息",
                "收回贷款本息",
                "本息",
            ],
            "contains_all": [],
            "none_of": [
                "本公司帐户",
                "本公司账户",
                "货款",
                "预付货款",
                "工程款",
                "劳务费",
                "技术服务费",
                "工资",
                "社保",
                "公积金",
                "税款",
                "纳税",
                "电子退库",
                "结息",
                "活期利息",
                "存款利息",
            ],
            "regex_any": [],
        },
    },
}
BANK_AUTO_TAG_SYSTEM_RULE = {
    "code": BANK_AUTO_TAG_INTERNAL_TRANSFER_CODE,
    "label": "内部往来款",
    "priority_label": "优先级 1",
    "source": "system",
    "status": "active",
    "editable": False,
    "archivable": False,
    "sortable": False,
}


def build_system_bank_transaction_tag_definitions() -> list[dict[str, Any]]:
    definitions: list[dict[str, Any]] = []
    for source in (
        BANK_TRANSACTION_CATEGORY_DEFINITIONS,
        BANK_TRANSACTION_AUTO_CATEGORY_DEFINITIONS,
        BANK_TRANSACTION_LEGACY_CATEGORY_DEFINITIONS,
    ):
        for code, definition in source.items():
            definitions.append(
                {
                    "code": str(code),
                    "label": str(definition.get("category_label") or code),
                    "path": [
                        str(item)
                        for item in list(definition.get("category_path") or [])
                        if str(item).strip()
                    ],
                    "source": "system",
                    "status": "active",
                    **_default_auto_tag_rule_fields(str(code)),
                }
            )
    return definitions


def _default_auto_tag_rule_fields(code: str) -> dict[str, Any]:
    default = DEFAULT_BANK_AUTO_TAG_RULES.get(code)
    if not isinstance(default, dict):
        return {}
    fields = {
        "priority": int(default["priority"]),
        "rule_code": str(default["rule_code"]),
        "rules": deepcopy(default["rules"]),
    }
    for key in (
        "direction",
        "account_scope",
        "output_primary_label",
        "output_sub_label",
        "turnover_role",
        "turnover_action_type",
        "stop_on_match",
        "review_required",
        "route_to",
    ):
        if key in default:
            fields[key] = deepcopy(default[key])
    return fields


def default_bank_transaction_tag_dictionary_payload() -> dict[str, Any]:
    return {
        "version": BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION,
        "definitions": build_system_bank_transaction_tag_definitions(),
    }


def bank_transaction_tag_dictionary_display_payload(
    payload: dict[str, Any] | None,
) -> dict[str, Any]:
    """Return only the tag metadata consumed by list-page clients."""
    source = payload if isinstance(payload, dict) else default_bank_transaction_tag_dictionary_payload()
    public_fields = (
        "code",
        "path",
        "label",
        "source",
        "status",
        "output_primary_label",
        "output_sub_label",
        "output_third_label",
        "turnover_role",
        "turnover_action_type",
        "turnover_family",
    )
    return {
        "version": int(source.get("version") or BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION),
        "definitions": [
            {field: definition[field] for field in public_fields if field in definition}
            for definition in list(source.get("definitions") or [])
            if isinstance(definition, dict) and str(definition.get("code") or "").strip()
        ],
    }


class BankTransactionCategoryError(ValueError):
    def __init__(
        self,
        error_code: str,
        message: str,
        *,
        transaction_id: str | None = None,
        expected_version: int | None = None,
        actual_version: int | None = None,
    ) -> None:
        super().__init__(message)
        self.error_code = error_code
        self.transaction_id = transaction_id
        self.expected_version = expected_version
        self.actual_version = actual_version


class BankTransactionCategoryValidationError(BankTransactionCategoryError):
    pass


class BankTransactionCategoryConflictError(BankTransactionCategoryError):
    pass


class BankAutoTagRulesValidationError(ValueError):
    def __init__(
        self,
        error_code: str,
        message: str,
        *,
        field_errors: list[dict[str, str]] | None = None,
        references: list[dict[str, Any]] | None = None,
    ) -> None:
        super().__init__(message)
        self.error_code = error_code
        self.field_errors = list(field_errors or [])
        self.references = [dict(reference) for reference in list(references or []) if isinstance(reference, dict)]


class BankTransactionCategoryService:
    def __init__(
        self,
        *,
        categories: dict[str, dict[str, Any]] | None = None,
        audit_log: list[dict[str, Any]] | None = None,
        tag_dictionary: dict[str, Any] | None = None,
        transaction_exists: Callable[[str], bool] | None = None,
    ) -> None:
        self._lock = RLock()
        self._transaction_exists = transaction_exists
        self._tag_dictionary_payload = self._normalize_tag_dictionary_payload(tag_dictionary)
        self._tag_definitions_by_code = self._build_tag_definition_index(self._tag_dictionary_payload)
        self._categories = self._normalize_categories(categories or {})
        self._audit_log = [
            deepcopy(entry)
            for entry in list(audit_log or [])
            if isinstance(entry, dict)
        ]
        self._snapshot_version_cache: str | None = None

    @classmethod
    def from_snapshot(
        cls,
        snapshot: dict[str, Any] | None,
        *,
        transaction_exists: Callable[[str], bool] | None = None,
    ) -> "BankTransactionCategoryService":
        if not snapshot:
            return cls(transaction_exists=transaction_exists)
        categories = snapshot.get("categories")
        audit_log = snapshot.get("audit_log")
        tag_dictionary = snapshot.get("tag_dictionary")
        return cls(
            categories=categories if isinstance(categories, dict) else {},
            audit_log=audit_log if isinstance(audit_log, list) else [],
            tag_dictionary=tag_dictionary if isinstance(tag_dictionary, dict) else None,
            transaction_exists=transaction_exists,
        )

    def snapshot(self) -> dict[str, Any]:
        with self._lock:
            return {
                "schema_version": BANK_TRANSACTION_CATEGORY_SCHEMA_VERSION,
                "categories": deepcopy(self._categories),
                "audit_log": deepcopy(self._audit_log),
                "tag_dictionary": deepcopy(self._tag_dictionary_payload),
            }

    def restore_snapshot(self, snapshot: dict[str, Any]) -> None:
        """Restore the process-local category cache after a durable write rollback."""
        categories = snapshot.get("categories") if isinstance(snapshot, dict) else None
        audit_log = snapshot.get("audit_log") if isinstance(snapshot, dict) else None
        tag_dictionary = snapshot.get("tag_dictionary") if isinstance(snapshot, dict) else None
        with self._lock:
            self._categories = self._normalize_categories(categories if isinstance(categories, dict) else {})
            self._audit_log = [
                deepcopy(entry)
                for entry in list(audit_log or [])
                if isinstance(entry, dict)
            ]
            self._tag_dictionary_payload = self._normalize_tag_dictionary_payload(
                tag_dictionary if isinstance(tag_dictionary, dict) else None
            )
            self._tag_definitions_by_code = self._build_tag_definition_index(self._tag_dictionary_payload)
            self._snapshot_version_cache = None

    def snapshot_version(self) -> str:
        """Return the canonical snapshot hash without copying the full state on every read."""
        with self._lock:
            if self._snapshot_version_cache is None:
                encoded = json.dumps(
                    {
                        "schema_version": BANK_TRANSACTION_CATEGORY_SCHEMA_VERSION,
                        "categories": self._categories,
                        "audit_log": self._audit_log,
                        "tag_dictionary": self._tag_dictionary_payload,
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                    default=str,
                ).encode("utf-8")
                self._snapshot_version_cache = hashlib.sha256(encoded).hexdigest()
            return self._snapshot_version_cache

    def tag_dictionary_payload(self) -> dict[str, Any]:
        with self._lock:
            return deepcopy(self._tag_dictionary_payload)

    def tag_dictionary_version(self) -> int:
        with self._lock:
            return int(self._tag_dictionary_payload.get("version") or BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION)

    def auto_tag_rule_version_label(self) -> str:
        return f"bank-auto-tag-rules:{self.tag_dictionary_version()}"

    def configure_tag_dictionary(self, payload: dict[str, Any] | None) -> None:
        normalized_payload = self._normalize_tag_dictionary_payload(payload)
        with self._lock:
            self._tag_dictionary_payload = normalized_payload
            self._tag_definitions_by_code = self._build_tag_definition_index(normalized_payload)
            self._snapshot_version_cache = None

    @classmethod
    def auto_tag_rules_payload(
        cls,
        tag_dictionary: dict[str, Any] | None,
        *,
        can_save: bool = True,
        read_model_status: str | None = None,
    ) -> dict[str, Any]:
        normalized = cls._normalize_tag_dictionary_payload(tag_dictionary)
        definitions = [
            dict(definition)
            for definition in list(normalized.get("definitions") or [])
            if cls._is_auto_tag_rule_definition(definition)
        ]
        active = [
            definition
            for definition in definitions
            if str(definition.get("status") or "active") == "active"
        ]
        archived = [
            definition
            for definition in definitions
            if str(definition.get("status") or "active") == "archived"
        ]
        active.sort(key=cls._auto_tag_rule_sort_key)
        archived.sort(key=lambda definition: (str(definition.get("label") or ""), str(definition.get("code") or "")))
        payload: dict[str, Any] = {
            "version": int(normalized.get("version") or BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION),
            "system_rule": dict(BANK_AUTO_TAG_SYSTEM_RULE),
            "active_rules": [
                cls._public_auto_tag_rule(definition, priority_index=index, sort_index=index - 1)
                for index, definition in enumerate(active, start=1)
            ],
            "archived_rules": [
                cls._public_auto_tag_rule(definition, priority_index=None, sort_index=None)
                for definition in archived
            ],
            "field_options": [dict(option) for option in BANK_AUTO_TAG_FIELD_OPTIONS],
            "turnover_third_label_options": [dict(option) for option in EXTERNAL_TURNOVER_THIRD_LABEL_OPTIONS],
            "turnover_action_type_options": [dict(option) for option in TURNOVER_ACTION_TYPE_OPTIONS],
            "permissions": {"can_save": bool(can_save)},
        }
        if read_model_status:
            payload["read_model_status"] = str(read_model_status)
        return payload

    @classmethod
    def normalize_auto_tag_rules_update(
        cls,
        value: Any,
        *,
        previous_tag_dictionary: dict[str, Any],
        references_by_code: dict[str, list[dict[str, Any]]] | None = None,
    ) -> dict[str, Any]:
        if not isinstance(value, dict):
            raise BankAutoTagRulesValidationError(
                "invalid_bank_auto_tag_rules_request",
                "自动标签规则请求必须是对象。",
            )
        if "system_rule" in value:
            raise BankAutoTagRulesValidationError(
                "invalid_bank_auto_tag_rules_request",
                "内部往来款为系统规则，不能在请求中提交或修改。",
            )
        if "expected_version" not in value:
            raise BankAutoTagRulesValidationError(
                "invalid_bank_auto_tag_rules_request",
                "expected_version is required.",
            )
        previous = cls._normalize_tag_dictionary_payload(previous_tag_dictionary)
        expected_version = cls._normalize_version(value.get("expected_version"))
        previous_version = int(previous.get("version") or BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION)
        if expected_version != previous_version:
            raise BankAutoTagRulesValidationError(
                "bank_transaction_tags_version_conflict",
                "规则已被其他用户更新，请刷新后重新编辑。",
            )

        active_items = value.get("active_rules")
        archived_items = value.get("archived_rules")
        if not isinstance(active_items, list) or not isinstance(archived_items, list):
            raise BankAutoTagRulesValidationError(
                "invalid_bank_auto_tag_rules_request",
                "active_rules 和 archived_rules 必须是数组。",
            )

        previous_definitions_by_code = cls._build_tag_definition_index(previous)
        previous_managed_codes = {
            code
            for code, definition in previous_definitions_by_code.items()
            if cls._is_auto_tag_rule_definition(definition)
        }
        next_managed_definitions: list[dict[str, Any]] = []
        seen_codes: set[str] = set()
        field_errors: list[dict[str, str]] = []

        previous_sort_orders = [
            cls._normalize_optional_sort_order(definition.get("sort_order"))
            for definition in previous_definitions_by_code.values()
            if cls._is_auto_tag_rule_definition(definition)
        ]
        next_new_sort_order = (max([value for value in previous_sort_orders if value is not None], default=0) + 1)

        for index, item in enumerate(active_items):
            priority = cls._normalize_submitted_auto_tag_priority(
                item.get("priority") if isinstance(item, dict) else None,
                path=f"active_rules[{index}].priority",
                field_errors=field_errors,
            )
            if priority is None:
                priority = 2
            definition = cls._normalize_auto_tag_rule_item(
                item,
                path_prefix=f"active_rules[{index}]",
                previous_definitions_by_code=previous_definitions_by_code,
                previous_managed_codes=previous_managed_codes,
                status="active",
                priority=priority,
                fallback_sort_order=next_new_sort_order + index,
                field_errors=field_errors,
            )
            code = definition["code"]
            if code in seen_codes:
                field_errors.append({"path": f"active_rules[{index}].code", "message": "标签 code 不能重复。"})
            seen_codes.add(code)
            next_managed_definitions.append(definition)

        for index, item in enumerate(archived_items):
            definition = cls._normalize_auto_tag_rule_item(
                item,
                path_prefix=f"archived_rules[{index}]",
                previous_definitions_by_code=previous_definitions_by_code,
                previous_managed_codes=previous_managed_codes,
                status="archived",
                priority=None,
                fallback_sort_order=None,
                field_errors=field_errors,
            )
            code = definition["code"]
            if code in seen_codes:
                field_errors.append({"path": f"archived_rules[{index}].code", "message": "标签 code 不能重复。"})
            seen_codes.add(code)
            next_managed_definitions.append(definition)

        if BANK_AUTO_TAG_INTERNAL_TRANSFER_CODE in seen_codes:
            raise BankAutoTagRulesValidationError(
                "invalid_bank_auto_tag_rules_request",
                "内部往来款为系统规则，不能修改、排序或停用。",
            )

        missing_codes = sorted(previous_managed_codes - seen_codes)
        if missing_codes:
            field_errors.append({
                "path": "active_rules",
                "message": "必须提交完整自动标签规则列表。",
            })

        cls._validate_auto_tag_duplicate_labels(next_managed_definitions, field_errors=field_errors)
        if field_errors:
            raise BankAutoTagRulesValidationError(
                "invalid_auto_tag_rule",
                "自动标签规则校验失败。",
                field_errors=field_errors,
            )

        newly_archived_codes = [
            definition["code"]
            for definition in next_managed_definitions
            if definition["status"] == "archived"
            and str(previous_definitions_by_code.get(definition["code"], {}).get("status") or "active") != "archived"
        ]
        references: list[dict[str, Any]] = []
        for code in newly_archived_codes:
            references.extend(list((references_by_code or {}).get(code) or []))
        if references:
            raise BankAutoTagRulesValidationError(
                "bank_transaction_tag_in_use_by_pending_invoice_filter",
                "该银行明细标签仍被下游规则引用，请先解除引用后再停用。",
                references=references,
            )

        next_by_code = dict(previous_definitions_by_code)
        for code in list(previous_managed_codes):
            if code in next_by_code:
                del next_by_code[code]
        for definition in next_managed_definitions:
            next_by_code[definition["code"]] = definition
        normalized_next = cls._normalize_tag_dictionary_payload(
            {
                "version": previous_version,
                "definitions": list(next_by_code.values()),
            }
        )
        changes = cls._auto_tag_rule_changes(previous, normalized_next)
        if changes["changed"]:
            normalized_next["version"] = previous_version + 1
        else:
            normalized_next["version"] = previous_version
        return {
            "tag_dictionary": normalized_next,
            "changes": changes,
            "old_version": previous_version,
            "new_version": int(normalized_next["version"]),
        }

    @classmethod
    def parse_auto_tag_rule_file_source(
        cls,
        source: Any,
        *,
        source_name: str | None = None,
        source_version: str | None = None,
    ) -> dict[str, Any]:
        source_payload, raw_rows = cls._auto_tag_file_rows(source)
        resolved_source_name = str(
            source_name
            or source_payload.get("source_name")
            or source_payload.get("source_workbook_name")
            or "bank_auto_tag_rules"
        ).strip()
        resolved_source_version = str(
            source_version
            or source_payload.get("source_version")
            or source_payload.get("schema_version")
            or BANK_AUTO_TAG_FILE_SCHEMA_VERSION
        ).strip()
        source_hash = str(source_payload.get("source_hash") or "").strip()
        if not source_hash:
            source_hash = "sha256:" + hashlib.sha256(
                json.dumps(raw_rows, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
            ).hexdigest()

        field_errors: list[dict[str, str]] = []
        active_rules: list[dict[str, Any]] = []
        skipped_rows = [
            dict(item)
            for item in list(source_payload.get("skipped_rows") or [])
            if isinstance(item, dict)
        ]
        seen_label_paths: set[tuple[str, str]] = set()

        for index, row in enumerate(raw_rows):
            normalized_row = cls._normalize_auto_tag_file_row(row)
            source_row = str(normalized_row.get("source_row") or index + 1)
            try:
                error_row_index = max(int(source_row) - 1, index)
            except (TypeError, ValueError):
                error_row_index = index
            primary_label = str(normalized_row.get("primary_label") or "").strip()
            sub_label = str(normalized_row.get("sub_label") or "").strip()
            if primary_label == BANK_AUTO_TAG_SYSTEM_RULE["label"]:
                if not any(item.get("source_row") == normalized_row.get("source_row") for item in skipped_rows):
                    skipped_rows.append(
                        {
                            "source_row": normalized_row.get("source_row") or index + 1,
                            "primary_label": primary_label,
                            "reason": "internal_transfer_system_rule",
                        }
                    )
                continue

            contains_any = cls._split_auto_tag_file_terms(normalized_row.get("contains"))
            contains_all = cls._split_auto_tag_file_terms(normalized_row.get("contains_all"))
            exact_any = cls._split_auto_tag_file_terms(normalized_row.get("exact"))
            none_of = cls._split_auto_tag_file_terms(normalized_row.get("none_of"))
            has_positive_condition = bool(contains_any or contains_all or exact_any)
            if not has_positive_condition:
                if primary_label or sub_label:
                    skipped_rows.append(
                        {
                            "source_row": normalized_row.get("source_row") or index + 1,
                            "primary_label": primary_label,
                            "sub_label": sub_label,
                            "reason": "missing_positive_condition",
                        }
                    )
                continue

            if not primary_label:
                field_errors.append({"path": f"rows[{error_row_index}].主标签", "message": "主标签不能为空。"})
            query_fields_label = str(normalized_row.get("query_fields") or "").strip()
            match_fields = BANK_AUTO_TAG_FILE_FIELD_MAPPINGS.get(query_fields_label)
            if match_fields is None:
                field_errors.append({
                    "path": f"rows[{error_row_index}].选择查询的项",
                    "message": f"未知的查询字段：{query_fields_label}",
                })
                match_fields = []
            label_path = (primary_label, sub_label)
            if primary_label and label_path in seen_label_paths:
                field_errors.append({
                    "path": f"rows[{error_row_index}].主标签",
                    "message": "文件规则主标签和子标签组合不能重复。",
                })
            seen_label_paths.add(label_path)
            if not primary_label or not match_fields:
                continue

            label = sub_label or primary_label
            active_rules.append(
                {
                    "label": label,
                    "output_primary_label": primary_label,
                    "output_sub_label": sub_label,
                    "direction": cls._auto_tag_file_direction(normalized_row.get("flow_type")),
                    "account_scope": deepcopy(DEFAULT_BANK_AUTO_TAG_ACCOUNT_SCOPE),
                    "rules": {
                        "match_fields": list(match_fields),
                        "exact_any": exact_any,
                        "contains_any": contains_any,
                        "contains_all": contains_all,
                        "none_of": none_of,
                        "regex_any": [],
                        "exact": exact_any,
                        "contains": contains_any,
                        "excludes": none_of,
                    },
                    "source_row": normalized_row.get("source_row") or source_row,
                }
            )

        if field_errors:
            raise BankAutoTagRulesValidationError(
                "invalid_bank_auto_tag_rule_file",
                "银行流水标签文件规则校验失败。",
                field_errors=field_errors,
            )
        return {
            "source": {
                "source_name": resolved_source_name,
                "source_version": resolved_source_version,
                "source_hash": source_hash,
                "schema_version": str(source_payload.get("schema_version") or BANK_AUTO_TAG_FILE_SCHEMA_VERSION),
                "field_mapping_version": str(
                    source_payload.get("field_mapping_version") or BANK_AUTO_TAG_FILE_FIELD_MAPPING_VERSION
                ),
            },
            "active_rules": active_rules,
            "skipped_rows": skipped_rows,
        }

    @classmethod
    def compare_auto_tag_rule_file_sources(cls, left: Any, right: Any) -> dict[str, Any]:
        left_rules = cls.parse_auto_tag_rule_file_source(left)["active_rules"]
        right_rules = cls.parse_auto_tag_rule_file_source(right)["active_rules"]
        left_comparable = [cls._auto_tag_file_rule_comparable(rule) for rule in left_rules]
        right_comparable = [cls._auto_tag_file_rule_comparable(rule) for rule in right_rules]
        if left_comparable == right_comparable:
            return {"matched": True, "diffs": []}
        diffs: list[dict[str, str]] = []
        max_len = max(len(left_comparable), len(right_comparable))
        for index in range(max_len):
            left_rule = left_comparable[index] if index < len(left_comparable) else None
            right_rule = right_comparable[index] if index < len(right_comparable) else None
            if left_rule != right_rule:
                diffs.append(
                    {
                        "path": f"rules[{index}]",
                        "message": "文件规则内容不一致。",
                    }
                )
        raise BankAutoTagRulesValidationError(
            "bank_auto_tag_rule_file_diff",
            "银行流水标签文件规则内容不一致。",
            field_errors=diffs,
        )

    @classmethod
    def normalize_auto_tag_rules_file_replacement(
        cls,
        source: Any,
        *,
        previous_tag_dictionary: dict[str, Any],
    ) -> dict[str, Any]:
        parsed = cls.parse_auto_tag_rule_file_source(source)
        previous = cls._normalize_tag_dictionary_payload(previous_tag_dictionary)
        previous_version = int(previous.get("version") or BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION)
        previous_definitions_by_code = cls._build_tag_definition_index(previous)
        previous_managed_codes = {
            code
            for code, definition in previous_definitions_by_code.items()
            if cls._is_auto_tag_rule_definition(definition)
        }
        reusable_by_label_path: dict[tuple[str, str], dict[str, Any]] = {}
        reusable_candidates = sorted(
            previous_definitions_by_code.values(),
            key=lambda definition: (
                str(definition.get("status") or "active") != "active",
                not cls._is_auto_tag_rule_definition(definition),
                str(definition.get("source") or "") != "custom",
                str(definition.get("code") or ""),
            ),
        )
        for definition in reusable_candidates:
            if not cls._is_auto_tag_rule_definition(definition) and str(definition.get("source") or "") != "custom":
                continue
            key = (
                str(definition.get("output_primary_label") or definition.get("label") or "").strip(),
                str(definition.get("output_sub_label") or "").strip(),
            )
            if key[0] and key not in reusable_by_label_path:
                reusable_by_label_path[key] = dict(definition)
            can_reuse_custom_label_only = (
                str(definition.get("source") or "") == "custom"
                and not cls._is_auto_tag_rule_definition(definition)
            )
            can_reuse_editable_system_label_only = cls._can_reuse_editable_system_code_by_label(definition)
            can_reuse_label_only = can_reuse_custom_label_only or can_reuse_editable_system_label_only
            if can_reuse_label_only:
                label = str(definition.get("label") or "").strip()
                label_only_key = ("", label)
                if label and label_only_key not in reusable_by_label_path:
                    reusable_by_label_path[label_only_key] = dict(definition)

        active_definitions: list[dict[str, Any]] = []
        reused_codes: list[str] = []
        added_codes: list[str] = []
        recovered_codes: list[str] = []
        active_codes: set[str] = set()
        active_label_paths: set[tuple[str, str]] = set()
        occupied_codes = dict(previous_definitions_by_code)
        for index, rule in enumerate(parsed["active_rules"]):
            primary_label = str(rule["output_primary_label"])
            sub_label = str(rule.get("output_sub_label") or "")
            label = sub_label or primary_label
            reusable = reusable_by_label_path.get((primary_label, sub_label))
            if reusable is None and sub_label:
                reusable = reusable_by_label_path.get(("", sub_label))
            if reusable is not None:
                code = str(reusable["code"])
                source_type = str(reusable.get("source") or "custom")
                rule_code = str(reusable.get("rule_code") or code)
                reused_codes.append(code)
            else:
                code = cls._generate_custom_auto_tag_code(label, rule["rules"], occupied_codes)
                source_type = "custom"
                rule_code = code
                added_codes.append(code)
                occupied_codes[code] = {"code": code}
            active_codes.add(code)
            active_label_paths.add((primary_label, sub_label))
            active_definitions.append(
                {
                    "code": code,
                    "label": label,
                    "path": ["自动识别", label],
                    "source": source_type if source_type in {"system", "custom"} else "custom",
                    "status": "active",
                    "priority": 2,
                    "sort_order": index + 1,
                    "direction": rule["direction"],
                    "account_scope": deepcopy(DEFAULT_BANK_AUTO_TAG_ACCOUNT_SCOPE),
                    "output_primary_label": primary_label,
                    "output_sub_label": sub_label,
                    "rules": cls._normalize_auto_tag_rule_conditions(rule["rules"], allow_invalid=False),
                    "rule_code": rule_code,
                }
            )

        next_sort_order = len(active_definitions) + 1
        for definition in previous_definitions_by_code.values():
            code = str(definition.get("code") or "").strip()
            if not code or code in active_codes:
                continue
            recovered_definition = cls._legacy_external_turnover_repair_definition(
                definition,
                sort_order=next_sort_order,
            )
            if recovered_definition is None:
                continue
            label_path_key = (
                str(recovered_definition.get("output_primary_label") or "").strip(),
                str(recovered_definition.get("output_sub_label") or "").strip(),
            )
            if label_path_key in active_label_paths:
                continue
            active_codes.add(code)
            active_label_paths.add(label_path_key)
            active_definitions.append(recovered_definition)
            recovered_codes.append(code)
            next_sort_order += 1

        archived_definitions: list[dict[str, Any]] = []
        for code in sorted(previous_managed_codes - active_codes):
            previous_definition = dict(previous_definitions_by_code[code])
            previous_definition["status"] = "archived"
            previous_definition["priority"] = cls._normalize_optional_priority(previous_definition.get("priority")) or 2
            archived_definitions.append(previous_definition)

        next_by_code = dict(previous_definitions_by_code)
        for code in previous_managed_codes:
            next_by_code.pop(code, None)
        for definition in [*active_definitions, *archived_definitions]:
            next_by_code[str(definition["code"])] = definition
        normalized_next = cls._normalize_tag_dictionary_payload(
            {
                "version": previous_version,
                "definitions": list(next_by_code.values()),
            }
        )
        changes = cls._auto_tag_rule_changes(previous, normalized_next)
        changes["source"] = dict(parsed["source"])
        changes["reused_codes"] = sorted(set(reused_codes))
        changes["added_codes"] = sorted(set(added_codes))
        changes["recovered_legacy_external_turnover_codes"] = sorted(set(recovered_codes))
        changes["skipped_rows"] = [dict(item) for item in parsed["skipped_rows"]]
        changes["changed"] = bool(
            changes["changed"]
            or changes["reused_codes"]
            or changes["added_codes"]
            or changes["recovered_legacy_external_turnover_codes"]
            or changes["skipped_rows"]
        )
        normalized_next["version"] = previous_version + 1 if changes["changed"] else previous_version
        return {
            "tag_dictionary": normalized_next,
            "changes": changes,
            "old_version": previous_version,
            "new_version": int(normalized_next["version"]),
        }

    @classmethod
    def _can_reuse_editable_system_code_by_label(cls, definition: dict[str, Any]) -> bool:
        code = str(definition.get("code") or "").strip()
        if str(definition.get("source") or "") != "system":
            return False
        if code == EXTERNAL_TURNOVER_CATEGORY_CODE or code == "fee":
            return True
        if code not in BANK_AUTO_TAG_EDITABLE_CODES:
            return False
        current_label = str(definition.get("label") or "").strip()
        default_label = str(cls.label_for(code) or "").strip()
        return bool(current_label and default_label and current_label != default_label)

    @classmethod
    def _legacy_external_turnover_repair_definition(
        cls,
        definition: dict[str, Any],
        *,
        sort_order: int,
    ) -> dict[str, Any] | None:
        if cls._is_auto_tag_rule_definition(definition):
            return None
        if str(definition.get("source") or "") != "custom":
            return None
        if str(definition.get("status") or "active") != "active":
            return None
        code = str(definition.get("code") or "").strip()
        label = str(definition.get("label") or "").strip()
        spec = LEGACY_EXTERNAL_TURNOVER_LABEL_REPAIR_SPECS.get(label)
        if not code or spec is None:
            return None
        primary_label = str(spec["primary_label"])
        direction = str(spec["direction"])
        action_type = infer_turnover_action_type(
            primary_label=primary_label,
            sub_label=label,
            direction=direction,
        )
        return {
            "code": code,
            "label": label,
            "path": ["自动识别", label],
            "source": "custom",
            "status": "active",
            "priority": 2,
            "sort_order": sort_order,
            "direction": direction,
            "account_scope": deepcopy(DEFAULT_BANK_AUTO_TAG_ACCOUNT_SCOPE),
            "output_primary_label": primary_label,
            "output_sub_label": label,
            "turnover_role": EXTERNAL_TURNOVER_ROLE,
            "turnover_action_type": action_type,
            "rules": cls._normalize_auto_tag_rule_conditions(
                {
                    "match_fields": list(BANK_AUTO_TAG_DEFAULT_TEXT_FIELDS),
                    "contains_any": list(spec["contains_any"]),
                    "contains_all": [],
                    "exact_any": [],
                    "regex_any": [],
                    "none_of": [],
                },
                allow_invalid=False,
            ),
            "rule_code": str(definition.get("rule_code") or code).strip() or code,
        }

    def get(self, transaction_id: str) -> dict[str, Any]:
        transaction_key = self._normalize_transaction_id(transaction_id)
        with self._lock:
            return self._public_record(transaction_key, self._categories.get(transaction_key))

    def bulk_get(self, transaction_ids: list[str]) -> dict[str, dict[str, Any]]:
        normalized_ids = [
            self._normalize_transaction_id(transaction_id)
            for transaction_id in list(transaction_ids or [])
            if str(transaction_id or "").strip()
        ]
        with self._lock:
            return {
                transaction_id: self._public_record(transaction_id, self._categories.get(transaction_id))
                for transaction_id in normalized_ids
            }

    def apply_updates(self, updates: list[dict[str, Any]], *, actor: str) -> dict[str, Any]:
        return self._apply_updates(updates, actor=actor)

    def apply_turnover_updates(self, updates: list[dict[str, Any]], *, actor: str) -> dict[str, Any]:
        return self._apply_updates(
            updates,
            actor=actor,
            source="turnover_ledger",
            allowed_category_codes=set(BANK_TRANSACTION_CATEGORY_DEFINITIONS),
            invalid_category_error="invalid_turnover_category_code",
            invalid_category_message="Only turnover leaf category codes can be selected from turnover ledger.",
        )

    def assign_manual_category(
        self,
        *,
        transaction_id: str,
        category_code: str,
        actor: str,
        category_primary_label: str | None = None,
        category_sub_label: str | None = None,
        category_third_label: str | None = None,
        category_label_path: list[str] | None = None,
        turnover_action_type: str | None = None,
        turnover_family: str | None = None,
    ) -> dict[str, Any]:
        return self._apply_updates(
            [
                {
                    "transaction_id": transaction_id,
                    "category_code": category_code,
                    "manual_assignment": True,
                    "category_primary_label": category_primary_label,
                    "category_sub_label": category_sub_label,
                    "category_third_label": category_third_label,
                    "category_label_path": category_label_path,
                    "turnover_action_type": turnover_action_type,
                    "turnover_family": turnover_family,
                }
            ],
            actor=actor,
            source="manual",
        )

    def confirm_auto_category(
        self,
        *,
        transaction_id: str,
        category_code: str,
        candidate_category_codes: list[str],
        rule_version: str,
        actor: str,
        category_primary_label: str | None = None,
        category_sub_label: str | None = None,
        category_third_label: str | None = None,
        category_label_path: list[str] | None = None,
        turnover_action_type: str | None = None,
        turnover_family: str | None = None,
    ) -> dict[str, Any]:
        normalized_transaction_id = self._normalize_transaction_id(transaction_id)
        normalized_category_code = str(category_code or "").strip()
        normalized_candidates = [
            str(code).strip()
            for code in list(candidate_category_codes or [])
            if str(code or "").strip()
        ]
        if not normalized_transaction_id:
            raise BankTransactionCategoryValidationError("unknown_transaction_id", "transaction_id is required.")
        if self._transaction_exists is not None and not self._transaction_exists(normalized_transaction_id):
            raise BankTransactionCategoryValidationError(
                "unknown_transaction_id",
                f"Unknown bank transaction id: {normalized_transaction_id}",
                transaction_id=normalized_transaction_id,
            )
        if normalized_category_code not in normalized_candidates:
            raise BankTransactionCategoryValidationError(
                "invalid_category_confirmation_candidate",
                "只能选择当前自动规则命中的候选标签。",
                transaction_id=normalized_transaction_id,
            )
        definition = self._tag_definitions_by_code.get(normalized_category_code)
        if not isinstance(definition, dict) or str(definition.get("status") or "active") == "archived":
            raise BankTransactionCategoryValidationError(
                "invalid_category_code",
                f"Invalid bank transaction category code: {normalized_category_code}",
                transaction_id=normalized_transaction_id,
            )
        return self._apply_updates(
            [
                {
                    "transaction_id": normalized_transaction_id,
                    "category_code": normalized_category_code,
                    "candidate_category_codes": normalized_candidates,
                    "rule_version": str(rule_version or "").strip(),
                    "category_primary_label": category_primary_label,
                    "category_sub_label": category_sub_label,
                    "category_third_label": category_third_label,
                    "category_label_path": category_label_path,
                    "turnover_action_type": turnover_action_type,
                    "turnover_family": turnover_family,
                }
            ],
            actor=actor,
            source="auto_confirmation",
        )

    def clear_manual_category(self, *, transaction_id: str, actor: str) -> dict[str, Any]:
        normalized_transaction_id = self._normalize_transaction_id(transaction_id)
        if not normalized_transaction_id:
            raise BankTransactionCategoryValidationError("unknown_transaction_id", "transaction_id is required.")
        if self._transaction_exists is not None and not self._transaction_exists(normalized_transaction_id):
            raise BankTransactionCategoryValidationError(
                "unknown_transaction_id",
                f"Unknown bank transaction id: {normalized_transaction_id}",
                transaction_id=normalized_transaction_id,
            )
        current = self.get(normalized_transaction_id)
        if str(current.get("source") or "") != "manual" or not bool(current.get("manual_assignment")):
            raise BankTransactionCategoryValidationError(
                "invalid_manual_category_clear_target",
                "只能撤销从待分类状态人工添加的标签。",
                transaction_id=normalized_transaction_id,
            )
        return self._apply_updates(
            [{"transaction_id": normalized_transaction_id, "category_code": None, "manual_assignment": True}],
            actor=actor,
            source="manual",
        )

    def revoke_auto_category_confirmation(self, *, transaction_id: str, actor: str) -> dict[str, Any]:
        normalized_transaction_id = self._normalize_transaction_id(transaction_id)
        if not normalized_transaction_id:
            raise BankTransactionCategoryValidationError("unknown_transaction_id", "transaction_id is required.")
        if self._transaction_exists is not None and not self._transaction_exists(normalized_transaction_id):
            raise BankTransactionCategoryValidationError(
                "unknown_transaction_id",
                f"Unknown bank transaction id: {normalized_transaction_id}",
                transaction_id=normalized_transaction_id,
            )
        current = self.get(normalized_transaction_id)
        if str(current.get("source") or "") not in {"auto_confirmation", "auto_confirmation_revoked"}:
            raise BankTransactionCategoryValidationError(
                "invalid_category_confirmation_revoke_target",
                "当前流水没有可撤销的标签确认。",
                transaction_id=normalized_transaction_id,
            )
        return self._apply_updates(
            [{"transaction_id": normalized_transaction_id, "category_code": None}],
            actor=actor,
            source="auto_confirmation_revoked",
        )

    def _apply_updates(
        self,
        updates: list[dict[str, Any]],
        *,
        actor: str,
        source: str = "manual",
        allowed_category_codes: set[str] | None = None,
        invalid_category_error: str = "invalid_category_code",
        invalid_category_message: str = "Invalid bank transaction category code.",
    ) -> dict[str, Any]:
        normalized_actor = str(actor or "").strip()
        if not normalized_actor:
            raise BankTransactionCategoryValidationError(
                "permission_denied",
                "actor is required to update bank transaction categories.",
            )
        if not isinstance(updates, list) or not updates:
            raise BankTransactionCategoryValidationError(
                "invalid_category_update",
                "updates must be a non-empty array.",
            )

        timestamp = datetime.now(UTC).isoformat()
        with self._lock:
            normalized_updates = [self._normalize_update(update) for update in updates]
            if allowed_category_codes is not None:
                for update in normalized_updates:
                    category_code = update.get("category_code")
                    if category_code not in allowed_category_codes:
                        raise BankTransactionCategoryValidationError(
                            invalid_category_error,
                            invalid_category_message,
                            transaction_id=update["transaction_id"],
                        )
            transaction_ids = [update["transaction_id"] for update in normalized_updates]
            if len(set(transaction_ids)) != len(transaction_ids):
                raise BankTransactionCategoryValidationError(
                    "invalid_category_update",
                    "duplicate transaction_id in updates.",
                )

            for update in normalized_updates:
                transaction_id = update["transaction_id"]
                if self._transaction_exists is not None and not self._transaction_exists(transaction_id):
                    raise BankTransactionCategoryValidationError(
                        "unknown_transaction_id",
                        f"Unknown bank transaction id: {transaction_id}",
                        transaction_id=transaction_id,
                    )
                actual_version = self._current_version(transaction_id)
                expected_version = update.get("expected_version")
                if expected_version is not None and expected_version != actual_version:
                    raise BankTransactionCategoryConflictError(
                        "category_version_conflict",
                        "Bank transaction category version conflict.",
                        transaction_id=transaction_id,
                        expected_version=expected_version,
                        actual_version=actual_version,
                    )

            updated_categories: list[dict[str, Any]] = []
            audit_entries: list[dict[str, Any]] = []
            for update in normalized_updates:
                transaction_id = update["transaction_id"]
                category_code = update["category_code"]
                existing = self._categories.get(transaction_id)
                previous_code = existing.get("category_code") if isinstance(existing, dict) else None
                previous_source = str(existing.get("source") or "") if isinstance(existing, dict) else ""
                if (
                    existing is not None
                    and previous_code == category_code
                    and previous_source == source
                    and self._stored_category_semantics(existing) == self._submitted_category_semantics(update)
                ):
                    updated_categories.append(self._public_record(transaction_id, existing))
                    continue

                next_version = self._current_version(transaction_id) + 1
                label_fields = self._label_fields_for_update(category_code, update)
                record = {
                    "transaction_id": transaction_id,
                    "category_code": category_code,
                    "category_label": label_fields.get("category_label") or self._label_for(category_code),
                    "category_path": self._path_for(category_code),
                    "category_primary_label": label_fields.get("category_primary_label"),
                    "category_sub_label": label_fields.get("category_sub_label"),
                    "category_third_label": label_fields.get("category_third_label"),
                    "category_label_path": list(label_fields.get("category_label_path") or []),
                    "turnover_role": label_fields.get("turnover_role") or "",
                    "turnover_action_type": label_fields.get("turnover_action_type"),
                    "turnover_family": label_fields.get("turnover_family"),
                    "source": source,
                    "updated_by": normalized_actor,
                    "updated_at": timestamp,
                    "version": next_version,
                }
                if update.get("manual_assignment"):
                    record["manual_assignment"] = True
                if update.get("candidate_category_codes"):
                    record["candidate_category_codes"] = list(update.get("candidate_category_codes") or [])
                if update.get("rule_version"):
                    record["rule_version"] = str(update.get("rule_version") or "")
                self._categories[transaction_id] = record
                updated_categories.append(self._public_record(transaction_id, record))
                audit_entries.append(
                    {
                        "transaction_id": transaction_id,
                        "previous_category_code": previous_code,
                        "category_code": category_code,
                        "category_label_path": list(label_fields.get("category_label_path") or []),
                        "category_third_label": label_fields.get("category_third_label"),
                        "turnover_action_type": label_fields.get("turnover_action_type"),
                        "turnover_family": label_fields.get("turnover_family"),
                        "source": source,
                        "candidate_category_codes": list(update.get("candidate_category_codes") or []),
                        "rule_version": str(update.get("rule_version") or ""),
                        "manual_assignment": bool(update.get("manual_assignment")),
                        "updated_by": normalized_actor,
                        "updated_at": timestamp,
                        "version": next_version,
                    }
                )

            self._audit_log.extend(audit_entries)
            if audit_entries:
                self._snapshot_version_cache = None
            return {
                "changed": bool(audit_entries),
                "updated_transaction_ids": [entry["transaction_id"] for entry in updated_categories],
                "updated_categories": [
                    {
                        "transaction_id": entry["transaction_id"],
                        "category_code": entry["category_code"],
                        "category_label": entry["category_label"],
                        "category_path": entry["category_path"],
                        "version": entry["category_version"],
                    }
                    for entry in updated_categories
                ],
            }

    def category_counts(self, transaction_ids: list[str]) -> dict[str, int]:
        counts = {key: 0 for key in self.tag_count_keys()}
        normalized_ids = [
            self._normalize_transaction_id(transaction_id)
            for transaction_id in list(transaction_ids or [])
            if str(transaction_id or "").strip()
        ]
        with self._lock:
            for transaction_id in normalized_ids:
                category_code = self._categories.get(transaction_id, {}).get("category_code")
                if self.has_tag_definition(category_code):
                    counts.setdefault(str(category_code), 0)
                    counts[str(category_code)] += 1
                else:
                    counts["uncategorized"] += 1
        return counts

    @staticmethod
    def label_for(category_code: str | None) -> str | None:
        if category_code is None:
            return None
        return BANK_TRANSACTION_CATEGORY_LABELS.get(category_code)

    @staticmethod
    def path_for(category_code: str | None) -> list[str]:
        if category_code is None:
            return []
        definition = BANK_TRANSACTION_CATEGORY_DEFINITIONS.get(category_code)
        if definition is None:
            definition = BANK_TRANSACTION_AUTO_CATEGORY_DEFINITIONS.get(category_code)
        if definition is None:
            definition = BANK_TRANSACTION_LEGACY_CATEGORY_DEFINITIONS.get(category_code)
        if definition is None:
            return []
        return list(definition["category_path"])

    @classmethod
    def _is_auto_tag_rule_definition(cls, definition: dict[str, Any]) -> bool:
        code = str(definition.get("code") or "").strip()
        if code == BANK_AUTO_TAG_INTERNAL_TRANSFER_CODE:
            return False
        return code in BANK_AUTO_TAG_EDITABLE_CODES or isinstance(definition.get("rules"), dict)

    @classmethod
    def _public_auto_tag_rule(
        cls,
        definition: dict[str, Any],
        *,
        priority_index: int | None,
        sort_index: int | None,
    ) -> dict[str, Any]:
        code = str(definition.get("code") or "").strip()
        rules = cls._normalize_auto_tag_rule_conditions(definition.get("rules"), allow_invalid=True)
        priority = cls._normalize_optional_priority(definition.get("priority")) or 2
        sort_order = cls._normalize_optional_sort_order(definition.get("sort_order"))
        if sort_order is None and sort_index is not None:
            sort_order = sort_index
        payload: dict[str, Any] = {
            "code": code,
            "label": str(definition.get("label") or code),
            "status": str(definition.get("status") or "active"),
            "source": str(definition.get("source") or "custom"),
            "direction": cls._normalize_auto_tag_direction(definition.get("direction")),
            "account_scope": cls._normalize_auto_tag_account_scope(definition.get("account_scope")),
            "output_primary_label": str(definition.get("output_primary_label") or definition.get("label") or code),
            "output_sub_label": str(definition.get("output_sub_label") or "").strip(),
            "rules": rules,
            "rule_code": str(definition.get("rule_code") or code),
            "rule_summary": cls._auto_tag_rule_summary(rules, archived=priority_index is None),
            "editable": True,
            "archivable": priority_index is not None,
            "sortable": priority_index is not None,
        }
        if (
            is_external_turnover_primary_label(payload["output_primary_label"])
            or definition.get("turnover_action_type")
        ):
            action_type = normalize_turnover_action_type(definition.get("turnover_action_type"))
            payload["turnover_role"] = EXTERNAL_TURNOVER_ROLE
            payload["turnover_action_type"] = action_type
        if priority_index is not None:
            payload["priority"] = priority
            payload["priority_label"] = f"优先级 {priority}"
            payload["sort_order"] = sort_order or priority_index
        elif definition.get("priority") is not None:
            payload["priority"] = priority
            payload["priority_label"] = f"优先级 {priority}"
            if sort_order is not None:
                payload["sort_order"] = sort_order
        return payload

    @classmethod
    def _normalize_auto_tag_rule_item(
        cls,
        item: Any,
        *,
        path_prefix: str,
        previous_definitions_by_code: dict[str, dict[str, Any]],
        previous_managed_codes: set[str],
        status: str,
        priority: int | None,
        fallback_sort_order: int | None,
        field_errors: list[dict[str, str]],
    ) -> dict[str, Any]:
        if not isinstance(item, dict):
            field_errors.append({"path": path_prefix, "message": "规则必须是对象。"})
            item = {}
        raw_code = item.get("code")
        code = str(raw_code or "").strip()
        is_new = not code
        if code == BANK_AUTO_TAG_INTERNAL_TRANSFER_CODE:
            field_errors.append({"path": f"{path_prefix}.code", "message": "内部往来款不能编辑。"})
        if not is_new and code not in previous_managed_codes:
            field_errors.append({"path": f"{path_prefix}.code", "message": "不能提交未知或非自动规则标签 code。"})
        if is_new and raw_code not in (None, ""):
            field_errors.append({"path": f"{path_prefix}.code", "message": "新增标签不能由前端指定 code。"})
        if status == "archived" and is_new:
            field_errors.append({"path": f"{path_prefix}.code", "message": "停用规则必须包含已有标签 code。"})

        raw_label = str(item.get("label") or "").strip()
        previous_label = ""
        if not is_new:
            previous_label = str(previous_definitions_by_code.get(code, {}).get("label") or "").strip()
        output_primary_label = str(item.get("output_primary_label") or "").strip()
        if raw_label and raw_label != previous_label and not str(item.get("output_sub_label") or "").strip():
            output_primary_label = raw_label
        output_sub_label = str(item.get("output_sub_label") or "").strip()
        if not output_primary_label:
            field_errors.append({"path": f"{path_prefix}.output_primary_label", "message": "主标签名称不能为空。"})
            output_primary_label = str(item.get("label") or code or "未命名标签").strip()
        label = output_sub_label or output_primary_label
        raw_output_third_label = str(item.get("output_third_label") or item.get("category_third_label") or "").strip()
        raw_turnover_action_type = str(item.get("turnover_action_type") or "").strip()
        previous_definition = previous_definitions_by_code.get(code, {}) if not is_new else {}
        external_turnover_rule = (
            is_external_turnover_definition(previous_definition)
            or is_external_turnover_primary_label(output_primary_label)
            or str(item.get("turnover_role") or "").strip() == EXTERNAL_TURNOVER_ROLE
        )
        configured_external_turnover_rule = (
            is_external_turnover_primary_label(output_primary_label)
            or bool(raw_output_third_label)
            or bool(raw_turnover_action_type)
            or str(item.get("turnover_role") or "").strip() == EXTERNAL_TURNOVER_ROLE
        )
        turnover_action_type = ""
        if external_turnover_rule and configured_external_turnover_rule:
            turnover_action_type = normalize_turnover_action_type(raw_turnover_action_type)
            if raw_turnover_action_type and not turnover_action_type:
                field_errors.append({"path": f"{path_prefix}.turnover_action_type", "message": "台账动作类型无效。"})
            if not turnover_action_type:
                turnover_action_type = infer_turnover_action_type(
                    primary_label=output_primary_label,
                    sub_label=output_sub_label,
                    direction=item.get("direction"),
                )
            if not turnover_action_type:
                field_errors.append({"path": f"{path_prefix}.turnover_action_type", "message": "外部往来款规则必须配置台账动作类型。"})
        else:
            if raw_output_third_label:
                field_errors.append({"path": f"{path_prefix}.output_third_label", "message": "只有外部往来款付款/收款规则可以配置子子标签。"})
            if raw_turnover_action_type:
                field_errors.append({"path": f"{path_prefix}.turnover_action_type", "message": "只有外部往来款付款/收款规则可以配置台账动作类型。"})

        rules = cls._normalize_auto_tag_rule_conditions(item.get("rules"), allow_invalid=False)
        direction = cls._normalize_auto_tag_direction(item.get("direction"))
        account_scope = cls._normalize_auto_tag_account_scope(item.get("account_scope"))
        raw_match_fields = item.get("rules", {}).get("match_fields") if isinstance(item.get("rules"), dict) else []
        invalid_fields = [
            str(field or "").strip()
            for field in list(raw_match_fields or [])
            if str(field or "").strip() and str(field or "").strip() not in BANK_AUTO_TAG_ALLOWED_FIELDS
        ]
        if invalid_fields:
            field_errors.append({"path": f"{path_prefix}.rules.match_fields", "message": "匹配字段不在允许范围内。"})
        if status == "active":
            if not rules["match_fields"]:
                field_errors.append({"path": f"{path_prefix}.rules.match_fields", "message": "匹配字段不能为空。"})
            if not (rules["exact_any"] or rules["contains_any"] or rules["contains_all"] or rules["regex_any"]):
                field_errors.append({"path": f"{path_prefix}.rules.contains", "message": "精确命中字样和包含字样不能同时为空。"})
        if not rules["match_fields"]:
            rules["match_fields"] = ["all_text"] if status == "active" else []

        if is_new:
            code = cls._generate_custom_auto_tag_code(label, rules, previous_definitions_by_code)
            source = "custom"
            path = ["自动识别", label]
            rule_code = code
        else:
            previous = previous_definitions_by_code.get(code, {})
            source = str(previous.get("source") or "custom")
            path = list(previous.get("path") or ["自动识别", label])
            if path and len(path) >= 2 and path[0] == "自动识别":
                path = ["自动识别", label]
            rule_code = str(previous.get("rule_code") or code).strip() or code

        definition: dict[str, Any] = {
            "code": code,
            "label": label,
            "path": path,
            "source": source if source in {"system", "custom"} else "custom",
            "status": status,
            "direction": direction,
            "account_scope": account_scope,
            "output_primary_label": output_primary_label,
            "output_sub_label": output_sub_label,
            "rules": rules,
            "rule_code": rule_code,
        }
        if configured_external_turnover_rule:
            definition["turnover_role"] = EXTERNAL_TURNOVER_ROLE
            definition["turnover_action_type"] = turnover_action_type
        for system_key in ("stop_on_match", "review_required", "route_to"):
            previous_value = previous_definitions_by_code.get(code, {}).get(system_key)
            if previous_value is not None:
                definition[system_key] = deepcopy(previous_value)
        if priority is not None:
            definition["priority"] = priority
        sort_order = cls._normalize_optional_sort_order(item.get("sort_order")) if isinstance(item, dict) else None
        if sort_order is None and not is_new:
            sort_order = cls._normalize_optional_sort_order(previous_definitions_by_code.get(code, {}).get("sort_order"))
        if sort_order is None:
            sort_order = fallback_sort_order
        if sort_order is not None:
            definition["sort_order"] = sort_order
        return definition

    @classmethod
    def _generate_custom_auto_tag_code(
        cls,
        label: str,
        rules: dict[str, list[str]],
        previous_definitions_by_code: dict[str, dict[str, Any]],
    ) -> str:
        seed = json.dumps(
            {"label": label, "rules": rules},
            ensure_ascii=False,
            sort_keys=True,
        )
        base = f"custom_{hashlib.sha1(seed.encode('utf-8')).hexdigest()[:12]}"
        if base not in previous_definitions_by_code:
            return base
        suffix = 2
        while f"{base}_{suffix}" in previous_definitions_by_code:
            suffix += 1
        return f"{base}_{suffix}"

    @staticmethod
    def _validate_auto_tag_duplicate_labels(
        definitions: list[dict[str, Any]],
        *,
        field_errors: list[dict[str, str]],
    ) -> None:
        seen_active_label_paths: dict[tuple[str, str], int] = {
            (str(BANK_AUTO_TAG_SYSTEM_RULE["label"]).strip(), ""): -1
        }
        for index, definition in enumerate(definitions):
            status = str(definition.get("status") or "active")
            if status != "active":
                continue
            primary_label = str(definition.get("output_primary_label") or "").strip()
            sub_label = str(definition.get("output_sub_label") or "").strip()
            if not primary_label:
                continue
            label_path = (primary_label, sub_label)
            if label_path in seen_active_label_paths:
                field_errors.append({"path": f"active_rules[{index}].output_sub_label", "message": "主标签名称和子标签名称组合不能重复。"})
                continue
            seen_active_label_paths[label_path] = index

    @classmethod
    def _auto_tag_rule_summary(cls, rules: dict[str, list[str]], *, archived: bool) -> str:
        if archived and not (rules.get("exact_any") or rules.get("contains_any") or rules.get("contains_all") or rules.get("none_of") or rules.get("regex_any")):
            return "已停用"
        field_labels = [
            BANK_AUTO_TAG_FIELD_LABELS.get(field, field)
            for field in list(rules.get("match_fields") or [])
        ]
        prefix = "、".join(field_labels) if field_labels else "未选择字段"
        parts: list[str] = []
        if rules.get("exact_any"):
            parts.append(f"{prefix}精确命中：{'、'.join(rules['exact_any'])}")
        if rules.get("contains_any"):
            parts.append(f"{prefix}包含任一：{'、'.join(rules['contains_any'])}")
        if rules.get("contains_all"):
            parts.append(f"同时包含：{'、'.join(rules['contains_all'])}")
        if rules.get("none_of"):
            parts.append(f"排除：{'、'.join(rules['none_of'])}")
        if rules.get("regex_any"):
            parts.append(f"正则：{'、'.join(rules['regex_any'])}")
        return "；".join(parts) if parts else ("已停用" if archived else "未配置规则")

    @classmethod
    def _auto_tag_rule_changes(cls, previous: dict[str, Any], next_payload: dict[str, Any]) -> dict[str, Any]:
        previous_rules = cls.auto_tag_rules_payload(previous)
        next_rules = cls.auto_tag_rules_payload(next_payload)
        ignored_fingerprint_keys = {
            "priority_label",
            "rule_summary",
            "editable",
            "archivable",
            "sortable",
        }

        def rule_fingerprint(rule: dict[str, Any]) -> dict[str, Any]:
            return {
                key: deepcopy(value)
                for key, value in rule.items()
                if key not in ignored_fingerprint_keys
            }

        previous_by_code = {
            rule["code"]: rule
            for rule in [*previous_rules["active_rules"], *previous_rules["archived_rules"]]
        }
        next_by_code = {
            rule["code"]: rule
            for rule in [*next_rules["active_rules"], *next_rules["archived_rules"]]
        }
        added = [
            {"code": code, "label": next_by_code[code]["label"]}
            for code in sorted(set(next_by_code) - set(previous_by_code))
        ]
        renamed = [
            {
                "code": code,
                "old_label": previous_by_code[code]["label"],
                "new_label": next_by_code[code]["label"],
                "old_output_primary_label": previous_by_code[code].get("output_primary_label"),
                "new_output_primary_label": next_by_code[code].get("output_primary_label"),
                "old_output_sub_label": previous_by_code[code].get("output_sub_label"),
                "new_output_sub_label": next_by_code[code].get("output_sub_label"),
                "old_turnover_action_type": previous_by_code[code].get("turnover_action_type"),
                "new_turnover_action_type": next_by_code[code].get("turnover_action_type"),
            }
            for code in sorted(set(previous_by_code).intersection(next_by_code))
            if (
                previous_by_code[code]["status"] == next_by_code[code]["status"]
                and (
                    previous_by_code[code]["label"] != next_by_code[code]["label"]
                    or (
                        previous_by_code[code].get("output_primary_label")
                        != next_by_code[code].get("output_primary_label")
                    )
                    or (
                        previous_by_code[code].get("output_sub_label")
                        != next_by_code[code].get("output_sub_label")
                    )
                    or (
                        previous_by_code[code].get("turnover_action_type")
                        != next_by_code[code].get("turnover_action_type")
                    )
                )
            )
        ]
        archived = [
            code
            for code in sorted(set(previous_by_code).intersection(next_by_code))
            if previous_by_code[code]["status"] != "archived" and next_by_code[code]["status"] == "archived"
        ]
        reenabled = [
            code
            for code in sorted(set(previous_by_code).intersection(next_by_code))
            if previous_by_code[code]["status"] == "archived" and next_by_code[code]["status"] == "active"
        ]
        previous_order = [(rule["code"], rule.get("priority")) for rule in previous_rules["active_rules"]]
        next_order = [(rule["code"], rule.get("priority")) for rule in next_rules["active_rules"]]
        priority_changes = [] if previous_order == next_order else [{"old_order": previous_order, "new_order": next_order}]
        rule_changes = [
            {
                "code": code,
                "old_summary": previous_by_code[code].get("rule_summary", ""),
                "new_summary": next_by_code[code].get("rule_summary", ""),
            }
            for code in sorted(set(previous_by_code).intersection(next_by_code))
            if previous_by_code[code].get("rules") != next_by_code[code].get("rules")
        ]
        rule_payload_changes = [
            {"code": code}
            for code in sorted(set(previous_by_code).intersection(next_by_code))
            if rule_fingerprint(previous_by_code[code]) != rule_fingerprint(next_by_code[code])
        ]
        changed = bool(added or archived or reenabled or rule_payload_changes)
        return {
            "changed": changed,
            "added_tags": added,
            "renamed_tags": renamed,
            "archived_codes": archived,
            "reenabled_codes": reenabled,
            "priority_changes": priority_changes,
            "rule_changes": rule_changes,
            "rule_payload_changes": rule_payload_changes,
        }

    @classmethod
    def _build_tag_definition_index(cls, payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
        return {
            str(definition["code"]): dict(definition)
            for definition in list(payload.get("definitions") or [])
            if isinstance(definition, dict) and str(definition.get("code") or "").strip()
        }

    @classmethod
    def _normalize_tag_dictionary_payload(cls, payload: dict[str, Any] | None) -> dict[str, Any]:
        raw_payload = payload if isinstance(payload, dict) else {}
        version = cls._normalize_version(
            raw_payload.get("version", BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION)
        )
        if version <= 0:
            version = BANK_TRANSACTION_TAG_DICTIONARY_INITIAL_VERSION

        definitions_by_code: dict[str, dict[str, Any]] = {
            definition["code"]: definition
            for definition in build_system_bank_transaction_tag_definitions()
        }
        raw_definitions = raw_payload.get("definitions")
        if raw_definitions is None:
            raw_definitions = raw_payload.get("tags")
        for item in list(raw_definitions or []):
            if not isinstance(item, dict):
                continue
            definition = cls._normalize_tag_definition(item)
            code = definition["code"]
            if definition["source"] == "system" and code in definitions_by_code:
                next_definition = {
                    **definitions_by_code[code],
                    "status": definition["status"],
                }
                if code in BANK_AUTO_TAG_EDITABLE_CODES:
                    next_definition["label"] = definition["label"]
                    if definition.get("priority") is not None:
                        next_definition["priority"] = definition["priority"]
                    if isinstance(definition.get("rules"), dict):
                        next_definition["rules"] = deepcopy(definition["rules"])
                    if definition.get("rule_code"):
                        next_definition["rule_code"] = definition["rule_code"]
                    for key in (
                        "direction",
                        "account_scope",
                        "output_primary_label",
                        "output_sub_label",
                        "turnover_role",
                        "turnover_action_type",
                        "sort_order",
                        "stop_on_match",
                        "review_required",
                        "route_to",
                    ):
                        if key in definition:
                            next_definition[key] = deepcopy(definition[key])
                definitions_by_code[code] = next_definition
                continue
            if code in definitions_by_code and definitions_by_code[code]["source"] == "system":
                continue
            definitions_by_code[code] = definition
        return {
            "version": version,
            "definitions": sorted(definitions_by_code.values(), key=lambda item: (item["source"] != "system", item["code"])),
        }

    @classmethod
    def _normalize_legacy_auto_tag_priority_sequence(cls, definitions_by_code: dict[str, dict[str, Any]]) -> None:
        for definition in list(definitions_by_code.values()):
            if not cls._is_auto_tag_rule_definition(definition):
                continue
            if str(definition.get("status") or "active") != "active":
                continue
            priority = cls._normalize_optional_priority(definition.get("priority"))
            sort_order = cls._normalize_optional_sort_order(definition.get("sort_order"))
            if priority is None or sort_order is None:
                continue
            if priority < 10 or priority != sort_order * 10:
                continue
            code = str(definition.get("code") or "")
            normalized = dict(definition)
            normalized["priority"] = 2
            definitions_by_code[code] = normalized

    @classmethod
    def _normalize_tag_definition(cls, item: dict[str, Any]) -> dict[str, Any]:
        label = str(item.get("label") or item.get("category_label") or "").strip()
        path = [
            str(value).strip()
            for value in list(item.get("path") or item.get("category_path") or [])
            if str(value).strip()
        ]
        code = str(item.get("code") or item.get("category_code") or "").strip()
        if not code:
            seed = json.dumps({"label": label, "path": path}, ensure_ascii=False, sort_keys=True)
            code = f"custom_{hashlib.sha1(seed.encode('utf-8')).hexdigest()[:12]}"
        if not BANK_TRANSACTION_TAG_CODE_RE.match(code):
            code = f"custom_{hashlib.sha1(code.encode('utf-8')).hexdigest()[:12]}"
        source = str(item.get("source") or "custom").strip()
        if source not in {"system", "custom"}:
            source = "custom"
        status = str(item.get("status") or "active").strip()
        if status not in {"active", "archived"}:
            status = "active"
        definition = {
            "code": code,
            "label": label or code,
            "path": path,
            "source": source,
            "status": status,
        }
        priority = cls._normalize_optional_priority(item.get("priority"))
        default_fields = _default_auto_tag_rule_fields(code)
        if priority is None and default_fields:
            priority = int(default_fields["priority"])
        if priority is not None:
            definition["priority"] = priority
        sort_order = cls._normalize_optional_sort_order(item.get("sort_order"))
        if sort_order is not None:
            definition["sort_order"] = sort_order
        rules = item.get("rules")
        if not isinstance(rules, dict) and default_fields:
            rules = default_fields["rules"]
        if isinstance(rules, dict):
            definition["direction"] = cls._normalize_auto_tag_direction(item.get("direction", default_fields.get("direction")))
            definition["account_scope"] = cls._normalize_auto_tag_account_scope(
                item.get("account_scope", default_fields.get("account_scope"))
            )
            output_primary_label = str(item.get("output_primary_label") or label or code).strip()
            if output_primary_label:
                definition["output_primary_label"] = output_primary_label
            output_sub_label = str(item.get("output_sub_label") or "").strip()
            if output_sub_label:
                definition["output_sub_label"] = output_sub_label
            raw_action_type = str(item.get("turnover_action_type") or "").strip()
            external_turnover_definition = (
                is_external_turnover_primary_label(output_primary_label)
                or bool(raw_action_type)
                or str(item.get("turnover_role") or "").strip() == EXTERNAL_TURNOVER_ROLE
            )
            if external_turnover_definition:
                action_type = normalize_turnover_action_type(raw_action_type) or infer_turnover_action_type(
                    primary_label=output_primary_label,
                    sub_label=output_sub_label,
                    direction=item.get("direction"),
                )
                definition["turnover_role"] = EXTERNAL_TURNOVER_ROLE
                if action_type:
                    definition["turnover_action_type"] = action_type
            derived_label = output_sub_label or output_primary_label
            if derived_label:
                definition["label"] = derived_label
                if not path or (len(path) >= 2 and path[0] == "自动识别"):
                    definition["path"] = ["自动识别", derived_label]
            definition["rules"] = cls._normalize_auto_tag_rule_conditions(rules, allow_invalid=True)
            rule_code = str(item.get("rule_code") or default_fields.get("rule_code") or "").strip()
            if rule_code:
                definition["rule_code"] = rule_code
            for key in ("stop_on_match", "review_required", "route_to"):
                if key in item:
                    definition[key] = deepcopy(item[key])
                elif key in default_fields:
                    definition[key] = deepcopy(default_fields[key])
        return definition

    @classmethod
    def _normalize_auto_tag_rule_conditions(cls, value: Any, *, allow_invalid: bool = False) -> dict[str, list[str]]:
        raw = value if isinstance(value, dict) else {}
        match_fields = cls._normalize_match_fields(raw.get("match_fields"), allow_invalid=allow_invalid)
        exact_any = cls._normalize_term_list(raw.get("exact_any", raw.get("exact")))
        contains_any = cls._normalize_term_list(raw.get("contains_any", raw.get("contains")))
        none_of = cls._normalize_term_list(raw.get("none_of", raw.get("excludes")))
        contains_all = cls._normalize_term_list(raw.get("contains_all"))
        regex_any = cls._normalize_term_list(raw.get("regex_any"))
        return {
            "match_fields": match_fields,
            "exact_any": exact_any,
            "contains_any": contains_any,
            "contains_all": contains_all,
            "none_of": none_of,
            "regex_any": regex_any,
            # Backward-compatible aliases for older clients and persisted payloads.
            "exact": exact_any,
            "contains": contains_any,
            "excludes": none_of,
        }

    @staticmethod
    def _normalize_auto_tag_direction(value: Any) -> str:
        direction = str(value or DEFAULT_BANK_AUTO_TAG_DIRECTION).strip()
        return direction if direction in BANK_AUTO_TAG_ALLOWED_DIRECTIONS else DEFAULT_BANK_AUTO_TAG_DIRECTION

    @staticmethod
    def _normalize_auto_tag_account_scope(value: Any) -> dict[str, Any]:
        raw = value if isinstance(value, dict) else {}
        scope_type = str(raw.get("type") or raw.get("mode") or "any").strip()
        if scope_type not in BANK_AUTO_TAG_ALLOWED_ACCOUNT_SCOPE_TYPES:
            scope_type = "any"
        values = BankTransactionCategoryService._normalize_term_list(raw.get("values"))
        if scope_type == "any":
            values = []
        return {"type": scope_type, "values": values}

    @staticmethod
    def _normalize_term_list(value: Any) -> list[str]:
        terms: list[str] = []
        seen: set[str] = set()
        raw_values = value if isinstance(value, list) else []
        for item in raw_values:
            text = str(item or "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            terms.append(text)
        return terms

    @classmethod
    def _auto_tag_file_rows(cls, source: Any) -> tuple[dict[str, Any], list[Any]]:
        if isinstance(source, (str, Path)):
            source_path = Path(source)
            if source_path.suffix.lower() == ".xlsx":
                return cls._xlsx_auto_tag_file_rows(source_path)
        if isinstance(source, dict):
            if isinstance(source.get("rules"), list):
                return dict(source), list(source.get("rules") or [])
            if isinstance(source.get("rows"), list):
                return dict(source), cls._worksheet_auto_tag_file_rows(source.get("rows"))
            return dict(source), []
        if isinstance(source, list):
            if source and isinstance(source[0], (list, tuple)):
                return {}, cls._worksheet_auto_tag_file_rows(source)
            return {}, list(source)
        return {}, []

    @classmethod
    def _xlsx_auto_tag_file_rows(cls, source_path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is declared in backend requirements.
            raise BankAutoTagRulesValidationError(
                "bank_auto_tag_rule_xlsx_unavailable",
                "当前运行环境缺少 openpyxl，无法读取银行流水标签 Excel 文件。",
            ) from exc

        workbook = load_workbook(source_path, data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return (
            {
                "source_name": source_path.name,
                "source_workbook_name": source_path.name,
            },
            cls._worksheet_auto_tag_file_rows(rows),
        )

    @classmethod
    def _worksheet_auto_tag_file_rows(cls, rows: Any) -> list[dict[str, Any]]:
        raw_rows = list(rows or [])
        if not raw_rows:
            return []
        header_row_index = cls._auto_tag_file_header_row_index(raw_rows)
        headers = [str(value or "").strip() for value in list(raw_rows[header_row_index] or [])]
        normalized_rows: list[dict[str, Any]] = []
        last_flow_type = ""
        last_primary_label = ""
        for row_index, row in enumerate(raw_rows[header_row_index + 1 :], start=header_row_index + 1):
            values = list(row or [])
            if not any(str(value or "").strip() for value in values):
                continue
            raw = {
                headers[index]: values[index] if index < len(values) else ""
                for index in range(len(headers))
                if headers[index]
            }
            normalized = cls._normalize_auto_tag_file_row(raw)
            if normalized.get("flow_type"):
                last_flow_type = str(normalized["flow_type"])
            else:
                normalized["flow_type"] = last_flow_type
            if normalized.get("primary_label"):
                last_primary_label = str(normalized["primary_label"])
            else:
                normalized["primary_label"] = last_primary_label
            if not normalized.get("source_row"):
                normalized["source_row"] = row_index + 1
            normalized_rows.append(normalized)
        return normalized_rows

    @classmethod
    def _auto_tag_file_header_row_index(cls, rows: list[Any]) -> int:
        required_fields = ("flow_type", "primary_label", "sub_label", "query_fields", "contains")
        for index, row in enumerate(rows):
            cells = {str(value or "").strip() for value in list(row or [])}
            matched_fields = {
                field
                for field, aliases in BANK_AUTO_TAG_FILE_HEADER_ALIASES.items()
                if any(alias in cells for alias in aliases)
            }
            if all(field in matched_fields for field in required_fields):
                return index
        return 0

    @classmethod
    def _normalize_auto_tag_file_row(cls, row: Any) -> dict[str, Any]:
        raw = row if isinstance(row, dict) else {}
        normalized: dict[str, Any] = {}
        for field, aliases in BANK_AUTO_TAG_FILE_HEADER_ALIASES.items():
            for alias in aliases:
                if alias in raw:
                    normalized[field] = raw.get(alias)
                    break
            else:
                normalized[field] = ""
        return {
            key: (str(value).strip() if value is not None else "")
            for key, value in normalized.items()
        }

    @staticmethod
    def _split_auto_tag_file_terms(value: Any) -> list[str]:
        terms: list[str] = []
        seen: set[str] = set()
        for item in _BANK_AUTO_TAG_FILE_TERM_SPLIT_RE.split(str(value or "")):
            term = item.strip()
            if not term or term in seen:
                continue
            seen.add(term)
            terms.append(term)
        return terms

    @staticmethod
    def _auto_tag_file_direction(value: Any) -> str:
        text = str(value or "").strip()
        if text == "支出":
            return "expense"
        if text == "收入":
            return "income"
        return "any"

    @staticmethod
    def _auto_tag_file_rule_comparable(rule: dict[str, Any]) -> dict[str, Any]:
        rules = rule.get("rules") if isinstance(rule.get("rules"), dict) else {}
        return {
            "output_primary_label": str(rule.get("output_primary_label") or ""),
            "output_sub_label": str(rule.get("output_sub_label") or ""),
            "direction": str(rule.get("direction") or "any"),
            "match_fields": list(rules.get("match_fields") or []),
            "exact_any": list(rules.get("exact_any") or []),
            "contains_any": list(rules.get("contains_any") or []),
            "contains_all": list(rules.get("contains_all") or []),
            "none_of": list(rules.get("none_of") or []),
        }

    @classmethod
    def _auto_tag_rule_sort_key(cls, definition: dict[str, Any]) -> tuple[int, int, str]:
        priority = cls._normalize_optional_priority(definition.get("priority")) or 2
        if priority < 2:
            priority = 2
        sort_order = cls._normalize_optional_sort_order(definition.get("sort_order"))
        if sort_order is None:
            sort_order = 10_000
        return (priority, sort_order, str(definition.get("code") or ""))

    @classmethod
    def _normalize_submitted_auto_tag_priority(
        cls,
        value: Any,
        *,
        path: str,
        field_errors: list[dict[str, str]],
    ) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            field_errors.append({"path": path, "message": "普通规则优先级必须是大于等于 2 的整数。"})
            return None
        text = str(value).strip()
        if not re.fullmatch(r"[+-]?\d+", text):
            field_errors.append({"path": path, "message": "普通规则优先级必须是大于等于 2 的整数。"})
            return None
        priority = int(text)
        if priority < 2:
            field_errors.append({"path": path, "message": "普通规则优先级必须大于等于 2。"})
            return None
        return priority

    @staticmethod
    def _normalize_match_fields(value: Any, *, allow_invalid: bool = False) -> list[str]:
        fields: list[str] = []
        seen: set[str] = set()
        raw_values = value if isinstance(value, list) else []
        for item in raw_values:
            field = str(item or "").strip()
            if not field or field in seen:
                continue
            if allow_invalid or field in BANK_AUTO_TAG_ALLOWED_FIELDS:
                seen.add(field)
                fields.append(field)
        return fields

    @staticmethod
    def _normalize_optional_priority(value: Any) -> int | None:
        if value in (None, ""):
            return None
        try:
            return max(int(value), 0)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _normalize_optional_sort_order(value: Any) -> int | None:
        if value in (None, ""):
            return None
        try:
            return max(int(value), 0)
        except (TypeError, ValueError):
            return None

    def _stored_category_semantics(self, record: dict[str, Any]) -> tuple[Any, ...]:
        return (
            str(record.get("category_primary_label") or ""),
            str(record.get("category_sub_label") or ""),
            str(record.get("category_third_label") or ""),
            tuple(str(item) for item in list(record.get("category_label_path") or [])),
            str(record.get("turnover_action_type") or ""),
            str(record.get("turnover_family") or ""),
        )

    def _submitted_category_semantics(self, update: dict[str, Any]) -> tuple[Any, ...]:
        fields = self._label_fields_for_update(update.get("category_code"), update)
        return (
            str(fields.get("category_primary_label") or ""),
            str(fields.get("category_sub_label") or ""),
            str(fields.get("category_third_label") or ""),
            tuple(str(item) for item in list(fields.get("category_label_path") or [])),
            str(fields.get("turnover_action_type") or ""),
            str(fields.get("turnover_family") or ""),
        )

    def _label_fields_for_update(self, category_code: str | None, update: dict[str, Any]) -> dict[str, Any]:
        base = self._label_fields_for(category_code)
        definition = self._tag_definitions_by_code.get(category_code) if category_code is not None else None
        update_path = [
            str(item).strip()
            for item in list(update.get("category_label_path") or [])
            if str(item).strip()
        ]
        primary_label = str(update.get("category_primary_label") or base.get("category_primary_label") or "").strip()
        sub_label = str(update.get("category_sub_label") or base.get("category_sub_label") or "").strip()
        third_label = (
            normalize_external_third_label(update.get("category_third_label"))
            or normalize_external_third_label(base.get("category_third_label"))
        )
        if update_path:
            primary_label = primary_label or (update_path[0] if len(update_path) >= 1 else "")
            sub_label = sub_label or (update_path[1] if len(update_path) >= 2 else "")
            if len(update_path) >= 3:
                third_label = third_label or normalize_external_third_label(update_path[2])
        external_turnover = (
            is_external_turnover_definition(definition)
            or is_external_turnover_primary_label(primary_label)
            or bool(normalize_turnover_action_type(update.get("turnover_action_type")))
        )
        action_type = None
        turnover_family = None
        turnover_role = ""
        if external_turnover:
            action_type = (
                normalize_turnover_action_type(update.get("turnover_action_type"))
                or normalize_turnover_action_type((definition or {}).get("turnover_action_type"))
                or infer_turnover_action_type(
                    primary_label=primary_label,
                    sub_label=sub_label,
                    direction=(definition or {}).get("direction"),
                )
            )
            turnover_family = (
                str(update.get("turnover_family") or "").strip()
                or str((definition or {}).get("turnover_family") or "").strip()
                or turnover_family_for_third_label(third_label)
            )
            turnover_role = EXTERNAL_TURNOVER_ROLE
        label_path = update_path or turnover_label_path(primary_label, sub_label, third_label if external_turnover else "")
        category_label = str(self._label_for(category_code) or "").strip()
        if external_turnover:
            category_label = sub_label or primary_label or category_label
        return {
            "category_label": category_label or None,
            "category_primary_label": primary_label or None,
            "category_sub_label": sub_label or None,
            "category_third_label": third_label or None,
            "category_label_path": label_path,
            "turnover_role": turnover_role,
            "turnover_action_type": action_type,
            "turnover_family": turnover_family or None,
        }

    def _normalize_categories(self, categories: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
        normalized: dict[str, dict[str, Any]] = {}
        for transaction_id, record in categories.items():
            if not isinstance(record, dict):
                continue
            normalized_id = self._normalize_transaction_id(transaction_id or record.get("transaction_id"))
            if not normalized_id:
                continue
            category_code = self._normalize_category_code_current(record.get("category_code"))
            category_label = str(record.get("category_label") or self._label_for(category_code) or "").strip()
            category_path = [
                str(item).strip()
                for item in list(record.get("category_path") or self._path_for(category_code))
                if str(item).strip()
            ]
            category_primary_label = str(record.get("category_primary_label") or "").strip() or None
            category_sub_label = str(record.get("category_sub_label") or "").strip() or None
            category_third_label = normalize_external_third_label(record.get("category_third_label")) or None
            category_label_path = [
                str(item).strip()
                for item in list(record.get("category_label_path") or [])
                if str(item).strip()
            ]
            if not category_label_path:
                category_label_path = turnover_label_path(category_primary_label, category_sub_label, category_third_label)
            turnover_action_type = normalize_turnover_action_type(record.get("turnover_action_type")) or None
            turnover_family = str(record.get("turnover_family") or "").strip() or turnover_family_for_third_label(category_third_label)
            normalized[normalized_id] = {
                "transaction_id": normalized_id,
                "category_code": category_code,
                "category_label": category_label or None,
                "category_path": category_path,
                "category_primary_label": category_primary_label,
                "category_sub_label": category_sub_label,
                "category_third_label": category_third_label,
                "category_label_path": category_label_path,
                "turnover_role": str(record.get("turnover_role") or "").strip() or (EXTERNAL_TURNOVER_ROLE if turnover_action_type else ""),
                "turnover_action_type": turnover_action_type,
                "turnover_family": turnover_family or None,
                "source": str(record.get("source") or "manual").strip() or "manual",
                "manual_assignment": bool(record.get("manual_assignment")),
                "updated_by": str(record.get("updated_by") or "").strip(),
                "updated_at": str(record.get("updated_at") or "").strip(),
                "version": self._normalize_version(record.get("version")),
                "candidate_category_codes": [
                    str(code).strip()
                    for code in list(record.get("candidate_category_codes") or [])
                    if str(code or "").strip()
                ],
                "rule_version": str(record.get("rule_version") or "").strip(),
            }
        return normalized

    def _normalize_update(self, update: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(update, dict):
            raise BankTransactionCategoryValidationError(
                "invalid_category_update",
                "each update must be an object.",
            )
        transaction_id = self._normalize_transaction_id(update.get("transaction_id"))
        if not transaction_id:
            raise BankTransactionCategoryValidationError(
                "unknown_transaction_id",
                "transaction_id is required.",
            )
        raw_code = update.get("category_code")
        category_code = None if raw_code is None else str(raw_code).strip()
        if category_code == "":
            category_code = None
        if category_code is not None:
            definition = self._tag_definitions_by_code.get(category_code)
            if not isinstance(definition, dict):
                raise BankTransactionCategoryValidationError(
                    "invalid_category_code",
                    f"Invalid bank transaction category code: {category_code}",
                    transaction_id=transaction_id,
                )
            if definition.get("status") == "archived":
                raise BankTransactionCategoryValidationError(
                    "archived_category_code",
                    f"Archived bank transaction category code cannot be selected: {category_code}",
                        transaction_id=transaction_id,
                )
        category_label_path = [
            str(item).strip()
            for item in list(update.get("category_label_path") or [])
            if str(item).strip()
        ]
        category_primary_label = str(update.get("category_primary_label") or "").strip()
        category_sub_label = str(update.get("category_sub_label") or "").strip()
        category_third_label = normalize_external_third_label(update.get("category_third_label"))
        if category_label_path:
            category_primary_label = category_primary_label or (category_label_path[0] if len(category_label_path) >= 1 else "")
            category_sub_label = category_sub_label or (category_label_path[1] if len(category_label_path) >= 2 else "")
            category_third_label = category_third_label or (
                normalize_external_third_label(category_label_path[2]) if len(category_label_path) >= 3 else ""
            )
        turnover_action_type = normalize_turnover_action_type(update.get("turnover_action_type"))
        turnover_family = str(update.get("turnover_family") or "").strip() or turnover_family_for_third_label(category_third_label)
        expected_version = update.get("expected_version")
        if expected_version is not None:
            expected_version = self._normalize_version(expected_version)
        return {
            "transaction_id": transaction_id,
            "category_code": category_code,
            "expected_version": expected_version,
            "candidate_category_codes": [
                str(code).strip()
                for code in list(update.get("candidate_category_codes") or [])
                if str(code or "").strip()
            ],
            "rule_version": str(update.get("rule_version") or "").strip(),
            "manual_assignment": bool(update.get("manual_assignment")),
            "category_primary_label": category_primary_label or None,
            "category_sub_label": category_sub_label or None,
            "category_third_label": category_third_label or None,
            "category_label_path": category_label_path,
            "turnover_action_type": turnover_action_type or None,
            "turnover_family": turnover_family or None,
        }

    @staticmethod
    def _normalize_transaction_id(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def _normalize_category_code(value: Any) -> str | None:
        if value is None:
            return None
        category_code = str(value or "").strip()
        if not category_code:
            return None
        if category_code not in BANK_TRANSACTION_CATEGORY_LABELS:
            return None
        return category_code

    def _normalize_category_code_current(self, value: Any) -> str | None:
        if value is None:
            return None
        category_code = str(value or "").strip()
        if not category_code:
            return None
        if category_code not in self._tag_definitions_by_code:
            return category_code if category_code.startswith("custom_") else None
        return category_code

    @staticmethod
    def _normalize_version(value: Any) -> int:
        try:
            return max(int(value or 0), 0)
        except (TypeError, ValueError):
            return 0

    def _current_version(self, transaction_id: str) -> int:
        return self._normalize_version(self._categories.get(transaction_id, {}).get("version"))

    def _label_for(self, category_code: str | None) -> str | None:
        if category_code is None:
            return None
        definition = self._tag_definitions_by_code.get(category_code)
        if isinstance(definition, dict):
            return str(definition.get("label") or category_code)
        return self.label_for(category_code)

    def _path_for(self, category_code: str | None) -> list[str]:
        if category_code is None:
            return []
        definition = self._tag_definitions_by_code.get(category_code)
        if isinstance(definition, dict):
            return [
                str(item)
                for item in list(definition.get("path") or [])
                if str(item).strip()
            ]
        return self.path_for(category_code)

    def _label_fields_for(self, category_code: str | None) -> dict[str, Any]:
        label = self._label_for(category_code)
        if category_code is None:
            return {
                "category_primary_label": None,
                "category_sub_label": None,
                "category_third_label": None,
                "category_label_path": [],
            }
        definition = self._tag_definitions_by_code.get(category_code)
        primary_label = ""
        sub_label = ""
        third_label = ""
        if isinstance(definition, dict):
            primary_label = str(definition.get("output_primary_label") or "").strip()
            sub_label = str(definition.get("output_sub_label") or "").strip()
            third_label = normalize_external_third_label(definition.get("output_third_label"))
        if not primary_label:
            path = self._path_for(category_code)
            if len(path) >= 2 and path[0] != "自动识别":
                primary_label = str(path[-2] or "").strip()
                if not sub_label:
                    sub_label = str(path[-1] or "").strip()
            else:
                primary_label = str(label or "").strip()
        label_path = turnover_label_path(primary_label, sub_label, third_label)
        return {
            "category_primary_label": primary_label or None,
            "category_sub_label": sub_label or None,
            "category_third_label": third_label or None,
            "category_label_path": label_path,
        }

    def category_semantics_for_code(self, category_code: str | None) -> dict[str, Any]:
        normalized_category_code = str(category_code or "").strip()
        label_fields = self._label_fields_for(normalized_category_code or None)
        definition = self._tag_definitions_by_code.get(normalized_category_code)
        external_turnover = (
            is_external_turnover_definition(definition)
            or is_external_turnover_primary_label(label_fields.get("category_primary_label"))
        )
        turnover_role = ""
        turnover_action_type = None
        turnover_family = None
        if external_turnover:
            turnover_role = EXTERNAL_TURNOVER_ROLE
            turnover_action_type = (
                normalize_turnover_action_type((definition or {}).get("turnover_action_type"))
                or infer_turnover_action_type(
                    primary_label=label_fields.get("category_primary_label"),
                    sub_label=label_fields.get("category_sub_label"),
                    direction=(definition or {}).get("direction"),
                )
                or None
            )
            turnover_family = (
                str((definition or {}).get("turnover_family") or "").strip()
                or turnover_family_for_third_label(label_fields.get("category_third_label"))
                or None
            )
        return {
            "category_code": normalized_category_code or None,
            "category_label": self._label_for(normalized_category_code or None),
            "category_path": self._path_for(normalized_category_code or None),
            **label_fields,
            "turnover_role": turnover_role,
            "turnover_action_type": turnover_action_type,
            "turnover_family": turnover_family,
        }

    def tag_count_keys(self) -> list[str]:
        with self._lock:
            active_codes = [
                str(code)
                for code, definition in self._tag_definitions_by_code.items()
                if definition.get("status") == "active"
            ]
        return [*active_codes, "uncategorized"]

    def has_tag_definition(self, category_code: str | None) -> bool:
        if category_code is None:
            return False
        with self._lock:
            return str(category_code) in self._tag_definitions_by_code

    def _public_record(self, transaction_id: str, record: dict[str, Any] | None) -> dict[str, Any]:
        category_code = record.get("category_code") if isinstance(record, dict) else None
        version = self._normalize_version(record.get("version")) if isinstance(record, dict) else 0
        computed_label_fields = self._label_fields_for(category_code)
        category_primary_label = (
            str(record.get("category_primary_label") or "").strip()
            if isinstance(record, dict)
            else ""
        ) or computed_label_fields.get("category_primary_label")
        category_sub_label = (
            str(record.get("category_sub_label") or "").strip()
            if isinstance(record, dict)
            else ""
        ) or computed_label_fields.get("category_sub_label")
        category_third_label = (
            normalize_external_third_label(record.get("category_third_label"))
            if isinstance(record, dict)
            else ""
        ) or computed_label_fields.get("category_third_label")
        category_label_path = (
            [
                str(item).strip()
                for item in list(record.get("category_label_path") or [])
                if str(item).strip()
            ]
            if isinstance(record, dict)
            else []
        ) or turnover_label_path(category_primary_label, category_sub_label, category_third_label)
        category_label = (
            str(record.get("category_label") or "").strip()
            if isinstance(record, dict)
            else ""
        ) or self._label_for(category_code)
        computed_semantics = self.category_semantics_for_code(category_code)
        return {
            "transaction_id": transaction_id,
            "category_code": category_code,
            "category_label": category_label,
            "category_primary_label": category_primary_label,
            "category_sub_label": category_sub_label,
            "category_third_label": category_third_label,
            "category_label_path": category_label_path,
            "category_path": (
                [
                    str(item).strip()
                    for item in list(record.get("category_path") or [])
                    if str(item).strip()
                ]
                if isinstance(record, dict)
                else []
            ) or self._path_for(category_code),
            "turnover_role": (
                str(record.get("turnover_role") or "").strip()
                if isinstance(record, dict)
                else ""
            ) or str(computed_semantics.get("turnover_role") or ""),
            "turnover_action_type": (
                str(record.get("turnover_action_type") or "").strip()
                if isinstance(record, dict) and record.get("turnover_action_type")
                else None
            ) or computed_semantics.get("turnover_action_type"),
            "turnover_family": (
                str(record.get("turnover_family") or "").strip()
                if isinstance(record, dict) and record.get("turnover_family")
                else None
            ) or computed_semantics.get("turnover_family"),
            "category_version": version,
            "source": str(record.get("source") or "") if isinstance(record, dict) else "",
            "updated_by": str(record.get("updated_by") or "") if isinstance(record, dict) else "",
            "updated_at": str(record.get("updated_at") or "") if isinstance(record, dict) else "",
            "confirmed_candidate_category_codes": list(record.get("candidate_category_codes") or []) if isinstance(record, dict) else [],
            "category_rule_version": str(record.get("rule_version") or "") if isinstance(record, dict) else "",
            "manual_assignment": bool(record.get("manual_assignment")) if isinstance(record, dict) else False,
        }
