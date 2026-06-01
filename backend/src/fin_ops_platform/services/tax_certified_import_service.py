from __future__ import annotations

from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import re
from typing import Any
import warnings

from openpyxl import load_workbook


MONTH_FROM_FILENAME_RE = re.compile(r"(\d{4})年(\d{1,2})月")
COMPACT_MONTH_RE = re.compile(r"^\d{6}$")
CERTIFIED_SELECTION_STATUSES = {"已勾选", "已认证"}
ALLOWED_INVOICE_STATUSES = {"正常"}


@dataclass(slots=True)
class UploadedCertifiedImportFile:
    file_name: str
    content: bytes


@dataclass(slots=True)
class TaxCertifiedInvoiceRecord:
    id: str
    unique_key: str
    month: str
    source_file_name: str
    source_row_number: int
    taxpayer_tax_no: str | None
    taxpayer_name: str | None
    digital_invoice_no: str | None
    invoice_code: str | None
    invoice_no: str | None
    issue_date: str | None
    seller_tax_no: str | None
    seller_name: str | None
    amount: str | None
    tax_amount: str | None
    deductible_tax_amount: str | None
    selection_status: str | None
    invoice_status: str | None
    selection_time: str | None
    invoice_source: str | None = None
    invoice_kind: str | None = None
    risk_level: str | None = None
    batch_id: str | None = None
    imported_at: datetime = field(default_factory=lambda: datetime.now(UTC))


@dataclass(slots=True)
class TaxCertifiedImportPreviewRowResult:
    id: str
    month: str
    source_file_name: str
    source_row_number: int
    row_status: str
    error_message: str | None = None
    unique_key: str | None = None
    taxpayer_tax_no: str | None = None
    taxpayer_name: str | None = None
    digital_invoice_no: str | None = None
    invoice_code: str | None = None
    invoice_no: str | None = None
    issue_date: str | None = None
    seller_tax_no: str | None = None
    seller_name: str | None = None
    amount: str | None = None
    tax_amount: str | None = None
    deductible_tax_amount: str | None = None
    selection_status: str | None = None
    invoice_status: str | None = None
    selection_time: str | None = None
    invoice_source: str | None = None
    invoice_kind: str | None = None
    risk_level: str | None = None
    match_status: str = "unknown"
    matched_plan_id: str | None = None
    dedupe_status: str = "not_applicable"


@dataclass(slots=True)
class TaxCertifiedImportPreviewFile:
    id: str
    file_name: str
    month: str
    recognized_count: int
    invalid_count: int
    rows: list[TaxCertifiedInvoiceRecord] = field(default_factory=list)
    row_results: list[TaxCertifiedImportPreviewRowResult] = field(default_factory=list)


@dataclass(slots=True)
class TaxCertifiedImportSession:
    id: str
    imported_by: str
    file_count: int
    status: str
    files: list[TaxCertifiedImportPreviewFile]
    created_at: datetime = field(default_factory=lambda: datetime.now(UTC))


@dataclass(slots=True)
class TaxCertifiedImportBatch:
    id: str
    session_id: str
    imported_by: str
    file_count: int
    months: list[str]
    persisted_record_count: int
    created_at: datetime = field(default_factory=lambda: datetime.now(UTC))


class TaxCertifiedImportService:
    def __init__(self, *, state_store: Any | None = None) -> None:
        self._state_store = state_store
        self._session_counter = 0
        self._file_counter = 0
        self._batch_counter = 0
        self._sessions: dict[str, TaxCertifiedImportSession] = {}
        self._batches: dict[str, TaxCertifiedImportBatch] = {}
        self._records: dict[str, TaxCertifiedInvoiceRecord] = {}

        snapshot = self._state_store.load_tax_certified_imports() if self._state_store is not None else {}
        if snapshot:
            self._hydrate(snapshot)

    def preview_files(
        self,
        *,
        imported_by: str,
        uploads: list[UploadedCertifiedImportFile],
    ) -> TaxCertifiedImportSession:
        session = TaxCertifiedImportSession(
            id=self._next_session_id(),
            imported_by=imported_by,
            file_count=len(uploads),
            status="preview_ready",
            files=[],
        )
        for upload in uploads:
            session.files.append(self._preview_single_file(upload))
        self._sessions[session.id] = session
        self._persist()
        return session

    def get_session(self, session_id: str) -> TaxCertifiedImportSession:
        return self._sessions[session_id]

    def confirm_session(self, session_id: str) -> TaxCertifiedImportBatch:
        session = self._sessions[session_id]
        existing_batch = self._batch_for_session(session.id)
        if existing_batch is not None:
            return existing_batch
        persisted_record_count = 0
        months: list[str] = []
        batch_id = self._next_batch_id()
        for preview_file in session.files:
            if preview_file.month not in months:
                months.append(preview_file.month)
            for record in preview_file.rows:
                record.batch_id = batch_id
                self._records[record.unique_key] = record
                persisted_record_count += 1
        batch = TaxCertifiedImportBatch(
            id=batch_id,
            session_id=session.id,
            imported_by=session.imported_by,
            file_count=session.file_count,
            months=months,
            persisted_record_count=persisted_record_count,
        )
        self._batches[batch.id] = batch
        session.status = "confirmed"
        self._persist()
        return batch

    def _batch_for_session(self, session_id: str) -> TaxCertifiedImportBatch | None:
        for batch in self._batches.values():
            if batch.session_id == session_id:
                return batch
        return None

    def list_records_for_month(self, month: str) -> list[TaxCertifiedInvoiceRecord]:
        return sorted(
            (record for record in self._records.values() if record.month == month),
            key=lambda item: (item.source_file_name, item.source_row_number, item.invoice_no or "", item.id),
        )

    def snapshot(self) -> dict[str, Any]:
        return {
            "session_counter": self._session_counter,
            "file_counter": self._file_counter,
            "batch_counter": self._batch_counter,
            "sessions": self._sessions,
            "batches": self._batches,
            "records": self._records,
        }

    def source_version(self) -> str:
        return (
            f"sessions:{len(self._sessions)}:{self._session_counter}|"
            f"batches:{len(self._batches)}:{self._batch_counter}|"
            f"records:{len(self._records)}"
        )

    def _hydrate(self, snapshot: dict[str, Any]) -> None:
        self._session_counter = int(snapshot.get("session_counter", 0))
        self._file_counter = int(snapshot.get("file_counter", 0))
        self._batch_counter = int(snapshot.get("batch_counter", 0))
        self._records = {
            str(key): self._record_from_snapshot(value)
            for key, value in dict(snapshot.get("records") or {}).items()
        }
        self._sessions = {
            str(key): self._session_from_snapshot(value)
            for key, value in dict(snapshot.get("sessions") or {}).items()
        }
        self._batches = {
            str(key): self._batch_from_snapshot(value)
            for key, value in dict(snapshot.get("batches") or {}).items()
        }

    @classmethod
    def _session_from_snapshot(cls, value: Any) -> TaxCertifiedImportSession:
        if isinstance(value, TaxCertifiedImportSession):
            return value
        raw = dict(value) if isinstance(value, dict) else {}
        raw["files"] = [cls._file_from_snapshot(item) for item in raw.get("files") or []]
        raw["created_at"] = cls._datetime_from_snapshot(raw.get("created_at"))
        return TaxCertifiedImportSession(
            id=str(raw.get("id") or raw.get("session_id") or ""),
            imported_by=str(raw.get("imported_by") or ""),
            file_count=int(raw.get("file_count") or len(raw["files"])),
            status=str(raw.get("status") or "preview_ready"),
            files=raw["files"],
            created_at=raw["created_at"],
        )

    @classmethod
    def _file_from_snapshot(cls, value: Any) -> TaxCertifiedImportPreviewFile:
        if isinstance(value, TaxCertifiedImportPreviewFile):
            return value
        raw = dict(value) if isinstance(value, dict) else {}
        rows = [cls._record_from_snapshot(item) for item in raw.get("rows") or []]
        row_results = [cls._row_result_from_snapshot(item) for item in raw.get("row_results") or []]
        if not row_results:
            row_results = [cls._row_result_from_record(row) for row in rows]
        return TaxCertifiedImportPreviewFile(
            id=str(raw.get("id") or raw.get("file_id") or raw.get("file_name") or ""),
            file_name=str(raw.get("file_name") or ""),
            month=str(raw.get("month") or ""),
            recognized_count=int(raw.get("recognized_count") or len(rows)),
            invalid_count=int(raw.get("invalid_count") or 0),
            rows=rows,
            row_results=row_results,
        )

    @classmethod
    def _batch_from_snapshot(cls, value: Any) -> TaxCertifiedImportBatch:
        if isinstance(value, TaxCertifiedImportBatch):
            return value
        raw = dict(value) if isinstance(value, dict) else {}
        return TaxCertifiedImportBatch(
            id=str(raw.get("id") or raw.get("batch_id") or ""),
            session_id=str(raw.get("session_id") or ""),
            imported_by=str(raw.get("imported_by") or ""),
            file_count=int(raw.get("file_count") or 0),
            months=[str(item) for item in raw.get("months") or []],
            persisted_record_count=int(raw.get("persisted_record_count") or raw.get("record_count") or 0),
            created_at=cls._datetime_from_snapshot(raw.get("created_at")),
        )

    @classmethod
    def _record_from_snapshot(cls, value: Any) -> TaxCertifiedInvoiceRecord:
        if isinstance(value, TaxCertifiedInvoiceRecord):
            return value
        raw = dict(value) if isinstance(value, dict) else {}
        return TaxCertifiedInvoiceRecord(
            id=str(raw.get("id") or raw.get("record_id") or raw.get("unique_key") or ""),
            unique_key=str(raw.get("unique_key") or ""),
            month=str(raw.get("month") or ""),
            source_file_name=str(raw.get("source_file_name") or ""),
            source_row_number=int(raw.get("source_row_number") or 0),
            taxpayer_tax_no=cls._string_or_none(raw.get("taxpayer_tax_no")),
            taxpayer_name=cls._string_or_none(raw.get("taxpayer_name")),
            digital_invoice_no=cls._string_or_none(raw.get("digital_invoice_no")),
            invoice_code=cls._string_or_none(raw.get("invoice_code")),
            invoice_no=cls._string_or_none(raw.get("invoice_no")),
            issue_date=cls._string_or_none(raw.get("issue_date")),
            seller_tax_no=cls._string_or_none(raw.get("seller_tax_no")),
            seller_name=cls._string_or_none(raw.get("seller_name")),
            amount=cls._string_or_none(raw.get("amount")),
            tax_amount=cls._string_or_none(raw.get("tax_amount")),
            deductible_tax_amount=cls._string_or_none(raw.get("deductible_tax_amount")),
            selection_status=cls._string_or_none(raw.get("selection_status")),
            invoice_status=cls._string_or_none(raw.get("invoice_status")),
            selection_time=cls._string_or_none(raw.get("selection_time")),
            invoice_source=cls._string_or_none(raw.get("invoice_source")),
            invoice_kind=cls._string_or_none(raw.get("invoice_kind")),
            risk_level=cls._string_or_none(raw.get("risk_level")),
            batch_id=cls._string_or_none(raw.get("batch_id")),
            imported_at=cls._datetime_from_snapshot(raw.get("imported_at")),
        )

    @classmethod
    def _row_result_from_snapshot(cls, value: Any) -> TaxCertifiedImportPreviewRowResult:
        if isinstance(value, TaxCertifiedImportPreviewRowResult):
            return value
        raw = dict(value) if isinstance(value, dict) else {}
        return TaxCertifiedImportPreviewRowResult(
            id=str(raw.get("id") or raw.get("unique_key") or raw.get("row_id") or ""),
            month=str(raw.get("month") or ""),
            source_file_name=str(raw.get("source_file_name") or ""),
            source_row_number=int(raw.get("source_row_number") or 0),
            row_status=str(raw.get("row_status") or "recognized"),
            error_message=cls._string_or_none(raw.get("error_message")),
            unique_key=cls._string_or_none(raw.get("unique_key")),
            taxpayer_tax_no=cls._string_or_none(raw.get("taxpayer_tax_no")),
            taxpayer_name=cls._string_or_none(raw.get("taxpayer_name")),
            digital_invoice_no=cls._string_or_none(raw.get("digital_invoice_no")),
            invoice_code=cls._string_or_none(raw.get("invoice_code")),
            invoice_no=cls._string_or_none(raw.get("invoice_no")),
            issue_date=cls._string_or_none(raw.get("issue_date")),
            seller_tax_no=cls._string_or_none(raw.get("seller_tax_no")),
            seller_name=cls._string_or_none(raw.get("seller_name")),
            amount=cls._string_or_none(raw.get("amount")),
            tax_amount=cls._string_or_none(raw.get("tax_amount")),
            deductible_tax_amount=cls._string_or_none(raw.get("deductible_tax_amount")),
            selection_status=cls._string_or_none(raw.get("selection_status")),
            invoice_status=cls._string_or_none(raw.get("invoice_status")),
            selection_time=cls._string_or_none(raw.get("selection_time")),
            invoice_source=cls._string_or_none(raw.get("invoice_source")),
            invoice_kind=cls._string_or_none(raw.get("invoice_kind")),
            risk_level=cls._string_or_none(raw.get("risk_level")),
            match_status=str(raw.get("match_status") or "unknown"),
            matched_plan_id=cls._string_or_none(raw.get("matched_plan_id")),
            dedupe_status=str(raw.get("dedupe_status") or "not_applicable"),
        )

    @staticmethod
    def _row_result_from_record(record: TaxCertifiedInvoiceRecord) -> TaxCertifiedImportPreviewRowResult:
        return TaxCertifiedImportPreviewRowResult(
            id=record.id,
            unique_key=record.unique_key,
            month=record.month,
            source_file_name=record.source_file_name,
            source_row_number=record.source_row_number,
            row_status="recognized",
            taxpayer_tax_no=record.taxpayer_tax_no,
            taxpayer_name=record.taxpayer_name,
            digital_invoice_no=record.digital_invoice_no,
            invoice_code=record.invoice_code,
            invoice_no=record.invoice_no,
            issue_date=record.issue_date,
            seller_tax_no=record.seller_tax_no,
            seller_name=record.seller_name,
            amount=record.amount,
            tax_amount=record.tax_amount,
            deductible_tax_amount=record.deductible_tax_amount,
            selection_status=record.selection_status,
            invoice_status=record.invoice_status,
            selection_time=record.selection_time,
            invoice_source=record.invoice_source,
            invoice_kind=record.invoice_kind,
            risk_level=record.risk_level,
            dedupe_status="new",
        )

    @staticmethod
    def _datetime_from_snapshot(value: Any) -> datetime:
        if isinstance(value, datetime):
            return value
        if value:
            try:
                return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
            except ValueError:
                pass
        return datetime.now(UTC)

    @staticmethod
    def _string_or_none(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _persist(self) -> None:
        if self._state_store is None:
            return
        self._state_store.save_tax_certified_imports(self.snapshot())

    def _preview_single_file(self, upload: UploadedCertifiedImportFile) -> TaxCertifiedImportPreviewFile:
        taxpayer_tax_no, taxpayer_name, month, rows = self._read_template(upload)
        records: list[TaxCertifiedInvoiceRecord] = []
        row_results: list[TaxCertifiedImportPreviewRowResult] = []
        seen_unique_keys: set[str] = set()
        invalid_count = 0
        for row_number, mapped in rows:
            if not self._has_meaningful_row(mapped):
                continue
            record, error_message = self._build_record_or_error(
                upload=upload,
                month=month,
                taxpayer_tax_no=taxpayer_tax_no,
                taxpayer_name=taxpayer_name,
                row_number=row_number,
                mapped=mapped,
            )
            if record is None:
                invalid_count += 1
                row_results.append(
                    self._invalid_row_result(
                        upload=upload,
                        month=month,
                        row_number=row_number,
                        mapped=mapped,
                        error_message=error_message or "无法识别为可导入的已认证发票记录。",
                    )
                )
                continue
            row_result = self._row_result_from_record(record)
            if record.unique_key in self._records or record.unique_key in seen_unique_keys:
                row_result.dedupe_status = "duplicate"
            seen_unique_keys.add(record.unique_key)
            records.append(record)
            row_results.append(row_result)
        return TaxCertifiedImportPreviewFile(
            id=self._next_file_id(),
            file_name=upload.file_name,
            month=month,
            recognized_count=len(records),
            invalid_count=invalid_count,
            rows=records,
            row_results=row_results,
        )

    def _read_template(
        self,
        upload: UploadedCertifiedImportFile,
    ) -> tuple[str | None, str | None, str, list[tuple[int, dict[str, Any]]]]:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(BytesIO(upload.content), data_only=True, read_only=True)
        if "发票" not in workbook.sheetnames:
            raise ValueError("已认证发票模板缺少“发票”sheet。")
        sheet = workbook["发票"]
        rows = list(sheet.iter_rows(values_only=True))
        taxpayer_tax_no = _normalize_text(_cell_value(rows, 1, 1))
        taxpayer_name = _normalize_text(_cell_value(rows, 1, 8))
        month = self._resolve_month(upload.file_name, rows)
        header_row_index = self._find_header_row(rows)
        header = [_normalize_text(cell) or "" for cell in rows[header_row_index]]
        data_rows: list[tuple[int, dict[str, Any]]] = []
        for row_index in range(header_row_index + 1, len(rows)):
            row = rows[row_index]
            mapped = {
                header[column_index]: row[column_index] if column_index < len(row) else None
                for column_index in range(len(header))
                if header[column_index]
            }
            data_rows.append((row_index + 1, mapped))
        return taxpayer_tax_no, taxpayer_name, month, data_rows

    def _resolve_month(self, file_name: str, rows: list[tuple[Any, ...]]) -> str:
        raw_month = _normalize_text(_cell_value(rows, 1, 5))
        if raw_month and COMPACT_MONTH_RE.match(raw_month):
            return f"{raw_month[:4]}-{raw_month[4:6]}"
        file_match = MONTH_FROM_FILENAME_RE.search(file_name)
        if file_match:
            return f"{file_match.group(1)}-{int(file_match.group(2)):02d}"
        raise ValueError("无法识别已认证发票所属月份。")

    @staticmethod
    def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
        required_headers = {"序号", "勾选状态", "数电发票号码", "发票号码", "开票日期", "销售方纳税人识别号", "税额", "发票状态"}
        for index, row in enumerate(rows):
            normalized = {_normalize_text(cell) for cell in row if _normalize_text(cell)}
            if required_headers.issubset(normalized):
                return index
        raise ValueError("无法识别已认证发票模板表头。")

    @staticmethod
    def _has_meaningful_row(mapped: dict[str, Any]) -> bool:
        return any(_normalize_text(value) for value in mapped.values())

    def _build_record(
        self,
        *,
        upload: UploadedCertifiedImportFile,
        month: str,
        taxpayer_tax_no: str | None,
        taxpayer_name: str | None,
        row_number: int,
        mapped: dict[str, Any],
    ) -> TaxCertifiedInvoiceRecord | None:
        record, _error_message = self._build_record_or_error(
            upload=upload,
            month=month,
            taxpayer_tax_no=taxpayer_tax_no,
            taxpayer_name=taxpayer_name,
            row_number=row_number,
            mapped=mapped,
        )
        return record

    def _build_record_or_error(
        self,
        *,
        upload: UploadedCertifiedImportFile,
        month: str,
        taxpayer_tax_no: str | None,
        taxpayer_name: str | None,
        row_number: int,
        mapped: dict[str, Any],
    ) -> tuple[TaxCertifiedInvoiceRecord | None, str | None]:
        selection_status = _normalize_text(mapped.get("勾选状态"))
        invoice_status = _normalize_text(mapped.get("发票状态"))
        deductible_tax_amount = _to_money_string(mapped.get("有效抵扣税额"))
        if selection_status not in CERTIFIED_SELECTION_STATUSES:
            return None, f"勾选状态为“{selection_status or '空'}”，不是可导入的已勾选/已认证记录。"
        if invoice_status not in ALLOWED_INVOICE_STATUSES:
            return None, f"发票状态为“{invoice_status or '空'}”，不是正常发票。"
        if deductible_tax_amount is not None and Decimal(deductible_tax_amount) <= Decimal("0.00"):
            return None, "有效抵扣税额必须大于 0。"

        digital_invoice_no = _normalize_text(mapped.get("数电发票号码"))
        invoice_code = _normalize_text(mapped.get("发票代码"))
        invoice_no = _normalize_text(mapped.get("发票号码"))
        seller_tax_no = _normalize_text(mapped.get("销售方纳税人识别号"))
        seller_name = _normalize_text(mapped.get("销售方纳税人名称"))
        issue_date = _to_date_string(mapped.get("开票日期"))
        tax_amount = _to_money_string(mapped.get("税额"))
        unique_key = _build_unique_key(
            digital_invoice_no=digital_invoice_no,
            invoice_code=invoice_code,
            invoice_no=invoice_no,
            seller_tax_no=seller_tax_no,
            seller_name=seller_name,
            issue_date=issue_date,
            tax_amount=tax_amount,
        )

        return TaxCertifiedInvoiceRecord(
            id=unique_key,
            unique_key=unique_key,
            month=month,
            source_file_name=upload.file_name,
            source_row_number=row_number,
            taxpayer_tax_no=taxpayer_tax_no,
            taxpayer_name=taxpayer_name,
            digital_invoice_no=digital_invoice_no,
            invoice_code=invoice_code,
            invoice_no=invoice_no,
            issue_date=issue_date,
            seller_tax_no=seller_tax_no,
            seller_name=seller_name,
            amount=_to_money_string(mapped.get("金额")),
            tax_amount=tax_amount,
            deductible_tax_amount=deductible_tax_amount,
            selection_status=selection_status,
            invoice_status=invoice_status,
            selection_time=_to_datetime_string(mapped.get("勾选时间")),
            invoice_source=_normalize_text(mapped.get("发票来源")),
            invoice_kind=_normalize_text(mapped.get("发票票种")),
            risk_level=_normalize_text(mapped.get("发票风险等级")),
        ), None

    def _invalid_row_result(
        self,
        *,
        upload: UploadedCertifiedImportFile,
        month: str,
        row_number: int,
        mapped: dict[str, Any],
        error_message: str,
    ) -> TaxCertifiedImportPreviewRowResult:
        return TaxCertifiedImportPreviewRowResult(
            id=f"{upload.file_name}:{row_number}",
            month=month,
            source_file_name=upload.file_name,
            source_row_number=row_number,
            row_status="invalid",
            error_message=error_message,
            digital_invoice_no=_normalize_text(mapped.get("数电发票号码")),
            invoice_code=_normalize_text(mapped.get("发票代码")),
            invoice_no=_normalize_text(mapped.get("发票号码")),
            issue_date=_to_date_string(mapped.get("开票日期")),
            seller_tax_no=_normalize_text(mapped.get("销售方纳税人识别号")),
            seller_name=_normalize_text(mapped.get("销售方纳税人名称")),
            amount=_to_money_string(mapped.get("金额")),
            tax_amount=_to_money_string(mapped.get("税额")),
            deductible_tax_amount=_to_money_string(mapped.get("有效抵扣税额")),
            selection_status=_normalize_text(mapped.get("勾选状态")),
            invoice_status=_normalize_text(mapped.get("发票状态")),
            selection_time=_to_datetime_string(mapped.get("勾选时间")),
            invoice_source=_normalize_text(mapped.get("发票来源")),
            invoice_kind=_normalize_text(mapped.get("发票票种")),
            risk_level=_normalize_text(mapped.get("发票风险等级")),
        )

    def _next_session_id(self) -> str:
        self._session_counter += 1
        return f"tax-certified-session-{self._session_counter:04d}"

    def _next_file_id(self) -> str:
        self._file_counter += 1
        return f"tax-certified-file-{self._file_counter:04d}"

    def _next_batch_id(self) -> str:
        self._batch_counter += 1
        return f"tax-certified-batch-{self._batch_counter:04d}"


def _cell_value(rows: list[tuple[Any, ...]], row_index: int, column_index: int) -> Any:
    if row_index >= len(rows):
        return None
    row = rows[row_index]
    if column_index >= len(row):
        return None
    return row[column_index]


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _to_date_string(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    normalized = _normalize_text(value)
    if normalized is None:
        return None
    return normalized[:10] if len(normalized) >= 10 else normalized


def _to_datetime_string(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    normalized = _normalize_text(value)
    return normalized


def _to_money_string(value: Any) -> str | None:
    if value is None:
        return None
    normalized = _normalize_text(value)
    if normalized is None:
        return None
    normalized = normalized.replace(",", "")
    try:
        decimal_value = Decimal(normalized)
    except InvalidOperation:
        return None
    return f"{decimal_value:.2f}"


def _build_unique_key(
    *,
    digital_invoice_no: str | None,
    invoice_code: str | None,
    invoice_no: str | None,
    seller_tax_no: str | None,
    seller_name: str | None,
    issue_date: str | None,
    tax_amount: str | None,
) -> str:
    if digital_invoice_no:
        return f"digital:{digital_invoice_no}"
    if invoice_code and invoice_no:
        return f"invoice:{invoice_code}:{invoice_no}"
    seller_part = seller_tax_no or seller_name or "unknown-seller"
    issue_part = issue_date or "unknown-date"
    tax_part = tax_amount or "0.00"
    return f"fallback:{seller_part}:{issue_part}:{tax_part}"
