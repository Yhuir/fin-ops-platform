from __future__ import annotations

from collections import defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook

from fin_ops_platform.domain.enums import TransactionDirection
from fin_ops_platform.services.imports import ImportNormalizationService
from fin_ops_platform.services.live_workbench_service import format_decimal
from fin_ops_platform.services.project_detail_export_service import ProjectDetailExportService


ZERO = Decimal("0.00")
EXCLUDED_COST_EXPENSE_TYPES = {"借款", "还款"}
OA_INVOICE_OFFSET_AUTO_MATCH_CODE = "oa_invoice_offset_auto_match"
OA_INVOICE_OFFSET_TAG = "冲"


class CostStatisticsService:
    def __init__(
        self,
        import_service: ImportNormalizationService,
        *,
        grouped_workbench_loader: Callable[[str], dict[str, Any]],
        row_detail_loader: Callable[[str], dict[str, Any]],
        raw_workbench_loader: Callable[[str], dict[str, Any]] | None = None,
    ) -> None:
        self._import_service = import_service
        self._grouped_workbench_loader = grouped_workbench_loader
        self._row_detail_loader = row_detail_loader
        self._project_detail_export_service = ProjectDetailExportService(
            import_service,
            grouped_workbench_loader=grouped_workbench_loader,
            raw_workbench_loader=raw_workbench_loader,
        )

    def get_month_statistics(self, month: str) -> dict[str, Any]:
        entries = self._build_cost_entries(month)
        grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
        for entry in entries:
            key = (entry["project_name"], entry["expense_type"], entry["expense_content"])
            bucket = grouped.setdefault(
                key,
                {
                    "project_name": entry["project_name"],
                    "expense_type": entry["expense_type"],
                    "expense_content": entry["expense_content"],
                    "amount": ZERO,
                    "transaction_count": 0,
                    "sample_transaction_ids": [],
                },
            )
            bucket["amount"] += entry["amount_decimal"]
            bucket["transaction_count"] += 1
            if entry["transaction_id"] not in bucket["sample_transaction_ids"]:
                bucket["sample_transaction_ids"].append(entry["transaction_id"])

        rows = [
            {
                "project_name": bucket["project_name"],
                "expense_type": bucket["expense_type"],
                "expense_content": bucket["expense_content"],
                "amount": format_decimal(bucket["amount"]),
                "transaction_count": bucket["transaction_count"],
                "sample_transaction_ids": list(bucket["sample_transaction_ids"]),
            }
            for bucket in sorted(
                grouped.values(),
                key=lambda item: (item["project_name"], item["expense_type"], item["expense_content"]),
            )
        ]
        return {
            "month": month,
            "summary": {
                "row_count": len(rows),
                "transaction_count": len(entries),
                "total_amount": format_decimal(sum((entry["amount_decimal"] for entry in entries), start=ZERO)),
            },
            "rows": rows,
        }

    def get_explorer(self, month: str) -> dict[str, Any]:
        entries = self._build_cost_entries(month)
        sorted_entries = sorted(entries, key=lambda item: (item["trade_time"], item["transaction_id"]), reverse=True)

        project_groups: dict[str, dict[str, Any]] = {}
        expense_type_groups: dict[str, dict[str, Any]] = {}

        for entry in sorted_entries:
            project_bucket = project_groups.setdefault(
                entry["project_name"],
                {
                    "project_name": entry["project_name"],
                    "total_amount": ZERO,
                    "transaction_count": 0,
                    "expense_types": set(),
                },
            )
            project_bucket["total_amount"] += entry["amount_decimal"]
            project_bucket["transaction_count"] += 1
            project_bucket["expense_types"].add(entry["expense_type"])

            expense_bucket = expense_type_groups.setdefault(
                entry["expense_type"],
                {
                    "expense_type": entry["expense_type"],
                    "total_amount": ZERO,
                    "transaction_count": 0,
                    "projects": set(),
                },
            )
            expense_bucket["total_amount"] += entry["amount_decimal"]
            expense_bucket["transaction_count"] += 1
            expense_bucket["projects"].add(entry["project_name"])

        return {
            "month": month,
            "summary": self._summary_payload(sorted_entries),
            "time_rows": [self._serialize_cost_entry(entry) for entry in sorted_entries],
            "project_rows": [
                {
                    "project_name": bucket["project_name"],
                    "total_amount": format_decimal(bucket["total_amount"]),
                    "transaction_count": bucket["transaction_count"],
                    "expense_type_count": len(bucket["expense_types"]),
                }
                for bucket in sorted(
                    project_groups.values(),
                    key=lambda item: (-item["total_amount"], item["project_name"]),
                )
            ],
            "expense_type_rows": [
                {
                    "expense_type": bucket["expense_type"],
                    "total_amount": format_decimal(bucket["total_amount"]),
                    "transaction_count": bucket["transaction_count"],
                    "project_count": len(bucket["projects"]),
                }
                for bucket in sorted(
                    expense_type_groups.values(),
                    key=lambda item: (-item["total_amount"], item["expense_type"]),
                )
            ],
        }

    def get_project_statistics(self, month: str, project_name: str) -> dict[str, Any]:
        entries = [entry for entry in self._build_cost_entries(month) if entry["project_name"] == project_name]
        rows = [
            {
                "transaction_id": entry["transaction_id"],
                "trade_time": entry["trade_time"],
                "direction": entry["direction"],
                "expense_type": entry["expense_type"],
                "expense_content": entry["expense_content"],
                "amount": format_decimal(entry["amount_decimal"]),
                "counterparty_name": entry["counterparty_name"],
                "payment_account_label": entry["payment_account_label"],
            }
            for entry in sorted(entries, key=lambda item: (item["trade_time"], item["transaction_id"]))
        ]
        return {
            "month": month,
            "project_name": project_name,
            "summary": {
                "row_count": len(rows),
                "transaction_count": len(entries),
                "total_amount": format_decimal(sum((entry["amount_decimal"] for entry in entries), start=ZERO)),
            },
            "rows": rows,
        }

    def get_expense_type_statistics(self, month: str, expense_type: str) -> dict[str, Any]:
        entries = [entry for entry in self._build_cost_entries(month) if entry["expense_type"] == expense_type]
        rows = [
            {
                "transaction_id": entry["transaction_id"],
                "trade_time": entry["trade_time"],
                "direction": entry["direction"],
                "project_name": entry["project_name"],
                "expense_type": entry["expense_type"],
                "expense_content": entry["expense_content"],
                "amount": format_decimal(entry["amount_decimal"]),
                "counterparty_name": entry["counterparty_name"],
                "payment_account_label": entry["payment_account_label"],
            }
            for entry in sorted(entries, key=lambda item: (item["trade_time"], item["transaction_id"]), reverse=True)
        ]
        return {
            "month": month,
            "expense_type": expense_type,
            "summary": self._summary_payload(entries),
            "rows": rows,
        }

    def get_export_preview(
        self,
        *,
        month: str,
        view: str,
        project_name: str | None = None,
        project_names: list[str] | None = None,
        expense_types: list[str] | None = None,
        start_month: str | None = None,
        end_month: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        aggregate_by: str | None = None,
    ) -> dict[str, Any]:
        normalized_expense_types = self._normalize_expense_types(expense_types)
        normalized_project_names = self._normalize_project_names(project_names or ([project_name] if project_name else []))
        ordered_project_names = sorted(normalized_project_names)
        resolved_aggregate_by = self._normalize_project_aggregate_by(aggregate_by)
        if view == "time":
            entries = self._build_filtered_entries(
                month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
            )
            scope_label = self._build_scope_label(
                month=month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
            )
            return self._build_preview_payload(
                view=view,
                file_name=self._build_filename(month=scope_label, view=view),
                scope_label=scope_label,
                sheet_names=["按时间统计"],
                columns=["时间", "项目名称", "费用类型", "金额", "费用内容", "资金方向", "对方户名", "支付账户"],
                rows=[
                    [
                        entry["trade_time"],
                        entry["project_name"],
                        entry["expense_type"],
                        format_decimal(entry["amount_decimal"]),
                        entry["expense_content"],
                        entry["direction"],
                        entry["counterparty_name"],
                        entry["payment_account_label"],
                    ]
                    for entry in entries
                ],
                total_amount=format_decimal(sum((entry["amount_decimal"] for entry in entries), start=ZERO)),
            )
        if view == "project":
            if not normalized_project_names:
                raise ValueError("project_name is required for project export preview")
            if resolved_aggregate_by is not None or len(normalized_project_names) > 1:
                entries = self._build_filtered_entries(
                    "all",
                    start_month=start_month,
                    end_month=end_month,
                    start_date=start_date,
                    end_date=end_date,
                    project_names=normalized_project_names,
                    expense_types=normalized_expense_types,
                )
                rows = self._build_project_aggregate_rows(entries, aggregate_by=resolved_aggregate_by or "month")
                scope_label = self._build_scope_label(
                    month="all",
                    start_month=start_month,
                    end_month=end_month,
                    start_date=start_date,
                    end_date=end_date,
                )
                return self._build_preview_payload(
                    view=view,
                    file_name=self._build_filename(
                        month=scope_label,
                        view=view,
                        project_names=ordered_project_names,
                        aggregate_by=resolved_aggregate_by or "month",
                    ),
                    scope_label=scope_label,
                    sheet_names=["按项目统计"],
                    columns=["统计周期", "项目名称", "费用类型", "金额", "费用内容", "支出笔数"],
                    rows=[
                        [
                          row["period_label"],
                          row["project_name"],
                          row["expense_type"],
                          row["amount"],
                          row["expense_content"],
                          str(row["transaction_count"]),
                        ]
                        for row in rows
                    ],
                    total_amount=format_decimal(sum((row["amount_decimal"] for row in rows), start=ZERO)),
                )
            payload = self._project_detail_export_service.build_export_payload_with_options(
                month=month,
                project_name=ordered_project_names[0],
                expense_types=sorted(normalized_expense_types),
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
                include_oa_details=True,
                include_invoice_details=True,
                include_exception_rows=True,
                include_ignored_rows=True,
                include_expense_content_summary=True,
                sort_by="time",
            )
            return self._build_preview_payload(
                view=view,
                file_name=self._build_filename(
                    month=payload["summary"]["scope_label"],
                    view=view,
                    project_name=ordered_project_names[0],
                ),
                scope_label=payload["summary"]["scope_label"],
                sheet_names=self._project_sheet_names(payload["options"]),
                columns=["时间", "资金方向", "费用类型", "金额", "费用内容", "对方户名", "支付账户"],
                rows=[
                    [
                        row["trade_time"],
                        row["direction"],
                        row["expense_type"],
                        row["amount"],
                        row["expense_content"],
                        row["counterparty_name"],
                        row["payment_account_label"],
                    ]
                    for row in payload["transaction_rows"]
                ],
                total_amount=payload["summary"]["total_amount"],
            )
        if view == "expense_type":
            if not normalized_expense_types:
                raise ValueError("expense_type is required for expense_type export preview")
            entries = self._build_filtered_entries(
                month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
                expense_types=normalized_expense_types,
            )
            scope_label = self._build_scope_label(
                month=month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
            )
            expense_label = self._build_expense_type_label(normalized_expense_types)
            return self._build_preview_payload(
                view=view,
                file_name=self._build_filename(month=scope_label, view=view, expense_type=expense_label),
                scope_label=scope_label,
                sheet_names=["按费用类型统计"],
                columns=["时间", "项目名称", "资金方向", "金额", "费用内容", "对方户名", "支付账户"],
                rows=[
                    [
                        entry["trade_time"],
                        entry["project_name"],
                        entry["direction"],
                        format_decimal(entry["amount_decimal"]),
                        entry["expense_content"],
                        entry["counterparty_name"],
                        entry["payment_account_label"],
                    ]
                    for entry in entries
                ],
                total_amount=format_decimal(sum((entry["amount_decimal"] for entry in entries), start=ZERO)),
            )
        raise ValueError("view must be time, project, or expense_type.")

    def get_transaction_detail(self, transaction_id: str) -> dict[str, Any]:
        transaction = self._import_service.get_transaction(transaction_id)
        month = (transaction.txn_date or "")[:7]
        if not month:
            raise KeyError(transaction_id)
        entry = next(
            (candidate for candidate in self._build_cost_entries(month) if candidate["transaction_id"] == transaction_id),
            None,
        )
        if entry is None:
            raise KeyError(transaction_id)
        raw_detail = self._row_detail_loader(transaction_id)
        row_detail = raw_detail.get("row", raw_detail) if isinstance(raw_detail, dict) else {}
        return {
            "month": month,
            "transaction": {
                "id": transaction_id,
                "project_name": entry["project_name"],
                "expense_type": entry["expense_type"],
                "expense_content": entry["expense_content"],
                "trade_time": entry["trade_time"],
                "direction": entry["direction"],
                "amount": format_decimal(entry["amount_decimal"]),
                "counterparty_name": entry["counterparty_name"],
                "payment_account_label": entry["payment_account_label"],
                "remark": entry["remark"],
                "oa_applicant": entry["oa_applicant"],
                "summary_fields": dict(row_detail.get("summary_fields", {})),
                "detail_fields": dict(row_detail.get("detail_fields", {})),
            },
        }

    def export_view(
        self,
        *,
        month: str,
        view: str,
        project_name: str | None = None,
        project_names: list[str] | None = None,
        expense_types: list[str] | None = None,
        transaction_id: str | None = None,
        start_month: str | None = None,
        end_month: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        aggregate_by: str | None = None,
        include_oa_details: bool = True,
        include_invoice_details: bool = True,
        include_exception_rows: bool = True,
        include_ignored_rows: bool = True,
        include_expense_content_summary: bool = True,
        sort_by: str = "time",
    ) -> tuple[str, bytes]:
        normalized_expense_types = self._normalize_expense_types(expense_types)
        normalized_project_names = self._normalize_project_names(project_names or ([project_name] if project_name else []))
        ordered_project_names = sorted(normalized_project_names)
        resolved_aggregate_by = self._normalize_project_aggregate_by(aggregate_by)
        if view == "time":
            entries = self._build_filtered_entries(
                month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
            )
            scope_label = self._build_scope_label(
                month=month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
            )
            payload = {
                "time_rows": [self._serialize_cost_entry(entry) for entry in entries],
            }
            workbook = self._build_time_workbook(payload)
            filename = self._build_filename(month=scope_label, view=view)
            return filename, self._serialize_workbook(workbook)
        if view == "month":
            payload = self.get_month_statistics(month)
            workbook = self._build_month_workbook(payload)
            filename = self._build_filename(month=month, view=view)
            return filename, self._serialize_workbook(workbook)
        if view == "project":
            if not normalized_project_names:
                raise ValueError("project_name is required for project export")
            if resolved_aggregate_by is not None or len(normalized_project_names) > 1:
                entries = self._build_filtered_entries(
                    "all",
                    start_month=start_month,
                    end_month=end_month,
                    start_date=start_date,
                    end_date=end_date,
                    project_names=normalized_project_names,
                    expense_types=normalized_expense_types,
                )
                rows = self._build_project_aggregate_rows(entries, aggregate_by=resolved_aggregate_by or "month")
                workbook = self._build_project_aggregate_workbook(rows)
                scope_label = self._build_scope_label(
                    month="all",
                    start_month=start_month,
                    end_month=end_month,
                    start_date=start_date,
                    end_date=end_date,
                )
                filename = self._build_filename(
                    month=scope_label,
                    view=view,
                    project_names=ordered_project_names,
                    aggregate_by=resolved_aggregate_by or "month",
                )
                return filename, self._serialize_workbook(workbook)
            payload = self._project_detail_export_service.build_export_payload_with_options(
                month=month,
                project_name=ordered_project_names[0],
                expense_types=sorted(normalized_expense_types),
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
                include_oa_details=include_oa_details,
                include_invoice_details=include_invoice_details,
                include_exception_rows=include_exception_rows,
                include_ignored_rows=include_ignored_rows,
                include_expense_content_summary=include_expense_content_summary,
                sort_by=sort_by,
            )
            workbook = self._project_detail_export_service.build_workbook(payload)
            filename = self._build_filename(
                month=payload["summary"]["scope_label"],
                view=view,
                project_name=ordered_project_names[0],
            )
            return filename, self._serialize_workbook(workbook)
        if view == "expense_type":
            if not normalized_expense_types:
                raise ValueError("expense_type is required for expense_type export")
            entries = self._build_filtered_entries(
                month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
                expense_types=normalized_expense_types,
            )
            scope_label = self._build_scope_label(
                month=month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
            )
            expense_label = self._build_expense_type_label(normalized_expense_types)
            payload = {
                "rows": [
                    {
                        "trade_time": entry["trade_time"],
                        "project_name": entry["project_name"],
                        "amount": format_decimal(entry["amount_decimal"]),
                        "expense_content": entry["expense_content"],
                        "direction": entry["direction"],
                        "counterparty_name": entry["counterparty_name"],
                        "payment_account_label": entry["payment_account_label"],
                    }
                    for entry in sorted(entries, key=lambda item: (item["trade_time"], item["transaction_id"]), reverse=True)
                ]
            }
            workbook = self._build_expense_type_workbook(payload)
            filename = self._build_filename(month=scope_label, view=view, expense_type=expense_label)
            return filename, self._serialize_workbook(workbook)
        if view == "transaction":
            if not transaction_id:
                raise ValueError("transaction_id is required for transaction export")
            payload = self.get_transaction_detail(transaction_id)
            workbook = self._build_transaction_workbook(payload)
            filename = self._build_filename(
                month=payload["month"],
                view=view,
                project_name=payload["transaction"]["project_name"],
                transaction_id=transaction_id,
            )
            return filename, self._serialize_workbook(workbook)
        raise ValueError(f"unsupported export view: {view}")

    def _build_cost_entries(self, month: str) -> list[dict[str, Any]]:
        entries: list[dict[str, Any]] = []
        for scoped_month in self._resolve_target_months(month):
            payload = self._grouped_workbench_loader(scoped_month)
            groups = list(((payload.get("paired") or {}).get("groups") or []))
            for group in groups:
                oa_rows = list(group.get("oa_rows") or [])
                bank_rows = list(group.get("bank_rows") or [])
                if not oa_rows or not bank_rows:
                    continue
                context = self._resolve_group_cost_context(oa_rows)
                if context is None:
                    continue
                for bank_row in bank_rows:
                    amount = self._extract_outflow_amount(bank_row)
                    if amount is None:
                        continue
                    entries.append(
                        {
                            "group_id": str(group.get("group_id", "")),
                            "transaction_id": str(bank_row["id"]),
                            "trade_time": str(bank_row.get("trade_time") or bank_row.get("pay_receive_time") or ""),
                            "counterparty_name": str(bank_row.get("counterparty_name") or ""),
                            "payment_account_label": str(bank_row.get("payment_account_label") or ""),
                            "direction": str(bank_row.get("direction") or "支出"),
                            "remark": str(bank_row.get("remark") or ""),
                            "project_name": context["project_name"],
                            "expense_type": context["expense_type"],
                            "expense_content": context["expense_content"],
                            "oa_applicant": context["oa_applicant"],
                            "amount_decimal": amount,
                        }
                    )
        return entries

    def _build_filtered_entries(
        self,
        month: str,
        *,
        start_month: str | None = None,
        end_month: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        project_name: str | None = None,
        project_names: set[str] | list[str] | None = None,
        expense_types: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        normalized_project_name = self._clean_text(project_name)
        normalized_project_names = self._normalize_project_names(project_names)
        normalized_expense_types = self._normalize_expense_types(expense_types)
        entries = self._build_cost_entries_for_months(
            self._resolve_target_months(
                month,
                start_month=start_month,
                end_month=end_month,
                start_date=start_date,
                end_date=end_date,
            )
        )
        entries = self._filter_entries_by_date_range(entries, start_date=start_date, end_date=end_date)
        if normalized_project_name:
            entries = [entry for entry in entries if entry["project_name"] == normalized_project_name]
        if normalized_project_names:
            entries = [entry for entry in entries if entry["project_name"] in normalized_project_names]
        if normalized_expense_types:
            entries = [entry for entry in entries if entry["expense_type"] in normalized_expense_types]
        return sorted(entries, key=lambda item: (item["trade_time"], item["transaction_id"]), reverse=True)

    def _build_cost_entries_for_months(self, months: list[str]) -> list[dict[str, Any]]:
        entries: list[dict[str, Any]] = []
        for scoped_month in months:
            entries.extend(self._build_cost_entries(scoped_month))
        return entries

    def _resolve_target_months(
        self,
        month: str,
        *,
        start_month: str | None = None,
        end_month: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
    ) -> list[str]:
        normalized = (month or "").strip()
        if not start_month and start_date:
            start_month = start_date[:7]
        if not end_month and end_date:
            end_month = end_date[:7]
        if start_month and end_month and start_month > end_month:
            start_month, end_month = end_month, start_month
        if normalized and normalized.lower() != "all":
            months = [normalized]
        else:
            months = sorted(
                {
                    (transaction.txn_date or "")[:7]
                    for transaction in self._import_service.list_transactions()
                    if (transaction.txn_date or "")[:7]
                },
                reverse=True,
            )
        if start_month:
            months = [item for item in months if item >= start_month]
        if end_month:
            months = [item for item in months if item <= end_month]
        return months

    def _build_month_workbook(self, payload: dict[str, Any]) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "月份汇总"
        headers = ["项目名称", "费用类型", "金额", "费用内容", "支出笔数"]
        sheet.append(headers)
        for row in payload["rows"]:
            sheet.append(
                [
                    row["project_name"],
                    row["expense_type"],
                    row["amount"],
                    row["expense_content"],
                    row["transaction_count"],
                ]
            )
        self._apply_sheet_layout(sheet, [28, 24, 14, 36, 12])
        return workbook

    def _build_time_workbook(self, payload: dict[str, Any]) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "按时间统计"
        headers = ["时间", "项目名称", "费用类型", "金额", "费用内容", "资金方向", "对方户名", "支付账户"]
        sheet.append(headers)
        for row in payload["time_rows"]:
            sheet.append(
                [
                    row["trade_time"],
                    row["project_name"],
                    row["expense_type"],
                    row["amount"],
                    row["expense_content"],
                    row["direction"],
                    row["counterparty_name"],
                    row["payment_account_label"],
                ]
            )
        self._apply_sheet_layout(sheet, [22, 28, 24, 14, 36, 12, 24, 24])
        return workbook

    def _build_project_workbook(self, payload: dict[str, Any]) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "项目明细"
        headers = ["时间", "费用类型", "金额", "费用内容", "对方户名", "支付账户"]
        sheet.append(headers)
        for row in payload["rows"]:
            sheet.append(
                [
                    row["trade_time"],
                    row["expense_type"],
                    row["amount"],
                    row["expense_content"],
                    row["counterparty_name"],
                    row["payment_account_label"],
                ]
            )
        self._apply_sheet_layout(sheet, [22, 24, 14, 34, 28, 24])
        return workbook

    def _build_project_aggregate_workbook(self, rows: list[dict[str, Any]]) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "按项目统计"
        headers = ["统计周期", "项目名称", "费用类型", "金额", "费用内容", "支出笔数"]
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    row["period_label"],
                    row["project_name"],
                    row["expense_type"],
                    row["amount"],
                    row["expense_content"],
                    row["transaction_count"],
                ]
            )
        self._apply_sheet_layout(sheet, [16, 28, 24, 14, 36, 12])
        return workbook

    def _build_expense_type_workbook(self, payload: dict[str, Any]) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "按费用类型统计"
        headers = ["时间", "项目名称", "金额", "费用内容", "资金方向", "对方户名", "支付账户"]
        sheet.append(headers)
        for row in payload["rows"]:
            sheet.append(
                [
                    row["trade_time"],
                    row["project_name"],
                    row["amount"],
                    row["expense_content"],
                    row["direction"],
                    row["counterparty_name"],
                    row["payment_account_label"],
                ]
            )
        self._apply_sheet_layout(sheet, [22, 28, 14, 36, 12, 24, 24])
        return workbook

    def _build_transaction_workbook(self, payload: dict[str, Any]) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "流水详情"
        transaction = payload["transaction"]
        sheet.append(["字段", "值"])
        base_rows = [
            ("交易ID", transaction["id"]),
            ("月份", payload["month"]),
            ("项目名称", transaction["project_name"]),
            ("费用类型", transaction["expense_type"]),
            ("费用内容", transaction["expense_content"]),
            ("交易时间", transaction["trade_time"]),
            ("资金方向", transaction["direction"]),
            ("金额", transaction["amount"]),
            ("对方户名", transaction["counterparty_name"]),
            ("OA提交人", transaction["oa_applicant"]),
            ("支付账户", transaction["payment_account_label"]),
            ("备注", transaction["remark"]),
        ]
        for item in base_rows:
            sheet.append(list(item))
        sheet.append([])
        sheet.append(["摘要字段", "值"])
        for key, value in transaction["summary_fields"].items():
            sheet.append([key, value])
        sheet.append([])
        sheet.append(["详细字段", "值"])
        for key, value in transaction["detail_fields"].items():
            sheet.append([key, value])
        self._apply_sheet_layout(sheet, [22, 52])
        return workbook

    @staticmethod
    def _apply_sheet_layout(sheet: Any, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

    @staticmethod
    def _serialize_workbook(workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _sanitize_filename_part(value: str) -> str:
        sanitized = value.strip().replace("/", "-").replace("\\", "-").replace(":", "：")
        return sanitized[:80] if len(sanitized) > 80 else sanitized

    def _build_filename(
        self,
        *,
        month: str,
        view: str,
        project_name: str | None = None,
        project_names: list[str] | None = None,
        aggregate_by: str | None = None,
        expense_type: str | None = None,
        transaction_id: str | None = None,
    ) -> str:
        month_segment = "全部期间" if (month or "").strip().lower() == "all" else month
        if view == "time":
            return f"成本统计_{month_segment}_按时间统计.xlsx"
        if view == "month":
            return f"成本统计_{month_segment}_月份汇总.xlsx"
        if view == "project":
            if aggregate_by is not None:
                project_label = self._build_project_export_label(project_names or ([project_name] if project_name else []))
                return f"成本统计_{month_segment}_按项目统计_按{'月' if aggregate_by == 'month' else '年'}_{project_label}.xlsx"
            project_segment = self._sanitize_filename_part(project_name or "未命名项目")
            return f"成本统计_{month_segment}_项目明细_{project_segment}.xlsx"
        if view == "expense_type":
            expense_segment = self._sanitize_filename_part(expense_type or "未命名费用类型")
            return f"成本统计_{month_segment}_按费用类型统计_{expense_segment}.xlsx"
        project_segment = self._sanitize_filename_part(project_name or "未命名项目")
        transaction_segment = self._sanitize_filename_part(transaction_id or "unknown")
        return f"成本统计_{month_segment}_流水详情_{project_segment}_{transaction_segment}.xlsx"

    @staticmethod
    def _build_scope_label(
        *,
        month: str,
        start_month: str | None = None,
        end_month: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
    ) -> str:
        if start_date and end_date:
            return f"{start_date}至{end_date}"
        if start_month and end_month:
            return f"{start_month}至{end_month}"
        if str(month).strip().lower() == "all":
            return "全部期间"
        return month or "—"

    def _build_project_aggregate_rows(
        self,
        entries: list[dict[str, Any]],
        *,
        aggregate_by: str,
    ) -> list[dict[str, Any]]:
        buckets: dict[tuple[str, str, str, str], dict[str, Any]] = {}
        for entry in entries:
            period_label = (entry["trade_time"] or "")[:7] if aggregate_by == "month" else (entry["trade_time"] or "")[:4]
            key = (period_label, entry["project_name"], entry["expense_type"], entry["expense_content"])
            bucket = buckets.setdefault(
                key,
                {
                    "period_label": period_label or "—",
                    "project_name": entry["project_name"],
                    "expense_type": entry["expense_type"],
                    "expense_content": entry["expense_content"],
                    "amount_decimal": ZERO,
                    "transaction_count": 0,
                },
            )
            bucket["amount_decimal"] += entry["amount_decimal"]
            bucket["transaction_count"] += 1
        return [
            {
                **bucket,
                "amount": format_decimal(bucket["amount_decimal"]),
            }
            for bucket in sorted(
                buckets.values(),
                key=lambda item: (item["period_label"], item["project_name"], item["expense_type"], item["expense_content"]),
            )
        ]

    @staticmethod
    def _filter_entries_by_date_range(
        entries: list[dict[str, Any]],
        *,
        start_date: str | None = None,
        end_date: str | None = None,
    ) -> list[dict[str, Any]]:
        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date
        if not start_date and not end_date:
            return entries
        filtered: list[dict[str, Any]] = []
        for entry in entries:
            trade_date = str(entry.get("trade_time") or "")[:10]
            if not trade_date:
                continue
            if start_date and trade_date < start_date:
                continue
            if end_date and trade_date > end_date:
                continue
            filtered.append(entry)
        return filtered

    @staticmethod
    def _build_expense_type_label(expense_types: set[str]) -> str:
        ordered = sorted(expense_types)
        if not ordered:
            return "未命名费用类型"
        if len(ordered) == 1:
            return ordered[0]
        return f"{ordered[0]}等{len(ordered)}类"

    @staticmethod
    def _project_sheet_names(options: dict[str, Any]) -> list[str]:
        sheet_names = ["导出说明", "项目汇总", "按费用类型汇总"]
        if options.get("include_expense_content_summary", True):
            sheet_names.append("按费用内容汇总")
        sheet_names.append("流水明细")
        if options.get("include_oa_details", True):
            sheet_names.append("OA关联明细")
        if options.get("include_invoice_details", True):
            sheet_names.append("发票关联明细")
        if options.get("include_exception_rows", True) or options.get("include_ignored_rows", True):
            sheet_names.append("异常与未闭环")
        return sheet_names

    @staticmethod
    def _build_preview_payload(
        *,
        view: str,
        file_name: str,
        scope_label: str,
        sheet_names: list[str],
        columns: list[str],
        rows: list[list[str]],
        total_amount: str,
    ) -> dict[str, Any]:
        return {
            "view": view,
            "file_name": file_name,
            "scope_label": scope_label,
            "summary": {
                "row_count": len(rows),
                "transaction_count": len(rows),
                "total_amount": total_amount,
                "sheet_count": len(sheet_names),
            },
            "sheet_names": sheet_names,
            "columns": columns,
            "rows": rows[:8],
        }

    def _resolve_group_cost_context(self, oa_rows: list[dict[str, Any]]) -> dict[str, str] | None:
        contexts: set[tuple[str, str, str, str]] = set()
        for row in oa_rows:
            if self._is_cost_excluded_oa_row(row):
                continue
            project_name = self._clean_text(row.get("project_name"))
            expense_type = self._clean_text(row.get("expense_type"))
            expense_content = self._clean_text(row.get("expense_content")) or self._clean_text(row.get("reason"))
            applicant = self._clean_text(row.get("applicant"))
            detail_fields = row.get("detail_fields")
            if isinstance(detail_fields, dict):
                if not expense_type:
                    expense_type = self._clean_text(detail_fields.get("费用类型"))
                if not expense_content:
                    expense_content = self._clean_text(detail_fields.get("费用内容"))
                if not applicant:
                    applicant = self._clean_text(detail_fields.get("申请人"))
            if expense_type in EXCLUDED_COST_EXPENSE_TYPES:
                continue
            if not (project_name and expense_type and expense_content):
                continue
            contexts.add((project_name, expense_type, expense_content, applicant))
        if len(contexts) != 1:
            return None
        project_name, expense_type, expense_content, applicant = next(iter(contexts))
        return {
            "project_name": project_name,
            "expense_type": expense_type,
            "expense_content": expense_content,
            "oa_applicant": applicant or "—",
        }

    @staticmethod
    def _is_cost_excluded_oa_row(row: dict[str, Any]) -> bool:
        if bool(row.get("cost_excluded")):
            return True
        tags = {str(tag).strip() for tag in list(row.get("tags") or []) if str(tag).strip()}
        if OA_INVOICE_OFFSET_TAG in tags:
            return True
        relation = row.get("oa_bank_relation")
        if isinstance(relation, dict) and str(relation.get("code", "")) == OA_INVOICE_OFFSET_AUTO_MATCH_CODE:
            return True
        return False

    def _extract_outflow_amount(self, bank_row: dict[str, Any]) -> Decimal | None:
        transaction_id = str(bank_row.get("id", ""))
        try:
            transaction = self._import_service.get_transaction(transaction_id)
        except KeyError:
            transaction = None

        if transaction is not None:
            if transaction.txn_direction != TransactionDirection.OUTFLOW:
                return None
            return transaction.amount

        debit_amount = self._parse_decimal(bank_row.get("debit_amount"))
        credit_amount = self._parse_decimal(bank_row.get("credit_amount"))
        if credit_amount not in (None, ZERO):
            return None
        if debit_amount in (None, ZERO):
            return None
        return debit_amount

    @staticmethod
    def _parse_decimal(value: Any) -> Decimal | None:
        if value in (None, "", "—", "--"):
            return None
        text = str(value).replace(",", "").strip()
        if not text:
            return None
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _clean_text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text in {"-", "--", "—", "——"}:
            return ""
        return text

    @staticmethod
    def _normalize_expense_types(expense_types: set[str] | list[str] | None) -> set[str]:
        return {str(item).strip() for item in (expense_types or []) if str(item).strip()}

    @staticmethod
    def _normalize_project_names(project_names: set[str] | list[str] | None) -> set[str]:
        return {str(item).strip() for item in (project_names or []) if str(item).strip()}

    @staticmethod
    def _normalize_project_aggregate_by(aggregate_by: str | None) -> str | None:
        if aggregate_by in {"month", "year"}:
            return aggregate_by
        return None

    def _build_project_export_label(self, project_names: list[str]) -> str:
        ordered = [self._sanitize_filename_part(project_name) for project_name in project_names if project_name]
        if not ordered:
            return "未命名项目"
        if len(ordered) == 1:
            return ordered[0]
        return f"{ordered[0]}等{len(ordered)}个项目"

    def _summary_payload(self, entries: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            "row_count": len(entries),
            "transaction_count": len(entries),
            "total_amount": format_decimal(sum((entry["amount_decimal"] for entry in entries), start=ZERO)),
        }

    def _serialize_cost_entry(self, entry: dict[str, Any]) -> dict[str, Any]:
        return {
            "transaction_id": entry["transaction_id"],
            "trade_time": entry["trade_time"],
            "direction": entry["direction"],
            "project_name": entry["project_name"],
            "expense_type": entry["expense_type"],
            "expense_content": entry["expense_content"],
            "amount": format_decimal(entry["amount_decimal"]),
            "counterparty_name": entry["counterparty_name"],
            "payment_account_label": entry["payment_account_label"],
            "remark": entry["remark"],
            "oa_applicant": entry["oa_applicant"],
        }
